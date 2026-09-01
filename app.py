from pathlib import Path
import numpy as np
import pandas as pd
import plotly.express as px
import streamlit as st
from io import BytesIO

try:
    import duckdb
except ImportError:
    duckdb = None

def df_to_excel_bytes(df, sheet_name="Sheet1"):
    """Downloadable Excel with basic header styling."""
    bio = BytesIO()
    try:
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        with pd.ExcelWriter(bio, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name=sheet_name)
            ws = writer.sheets[sheet_name]
            header_fill = PatternFill("solid", fgColor="1E3A8A")
            header_font = Font(bold=True, color="FFFFFF")
            thin = Border(
                left=Side(style="thin", color="CBD5E1"),
                right=Side(style="thin", color="CBD5E1"),
                top=Side(style="thin", color="CBD5E1"),
                bottom=Side(style="thin", color="CBD5E1"),
            )
            green = PatternFill("solid", fgColor="DCFCE7")
            red = PatternFill("solid", fgColor="FEE2E2")
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
                for cell in row:
                    cell.border = thin
                    cell.alignment = Alignment(horizontal="center")
                    val = cell.value
                    if isinstance(val, (int, float)) and val != 0:
                        hdr = str(ws.cell(1, cell.column).value or "")
                        if "VAR" in hdr.upper() or "%" in hdr or "PCT" in hdr:
                            cell.fill = green if val > 0 else red
            for col in ws.columns:
                width = min(18, max(8, max(len(str(c.value or "")) for c in col) + 2))
                ws.column_dimensions[col[0].column_letter].width = width
        bio.seek(0)
        return bio.getvalue()
    except Exception:
        bio = BytesIO()
        df.to_excel(bio, index=False)
        bio.seek(0)
        return bio.getvalue()



def excel_with_title(df, sheet_name="Sheet1", report_title=""):
    """CSV-style dataframe to Excel with optional title row matching filter heading."""
    bio = BytesIO()
    try:
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        with pd.ExcelWriter(bio, engine="openpyxl") as writer:
            start = 1
            if report_title:
                # write empty then overlay title
                pd.DataFrame([[report_title]]).to_excel(writer, index=False, header=False, sheet_name=sheet_name[:31], startrow=0)
                start = 2
            df.to_excel(writer, index=False, sheet_name=sheet_name[:31], startrow=start)
            ws = writer.sheets[sheet_name[:31]]
            if report_title:
                ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(1, len(df.columns)))
                ws.cell(1, 1).font = Font(bold=True, size=12, color="1E3A8A")
            hdr_row = start + 1
            header_fill = PatternFill("solid", fgColor="1E3A8A")
            header_font = Font(bold=True, color="FFFFFF")
            green = PatternFill("solid", fgColor="DCFCE7")
            red = PatternFill("solid", fgColor="FEE2E2")
            for cell in ws[hdr_row]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
            for row in ws.iter_rows(min_row=hdr_row + 1, max_row=ws.max_row, max_col=ws.max_column):
                for cell in row:
                    cell.alignment = Alignment(horizontal="center")
                    val = cell.value
                    hdr = str(ws.cell(hdr_row, cell.column).value or "").upper()
                    if isinstance(val, (int, float)) and val != 0:
                        if "VAR" in hdr or "%" in hdr or "PCT" in hdr:
                            cell.fill = green if val > 0 else red
        bio.seek(0)
        return bio.getvalue()
    except Exception:
        return df_to_excel_bytes(df, sheet_name)

def act_vs_act_to_excel_bytes(merged_df, group_col, prefix="GROSS", pax_heading="TOTAL PASSENGERS", sheet_name="Report", report_title=""):
    """Excel matching on-screen ACT VS ACT / Product layout (headers, formats, colours)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = str(sheet_name)[:31]
    thin = Border(
        left=Side(style="thin", color="94A3B8"),
        right=Side(style="thin", color="94A3B8"),
        top=Side(style="thin", color="94A3B8"),
        bottom=Side(style="thin", color="94A3B8"),
    )
    fills = {
        "left": PatternFill("solid", fgColor="0369A1"),
        "km": PatternFill("solid", fgColor="C2410C"),
        "earn": PatternFill("solid", fgColor="047857"),
        "tot": PatternFill("solid", fgColor="6B21A8"),
        "fpd": PatternFill("solid", fgColor="15803D"),
        "mhl": PatternFill("solid", fgColor="1D4ED8"),
        "sub": PatternFill("solid", fgColor="F1F5F9"),
        "green": PatternFill("solid", fgColor="DCFCE7"),
        "red": PatternFill("solid", fgColor="FEE2E2"),
        "total": PatternFill("solid", fgColor="E2EFDA"),
    }
    white = Font(bold=True, color="FFFFFF", size=10)
    dark = Font(bold=True, color="334155", size=9)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    has_sch = "sch_CY" in merged_df.columns

    groups = []
    if has_sch:
        groups.append(("NO OF SCHEDULES", 4, "km"))
    groups += [
        ("KILOMETERS (IN LKS.)", 4, "km"),
        (f"{prefix} EARNINGS (IN LKS.)", 4, "earn"),
        (f"{prefix} TOT EPK", 4, "tot"),
        (f"{prefix} FPD EPK", 4, "fpd"),
        (f"{prefix} MHL EPK", 4, "mhl"),
        (pax_heading, 4, "left"),
    ]

    ws.cell(1, 1, "S.No").fill = fills["left"]
    ws.cell(1, 1).font = white
    ws.cell(1, 1).alignment = center
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
    ws.cell(1, 2, group_col).fill = fills["left"]
    ws.cell(1, 2).font = white
    ws.cell(1, 2).alignment = center
    ws.merge_cells(start_row=1, start_column=2, end_row=2, end_column=2)

    col = 3
    for title, span, fk in groups:
        cell = ws.cell(1, col, title)
        cell.fill = fills[fk]
        cell.font = white
        cell.alignment = center
        if span > 1:
            ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + span - 1)
        for i in range(span):
            ws.cell(1, col + i).fill = fills[fk]
            ws.cell(1, col + i).font = white
            ws.cell(1, col + i).border = thin
        col += span

    col = 3
    for _ in range(len(groups)):
        for s in ["CY", "LY", "VAR", "% ▲/▼"]:
            cell = ws.cell(2, col, s)
            cell.fill = fills["sub"]
            cell.font = dark
            cell.alignment = center
            cell.border = thin
            col += 1

    def num(v, kind="num"):
        try:
            if v is None or (isinstance(v, float) and pd.isna(v)):
                return None
            fv = float(v)
            if kind == "pax":
                return int(round(fv)) if fv != 0 else None
            if kind == "sch":
                if fv == 0:
                    return None
                return int(round(fv)) if abs(fv - round(fv)) < 1e-6 else round(fv, 2)
            if kind in ("pct", "epk", "num"):
                return None if fv == 0 and kind != "epk" else round(fv, 2)
            return round(fv, 2)
        except Exception:
            return None

    def write_metric(r, c, val, kind, colour=False):
        cell = ws.cell(r, c, num(val, kind))
        cell.alignment = center
        cell.border = thin
        if colour and isinstance(cell.value, (int, float)) and cell.value is not None:
            if cell.value > 0:
                cell.fill = fills["green"]
                cell.font = Font(color="15803D", bold=True)
            elif cell.value < 0:
                cell.fill = fills["red"]
                cell.font = Font(color="B91C1C", bold=True)
        return c + 1

    sno = 0
    for _, row in merged_df.iterrows():
        r = ws.max_row + 1
        is_tot = str(row[group_col]) == "TOTAL"
        if not is_tot:
            sno += 1
        ws.cell(r, 1, "" if is_tot else sno).alignment = center
        ws.cell(r, 1).border = thin
        ws.cell(r, 2, row[group_col]).alignment = center
        ws.cell(r, 2).border = thin
        if is_tot:
            for cc in range(1, 3 + sum(g[1] for g in groups)):
                ws.cell(r, cc).fill = fills["total"]
                ws.cell(r, cc).font = Font(bold=True)
        c = 3
        if has_sch:
            c = write_metric(r, c, row.get("sch_CY"), "sch")
            c = write_metric(r, c, row.get("sch_LY"), "sch")
            c = write_metric(r, c, row.get("sch_VAR"), "sch", True)
            c = write_metric(r, c, row.get("sch_PCT"), "pct", True)
        for metric in ["kms", "earn_tot"]:
            c = write_metric(r, c, row.get(f"{metric}_CY"), "num")
            c = write_metric(r, c, row.get(f"{metric}_LY"), "num")
            c = write_metric(r, c, row.get(f"{metric}_VAR"), "num", True)
            c = write_metric(r, c, row.get(f"{metric}_PCT"), "pct", True)
        for epk in ["tot", "fpd", "mhl"]:
            c = write_metric(r, c, row.get(f"epk_{epk}_CY"), "epk")
            c = write_metric(r, c, row.get(f"epk_{epk}_LY"), "epk")
            c = write_metric(r, c, row.get(f"epk_{epk}_VAR"), "epk", True)
            c = write_metric(r, c, row.get(f"epk_{epk}_PCT"), "pct", True)
        c = write_metric(r, c, row.get("pax_CY"), "pax")
        c = write_metric(r, c, row.get("pax_LY"), "pax")
        c = write_metric(r, c, row.get("pax_VAR"), "pax", True)
        c = write_metric(r, c, row.get("pax_PCT"), "pct", True)

    for i in range(1, c):
        ws.column_dimensions[get_column_letter(i)].width = 11
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 12
    if report_title:
        ws.insert_rows(1)
        ws.cell(1, 1, report_title)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(8, c - 1))
        ws.cell(1, 1).font = Font(bold=True, size=12, color="1E3A8A")
        ws.row_dimensions[1].height = 20
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.getvalue()




def trends_dual_excel_bytes(df_net, df_gross, pax_heading="TOTAL PASSENGERS", report_title=""):
    """Export NET + GROSS on a SINGLE sheet with on-screen style headers/colours."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    thin = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1"),
    )
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    fills = {
        "left": PatternFill("solid", fgColor="0369A1"),
        "km": PatternFill("solid", fgColor="C2410C"),
        "earn": PatternFill("solid", fgColor="047857"),
        "tot": PatternFill("solid", fgColor="6B21A8"),
        "fpd": PatternFill("solid", fgColor="15803D"),
        "mhl": PatternFill("solid", fgColor="1D4ED8"),
        "or_tot": PatternFill("solid", fgColor="7C3AED"),
        "or_fpd": PatternFill("solid", fgColor="0D9488"),
        "or_mhl": PatternFill("solid", fgColor="2563EB"),
        "pax": PatternFill("solid", fgColor="0369A1"),
        "sub": PatternFill("solid", fgColor="F1F5F9"),
        "total": PatternFill("solid", fgColor="E2EFDA"),
        "pos": PatternFill("solid", fgColor="C6EFCE"),
        "neg": PatternFill("solid", fgColor="FFC7CE"),
        "section": PatternFill("solid", fgColor="1E3A8A"),
    }
    white = Font(bold=True, color="FFFFFF", size=10)
    sub_font = Font(bold=True, color="334155", size=9)
    data_font = Font(size=10)

    def write_block(ws, df, prefix, start_row, report_title=""):
        groups = [
            ("KILOMETERS (in lks.)", 4, "km"),
            (f"{prefix} EARNINGS (in lks.)", 4, "earn"),
            (f"{prefix} TOT EPK", 4, "tot"),
            (f"{prefix} FPD EPK", 4, "fpd"),
            (f"{prefix} MHL EPK", 4, "mhl"),
            ("TOT OR", 4, "or_tot"),
            ("FPD OR", 4, "or_fpd"),
            ("MHL OR", 4, "or_mhl"),
            (pax_heading, 4, "pax"),
        ]
        total_cols = 2 + sum(g[1] for g in groups)
        r = start_row
        # Section banner
        ws.cell(r, 1, f"{prefix} — {report_title}" if report_title else prefix)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=total_cols)
        ws.cell(r, 1).fill = fills["section"]
        ws.cell(r, 1).font = Font(bold=True, color="FFFFFF", size=12)
        ws.cell(r, 1).alignment = center
        r += 1
        r1, r2 = r, r + 1
        ws.cell(r1, 1, "S.No").fill = fills["left"]
        ws.cell(r1, 1).font = white
        ws.cell(r1, 1).alignment = center
        ws.merge_cells(start_row=r1, start_column=1, end_row=r2, end_column=1)
        ws.cell(r1, 2, "MONTH" if "Month" in (df.columns if df is not None else []) else "KEY").fill = fills["left"]
        # detect label col
        label_col = "Month" if df is not None and "Month" in df.columns else (
            "DEPOT" if df is not None and "DEPOT" in df.columns else (
                "PRODUCT" if df is not None and "PRODUCT" in df.columns else None
            )
        )
        ws.cell(r1, 2, label_col or "KEY").fill = fills["left"]
        ws.cell(r1, 2).font = white
        ws.cell(r1, 2).alignment = center
        ws.merge_cells(start_row=r1, start_column=2, end_row=r2, end_column=2)
        col = 3
        for title, span, key in groups:
            ws.cell(r1, col, title).fill = fills[key]
            ws.cell(r1, col).font = white
            ws.cell(r1, col).alignment = center
            if span > 1:
                ws.merge_cells(start_row=r1, start_column=col, end_row=r1, end_column=col + span - 1)
            for i, sub in enumerate(["CY", "LY", "VAR", "% ▲/▼"]):
                cell = ws.cell(r2, col + i, sub)
                cell.fill = fills["sub"]
                cell.font = sub_font
                cell.alignment = center
                cell.border = thin
            col += span
        for c in range(1, col):
            ws.cell(r1, c).border = thin
            ws.cell(r2, c).border = thin

        def put(rr, c, val, kind="num", colorize=False):
            cell = ws.cell(rr, c)
            cell.border = thin
            cell.alignment = center
            cell.font = data_font
            if val is None or (isinstance(val, float) and pd.isna(val)):
                cell.value = None
                return c + 1
            try:
                v = float(val)
            except Exception:
                cell.value = val
                return c + 1
            if kind == "pax":
                cell.value = int(round(v))
                cell.number_format = "#,##0"
            else:
                cell.value = v
                cell.number_format = "0.00"
            if colorize and not pd.isna(v) and v != 0:
                cell.fill = fills["pos"] if v > 0 else fills["neg"]
            return c + 1

        data_r = r2 + 1
        sno = 0
        if df is not None and len(df) > 0:
            for _, row in df.iterrows():
                lab = row.get(label_col, "") if label_col else ""
                is_tot = any(x in str(lab).upper() for x in ("UPTO", "TOTAL", "REGION"))
                if not is_tot:
                    sno += 1
                ws.cell(data_r, 1, "" if is_tot else sno).alignment = center
                ws.cell(data_r, 1).border = thin
                ws.cell(data_r, 2, lab).alignment = center
                ws.cell(data_r, 2).border = thin
                if is_tot:
                    for cc in range(1, total_cols + 1):
                        ws.cell(data_r, cc).fill = fills["total"]
                        ws.cell(data_r, cc).font = Font(bold=True, size=10)
                c = 3
                for metric in ["kms", "earn_tot"]:
                    c = put(data_r, c, row.get(f"{metric}_CY"), "num")
                    c = put(data_r, c, row.get(f"{metric}_LY"), "num")
                    c = put(data_r, c, row.get(f"{metric}_VAR"), "num", True)
                    c = put(data_r, c, row.get(f"{metric}_PCT"), "pct", True)
                for epk in ["tot", "fpd", "mhl"]:
                    c = put(data_r, c, row.get(f"epk_{epk}_CY"), "num")
                    c = put(data_r, c, row.get(f"epk_{epk}_LY"), "num")
                    c = put(data_r, c, row.get(f"epk_{epk}_VAR"), "num", True)
                    c = put(data_r, c, row.get(f"epk_{epk}_PCT"), "pct", True)
                for ot in ["tot", "fpd", "mhl"]:
                    c = put(data_r, c, row.get(f"or_{ot}_CY"), "num")
                    c = put(data_r, c, row.get(f"or_{ot}_LY"), "num")
                    c = put(data_r, c, row.get(f"or_{ot}_VAR"), "num", True)
                    c = put(data_r, c, row.get(f"or_{ot}_PCT"), "pct", True)
                c = put(data_r, c, row.get("pax_CY"), "pax")
                c = put(data_r, c, row.get("pax_LY"), "pax")
                c = put(data_r, c, row.get("pax_VAR"), "pax", True)
                c = put(data_r, c, row.get("pax_PCT"), "pct", True)
                data_r += 1
        return data_r + 1  # blank row gap

    wb = Workbook()
    ws = wb.active
    ws.title = "NET_GROSS"
    row = 1
    if df_net is not None and len(df_net) > 0:
        row = write_block(ws, df_net, "NET", row, report_title)
    if df_gross is not None and len(df_gross) > 0:
        row = write_block(ws, df_gross, "GROSS", row, report_title)
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 14
    for i in range(3, 40):
        ws.column_dimensions[get_column_letter(i)].width = 10
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.getvalue()


st.set_page_config(
    page_title="Historical Analysis Dashboard",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ========== CUSTOM CSS ==========
st.markdown(
    """
<style>
    .block-container { padding-top: 2rem !important; padding-bottom: 1rem !important; max-width: 100% !important; overflow: visible !important; }

    div[data-testid="stVerticalBlock"] {
        gap: 0.4rem !important;
    }

    /* Filter Panel */
    div[data-testid="stHorizontalBlock"]:has(div[data-testid="stSelectbox"]) {
        background: linear-gradient(135deg, #1e293b 0%, #0f172a 100%);
        padding: 10px 14px !important;
        border-radius: 8px;
        box-shadow: 0 2px 10px rgba(0, 0, 0, 0.15);
        border: 1px solid #334155;
        margin-top: 8px !important;
        margin-bottom: 12px !important;
        gap: 8px !important;
    }

    div[data-testid="stSelectbox"] label p {
        color: #FDFBF7 !important;
        font-family: 'Segoe UI', Roboto, sans-serif !important;
        font-weight: 700 !important;
        font-size: 12px !important;
        letter-spacing: 0.4px;
        text-transform: uppercase;
        margin-bottom: 4px !important;
        line-height: 1.2 !important;
        white-space: nowrap !important;
    }

    div[data-testid="stSelectbox"] div[data-baseweb="select"] > div {
        background-color: #ffffff !important;
        border: 1px solid #cbd5e1 !important;
        border-radius: 6px !important;
        color: #0f172a !important;
        font-weight: 700 !important;
        font-size: 12px !important;
        font-family: 'Segoe UI', Roboto, sans-serif !important;
        min-height: 32px !important;
        height: 32px !important;
        padding-top: 0px !important;
        padding-bottom: 0px !important;
        padding-left: 8px !important;
        line-height: 32px !important;
        box-shadow: 0 1px 2px rgba(0,0,0,0.05);
    }

    div[data-testid="stSelectbox"] div[data-baseweb="select"] svg {
        width: 14px !important;
        height: 14px !important;
    }

    div[data-testid="stTabs"] {
        margin-top: 4px !important;
        padding-top: 0px !important;
    }

    div[data-testid="stTabs"] div[role="tablist"] {
        gap: 6px;
        border-bottom: none !important;
        flex-wrap: wrap !important;
    }

    div[data-testid="stTabs"] button {
        font-family: 'Segoe UI', Roboto, sans-serif !important;
        font-weight: 700 !important;
        font-size: 12px !important;
        padding: 6px 14px !important;
        color: #ffffff !important;
        background-color: #e67e22 !important;
        border-radius: 4px !important;
        border: 1px solid #d35400 !important;
        margin-right: 2px !important;
    }
    div[data-testid="stTabs"] button p { color: #ffffff !important; }
    div[data-testid="stTabs"] button:hover {
        background-color: #f39c12 !important;
        color: #ffffff !important;
    }
    div[data-testid="stTabs"] button[aria-selected="true"] {
        color: #ffffff !important;
        background-color: #27ae60 !important;
        border-color: #1e8449 !important;
    }
    div[data-testid="stTabs"] button[aria-selected="true"] p { color: #ffffff !important; }
    div[data-testid="stTabs"] [data-baseweb="tab-highlight"] { display: none !important; }

    div[data-testid="stDownloadButton"] button {
        padding: 4px 12px !important;
        font-size: 12px !important;
        min-height: 32px !important;
    }

    hr { margin: 8px 0 !important; }
    h4 { margin: 8px 0 6px 0 !important; font-size: 16px !important; }

    .excel-table {
        border-collapse: collapse;
        width: 100%;
        font-family: 'Segoe UI', Arial, sans-serif;
        font-size: 12px;
        margin-top: 6px;
    }
    .excel-table th, .excel-table td {
        border: 1px solid #e2e8f0;
        padding: 4px 6px;
        text-align: center;
        white-space: nowrap;
        font-size: 12px;
    }
    .excel-table th { font-weight: 700; font-size: 11px; text-transform: uppercase; }
    .header-tot { background-color: #6b21a8; color: white; }
    .header-fpd { background-color: #15803d; color: white; }
    .header-mhl { background-color: #1d4ed8; color: white; }
    .header-sub { background-color: #f1f5f9; color: #334155; font-size: 10px; }
    .header-left { background-color: #0369a1; color: white; }
    .header-km { background-color: #c2410c; color: white; }
    .header-earn { background-color: #047857; color: white; }
    .pos { background-color: #dcfce7; color: #15803d; font-weight: 600; }
    .neg { background-color: #fee2e2; color: #b91c1c; font-weight: 600; }

    .title-bar {
        background: linear-gradient(90deg, #fef08a 0%, #fef9c3 100%);
        color: #854d0e;
        text-align: center;
        font-weight: 700;
        font-size: 14px;
        padding: 6px;
        border-radius: 4px;
        border: 1px solid #fef08a;
        margin-bottom: 8px;
        font-family: 'Segoe UI', Roboto, sans-serif !important;
    }

    .freeze-wrap, .op-wrap {
        max-height: 65vh;
        overflow: auto;
        border: 1px solid #e2e8f0;
        margin-bottom: 10px;
        position: relative;
        z-index: 0;
    }
    .freeze-table, .op-table {
        border-collapse: separate;
        border-spacing: 0;
        font-size: 12px;
        width: max-content;
    }
    .freeze-table th, .freeze-table td,
    .op-table th, .op-table td {
        border: 1px solid #e2e8f0;
        white-space: nowrap;
    }
    .freeze-table thead tr:nth-child(1) th { position: sticky; top: 0; z-index: 4; }
    .freeze-table thead tr:nth-child(2) th { position: sticky; top: 30px; z-index: 4; }
    .freeze-table thead tr:nth-child(3) th { position: sticky; top: 58px; z-index: 4; }
    .op-table thead tr:first-child th { position: sticky; top: 0; z-index: 4; }
    .op-table thead tr:nth-child(2) th { position: sticky; top: 34px; z-index: 4; }

    /* Scrollable tables – keep scrollbar clickable (avoid sticky overlay) */
    .table-scroll-fixable {
        max-height: 70vh;
        overflow: scroll !important;
        border: 1px solid #cbd5e1;
        width: 100%;
        position: relative;
        z-index: 0;
        isolation: isolate;
        -webkit-overflow-scrolling: touch;
        pointer-events: auto !important;
        margin-bottom: 10px;
        background: #fff;
    }
    .table-scroll-fixable table {
        position: relative;
        z-index: 0;
    }

    /* ---- Sidebar navigation (open / close via Streamlit << control) ---- */
    section[data-testid="stSidebar"] {
        background: linear-gradient(180deg, #f8fafc 0%, #eef2ff 45%, #f1f5f9 100%) !important;
        border-right: 1px solid #c7d2fe !important;
    }
    section[data-testid="stSidebar"] > div {
        padding-top: 0.6rem !important;
    }
    .nav-brand {
        background: linear-gradient(135deg, #1e3a8a 0%, #7c3aed 55%, #db2777 100%);
        color: #fff;
        font-weight: 800;
        font-size: 13px;
        letter-spacing: 0.6px;
        text-align: center;
        padding: 12px 10px;
        border-radius: 10px;
        margin: 4px 4px 14px 4px;
        box-shadow: 0 4px 12px rgba(30, 58, 138, 0.25);
        text-transform: uppercase;
    }
    .nav-brand small {
        display: block;
        font-size: 10px;
        font-weight: 600;
        opacity: 0.9;
        margin-top: 4px;
        letter-spacing: 0.3px;
        text-transform: none;
    }
    /* Radio list as menu rows */
    section[data-testid="stSidebar"] div[role="radiogroup"] {
        gap: 2px !important;
    }
    section[data-testid="stSidebar"] div[role="radiogroup"] label {
        background: transparent !important;
        border-radius: 8px !important;
        padding: 8px 12px !important;
        margin: 1px 4px !important;
        border: 1px solid transparent !important;
        transition: all 0.15s ease;
    }
    section[data-testid="stSidebar"] div[role="radiogroup"] label:hover {
        background: #e0e7ff !important;
        border-color: #c7d2fe !important;
    }
    section[data-testid="stSidebar"] div[role="radiogroup"] label[data-checked="true"],
    section[data-testid="stSidebar"] div[role="radiogroup"] label:has(input:checked) {
        background: linear-gradient(90deg, #e0e7ff 0%, #ddd6fe 100%) !important;
        border-color: #a5b4fc !important;
        box-shadow: inset 3px 0 0 #4f46e5;
    }
    section[data-testid="stSidebar"] div[role="radiogroup"] label p {
        font-family: 'Segoe UI', Roboto, sans-serif !important;
        font-size: 13px !important;
        font-weight: 600 !important;
        color: #334155 !important;
    }
    section[data-testid="stSidebar"] div[role="radiogroup"] label:has(input:checked) p {
        color: #1e3a8a !important;
        font-weight: 700 !important;
    }
    /* Hide radio circles for cleaner menu look */
    section[data-testid="stSidebar"] div[role="radiogroup"] label > div:first-child {
        display: none !important;
    }
    section[data-testid="stSidebar"] [data-testid="stMarkdownContainer"] p {
        color: #64748b;
    }
</style>
""",
    unsafe_allow_html=True,
)
# ========== PATH ==========
PARQUET_FILE = r"D:\dashboard\ser_wise.parquet"
SERVICE_MONTHLY_FILE = r"D:\dashboard\ser_wise.parquet"
# Dedicated service-monthly metrics file for Monthly boards 7 and 8.
# Both spellings are supported: ser_monthly.parquet / ser_montly.parquet.
MONTHLY_SERVICE_METRICS_FILE = r"D:\dashboard\ser_monthly.parquet"
MONTHLY_SERVICE_METRICS_ALTS = [
    Path(r"D:\Dashboard\ser_monthly.parquet"), Path(r"D:\dashboard\ser_monthly.parquet"),
    Path(r"D:\MONTHLY\ser_monthly.parquet"), Path(r"D:\Dashboard\ser_montly.parquet"),
    Path(r"D:\dashboard\ser_montly.parquet"), Path(r"D:\MONTHLY\ser_montly.parquet"),
    Path("ser_monthly.parquet"), Path("ser_montly.parquet"),
    Path(r"/home/workdir/attachments/ser_monthly.parquet"),
    Path(r"/home/workdir/attachments/ser_montly.parquet"),
]
# These tabs load from ser_wise.parquet (same folder as ser_wise)
SERVICE_MONTHLY_TABS = {
    "ACT VS ACT",
    "ACT vs ACT TRENDS",
    "Product wise",
    "Service-wise (SROS)",
    "Trends from 2024",
}


def _resolve_parquet(primary, extra_alts=None):
    path = Path(primary)
    if path.exists():
        return path
    name = Path(primary).name
    candidates = list(extra_alts or []) + [
        Path(name),
        Path(r"D:\Dashboard") / name,
        Path(r"D:\dashboard") / name,
        Path(r"D:\MONTHLY") / name,
        Path(r"/home/workdir/attachments") / name,
    ]
    for alt in candidates:
        p = Path(alt)
        if p.exists():
            return p
    return None


def _parquet_signature(path):
    try:
        p = Path(path)
        stt = p.stat()
        return (str(p.resolve()), int(stt.st_mtime_ns), int(stt.st_size))
    except Exception:
        return (str(path), 0, 0)


@st.cache_resource(show_spinner=False)
def _load_data_cached(parquet_path=None, required=True, signature=None):
    """Load performance parquet; normalize columns to app-standard names.

    parquet_path: optional override (ser_wise.parquet).
    required: if False, return empty DataFrame when missing (no st.stop).
    """
    if parquet_path:
        path = _resolve_parquet(parquet_path)
    else:
        path = _resolve_parquet(
            PARQUET_FILE,
            [
                Path(r"D:\Dashboard\ser_wise.parquet"),
                Path(r"D:\MONTHLY\ser_wise.parquet"),
                Path("ser_wise.parquet"),
                Path(r"/home/workdir/attachments/ser_wise.parquet"),
            ],
        )
    if path is None:
        if required:
            st.error(f"File not found: {parquet_path or PARQUET_FILE}")
            st.stop()
        return pd.DataFrame()
    # Read only columns used by the dashboard.  DuckDB scans Parquet directly and
    # avoids materialising the unused columns in the 51-column source file.
    # If DuckDB is unavailable, fall back to pandas for local compatibility.
    _dashboard_targets = {
        # Canonical names plus source aliases used by SER parquet.
        "Date", "Month_Name", "Month Name", "DEPOT", "REGION", "SER_NO",
        "PRODUCT", "PRODUCT_NAME", "ROUTEE", "ROUTE", "ROUTE_OLD",
        "Optd_KMs", "OPD_KMS", "OPTD_KMS", "DAY_SCH_KMS",
        "GE_TOT", "Gross Total", "GE_FPD", "Gross Fare Paid",
        "GE_MHL", "Gross MHL", "NE_TOT", "Net Total",
        "NE_FPD", "Net Fare Paid", "NE_MHL", "Net MHL",
        "PSNGR_TOT", "Passengers Total", "PSNGR_FPD", "Passengers Fare Paid",
        "PSNGR_MHL", "Passengers MHL", "MHL_NMHL", "MHL/NMHL",
        "RTC_HIRE", "RTC/HIRE", "Weekday", "NO_OF_SCHS", "NO.OF SCHs",
        "INTERSTATE", "TYPE", "D.TYPE", "NATURE", "SCH_DEP", "R/L", "LONG_TP.",
    }
    try:
        if duckdb is not None:
            con = duckdb.connect(database=":memory:")
            try:
                schema = con.execute(
                    "DESCRIBE SELECT * FROM read_parquet(?)", [str(path)]
                ).fetchall()
                actual_cols = [r[0] for r in schema]
                def _norm_src(c):
                    return str(c).strip().lower().replace("_", "").replace(" ", "")
                wanted_norm = {_norm_src(x) for x in _dashboard_targets}
                selected = [c for c in actual_cols if _norm_src(c) in wanted_norm]
                # Always include the full set when projection cannot identify enough
                # columns; this keeps compatibility with unusual parquet schemas.
                # Never silently omit core metrics. If aliases cannot be resolved,
                # read the source schema rather than creating blank metric columns.
                _required_norm = {
                    "date", "monthname", "depot", "product", "routee",
                    "opdkms", "optdkms", "grosstotal", "nettotal",
                    "passengerstotal", "mhlnmhl", "rtchire", "serno", "weekday"
                }
                _selected_norm = {_norm_src(c) for c in selected}
                if selected and len(_selected_norm & _required_norm) >= 7:
                    qcols = ", ".join('"' + str(c).replace('"', '""') + '"' for c in selected)
                    df = con.execute(
                        f"SELECT {qcols} FROM read_parquet(?)", [str(path)]
                    ).df()
                else:
                    df = con.execute("SELECT * FROM read_parquet(?)", [str(path)]).df()
            finally:
                con.close()
        else:
            df = pd.read_parquet(path)
    except Exception:
        # Keep the application compatible with Windows/local installations where
        # duckdb has not yet been installed.
        df = pd.read_parquet(path)

    # Map actual parquet columns -> names expected by the dashboard
    rename_map = {
        # keys are normalized (lower, no spaces/underscores)
        "monthname": "Month_Name",
        "month_name": "Month_Name",
        "date": "Date",
        "depot": "DEPOT",
        "region": "REGION",
        "serno": "SER_NO",
        "ser_no": "SER_NO",
        "product": "PRODUCT",
        "productname": "PRODUCT_NAME",
        "product_name": "PRODUCT_NAME",
        "routee": "ROUTEE",
        "route": "ROUTE",
        "routeold": "ROUTE_OLD",
        "mhl/nmhl": "MHL_NMHL",
        "mhlnmhl": "MHL_NMHL",
        "mhl_nmhl": "MHL_NMHL",
        "rtc/hire": "RTC_HIRE",
        "rtchire": "RTC_HIRE",
        "rtc_hire": "RTC_HIRE",
        "weekday": "Weekday",
        # kms
        "opdkms": "Optd_KMs",
        "opd_kms": "Optd_KMs",
        "optdkms": "Optd_KMs",
        "optd_kms": "Optd_KMs",
        "dayschkms": "DAY_SCH_KMS",
        # gross earnings
        "grosstotal": "GE_TOT",
        "gross total": "GE_TOT",
        "grossfarepaid": "GE_FPD",
        "gross fare paid": "GE_FPD",
        "grossmhl": "GE_MHL",
        "gross mhl": "GE_MHL",
        # net earnings
        "nettotal": "NE_TOT",
        "net total": "NE_TOT",
        "netfarepaid": "NE_FPD",
        "net fare paid": "NE_FPD",
        "netmhl": "NE_MHL",
        "net mhl": "NE_MHL",
        # passengers
        "passengerstotal": "PSNGR_TOT",
        "passengers total": "PSNGR_TOT",
        "passengersfarepaid": "PSNGR_FPD",
        "passengers fare paid": "PSNGR_FPD",
        "passengersmhl": "PSNGR_MHL",
        "passengers mhl": "PSNGR_MHL",
        # schedules count
        "no.ofschs": "NO_OF_SCHS",
        "noofschs": "NO_OF_SCHS",
    }

    def _norm(c):
        return str(c).strip().lower().replace("_", "").replace(" ", "")

    col_rename = {}
    for c in df.columns:
        key = _norm(c)
        key_sp = str(c).strip().lower()
        if key in rename_map:
            col_rename[c] = rename_map[key]
        elif key_sp in rename_map:
            col_rename[c] = rename_map[key_sp]
        # also try with slash kept
        elif str(c).strip().lower() in rename_map:
            col_rename[c] = rename_map[str(c).strip().lower()]

    # Explicit common labels from this parquet
    explicit = {
        "Month Name": "Month_Name",
        "OPD_KMS": "Optd_KMs",
        "Gross Total": "GE_TOT",
        "Gross Fare Paid": "GE_FPD",
        "Gross MHL": "GE_MHL",
        "Net Total": "NE_TOT",
        "Net Fare Paid": "NE_FPD",
        "Net MHL": "NE_MHL",
        "Passengers Total": "PSNGR_TOT",
        "Passengers Fare Paid": "PSNGR_FPD",
        "Passengers MHL": "PSNGR_MHL",
        "MHL/NMHL": "MHL_NMHL",
        "RTC/HIRE": "RTC_HIRE",
        "NO.OF SCHs": "NO_OF_SCHS",
        "SER_NO": "SER_NO",
        "DEPOT": "DEPOT",
        "PRODUCT": "PRODUCT",
        "ROUTE": "ROUTE",
        "ROUTEE": "ROUTEE",
        "Date": "Date",
        "Weekday": "Weekday",
    }
    for src, dst in explicit.items():
        if src in df.columns:
            col_rename[src] = dst

    df = df.rename(columns=col_rename)

    # Drop duplicate column names (keep first) — avoids DataFrame.str errors
    if df.columns.duplicated().any():
        df = df.loc[:, ~df.columns.duplicated()].copy()

    # Ensure required columns exist
    required = [
        "Date", "Month_Name", "DEPOT", "PRODUCT", "ROUTEE", "Optd_KMs",
        "GE_TOT", "GE_FPD", "GE_MHL", "NE_TOT", "NE_FPD", "NE_MHL",
        "PSNGR_TOT", "PSNGR_FPD", "PSNGR_MHL", "MHL_NMHL", "RTC_HIRE", "SER_NO", "Weekday",
    ]
    missing = [c for c in required if c not in df.columns]
    if missing:
        # soft-create empty columns so app does not crash
        for c in missing:
            df[c] = np.nan if c not in ("Month_Name", "DEPOT", "PRODUCT", "ROUTEE", "SER_NO", "Weekday", "MHL_NMHL", "RTC_HIRE") else ""
        # show once
        try:
            st.warning(f"Parquet missing columns (filled blank): {missing}")
        except Exception:
            pass

    if "Date" in df.columns and not pd.api.types.is_datetime64_any_dtype(df["Date"]):
        df["Date"] = pd.to_datetime(df["Date"], errors="coerce")

    # numeric coercions (only if Series, not duplicate DataFrame)
    for c in ["Optd_KMs", "GE_TOT", "GE_FPD", "GE_MHL", "NE_TOT", "NE_FPD", "NE_MHL",
              "PSNGR_TOT", "PSNGR_FPD", "PSNGR_MHL"]:
        if c in df.columns:
            s = df[c]
            if isinstance(s, pd.DataFrame):
                s = s.iloc[:, 0]
            df[c] = pd.to_numeric(s, errors="coerce").fillna(0)

    # string clean
    for c in ["DEPOT", "PRODUCT", "ROUTEE", "Month_Name", "MHL_NMHL", "RTC_HIRE", "SER_NO", "Weekday"]:
        if c in df.columns:
            s = df[c]
            if isinstance(s, pd.DataFrame):
                s = s.iloc[:, 0]
            df[c] = s.astype(str).str.strip().replace({"nan": "", "None": "", "<NA>": ""})

    # Date clean
    if "Date" in df.columns:
        s = df["Date"]
        if isinstance(s, pd.DataFrame):
            s = s.iloc[:, 0]
        df["Date"] = pd.to_datetime(s, errors="coerce")

    # Build ROUTEE once so downstream tabs never have to mutate the shared
    # cached DataFrame during a Streamlit rerun.
    if "ROUTEE" not in df.columns and "ROUTE" in df.columns:
        df["ROUTEE"] = df["ROUTE"].astype(str).str.strip()

    # IMPORTANT: do not convert dashboard dimensions to pandas Categorical here.
    # Many existing report functions intentionally use broad operations such as
    # merged.fillna(0).iterrows().  Pandas raises TypeError when 0 is assigned to
    # a categorical column whose categories do not contain 0.  Keeping these
    # columns as normal object/string columns preserves the original behaviour
    # and avoids breaking existing calculations, tables and Excel exports.
    # Memory reduction is achieved by DuckDB column projection rather than by
    # changing the dtypes used by the reporting logic.

    return df


def load_data(parquet_path=None, required=True):
    """Resolve Parquet and cache one shared DataFrame per file version."""
    if parquet_path:
        path = _resolve_parquet(parquet_path)
    else:
        path = _resolve_parquet(PARQUET_FILE, [Path(r"D:\Dashboard\ser_wise.parquet"), Path(r"D:\MONTHLY\ser_wise.parquet"), Path("ser_wise.parquet"), Path(r"/home/workdir/attachments/ser_wise.parquet")])
    sig = _parquet_signature(path) if path is not None else (str(parquet_path or PARQUET_FILE), 0, 0)
    return _load_data_cached(parquet_path, required, sig)

df_ser_wise = load_data()
df_service_monthly = df_ser_wise
# Default working frame; switched per tab after section is chosen
df = df_ser_wise


@st.cache_resource(show_spinner=False)
def _load_monthly_service_metrics_cached(path_str, signature=None):
    """Load the dedicated monthly service metrics parquet ONLY.

    This source is intentionally independent from ser_wise.parquet.  It must not
    require Date/Weekday or SMASTER identity fields.  Boards 7/8 get service
    identity/schedule information from SMASTER and performance metrics from
    this file.
    """
    p = Path(path_str)
    if not p.exists():
        return pd.DataFrame(), None
    try:
        if duckdb is not None:
            con = duckdb.connect(database=":memory:")
            try:
                dfm = con.execute("SELECT * FROM read_parquet(?)", [str(p)]).df()
            finally:
                con.close()
        else:
            dfm = pd.read_parquet(p)

        dfm.columns = [str(c).strip() for c in dfm.columns]

        def norm(c):
            return (str(c).strip().lower().replace("_", "")
                    .replace(" ", "").replace("/", "").replace(".", ""))

        # Source -> application-standard names.  Only monthly metric fields are
        # normalized here; master/schedule fields deliberately stay out.
        aliases = {
            "monthname": "Month_Name", "month": "Month_Name",
            "depot": "DEPOT",
            "serno": "SER_NO", "serviceno": "SER_NO", "servicenumber": "SER_NO",
            "opdkms": "Optd_KMs", "optdkms": "Optd_KMs", "operatedkms": "Optd_KMs",
            "grosstotal": "GE_TOT", "grosstotalamount": "GE_TOT",
            "grossfarepaid": "GE_FPD", "grossmhl": "GE_MHL",
            "nettotal": "NE_TOT", "netfarepaid": "NE_FPD", "netmhl": "NE_MHL",
            "passengerstotal": "PSNGR_TOT", "passengersfarepaid": "PSNGR_FPD",
            "passengersmhl": "PSNGR_MHL",
            "mhlnmhl": "MHL_NMHL",
            # Service-master attributes intentionally retained in the monthly
            # service metrics file because Boards 7/8 use these fields from
            # ser_monthly/ser_montly:
            "product": "PRODUCT",
            "schdep": "SCH_DEP",
            "sch.dep": "SCH_DEP",
            "scheduledpt": "SCH_DEP",
            "scheduledepot": "SCH_DEP",
            "schdepot": "SCH_DEP",
            "scheduledep": "SCH_DEP",
            "schdepo": "SCH_DEP",
            "rl": "R_L",
            "r/l": "R_L",
            "routelength": "R_L",
            "routelen": "R_L",
            "route.length": "R_L",
            "schkmsrl": "R_L",
            "route": "ROUTE",
            "routee": "ROUTEE",
            "rtchire": "RTC_HIRE",
            "days": "DAYS", "noofdays": "DAYS", "operateddays": "DAYS",
            "optdays": "DAYS", "optddays": "DAYS",
            "date": "Date",
        }
        ren = {}
        for c in dfm.columns:
            k = norm(c)
            if k in aliases:
                ren[c] = aliases[k]
        dfm = dfm.rename(columns=ren)
        if dfm.columns.duplicated().any():
            dfm = dfm.loc[:, ~dfm.columns.duplicated()].copy()

        # Month_Name is mandatory for monthly boards.  Derive it from Date or
        # Month + Year if the source uses those instead.
        if "Month_Name" not in dfm.columns:
            if "Date" in dfm.columns:
                dt = pd.to_datetime(dfm["Date"], errors="coerce")
                dfm["Month_Name"] = dt.dt.strftime("%b-%Y")
            else:
                cols = {norm(c): c for c in dfm.columns}
                mc = cols.get("month")
                yc = cols.get("year")
                if mc and yc:
                    mon = dfm[mc].astype(str).str.strip().str[:3].str.title()
                    yr = pd.to_numeric(dfm[yc], errors="coerce")
                    dfm["Month_Name"] = mon + "-" + yr.fillna(-1).astype(int).astype(str)
                    dfm.loc[yr.isna(), "Month_Name"] = ""

        # Normalize key fields.
        if "DEPOT" in dfm.columns:
            dfm["DEPOT"] = dfm["DEPOT"].astype(str).str.strip().str.upper()
        if "SER_NO" in dfm.columns:
            def _svc(v):
                if v is None or (isinstance(v, float) and pd.isna(v)):
                    return ""
                z = str(v).strip()
                if z.lower() in ("", "nan", "none"):
                    return ""
                if z.endswith(".0"):
                    z = z[:-2]
                try:
                    return str(int(float(z)))
                except Exception:
                    return z
            dfm["SER_NO"] = dfm["SER_NO"].map(_svc)

        if "Month_Name" in dfm.columns:
            dfm["Month_Name"] = dfm["Month_Name"].astype(str).str.strip()
        for c in ["PRODUCT", "SCH_DEP", "ROUTE", "ROUTEE", "RTC_HIRE"]:
            if c in dfm.columns:
                dfm[c] = dfm[c].astype(str).str.strip()
        # Fuzzy recover SCH_DEP / R_L if still missing after alias rename
        def _normk(c):
            return (str(c).strip().lower().replace("_", "")
                    .replace(" ", "").replace("/", "").replace(".", ""))
        if "SCH_DEP" not in dfm.columns:
            for c in list(dfm.columns):
                k = _normk(c)
                if ("sch" in k and "dep" in k) or k in ("schdep", "scheduledepot", "schdepot"):
                    dfm["SCH_DEP"] = dfm[c].astype(str)
                    break
        if "R_L" not in dfm.columns:
            for c in list(dfm.columns):
                k = _normk(c)
                if k in ("rl", "routelength", "routelen") or k.startswith("rl") and "epk" not in k:
                    dfm["R_L"] = pd.to_numeric(dfm[c], errors="coerce").fillna(0.0)
                    break
        if "R_L" in dfm.columns:
            dfm["R_L"] = pd.to_numeric(dfm["R_L"], errors="coerce").fillna(0.0)
        if "SCH_DEP" in dfm.columns:
            dfm["SCH_DEP"] = dfm["SCH_DEP"].astype(str).str.strip().replace(
                {"nan": "", "None": "", "none": "", "<NA>": ""}
            )
        for c in ["Optd_KMs", "GE_TOT", "GE_FPD", "GE_MHL", "NE_TOT", "NE_FPD",
                  "NE_MHL", "PSNGR_TOT", "PSNGR_FPD", "PSNGR_MHL", "DAYS"]:
            if c in dfm.columns:
                dfm[c] = pd.to_numeric(dfm[c], errors="coerce").fillna(0.0)

        return dfm, None
    except Exception as exc:
        return pd.DataFrame(), str(exc)


def load_monthly_service_metrics():
    p = _resolve_parquet(MONTHLY_SERVICE_METRICS_FILE, MONTHLY_SERVICE_METRICS_ALTS)
    if p is None:
        return pd.DataFrame(), None
    try:
        # _load_monthly_service_metrics_cached already returns (DataFrame, error).
        # Do not wrap that tuple again; Boards 7/8 expect (DataFrame, error).
        return _load_monthly_service_metrics_cached(str(p), _parquet_signature(p))
    except Exception as exc:
        return pd.DataFrame(), str(exc)


@st.cache_data(ttl=300, show_spinner=False)
def load_fleet_map(path: str = r"D:\Dashboard\FLEET.parquet"):
    """FLEET counts by (DEPOT, PRODUCT, MONTH).

    FLEET.parquet layout:
      LAST DAY = last day of the month the fleet belongs to  (e.g. 2026-07-31 → Jul-2026)
      Date     = first day of the NEXT month                 (e.g. 2026-08-01)

    Month key MUST come from LAST DAY (not Date), otherwise July fleet
    appears under August and values look wrong.

    Depot alias: TDR → TNDR.  PRODUCT=TOTAL excluded from depot/month totals.
    """
    p = Path(path)
    if not p.exists():
        for alt in [
            Path(r"D:\Dashboard\FLEET.parquet"),
            Path(r"D:\dashboard\FLEET.parquet"),
            Path(r"D:\MONTHLY\FLEET.parquet"),
            Path("FLEET.parquet"),
            Path(r"/home/workdir/attachments/FLEET.parquet"),
        ]:
            if alt.exists():
                p = alt
                break
        else:
            return {}, "FLEET.parquet not found"
    try:
        try:
            fdf = pd.read_parquet(p, engine="pyarrow")
        except Exception:
            fdf = pd.read_parquet(p)
        fdf.columns = [str(c).strip() for c in fdf.columns]

        def fc(cands):
            n = {
                str(c).strip().lower().replace(" ", "").replace("_", "").replace("/", ""): c
                for c in fdf.columns
            }
            for cand in cands:
                k = cand.lower().replace(" ", "").replace("_", "").replace("/", "")
                if k in n:
                    return n[k]
            for c in fdf.columns:
                cl = str(c).strip().lower().replace(" ", "").replace("_", "")
                for cand in cands:
                    if cand.lower().replace(" ", "").replace("_", "") in cl:
                        return c
            return None

        DEPOT_ALIAS = {"TDR": "TNDR", "TANDUR": "TNDR"}

        c_dep = fc(["DEPOT"])
        c_prod = fc(["PRODUCT"])
        c_mon = fc(["MONTH", "Month", "Month_Name", "Month Name"])
        c_year = fc(["YEAR", "Year"])
        c_date = fc(["Date", "DATE"])
        c_last = fc(["LAST DAY", "LASTDAY", "LAST_DAY"])
        c_fleet = fc(["FLEET", "Fleet", "FLEET_COUNT", "Buses", "BUSES", "NO_OF_FLEET"])
        if c_fleet is None:
            for c in fdf.columns:
                if not pd.api.types.is_numeric_dtype(fdf[c]):
                    continue
                if c in (c_dep, c_year, c_date, c_last):
                    continue
                if any(x in str(c).lower() for x in ("sch", "kms", "km")):
                    continue
                c_fleet = c
                break
        if c_fleet is None:
            return {}, f"No fleet column found. Columns: {list(fdf.columns)}"

        out = {
            "by_dpm": {}, "by_dm": {}, "by_m": {}, "by_d": {},
            "_source": str(p), "_rows": len(fdf),
        }

        for _, row in fdf.iterrows():
            try:
                val = float(pd.to_numeric(row[c_fleet], errors="coerce") or 0)
            except Exception:
                val = 0.0

            dep = str(row[c_dep]).strip().upper() if c_dep else "ALL"
            dep = DEPOT_ALIAS.get(dep, dep)
            prod = str(row[c_prod]).strip().upper() if c_prod else "ALL"

            # ---- month key (ONE correct key per row) ----
            mon_key = "ALL"

            if c_mon is not None:
                mraw = str(row[c_mon]).strip()
                if c_year is not None and str(row.get(c_year, "")).strip() not in ("", "nan", "None"):
                    try:
                        yr = int(float(row[c_year]))
                        mon_key = f"{mraw[:3].title()}-{yr}"
                    except Exception:
                        mon_key = mraw if mraw else "ALL"
                elif mraw and mraw.lower() not in ("nan", "none", ""):
                    mon_key = mraw

            elif c_last is not None:
                # Preferred: LAST DAY = actual month of fleet (Jul-31 → Jul-2026)
                dt = pd.to_datetime(row[c_last], errors="coerce")
                if pd.notna(dt):
                    mon_key = dt.strftime("%b-%Y")

            elif c_date is not None:
                # Fallback: Date is 1st of NEXT month → subtract 1 month
                dt = pd.to_datetime(row[c_date], errors="coerce")
                if pd.notna(dt):
                    dt_prev = dt - pd.DateOffset(months=1)
                    mon_key = dt_prev.strftime("%b-%Y")

            # by_dpm always (includes PRODUCT=TOTAL for explicit TOTAL lookup)
            out["by_dpm"][(dep, prod, mon_key)] = (
                out["by_dpm"].get((dep, prod, mon_key), 0) + val
            )
            # depot/month totals: skip PRODUCT=TOTAL to avoid double-count
            if prod not in ("TOTAL",):
                out["by_dm"][(dep, mon_key)] = out["by_dm"].get((dep, mon_key), 0) + val
                out["by_m"][mon_key] = out["by_m"].get(mon_key, 0) + val
            out["by_d"][dep] = out["by_d"].get(dep, 0) + val

        return out, None
    except Exception as e:
        return {}, str(e)


# Latest data date for display banner
try:
    _max_dt = pd.to_datetime(df["Date"], errors="coerce").max()
    DATA_UPDATED_ON = _max_dt.strftime("%d-%m-%Y") if pd.notna(_max_dt) else "N/A"
except Exception:
    DATA_UPDATED_ON = "N/A"



@st.cache_data(ttl=300, show_spinner=False)
def load_orf_map(path: str = r"D:\dashboard\ORF.xlsx"):
    """
    ORF.xlsx: DEPOT, PRODUCT, CY ORF, LY ORF
    Depot ORF = row where PRODUCT == TOTAL (e.g. BHEL + TOTAL = 7374.33)
    OR = (EPK * 10000) / Depot_ORF
    """
    p = Path(path)
    if not p.exists():
        for alt in [Path(r"D:\Dashboard\ORF.xlsx"), Path(r"D:\MONTHLY\ORF.xlsx"), Path("ORF.xlsx")]:
            if alt.exists():
                p = alt
                break
        else:
            return {}, {}, f"ORF file not found: {path}"
    try:
        s = None
        for hr in range(0, 8):
            trial = pd.read_excel(p, header=hr)
            cols_u = [str(c).strip().upper().replace(" ", "") for c in trial.columns]
            if any("DEPOT" in x for x in cols_u) and any("PRODUCT" in x for x in cols_u):
                s = trial
                break
        if s is None:
            s = pd.read_excel(p)
        s.columns = [str(c).strip() for c in s.columns]

        def find_col(*names):
            for c in s.columns:
                cu = str(c).strip().upper().replace(" ", "").replace("_", "")
                for n in names:
                    nu = n.upper().replace(" ", "").replace("_", "")
                    if cu == nu or nu in cu:
                        return c
            return None

        c_dep = find_col("DEPOT")
        c_prod = find_col("PRODUCT")
        c_cy = find_col("CYORF", "CY ORF", "ORF CY")
        c_ly = find_col("LYORF", "LY ORF", "ORF LY")
        if c_cy is None:
            c_cy = find_col("ORF")
        if c_ly is None:
            c_ly = c_cy
        if not c_dep or not c_prod or not c_cy:
            return {}, {}, f"ORF columns not found. Have: {list(s.columns)}"

        depot_orf = {}
        by_prod = {}
        for _, row in s.iterrows():
            dep = str(row[c_dep]).strip().upper()
            prod = str(row[c_prod]).strip().upper()
            if not dep or dep in ("NAN", "NONE", "NAT"):
                continue
            # keep REGION and TOTAL product rows for TOTAL-row ORF lookup
            try:
                cy = float(row[c_cy])
            except Exception:
                cy = np.nan
            try:
                ly = float(row[c_ly]) if c_ly is not None else cy
            except Exception:
                ly = np.nan
            if cy == 0:
                cy = np.nan
            if ly == 0:
                ly = np.nan
            if prod and prod not in ("NAN", "NONE", "NAT"):
                by_prod[(dep, prod)] = {"cy": cy, "ly": ly}
            # KEY: depot ORF from PRODUCT = TOTAL
            if prod == "TOTAL":
                depot_orf[dep] = {"cy": cy, "ly": ly}

        for dep in {d for (d, p) in by_prod.keys()}:
            if dep in depot_orf:
                continue
            cys = [v["cy"] for (d, p), v in by_prod.items() if d == dep and not pd.isna(v.get("cy"))]
            lys = [v["ly"] for (d, p), v in by_prod.items() if d == dep and not pd.isna(v.get("ly"))]
            depot_orf[dep] = {
                "cy": float(np.mean(cys)) if cys else np.nan,
                "ly": float(np.mean(lys)) if lys else np.nan,
            }
        return depot_orf, by_prod, None
    except Exception as e:
        return {}, {}, str(e)



def _resolve_smaster_path(primary=r"D:\Dashboard\SMASTER.parquet"):
    """Resolve SMASTER.parquet from known locations."""
    p = Path(primary)
    if p.exists():
        return p
    for alt in [
        Path(r"D:\Dashboard\SMASTER.parquet"),
        Path(r"D:\dashboard\SMASTER.parquet"),
        Path(r"D:\MONTHLY\SMASTER.parquet"),
        Path("SMASTER.parquet"),
        Path(r"/home/workdir/attachments/SMASTER.parquet"),
    ]:
        if alt.exists():
            return alt
    return p


@st.cache_resource(show_spinner=False)
def _load_smaster_cached(path, signature=None):
    p = Path(path)
    try:
        if duckdb is not None:
            con = duckdb.connect(database=":memory:")
            try:
                sdf = con.execute("SELECT * FROM read_parquet(?)", [str(p)]).df()
            finally:
                con.close()
        else:
            sdf = pd.read_parquet(p)
        sdf.columns = [str(c).strip() for c in sdf.columns]
        if "DATE" in sdf.columns and not pd.api.types.is_datetime64_any_dtype(sdf["DATE"]):
            sdf["DATE"] = pd.to_datetime(sdf["DATE"], errors="coerce")
        return sdf.dropna(axis=1, how="all"), None
    except Exception as e:
        return None, str(e)


def load_smaster(path):
    """Resolve SMASTER and cache one shared copy per file version."""
    p = Path(path)
    if not p.exists():
        resolved = _resolve_smaster_path(path)
        if resolved.exists():
            p = resolved
        else:
            return None, f"File not found: {path}"
    return _load_smaster_cached(str(p), _parquet_signature(p))

@st.cache_data(ttl=300, show_spinner=False)
def _load_sros_sch_map_pw(smaster_path: str, month_key: str, depot_key: str = "ALL"):
    """Product-wise schedule counts from SMASTER for a month (CY + LY pair)."""
    s, _sm_err = load_smaster(smaster_path)
    if s is None:
        return {}
    s = s.copy()
    s.columns = [str(c).strip() for c in s.columns]

    def fc(cands):
        n = {str(c).strip().lower().replace(" ", "").replace("_", ""): c for c in s.columns}
        for cand in cands:
            k = cand.lower().replace(" ", "").replace("_", "")
            if k in n:
                return n[k]
        return None

    cp, cs, cm, cyear, cd = fc(["PRODUCT"]), fc(["NoOfSchedules", "NoOfSchedule"]), fc(["MONTH"]), fc(["YEAR"]), fc(["DEPOT"])
    if not cp or not cs:
        return {}
    s["_P"] = s[cp].astype(str).str.strip()
    s["_SCH"] = pd.to_numeric(s[cs], errors="coerce").fillna(0)
    if cd and depot_key and str(depot_key).upper() not in ("ALL", "REGION", "NONE", ""):
        s = s[s[cd].astype(str).str.strip().str.upper() == str(depot_key).strip().upper()]
    if cm and cyear:
        _mon = s[cm].astype(str).str.strip().str[:3].str.title()
        _yr = pd.to_numeric(s[cyear], errors="coerce")
        s["_MK"] = _mon + "-" + _yr.fillna(-1).astype(int).astype(str)
        s.loc[_yr.isna(), "_MK"] = ""
    else:
        s["_MK"] = "ALL"
    cy_g = s[s["_MK"] == month_key].groupby("_P")["_SCH"].sum() if month_key else s.groupby("_P")["_SCH"].sum()
    ly_key = ""
    try:
        parts = str(month_key).split("-")
        ly_key = f"{parts[0]}-{int(parts[1]) - 1}"
    except Exception:
        pass
    ly_g = s[s["_MK"] == ly_key].groupby("_P")["_SCH"].sum() if ly_key else pd.Series(dtype=float)
    out = {}
    for p in set(list(cy_g.index) + list(ly_g.index)):
        out[str(p).strip()] = (float(cy_g.get(p, 0)), float(ly_g.get(p, 0)))
    return out


@st.cache_data(ttl=300, show_spinner=False)
def _load_sros_services_full(smaster_path: str, depot_key: str, month_key: str):
    """Service-level rows from SMASTER for Service-wise (SROS) tab."""
    s, _sm_err = load_smaster(smaster_path)
    if s is None:
        return pd.DataFrame()
    s = s.copy()
    s.columns = [str(c).strip() for c in s.columns]

    def fc(cands):
        n = {str(c).strip().lower().replace(" ", "").replace("_", ""): c for c in s.columns}
        for cand in cands:
            k = cand.lower().replace(" ", "").replace("_", "")
            if k in n:
                return n[k]
        return None

    c_svc = fc(["ServiceNo", "SER_NO", "SERVICE_NO"])
    c_dep = fc(["DEPOT"])
    c_prod = fc(["PRODUCT"])
    c_route = fc(["ROUTEE", "Routee", "routee"]) or fc(["ROUTEE"])
    c_sch = fc(["NoOfSchedules", "NoOfSchedule"])
    c_mon = fc(["MONTH"])
    c_year = fc(["YEAR"])
    if not c_svc:
        return pd.DataFrame()
    s = s.copy()
    s["_SVC"] = s[c_svc].map(lambda v: str(int(float(v))) if str(v).replace(".", "", 1).isdigit() else str(v).strip())
    s["_DEP"] = s[c_dep].astype(str).str.strip().str.upper() if c_dep else ""
    s["_PROD"] = s[c_prod].astype(str).str.strip() if c_prod else ""
    s["_ROUTE"] = s[c_route].astype(str).str.strip() if c_route else ""
    s["_SCH"] = pd.to_numeric(s[c_sch], errors="coerce").fillna(0) if c_sch else 0
    if c_mon and c_year:
        _mon = s[c_mon].astype(str).str.strip().str[:3].str.title()
        _yr = pd.to_numeric(s[c_year], errors="coerce")
        s["_MK"] = _mon + "-" + _yr.fillna(-1).astype(int).astype(str)
        s.loc[_yr.isna(), "_MK"] = ""
    else:
        s["_MK"] = "ALL"
    if depot_key and str(depot_key).upper() not in ("ALL", "REGION", ""):
        s = s[s["_DEP"] == str(depot_key).strip().upper()]
    return s


def add_or_columns_depot(merged, data_cy, data_ly, earn_tot_c, earn_fpd_c, earn_mhl_c, orf_map,
                         orf_by_prod=None, product_filter="ALL", depot_filter="ALL"):
    """
    OR = (EPK * 10000) / ORF

    Depot filter | Product filter | Depot-row ORF     | TOTAL-row ORF
    ALL          | ALL            | (DEPOT, TOTAL)    | (REGION, TOTAL)
    ALL          | AC-HBD         | (DEPOT, AC-HBD)  | (REGION, AC-HBD)
    BHEL         | ALL            | (BHEL, TOTAL)     | (BHEL, TOTAL)
    BHEL         | AC-HBD         | (BHEL, AC-HBD)   | (BHEL, AC-HBD)

    If REGION+PRODUCT missing in ORF, TOTAL OR falls back to kms-weighted avg of depot ORs.
    """
    merged = merged.copy()
    depot_orf = orf_map if isinstance(orf_map, dict) else {}
    by_prod = orf_by_prod if isinstance(orf_by_prod, dict) else {}
    prod_sel = str(product_filter).strip().upper() if product_filter else "ALL"
    dep_sel = str(depot_filter).strip().upper() if depot_filter else "ALL"
    prod_is_all = prod_sel in ("ALL", "NONE", "", "NAN")
    dep_is_all = dep_sel in ("ALL", "NONE", "", "NAN")

    def lookup(dep_key, prod_key, side):
        d = str(dep_key).strip().upper()
        p = str(prod_key).strip().upper()
        # try exact keys with minor product variants
        candidates = [p]
        if p not in ("TOTAL", "ALL"):
            candidates += [p.replace("-", " "), p.replace(" ", "-"), p.replace("-", "")]
        for pp in candidates:
            rec = by_prod.get((d, pp), {})
            val = rec.get(side, np.nan)
            if val is not None and not (isinstance(val, float) and np.isnan(val)):
                return float(val)
            # case-insensitive scan
            for (dd, pr), v in by_prod.items():
                if dd == d and pr.replace(" ", "").replace("-", "") == pp.replace(" ", "").replace("-", ""):
                    val = v.get(side, np.nan)
                    if val is not None and not (isinstance(val, float) and np.isnan(val)):
                        return float(val)
        if p == "TOTAL":
            rec2 = depot_orf.get(d, {})
            val2 = rec2.get(side, np.nan)
            if val2 is not None and not (isinstance(val2, float) and np.isnan(val2)):
                return float(val2)
        return np.nan

    def get_orf_for_row(dep, side):
        d = str(dep).strip().upper()
        is_total_row = d in ("TOTAL", "REGION")

        if is_total_row:
            # TOTAL row: REGION+PRODUCT when depot=ALL; else selected depot + product/TOTAL
            if dep_is_all and prod_is_all:
                return lookup("REGION", "TOTAL", side)
            if dep_is_all and not prod_is_all:
                return lookup("REGION", prod_sel, side)
            if not dep_is_all and prod_is_all:
                return lookup(dep_sel, "TOTAL", side)
            return lookup(dep_sel, prod_sel, side)

        # normal depot row
        if dep_is_all and prod_is_all:
            return lookup(d, "TOTAL", side)
        if dep_is_all and not prod_is_all:
            return lookup(d, prod_sel, side)
        if not dep_is_all and prod_is_all:
            return lookup(d, "TOTAL", side)
        return lookup(d, prod_sel, side)

    orf_cy = merged["DEPOT"].map(lambda d: get_orf_for_row(d, "cy"))
    orf_ly = merged["DEPOT"].map(lambda d: get_orf_for_row(d, "ly"))

    for kind in ["tot", "fpd", "mhl"]:
        epk_cy = pd.to_numeric(merged.get(f"epk_{kind}_CY"), errors="coerce")
        epk_ly = pd.to_numeric(merged.get(f"epk_{kind}_LY"), errors="coerce")
        oc = pd.to_numeric(orf_cy, errors="coerce")
        ol = pd.to_numeric(orf_ly, errors="coerce")
        merged[f"or_{kind}_CY"] = np.where(oc.fillna(0) != 0, (epk_cy * 10000) / oc, np.nan)
        merged[f"or_{kind}_LY"] = np.where(ol.fillna(0) != 0, (epk_ly * 10000) / ol, np.nan)
        merged[f"or_{kind}_VAR"] = merged[f"or_{kind}_CY"] - merged[f"or_{kind}_LY"]
        merged[f"or_{kind}_PCT"] = np.where(
            pd.to_numeric(merged[f"or_{kind}_LY"], errors="coerce").fillna(0) != 0,
            merged[f"or_{kind}_VAR"] * 100 / merged[f"or_{kind}_LY"],
            np.nan,
        )

    # TOTAL row fallback: if OR still blank, kms-weighted average of depot ORs
    special = merged["DEPOT"].astype(str).str.upper().isin(["TOTAL", "REGION"])
    if special.any():
        detail = merged[~special]
        w = pd.to_numeric(detail.get("kms_CY"), errors="coerce").fillna(0) if "kms_CY" in detail.columns else None
        for kind in ["tot", "fpd", "mhl"]:
            for side in ["CY", "LY"]:
                col = f"or_{kind}_{side}"
                cur = pd.to_numeric(merged.loc[special, col], errors="coerce")
                if cur.isna().all() or (cur.fillna(0) == 0).all():
                    vals = pd.to_numeric(detail[col], errors="coerce")
                    if w is not None and w.sum() > 0 and vals.notna().any():
                        m = vals.notna() & (w > 0)
                        avg = float(np.average(vals[m], weights=w[m])) if m.any() else np.nan
                    else:
                        avg = float(vals.mean()) if vals.notna().any() else np.nan
                    merged.loc[special, col] = avg
            merged.loc[special, f"or_{kind}_VAR"] = (
                merged.loc[special, f"or_{kind}_CY"] - merged.loc[special, f"or_{kind}_LY"]
            )
            merged.loc[special, f"or_{kind}_PCT"] = np.where(
                pd.to_numeric(merged.loc[special, f"or_{kind}_LY"], errors="coerce").fillna(0) != 0,
                merged.loc[special, f"or_{kind}_VAR"] * 100 / merged.loc[special, f"or_{kind}_LY"],
                np.nan,
            )
    return merged


def add_avu_epb_columns(merged, data_cy, data_ly, fleet_map, month_key, group_col="DEPOT", product_filter="ALL"):
    """AVU = kms / (fleet * days); EPB = earnings / (fleet * days).
    kms/earn in merged are in lakhs → convert back to absolute units.
    Fleet from FLEET.parquet (by_dm / by_m). Days = unique dates in side data.
    """
    merged = merged.copy()
    fmap = fleet_map if isinstance(fleet_map, dict) else {}
    by_dm = fmap.get("by_dm", {}) or {}
    by_m = fmap.get("by_m", {}) or {}
    by_dpm = fmap.get("by_dpm", {}) or {}
    prod_sel = str(product_filter or "ALL").strip().upper()
    prod_is_all = prod_sel in ("ALL", "NONE", "", "NAN")

    def _days(data):
        if data is None or len(data) == 0 or "Date" not in getattr(data, "columns", []):
            return 0
        return int(pd.to_datetime(data["Date"], errors="coerce").dropna().dt.normalize().nunique())

    days_cy = max(1, _days(data_cy))
    days_ly = max(1, _days(data_ly))

    def _month_keys(mk):
        keys = [str(mk).strip()] if mk else []
        raw = str(mk).strip() if mk else ""
        try:
            parts = raw.replace(" ", "").split("-")
            if len(parts) >= 2:
                mon = parts[0][:3].title()
                yr = parts[1]
                if yr.isdigit():
                    if len(yr) == 2:
                        keys += [f"{mon}-{yr}", f"{mon}-20{yr}", f"{mon}-19{yr}"]
                    elif len(yr) == 4:
                        keys += [f"{mon}-{yr}", f"{mon}-{yr[-2:]}"]
        except Exception:
            pass
        try:
            dt = pd.to_datetime(str(mk), errors="coerce")
            if pd.notna(dt):
                keys += [
                    dt.strftime("%b-%Y"),
                    dt.strftime("%b-%y"),
                    f"{dt.strftime('%b')}-{dt.year}",
                    f"{dt.strftime('%b').title()}-{dt.year}",
                ]
        except Exception:
            pass
        out, seen = [], set()
        for k in keys:
            if k and k not in seen:
                seen.add(k)
                out.append(k)
        return out

    mkeys = _month_keys(month_key)

    def fleet_for(dep):
        d = str(dep).strip().upper()
        if d in ("TOTAL", "REGION", "ALL", ""):
            for k in mkeys:
                if k in by_m and by_m[k]:
                    return float(by_m[k])
            # sum unique depot totals for matching month keys only
            tot = 0.0
            for (dd, mk), v in by_dm.items():
                if str(mk) in mkeys and str(dd).upper() not in ("TOTAL", "REGION", "ALL", ""):
                    tot += float(v or 0)
            if tot:
                return tot
            return float(sum(by_m.values()) if by_m else 0)
        if not prod_is_all:
            for k in mkeys:
                v = by_dpm.get((d, prod_sel, k), 0) or 0
                if v:
                    return float(v)
                # case-insensitive product scan
                for (dd, pr, mk), vv in by_dpm.items():
                    if str(dd).upper() == d and str(mk) == k and str(pr).upper() == prod_sel:
                        if vv:
                            return float(vv)
        for k in mkeys:
            v = by_dm.get((d, k), 0) or 0
            if v:
                return float(v)
            for (dd, mk), vv in by_dm.items():
                if str(dd).upper() == d and str(mk) == k and vv:
                    return float(vv)
        return 0.0

    for side, days in (("CY", days_cy), ("LY", days_ly)):
        fleets = merged[group_col].map(fleet_for)
        # kms / earn stored in lakhs in ACT tables
        kms_abs = pd.to_numeric(merged.get(f"kms_{side}"), errors="coerce").fillna(0) * 100000.0
        earn_abs = pd.to_numeric(merged.get(f"earn_tot_{side}"), errors="coerce").fillna(0) * 100000.0
        denom = fleets * float(days)
        merged[f"avu_{side}"] = np.where(denom > 0, kms_abs / denom, np.nan)
        merged[f"epb_{side}"] = np.where(denom > 0, earn_abs / denom, np.nan)

    for base in ("avu", "epb"):
        merged[f"{base}_VAR"] = merged[f"{base}_CY"] - merged[f"{base}_LY"]
        merged[f"{base}_PCT"] = np.where(
            pd.to_numeric(merged[f"{base}_LY"], errors="coerce").fillna(0) != 0,
            merged[f"{base}_VAR"] * 100 / merged[f"{base}_LY"],
            np.nan,
        )
    return merged


def render_act_table_with_or(merged, group_col, prefix, pax_heading):
    """HTML ACT VS ACT + OR + AVU + EPB; freeze S.No+group; clickable scrollbar."""
    # groups of 4: kms, earn, 3 epk, 3 or, avu, epb, pax = 11
    n_groups = 11
    html = [
        '<div class="table-scroll-fixable">',
        '<table class="excel-table" style="border-collapse:separate;border-spacing:0;width:max-content;">',
        "<thead>",
    ]
    html.append("<tr>")
    html.append(
        '<th rowspan="2" style="position:sticky;left:0;top:0;z-index:4;background:#0369a1;color:#fff;'
        'min-width:48px;padding:6px 4px;">S.No</th>'
    )
    html.append(
        f'<th rowspan="2" style="position:sticky;left:48px;top:0;z-index:4;background:#0369a1;color:#fff;'
        f'min-width:72px;padding:6px 4px;">{group_col}</th>'
    )
    groups = [
        ("KILOMETERS (in lks.)", "#c2410c"),
        (f"{prefix} EARNINGS (in lks.)", "#047857"),
        (f"{prefix} TOT EPK", "#6b21a8"),
        (f"{prefix} FPD EPK", "#15803d"),
        (f"{prefix} MHL EPK", "#1d4ed8"),
        ("TOT OR", "#7c3aed"),
        ("FPD OR", "#0d9488"),
        ("MHL OR", "#2563eb"),
        ("AVU", "#b45309"),
        ("EPB", "#be123c"),
        (pax_heading, "#0369a1"),
    ]
    for title, bg in groups:
        html.append(
            f'<th colspan="4" style="position:sticky;top:0;z-index:3;background:{bg};color:#fff;'
            f'padding:6px 4px;text-align:center;">{title}</th>'
        )
    html.append("</tr>")
    html.append("<tr>")
    for _ in range(n_groups):
        for s in ("CY", "LY", "VAR", "% ▲/▼"):
            html.append(
                f'<th style="position:sticky;top:32px;z-index:3;background:#f1f5f9;color:#334155;'
                f'padding:4px 3px;font-size:10px;">{s}</th>'
            )
    html.append("</tr></thead><tbody>")

    sno = 0
    for _, row in merged.iterrows():
        is_total = str(row[group_col]).upper() in ("TOTAL", "REGION")
        if not is_total:
            sno += 1
        row_bg = "#e2efda" if is_total else "#ffffff"
        left_bg = "#e2efda" if is_total else "#f8fafc"
        fw = "font-weight:700;" if is_total else ""
        html.append(f'<tr style="background:{row_bg};{fw}">')
        html.append(
            f'<td style="position:sticky;left:0;z-index:2;background:{left_bg};min-width:48px;'
            f'text-align:center;">{"" if is_total else sno}</td>'
        )
        html.append(
            f'<td style="position:sticky;left:48px;z-index:2;background:{left_bg};min-width:72px;'
            f'text-align:center;">{row[group_col]}</td>'
        )
        for metric in ["kms", "earn_tot"]:
            html.append(f'<td>{fmt(row[f"{metric}_CY"])}</td>')
            html.append(f'<td>{fmt(row[f"{metric}_LY"])}</td>')
            html.append(f'<td class="{var_class(row[f"{metric}_VAR"])}">{fmt(row[f"{metric}_VAR"])}</td>')
            html.append(f'<td class="{var_class(row[f"{metric}_PCT"])}">{fmt_growth(row[f"{metric}_PCT"])}</td>')
        for epk_type in ["tot", "fpd", "mhl"]:
            html.append(f'<td>{fmt(row[f"epk_{epk_type}_CY"])}</td>')
            html.append(f'<td>{fmt(row[f"epk_{epk_type}_LY"])}</td>')
            html.append(f'<td class="{var_class(row[f"epk_{epk_type}_VAR"])}">{fmt(row[f"epk_{epk_type}_VAR"])}</td>')
            html.append(f'<td class="{var_class(row[f"epk_{epk_type}_PCT"])}">{fmt_growth(row[f"epk_{epk_type}_PCT"])}</td>')
        for or_type in ["tot", "fpd", "mhl"]:
            html.append(f'<td>{fmt(row.get(f"or_{or_type}_CY"))}</td>')
            html.append(f'<td>{fmt(row.get(f"or_{or_type}_LY"))}</td>')
            html.append(f'<td class="{var_class(row.get(f"or_{or_type}_VAR"))}">{fmt(row.get(f"or_{or_type}_VAR"))}</td>')
            html.append(f'<td class="{var_class(row.get(f"or_{or_type}_PCT"))}">{fmt_growth(row.get(f"or_{or_type}_PCT"))}</td>')
        for metric in ["avu", "epb"]:
            html.append(f'<td>{fmt(row.get(f"{metric}_CY"))}</td>')
            html.append(f'<td>{fmt(row.get(f"{metric}_LY"))}</td>')
            html.append(f'<td class="{var_class(row.get(f"{metric}_VAR"))}">{fmt(row.get(f"{metric}_VAR"))}</td>')
            html.append(f'<td class="{var_class(row.get(f"{metric}_PCT"))}">{fmt_growth(row.get(f"{metric}_PCT"))}</td>')
        html.append(f'<td>{fmt_pax(row.get("pax_CY", 0))}</td>')
        html.append(f'<td>{fmt_pax(row.get("pax_LY", 0))}</td>')
        html.append(f'<td class="{var_class(row.get("pax_VAR", 0))}">{fmt_pax(row.get("pax_VAR", 0))}</td>')
        html.append(f'<td class="{var_class(row.get("pax_PCT", 0))}">{fmt_growth(row.get("pax_PCT", 0))}</td>')
        html.append("</tr>")
    html.append("</tbody></table></div>")
    return "".join(html)


def add_or_columns_product(merged, orf_map, orf_by_prod=None, depot_filter="ALL"):
    """OR for product-wise table: OR = (EPK * 10000) / ORF
    Product row ORF:
      depot ALL  -> (REGION, PRODUCT) else mean of (depots, PRODUCT)
      depot set  -> (DEPOT, PRODUCT)
    TOTAL product row:
      depot ALL  -> (REGION, TOTAL)
      depot set  -> (DEPOT, TOTAL)
    """
    merged = merged.copy()
    depot_orf = orf_map if isinstance(orf_map, dict) else {}
    by_prod = orf_by_prod if isinstance(orf_by_prod, dict) else {}
    dep_sel = str(depot_filter).strip().upper() if depot_filter else "ALL"
    dep_is_all = dep_sel in ("ALL", "NONE", "", "NAN", "REGION")

    def lookup(dep_key, prod_key, side):
        d = str(dep_key).strip().upper()
        p = str(prod_key).strip().upper()
        candidates = [p, p.replace("-", " "), p.replace(" ", "-"), p.replace("-", "")]
        for pp in candidates:
            rec = by_prod.get((d, pp), {})
            val = rec.get(side, np.nan)
            if val is not None and not (isinstance(val, float) and np.isnan(val)):
                return float(val)
            for (dd, pr), v in by_prod.items():
                if dd == d and pr.replace(" ", "").replace("-", "") == pp.replace(" ", "").replace("-", ""):
                    val = v.get(side, np.nan)
                    if val is not None and not (isinstance(val, float) and np.isnan(val)):
                        return float(val)
        if p == "TOTAL":
            val2 = depot_orf.get(d, {}).get(side, np.nan)
            if val2 is not None and not (isinstance(val2, float) and np.isnan(val2)):
                return float(val2)
        return np.nan

    def get_orf(prod, side):
        p = str(prod).strip().upper()
        if p in ("TOTAL", "REGION"):
            if dep_is_all:
                return lookup("REGION", "TOTAL", side)
            return lookup(dep_sel, "TOTAL", side)
        if dep_is_all:
            val = lookup("REGION", p, side)
            if not (isinstance(val, float) and np.isnan(val)):
                return val
            # mean across depots for this product
            vals = []
            for (dd, pr), v in by_prod.items():
                if pr.replace(" ", "").replace("-", "") == p.replace(" ", "").replace("-", "") and dd not in ("REGION", "TOTAL"):
                    x = v.get(side, np.nan)
                    if x is not None and not (isinstance(x, float) and np.isnan(x)):
                        vals.append(float(x))
            return float(np.mean(vals)) if vals else np.nan
        return lookup(dep_sel, p, side)

    orf_cy = merged["PRODUCT"].map(lambda p: get_orf(p, "cy"))
    orf_ly = merged["PRODUCT"].map(lambda p: get_orf(p, "ly"))

    for kind in ["tot", "fpd", "mhl"]:
        epk_cy = pd.to_numeric(merged.get(f"epk_{kind}_CY"), errors="coerce")
        epk_ly = pd.to_numeric(merged.get(f"epk_{kind}_LY"), errors="coerce")
        oc = pd.to_numeric(orf_cy, errors="coerce")
        ol = pd.to_numeric(orf_ly, errors="coerce")
        merged[f"or_{kind}_CY"] = np.where(oc.fillna(0) != 0, (epk_cy * 10000) / oc, np.nan)
        merged[f"or_{kind}_LY"] = np.where(ol.fillna(0) != 0, (epk_ly * 10000) / ol, np.nan)
        merged[f"or_{kind}_VAR"] = merged[f"or_{kind}_CY"] - merged[f"or_{kind}_LY"]
        merged[f"or_{kind}_PCT"] = np.where(
            pd.to_numeric(merged[f"or_{kind}_LY"], errors="coerce").fillna(0) != 0,
            merged[f"or_{kind}_VAR"] * 100 / merged[f"or_{kind}_LY"],
            np.nan,
        )
    return merged


def render_product_table_with_or(merged, prefix, pax_heading):
    """Product-wise table: S.No, PRODUCT, SCH(3), KMS, EARN, EPK x3, OR x3, PAX. Freeze top2 + first2."""
    def _fmt_sch(v):
        try:
            if pd.isna(v) or float(v) == 0:
                return ""
            return f"{int(round(float(v))):,}"
        except Exception:
            return ""

    n_metric_groups = 9  # sch is separate 3-col; then kms earn epk*3 or*3 pax = 1+1+3+3+1 = 9 groups of 4? 
    # Actually: SCH=3 cols, then 8 groups of 4 (kms, earn, 3 epk, 3 or, pax) = 3+32 = 35 + S.No + PRODUCT
    html = [
        '<div style="max-height:70vh;overflow:auto;border:1px solid #cbd5e1;width:100%;">',
        '<table class="excel-table" style="border-collapse:separate;border-spacing:0;width:max-content;">',
        "<thead><tr>",
    ]
    html.append(
        '<th rowspan="2" style="position:sticky;left:0;top:0;z-index:6;background:#0369a1;color:#fff;min-width:48px;">S.No</th>'
    )
    html.append(
        '<th rowspan="2" style="position:sticky;left:48px;top:0;z-index:6;background:#0369a1;color:#fff;min-width:90px;">PRODUCT</th>'
    )
    groups = [
        ("NO OF SCHEDULES", "#b45309", 3),
        ("KILOMETERS (in lks.)", "#c2410c", 4),
        (f"{prefix} EARNINGS (in lks.)", "#047857", 4),
        (f"{prefix} TOT EPK", "#6b21a8", 4),
        (f"{prefix} FPD EPK", "#15803d", 4),
        (f"{prefix} MHL EPK", "#1d4ed8", 4),
        ("TOT OR", "#7c3aed", 4),
        ("FPD OR", "#0d9488", 4),
        ("MHL OR", "#2563eb", 4),
        (pax_heading, "#0369a1", 4),
    ]
    for title, bg, span in groups:
        html.append(
            f'<th colspan="{span}" style="position:sticky;top:0;z-index:5;background:{bg};color:#fff;padding:6px 4px;">{title}</th>'
        )
    html.append("</tr><tr>")
    # SCH: CY LY VAR only
    for s in ("CY", "LY", "VAR"):
        html.append(f'<th style="position:sticky;top:32px;z-index:5;background:#f1f5f9;color:#334155;font-size:10px;">{s}</th>')
    for _ in range(9):
        for s in ("CY", "LY", "VAR", "% ▲/▼"):
            html.append(f'<th style="position:sticky;top:32px;z-index:5;background:#f1f5f9;color:#334155;font-size:10px;">{s}</th>')
    html.append("</tr></thead><tbody>")

    sno = 0
    for _, row in merged.iterrows():
        is_total = str(row["PRODUCT"]).upper() in ("TOTAL", "REGION")
        if not is_total:
            sno += 1
        row_bg = "#e2efda" if is_total else "#ffffff"
        left_bg = "#e2efda" if is_total else "#f8fafc"
        fw = "font-weight:700;" if is_total else ""
        html.append(f'<tr style="background:{row_bg};{fw}">')
        html.append(f'<td style="position:sticky;left:0;z-index:3;background:{left_bg};text-align:center;">{"" if is_total else sno}</td>')
        html.append(f'<td style="position:sticky;left:48px;z-index:3;background:{left_bg};text-align:center;">{row["PRODUCT"]}</td>')
        html.append(f'<td>{_fmt_sch(row.get("sch_CY"))}</td>')
        html.append(f'<td>{_fmt_sch(row.get("sch_LY"))}</td>')
        html.append(f'<td class="{var_class(row.get("sch_VAR"))}">{_fmt_sch(row.get("sch_VAR"))}</td>')
        for metric in ["kms", "earn_tot"]:
            html.append(f'<td>{fmt(row[f"{metric}_CY"])}</td>')
            html.append(f'<td>{fmt(row[f"{metric}_LY"])}</td>')
            html.append(f'<td class="{var_class(row[f"{metric}_VAR"])}">{fmt(row[f"{metric}_VAR"])}</td>')
            html.append(f'<td class="{var_class(row[f"{metric}_PCT"])}">{fmt_growth(row[f"{metric}_PCT"])}</td>')
        for epk_type in ["tot", "fpd", "mhl"]:
            html.append(f'<td>{fmt(row[f"epk_{epk_type}_CY"])}</td>')
            html.append(f'<td>{fmt(row[f"epk_{epk_type}_LY"])}</td>')
            html.append(f'<td class="{var_class(row[f"epk_{epk_type}_VAR"])}">{fmt(row[f"epk_{epk_type}_VAR"])}</td>')
            html.append(f'<td class="{var_class(row[f"epk_{epk_type}_PCT"])}">{fmt_growth(row[f"epk_{epk_type}_PCT"])}</td>')
        for or_type in ["tot", "fpd", "mhl"]:
            html.append(f'<td>{fmt(row.get(f"or_{or_type}_CY"))}</td>')
            html.append(f'<td>{fmt(row.get(f"or_{or_type}_LY"))}</td>')
            html.append(f'<td class="{var_class(row.get(f"or_{or_type}_VAR"))}">{fmt(row.get(f"or_{or_type}_VAR"))}</td>')
            html.append(f'<td class="{var_class(row.get(f"or_{or_type}_PCT"))}">{fmt_growth(row.get(f"or_{or_type}_PCT"))}</td>')
        html.append(f'<td>{fmt_pax(row.get("pax_CY", 0))}</td>')
        html.append(f'<td>{fmt_pax(row.get("pax_LY", 0))}</td>')
        html.append(f'<td class="{var_class(row.get("pax_VAR", 0))}">{fmt_pax(row.get("pax_VAR", 0))}</td>')
        html.append(f'<td class="{var_class(row.get("pax_PCT", 0))}">{fmt_growth(row.get("pax_PCT", 0))}</td>')
        html.append("</tr>")
    html.append("</tbody></table></div>")
    return "".join(html)






# ========== CASCADING FILTERS ==========
st.markdown("<!-- spacer for streamlit header -->", unsafe_allow_html=True)
# HEADING ABOVE FILTERS
st.markdown(
    f"""<div style="background:linear-gradient(90deg,#1e3a8a,#7c3aed,#db2777);color:#fff;text-align:center;padding:16px;border-radius:10px;margin:8px 0 12px 0;box-shadow:0 4px 12px rgba(0,0,0,0.2);">
        <div style="font-size:1.6rem;font-weight:800;letter-spacing:1px;text-transform:uppercase;">HISTORICAL ANALYSIS OF RANGAREDDY REGION</div>
        <div style="font-size:0.95rem;font-weight:600;margin-top:6px;opacity:0.95;">Data updated on {DATA_UPDATED_ON}</div>
    </div>""",
    unsafe_allow_html=True,
)

# Sidebar menu labels (display) → internal section keys used by elif branches
NAV_ITEMS = [
    ("Home", "Home"),
    ("Schedules", "Schedules"),
    ("Route Daywise", "Route Day-wise"),
    ("ACT VS ACT", "ACT VS ACT"),
    ("ACT vs ACT TRENDS", "ACT vs ACT TRENDS"),
    ("Day wise", "Day wise"),
    ("Product wise", "Product wise"),
    ("Service performance", "Service performance"),
    ("Service wise SROS", "Service-wise (SROS)"),
    ("Period Comparison", "Period Comparison"),
    ("Trends from 2024", "Trends from 2024"),
    ("Task", "Task"),
    ("DOR", "DOR"),
    ("Monthly files", "Monthly files"),
    ("MISSION RR", "MISSION RR"),
]
NAV_LABELS = [x[0] for x in NAV_ITEMS]
NAV_TO_SECTION = {x[0]: x[1] for x in NAV_ITEMS}
SECTIONS = [x[1] for x in NAV_ITEMS]

with st.sidebar:
    st.markdown(
        '<div class="nav-brand">RR Region Dashboard'
        f"<small>Data updated · {DATA_UPDATED_ON}</small></div>",
        unsafe_allow_html=True,
    )
    # Default selection
    if "main_section_label" not in st.session_state:
        st.session_state["main_section_label"] = "ACT vs ACT TRENDS"
    nav_label = st.radio(
        "Menu",
        NAV_LABELS,
        index=NAV_LABELS.index(st.session_state["main_section_label"])
        if st.session_state["main_section_label"] in NAV_LABELS
        else 0,
        key="sidebar_nav_radio",
        label_visibility="collapsed",
    )
    st.session_state["main_section_label"] = nav_label
    section = NAV_TO_SECTION.get(nav_label, nav_label)
    st.caption("Click « at top to close / open menu")

st.markdown(
    '<div style="height:4px;background:linear-gradient(90deg,#f59e0b,#ef4444,#8b5cf6);border-radius:2px;margin:6px 0 12px 0;"></div>',
    unsafe_allow_html=True,
)
st.markdown(
    '<div style="height:4px;background:#e67e22;margin:-8px 0 12px 0;border-radius:2px;"></div>',
    unsafe_allow_html=True,
)

# 1 ACT VS ACT  2 ACT vs ACT TRENDS  3 Product wise  4 Service-wise (SROS)  5 Trends from 2024
# → load from ser_wise.parquet (same folder as ser_wise)
if section in SERVICE_MONTHLY_TABS:
    if df_service_monthly is not None and len(df_service_monthly) > 0:
        df = df_service_monthly
    else:
        st.warning(
            f"**ser_wise.parquet** not found for **{section}**. "
            f"Using **ser_wise.parquet**. Put `ser_wise.parquet` in the same folder "
            f"as ser_wise (e.g. `D:\\dashboard\\ser_wise.parquet`)."
        )
        df = df_ser_wise
else:
    df = df_ser_wise

if section not in ("Schedules", "Home", "DOR", "MISSION RR", "Monthly files"):
    st.markdown("<div style='font-size:13px;font-weight:600;color:#334155;margin-bottom:2px;'>Filters</div>", unsafe_allow_html=True)

    def _find_col(frame, *cands):
        """Find a column by exact or fuzzy name; return None if missing."""
        if frame is None or len(getattr(frame, "columns", [])) == 0:
            return None
        norm = {str(c).strip().lower().replace(" ", "").replace("_", ""): c for c in frame.columns}
        for cand in cands:
            k = cand.lower().replace(" ", "").replace("_", "")
            if k in norm:
                return norm[k]
        for c in frame.columns:
            cl = str(c).strip().lower().replace(" ", "").replace("_", "")
            for cand in cands:
                if cand.lower().replace(" ", "").replace("_", "") in cl:
                    return c
        return None

    # Prefer exact column names used in ser_wise / service_monthly
    def _exact_or_find(frame, exact_names, *fuzzy):
        for n in exact_names:
            if n in frame.columns:
                return n
        return _find_col(frame, *exact_names, *fuzzy)

    col_depot = _exact_or_find(df, ["DEPOT", "Depot"], "DEPOT")
    col_mhl = _exact_or_find(df, ["MHL_NMHL", "MHL/NMHL"], "MHL_NMHL", "MHL/NMHL", "MHLNMHL", "MHL")
    col_rtc = _exact_or_find(df, ["RTC_HIRE", "RTC/HIRE"], "RTC_HIRE", "RTC/HIRE", "RTCHIRE", "RTC")
    col_product = _exact_or_find(df, ["PRODUCT", "Product"], "PRODUCT")
    # Cascading Route filter + all boards: ROUTEE column only
    col_route = None
    for rn in ("ROUTEE", "Routee", "routee"):
        if rn in df.columns:
            col_route = rn
            break
    if col_route is None:
        # create ROUTEE from ROUTE only if ROUTEE truly absent
        if "ROUTE" in df.columns:
            df["ROUTEE"] = df["ROUTE"].astype(str).str.strip()
            col_route = "ROUTEE"
    col_month = _exact_or_find(df, ["Month_Name", "MONTH_NAME", "Month"], "Month_Name", "MONTH_NAME", "Month", "MONTH")
    col_date = _exact_or_find(df, ["Date", "DATE"], "Date", "DATE", "TravelDate", "TripDate")

    temp = df

    AC_PRODUCTS = {
        "AC-HBD", "AC-SLP", "GRD+", "RJD", "e-GRD", "AC-HBD R", "AC-HBD H", "AC SLP", "AC HBD",
        "AC-HBD-R", "AC-HBD-H", "ACHBD", "ACSLP", "AC HBD R", "AC HBD H",
    }

    def _is_ac_product(p):
        raw = str(p).strip().upper()
        if not raw or raw in ("NAN", "NONE", "NAT", ""):
            return False
        compact = raw.replace(" ", "").replace("-", "").replace("_", "").replace("/", "")
        # known AC product codes
        known = {x.upper().replace(" ", "").replace("-", "").replace("_", "") for x in AC_PRODUCTS}
        if compact in known:
            return True
        # any product whose code starts with AC (but not words like "ACTIVE")
        if compact.startswith("AC") and len(compact) <= 12:
            return True
        if raw.startswith("AC-") or raw.startswith("AC ") or raw.startswith("AC/"):
            return True
        return False

    def _rtc_match_mask(series, rtc_val):
        s = series.astype(str).str.strip().str.upper()
        if rtc_val == "RTC":
            return (s == "RTC") | (s == "R") | s.str.startswith("RTC")
        if rtc_val == "HIRE":
            # H, HIRE, HIRED, HIRE BUS, etc.
            return (
                (s == "H")
                | s.str.startswith("HIRE")
                | s.str.contains(r"\bHIRE", regex=True, na=False)
            )
        return pd.Series(True, index=series.index)

    def _nat_sort_key(s):
        """Sort HYD1, HYD2, PKT… naturally."""
        import re
        parts = re.split(r"(\d+)", str(s).upper())
        return [int(p) if p.isdigit() else p for p in parts]

    c1, c2, c3, c4, c5, c6, c7, c8, c9, c10 = st.columns(10)

    with c1:
        # Depot list always from full frame (ser_wise / monthly) — HYD1, HYD2, PKT…
        if col_depot:
            _deps = [
                str(x).strip()
                for x in df[col_depot].dropna().unique()
                if str(x).strip() and str(x).strip().lower() not in ("nan", "region", "all", "none", "total", "")
            ]
            try:
                _deps = sorted(set(_deps), key=_nat_sort_key)
            except Exception:
                _deps = sorted(set(_deps), key=lambda z: str(z).upper())
            depot_opts = ["ALL", "REGION"] + _deps
        else:
            depot_opts = ["ALL"]
        depot = st.selectbox("DEPOT", depot_opts, index=0)
    if col_depot and depot not in ("ALL", "REGION"):
        temp = temp[temp[col_depot].astype(str).str.strip().str.upper() == str(depot).strip().upper()]

    with c2:
        if col_mhl:
            mhl_opts = ["ALL"] + sorted([str(x).strip() for x in temp[col_mhl].dropna().unique() if str(x).strip() and str(x).lower() != "nan"])
        else:
            mhl_opts = ["ALL"]
        mhl = st.selectbox("MHL / NMHL", mhl_opts, index=0)
    if col_mhl and mhl != "ALL":
        temp = temp[temp[col_mhl].astype(str).str.strip().str.upper() == str(mhl).strip().upper()]

    with c3:
        rtc_opts = ["ALL", "RTC", "HIRE"]
        rtc = st.selectbox("RTC / HIRE", rtc_opts, index=0)
    if col_rtc and rtc != "ALL":
        temp = temp[_rtc_match_mask(temp[col_rtc], rtc)]

    with c4:
        pax_opts = ["TOT", "FPD", "MHL"]
        passengers = st.selectbox("PASSENGERS", pax_opts, index=0)

    with c5:
        ac_opts = ["ALL", "AC", "NON-AC"]
        ac_type = st.selectbox("AC / NON-AC", ac_opts, index=0)
    if col_product and ac_type != "ALL":
        _ac_mask = temp[col_product].map(_is_ac_product)
        if ac_type == "AC":
            temp = temp[_ac_mask]
        else:
            temp = temp[~_ac_mask]

    with c6:
        if col_product:
            product_opts = ["ALL"] + sorted([str(x).strip() for x in temp[col_product].dropna().unique() if str(x).strip() and str(x).lower() != "nan"])
        else:
            product_opts = ["ALL"]
        product = st.selectbox("PRODUCT", product_opts, index=0)
    if col_product and product != "ALL":
        temp = temp[temp[col_product].astype(str).str.strip().str.upper() == str(product).strip().upper()]

    with c7:
        # Route options from ROUTEE on the cascading frame
        if col_route and col_route in temp.columns:
            route_opts = ["ALL"] + sorted(
                [str(x).strip() for x in temp[col_route].dropna().unique() if str(x).strip() and str(x).lower() not in ("nan", "none", "")],
                key=lambda z: str(z).upper(),
            )
        else:
            route_opts = ["ALL"]
        route = st.selectbox("ROUTE (ROUTEE)", route_opts, index=0)
    if col_route and route != "ALL":
        temp = temp[temp[col_route].astype(str).str.strip().str.upper() == str(route).strip().upper()]

    with c8:
        if col_month:
            raw_months = [x for x in temp[col_month].dropna().unique() if str(x).strip()]
            if not raw_months:
                raw_months = [x for x in df[col_month].dropna().unique() if str(x).strip()]
        else:
            raw_months = []
        def parse_month_key(m_str):
            try:
                return pd.to_datetime(m_str, format="%b-%y")
            except Exception:
                try:
                    return pd.to_datetime(m_str, format="%b-%Y")
                except Exception:
                    return pd.to_datetime(m_str, errors="coerce")
        month_opts = sorted(raw_months, key=parse_month_key, reverse=True) if raw_months else ["ALL"]
        month = st.selectbox("MONTH", month_opts, index=0)

    with c9:
        for_upto = st.selectbox("For / Upto", ["UPTO", "FOR"], index=0)

    with c10:
        net_gross = st.selectbox("NET / GROSS", ["Gross", "Net"], index=0)

    st.markdown("<div style='margin-top:2px;margin-bottom:2px;'></div>", unsafe_allow_html=True)

    # ========== GLOBAL FILTERING LOGIC ==========
    # Memoize CY/LY frames in session_state so switching tabs with the same
    # filters does not re-scan the full parquet (logic & calculations unchanged).
    _src_tag = "svc_monthly" if section in SERVICE_MONTHLY_TABS else "ser_wise"
    _filter_key = (
        _src_tag,
        str(depot), str(mhl), str(rtc), str(product), str(route),
        str(ac_type), str(month), str(for_upto), str(passengers), str(net_gross),
        len(df),
    )
    # Only cache the boolean mask/date metadata in session_state.  Storing full
    # CY/LY DataFrames here duplicates tens/hundreds of MB per user on Cloud.
    # Do not keep 450K-row CY/LY DataFrames in per-user session state.
    # Rebuild the small boolean masks on each rerun; this trades a cheap scan for
    # a large and persistent RAM allocation per Cloud user.
    _reuse = False

    # Resolve month / date columns safely (cheap; needed for aliases either way)
    _mcol = col_month if col_month else _find_col(df, "Month_Name", "MONTH_NAME", "Month", "MONTH")
    _dcol = col_date if col_date else _find_col(df, "Date", "DATE", "TravelDate")

    def _month_match(series, mval):
        return series.astype(str).str.strip() == str(mval).strip()

    if _reuse:
        base_mask = st.session_state["base_mask"]
        selected_max_date = st.session_state.get("selected_max_date", pd.NaT)
    else:
        base_mask = pd.Series(True, index=df.index)

        if depot not in ("ALL", "REGION") and col_depot:
            base_mask &= df[col_depot].astype(str).str.strip().str.upper() == str(depot).strip().upper()

        if mhl != "ALL" and col_mhl:
            base_mask &= df[col_mhl].astype(str).str.strip().str.upper() == str(mhl).strip().upper()

        if route != "ALL" and col_route:
            base_mask &= df[col_route].astype(str).str.strip().str.upper() == str(route).strip().upper()

        if product != "ALL" and col_product:
            base_mask &= df[col_product].astype(str).str.strip().str.upper() == str(product).strip().upper()

        if rtc != "ALL" and col_rtc:
            base_mask &= _rtc_match_mask(df[col_rtc], rtc)

        if ac_type != "ALL" and col_product:
            _ac_full = df[col_product].map(_is_ac_product)
            if ac_type == "AC":
                base_mask &= _ac_full
            else:
                base_mask &= ~_ac_full

        if _mcol and _dcol:
            selected_max_date = pd.to_datetime(df.loc[df[_mcol].astype(str).str.strip() == str(month).strip(), _dcol], errors="coerce").max()
        elif _dcol:
            selected_max_date = pd.to_datetime(df[_dcol], errors="coerce").max()
        else:
            selected_max_date = pd.NaT

        if pd.isna(selected_max_date):
            if _mcol:
                cy_mask = base_mask & _month_match(df[_mcol], month)
            else:
                cy_mask = base_mask.copy()
            ly_mask = pd.Series(False, index=df.index)
        else:
            cy_year = selected_max_date.year
            cy_month_num = selected_max_date.month
            fy_start_year = cy_year if cy_month_num >= 4 else cy_year - 1
            fy_start_date = pd.Timestamp(year=fy_start_year, month=4, day=1)

            if for_upto == "FOR":
                if _mcol:
                    cy_mask = base_mask & _month_match(df[_mcol], month)
                else:
                    cy_mask = base_mask.copy()
            else:
                if _dcol:
                    _dt = pd.to_datetime(df[_dcol], errors="coerce")
                    cy_mask = base_mask & (_dt >= fy_start_date) & (_dt <= selected_max_date)
                elif _mcol:
                    cy_mask = base_mask & _month_match(df[_mcol], month)
                else:
                    cy_mask = base_mask.copy()

            ly_max_date = selected_max_date - pd.DateOffset(years=1)
            ly_fy_start_date = fy_start_date - pd.DateOffset(years=1)

            if for_upto == "FOR":
                try:
                    mon_name, yr = str(month).split("-")[0], int(str(month).split("-")[1])
                    candidates = [f"{mon_name}-{yr-1}", f"{mon_name}-{str(yr-1)[-2:]}"]
                    if len(str(yr)) == 2:
                        candidates.append(f"{mon_name}-{2000+yr-1}")
                    if _mcol:
                        ly_mask = pd.Series(False, index=df.index)
                        for cand in candidates:
                            ly_mask = ly_mask | _month_match(df[_mcol], cand)
                        ly_mask = base_mask & ly_mask
                    else:
                        ly_mask = pd.Series(False, index=df.index)
                    # Cap LY at same day-of-month as available CY data (e.g. CY to 22-08-2026 → LY to 22-08-2025)
                    if _dcol and pd.notna(ly_max_date):
                        _dt_ly = pd.to_datetime(df[_dcol], errors="coerce")
                        ly_mask = ly_mask & (_dt_ly <= ly_max_date)
                    # Also cap CY to selected_max_date so average is consistent
                    if _dcol and pd.notna(selected_max_date):
                        _dt_cy = pd.to_datetime(df[_dcol], errors="coerce")
                        cy_mask = cy_mask & (_dt_cy <= selected_max_date)
                except Exception:
                    ly_mask = pd.Series(False, index=df.index)
            else:
                if _dcol:
                    _dt = pd.to_datetime(df[_dcol], errors="coerce")
                    ly_mask = base_mask & (_dt >= ly_fy_start_date) & (_dt <= ly_max_date)
                else:
                    ly_mask = pd.Series(False, index=df.index)

        cy_data = df[cy_mask].copy() if cy_mask is not None else df.iloc[0:0].copy()
        ly_data = df[ly_mask].copy() if ly_mask is not None else df.iloc[0:0].copy()


        st.session_state["_cy_ly_filter_key"] = _filter_key
        st.session_state["base_mask"] = base_mask
        st.session_state["selected_max_date"] = selected_max_date

    if net_gross == "Gross":

        earn_tot, earn_fpd, earn_mhl = "GE_TOT", "GE_FPD", "GE_MHL"

        prefix = "Gross"

    else:

        earn_tot, earn_fpd, earn_mhl = "NE_TOT", "NE_FPD", "NE_MHL"

        prefix = "Net"



    # Passenger column based on PASSENGERS filter

    if passengers == "FPD":

        pax_col = "PSNGR_FPD"

    elif passengers == "MHL":

        pax_col = "PSNGR_MHL"

    else:

        pax_col = "PSNGR_TOT"  # BOTH or TOT



    # Service column name for Service Performance tab only

    service_col = next((col for col in ["SER_NO", "SERVICE_NO", "SERVICE", "SERVICE_NUMBER", "ServiceNo"] if col in df.columns), None)



    # ========== HELPER FUNCTIONS ==========

    day_order = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]

    day_short = ["MON", "TUE", "WED", "THU", "FRI", "SAT", "SUN"]



    def weighted_epk(data, earn_col, group_cols=["DEPOT", "ROUTEE", "PRODUCT"]):

        if len(data) == 0:

            return pd.DataFrame()

        data = data.copy()

        # Normalize Weekday to Mon/Tue/... so pivot columns match day_order
        _wd_full = {
            "MONDAY": "Mon", "TUESDAY": "Tue", "WEDNESDAY": "Wed", "THURSDAY": "Thu",
            "FRIDAY": "Fri", "SATURDAY": "Sat", "SUNDAY": "Sun",
            "MON": "Mon", "TUE": "Tue", "WED": "Wed", "THU": "Thu",
            "FRI": "Fri", "SAT": "Sat", "SUN": "Sun",
        }
        need_from_date = True
        if "Weekday" in data.columns:
            s = data["Weekday"].astype(str).str.strip()
            non_empty = s[~s.str.lower().isin(["", "nan", "none", "nat", "<na>"])]
            if len(non_empty) > 0:
                need_from_date = False
                mapped = s.str.upper().map(_wd_full)
                fallback = s.str[:3].str.title()
                data["Weekday"] = mapped.fillna(fallback)
                data["Weekday"] = data["Weekday"].where(
                    data["Weekday"].isin(day_order),
                    fallback,
                )
        if need_from_date:
            if "Date" in data.columns:
                _dt = pd.to_datetime(data["Date"], errors="coerce")
                data["Weekday"] = _dt.dt.strftime("%a")
            else:
                data["Weekday"] = ""

        data = data[data["Weekday"].isin(day_order)].copy()
        if len(data) == 0:
            return pd.DataFrame()

        g = (

            data.groupby(group_cols + ["Weekday"])

            .agg(earnings=(earn_col, "sum"), kms=("Optd_KMs", "sum"))

            .reset_index()

        )

        g["epk"] = np.where(g["kms"] > 0, g["earnings"] / g["kms"], np.nan)

        pivot = g.pivot_table(

            index=group_cols,

            columns="Weekday",

            values="epk",

            aggfunc="first",

        ).reindex(columns=day_order)

        overall = data.groupby(group_cols).agg(

            earnings=(earn_col, "sum"), kms=("Optd_KMs", "sum")

        )

        # Align to the pivot index explicitly. This preserves the calculation
        # while avoiding pandas MultiIndex assignment/reindex errors.
        _overall_kms = pd.to_numeric(overall["kms"], errors="coerce")
        _overall_earn = pd.to_numeric(overall["earnings"], errors="coerce")
        _overall_epk = (_overall_earn / _overall_kms.replace(0, np.nan))
        _overall_epk = _overall_epk.reindex(pivot.index)
        pivot["UPTO"] = _overall_epk.to_numpy()

        return pivot.round(2)



    def fmt(v):

        if pd.isna(v) or v is None:

            return ""

        try:

            if float(v) == 0:

                return ""

        except Exception:

            pass

        return f"{v:,.2f}"



    def fmt_pax(v):

        """Passengers whole number, no decimals; blank if 0"""

        if pd.isna(v) or v is None:

            return ""

        try:

            iv = int(round(float(v)))

            if iv == 0:

                return ""

            return f"{iv:,}"

        except Exception:

            return ""



    def fmt_growth(v):

        """% Growth number only; blank if 0"""

        if pd.isna(v) or v is None:

            return ""

        try:

            if float(v) == 0:

                return ""

        except Exception:

            pass

        return f"{v:.2f}"



    def var_class(v):

        if pd.isna(v) or v is None:

            return ""

        if v > 0:

            return "pos"

        if v < 0:

            return "neg"

        return ""



    def build_act_vs_act_table(group_col, data_cy=None, data_ly=None, cy_label="CY", ly_label="LY"):
        if data_cy is None:
            data_cy = cy_data
        if data_ly is None:
            data_ly = ly_data
        if len(data_cy) == 0 and len(data_ly) == 0:
            return None, pd.DataFrame()
        def agg_summary(data):
            if len(data) == 0:
                return pd.DataFrame(columns=[group_col, "kms", "earn_tot", "earn_fpd", "earn_mhl", "pax"])
            g = (
                data.groupby(group_col)
                .agg(
                    kms=("Optd_KMs", "sum"),
                    earn_tot=(earn_tot, "sum"),
                    earn_fpd=(earn_fpd, "sum"),
                    earn_mhl=(earn_mhl, "sum"),
                    pax=(pax_col, "sum"),
                )
                .reset_index()
            )
            return g
        cy_sum = agg_summary(data_cy)
        ly_sum = agg_summary(data_ly)
        merged = cy_sum.merge(ly_sum, on=group_col, how="outer", suffixes=("_CY", "_LY"))
        for col in ["kms_CY", "kms_LY", "earn_tot_CY", "earn_tot_LY", "earn_fpd_CY", "earn_fpd_LY", "earn_mhl_CY", "earn_mhl_LY", "pax_CY", "pax_LY"]:
            merged[col] = merged.get(col, pd.Series(0, index=merged.index)).fillna(0)
        # Ensure numeric types (pax can become object/str after merge)
        for col in ["kms_CY", "kms_LY", "earn_tot_CY", "earn_tot_LY", "earn_fpd_CY", "earn_fpd_LY",
                    "earn_mhl_CY", "earn_mhl_LY", "pax_CY", "pax_LY"]:
            if col not in merged.columns:
                merged[col] = 0.0
            merged[col] = pd.to_numeric(merged[col], errors="coerce").fillna(0.0)
        merged["pax_VAR"] = merged["pax_CY"] - merged["pax_LY"]
        merged["pax_PCT"] = np.where(merged["pax_LY"] != 0, merged["pax_VAR"] * 100 / merged["pax_LY"], np.nan)
        # Passengers as whole numbers (no lakhs, no decimals)
        for col in ["pax_CY", "pax_LY", "pax_VAR"]:
            merged[col] = pd.to_numeric(merged[col], errors="coerce").fillna(0).round(0)
        merged["epk_tot_CY"] = np.where(merged["kms_CY"] > 0, merged["earn_tot_CY"] / merged["kms_CY"], np.nan)
        merged["epk_tot_LY"] = np.where(merged["kms_LY"] > 0, merged["earn_tot_LY"] / merged["kms_LY"], np.nan)
        merged["epk_fpd_CY"] = np.where(merged["kms_CY"] > 0, merged["earn_fpd_CY"] / merged["kms_CY"], np.nan)
        merged["epk_fpd_LY"] = np.where(merged["kms_LY"] > 0, merged["earn_fpd_LY"] / merged["kms_LY"], np.nan)
        merged["epk_mhl_CY"] = np.where(merged["kms_CY"] > 0, merged["earn_mhl_CY"] / merged["kms_CY"], np.nan)
        merged["epk_mhl_LY"] = np.where(merged["kms_LY"] > 0, merged["earn_mhl_LY"] / merged["kms_LY"], np.nan)
        merged["kms_VAR"] = merged["kms_CY"] - merged["kms_LY"]
        merged["kms_PCT"] = np.where(merged["kms_LY"] != 0, merged["kms_VAR"] * 100 / merged["kms_LY"], np.nan)
        merged["earn_tot_VAR"] = merged["earn_tot_CY"] - merged["earn_tot_LY"]
        merged["earn_tot_PCT"] = np.where(merged["earn_tot_LY"] != 0, merged["earn_tot_VAR"] * 100 / merged["earn_tot_LY"], np.nan)
        merged["epk_tot_VAR"] = merged["epk_tot_CY"] - merged["epk_tot_LY"]
        merged["epk_tot_PCT"] = np.where((merged["epk_tot_LY"].notna()) & (merged["epk_tot_LY"] != 0), merged["epk_tot_VAR"] * 100 / merged["epk_tot_LY"], np.nan)
        merged["epk_fpd_VAR"] = merged["epk_fpd_CY"] - merged["epk_fpd_LY"]
        merged["epk_fpd_PCT"] = np.where((merged["epk_fpd_LY"].notna()) & (merged["epk_fpd_LY"] != 0), merged["epk_fpd_VAR"] * 100 / merged["epk_fpd_LY"], np.nan)
        merged["epk_mhl_VAR"] = merged["epk_mhl_CY"] - merged["epk_mhl_LY"]
        merged["epk_mhl_PCT"] = np.where((merged["epk_mhl_LY"].notna()) & (merged["epk_mhl_LY"] != 0), merged["epk_mhl_VAR"] * 100 / merged["epk_mhl_LY"], np.nan)
        for col in ["kms_CY", "kms_LY", "kms_VAR", "earn_tot_CY", "earn_tot_LY", "earn_tot_VAR"]:
            merged[col] = merged[col] / 100000
        merged = merged.sort_values(group_col).reset_index(drop=True)
        t_kms_cy = data_cy["Optd_KMs"].sum() / 100000 if len(data_cy) else 0
        t_kms_ly = data_ly["Optd_KMs"].sum() / 100000 if len(data_ly) else 0
        t_earn_cy = data_cy[earn_tot].sum() / 100000 if len(data_cy) else 0
        t_earn_ly = data_ly[earn_tot].sum() / 100000 if len(data_ly) else 0
        t_epk_tot_cy = (data_cy[earn_tot].sum() / data_cy["Optd_KMs"].sum()) if len(data_cy) and data_cy["Optd_KMs"].sum() > 0 else np.nan
        t_epk_tot_ly = (data_ly[earn_tot].sum() / data_ly["Optd_KMs"].sum()) if len(data_ly) and data_ly["Optd_KMs"].sum() > 0 else np.nan
        t_epk_fpd_cy = (data_cy[earn_fpd].sum() / data_cy["Optd_KMs"].sum()) if len(data_cy) and data_cy["Optd_KMs"].sum() > 0 else np.nan
        t_epk_fpd_ly = (data_ly[earn_fpd].sum() / data_ly["Optd_KMs"].sum()) if len(data_ly) and data_ly["Optd_KMs"].sum() > 0 else np.nan
        t_epk_mhl_cy = (data_cy[earn_mhl].sum() / data_cy["Optd_KMs"].sum()) if len(data_cy) and data_cy["Optd_KMs"].sum() > 0 else np.nan
        t_epk_mhl_ly = (data_ly[earn_mhl].sum() / data_ly["Optd_KMs"].sum()) if len(data_ly) and data_ly["Optd_KMs"].sum() > 0 else np.nan
        kms_v = t_kms_cy - t_kms_ly
        earn_v = t_earn_cy - t_earn_ly
        epk_tot_v = (t_epk_tot_cy - t_epk_tot_ly) if pd.notna(t_epk_tot_cy) and pd.notna(t_epk_tot_ly) else np.nan
        epk_fpd_v = (t_epk_fpd_cy - t_epk_fpd_ly) if pd.notna(t_epk_fpd_cy) and pd.notna(t_epk_fpd_ly) else np.nan
        epk_mhl_v = (t_epk_mhl_cy - t_epk_mhl_ly) if pd.notna(t_epk_mhl_cy) and pd.notna(t_epk_mhl_ly) else np.nan
        total = {
            group_col: "TOTAL",
            "kms_CY": t_kms_cy,
            "kms_LY": t_kms_ly,
            "kms_VAR": kms_v,
            "kms_PCT": (kms_v * 100 / t_kms_ly) if t_kms_ly else np.nan,
            "earn_tot_CY": t_earn_cy,
            "earn_tot_LY": t_earn_ly,
            "earn_tot_VAR": earn_v,
            "earn_tot_PCT": (earn_v * 100 / t_earn_ly) if t_earn_ly else np.nan,
            "epk_tot_CY": t_epk_tot_cy,
            "epk_tot_LY": t_epk_tot_ly,
            "epk_tot_VAR": epk_tot_v,
            "epk_tot_PCT": (epk_tot_v * 100 / t_epk_tot_ly) if (pd.notna(t_epk_tot_ly) and t_epk_tot_ly != 0) else np.nan,
            "epk_fpd_CY": t_epk_fpd_cy,
            "epk_fpd_LY": t_epk_fpd_ly,
            "epk_fpd_VAR": epk_fpd_v,
            "epk_fpd_PCT": (epk_fpd_v * 100 / t_epk_fpd_ly) if (pd.notna(t_epk_fpd_ly) and t_epk_fpd_ly != 0) else np.nan,
            "epk_mhl_CY": t_epk_mhl_cy,
            "epk_mhl_LY": t_epk_mhl_ly,
            "epk_mhl_VAR": epk_mhl_v,
            "epk_mhl_PCT": (epk_mhl_v * 100 / t_epk_mhl_ly) if (pd.notna(t_epk_mhl_ly) and t_epk_mhl_ly != 0) else np.nan,
            "pax_CY": round(data_cy[pax_col].sum() if len(data_cy) else 0),
            "pax_LY": round(data_ly[pax_col].sum() if len(data_ly) else 0),
            "pax_VAR": 0,
            "pax_PCT": np.nan,
        }
        total["pax_VAR"] = total["pax_CY"] - total["pax_LY"]
        if total["pax_LY"]:
            total["pax_PCT"] = total["pax_VAR"] * 100 / total["pax_LY"]
        merged = pd.concat([merged, pd.DataFrame([total])], ignore_index=True)
        html = ['<div class="table-scroll"><table class="excel-table"><thead>']
        html.append("<tr>")
        html.append('<th class="header-left" rowspan="2">S.No</th>')
        html.append(f'<th class="header-left" rowspan="2">{group_col}</th>')
        html.append('<th class="header-km" colspan="4">KILOMETERS (in lks.)</th>')
        html.append(f'<th class="header-earn" colspan="4">{prefix} EARNINGS (in lks.)</th>')
        html.append(f'<th class="header-tot" colspan="4">{prefix} TOT EPK</th>')
        html.append(f'<th class="header-fpd" colspan="4">{prefix} FPD EPK</th>')
        html.append(f'<th class="header-mhl" colspan="4">{prefix} MHL EPK</th>')
        pax_heading = {"FPD": "FPD PASSENGERS", "MHL": "MHL PASSENGERS"}.get(passengers, "TOTAL PASSENGERS")
        html.append(f'<th class="header-left" colspan="4">{pax_heading}</th>')
        html.append("</tr><tr>")
        for _ in range(6):
            html.append(f'<th class="header-sub">{cy_label}</th>')
            html.append(f'<th class="header-sub">{ly_label}</th>')
            html.append('<th class="header-sub">VAR</th>')
            html.append('<th class="header-sub">% ▲/▼</th>')
        html.append("</tr>")
        for sno, (_, row) in enumerate(merged.iterrows(), 1):
            is_total = row[group_col] == "TOTAL"
            style = "font-weight:bold; background:#e2efda;" if is_total else ""
            html.append(f'<tr style="{style}">')
            html.append(f'<td>{"" if is_total else sno}</td>')
            html.append(f'<td>{row[group_col]}</td>')
            for metric in ["kms", "earn_tot"]:
                html.append(f'<td>{fmt(row[f"{metric}_CY"])}</td>')
                html.append(f'<td>{fmt(row[f"{metric}_LY"])}</td>')
                html.append(f'<td class="{var_class(row[f"{metric}_VAR"])}">{fmt(row[f"{metric}_VAR"])}</td>')
                html.append(f'<td class="{var_class(row[f"{metric}_PCT"])}">{fmt_growth(row[f"{metric}_PCT"])}</td>')
            for epk_type in ["tot", "fpd", "mhl"]:
                html.append(f'<td>{fmt(row[f"epk_{epk_type}_CY"])}</td>')
                html.append(f'<td>{fmt(row[f"epk_{epk_type}_LY"])}</td>')
                html.append(f'<td class="{var_class(row[f"epk_{epk_type}_VAR"])}">{fmt(row[f"epk_{epk_type}_VAR"])}</td>')
                html.append(f'<td class="{var_class(row[f"epk_{epk_type}_PCT"])}">{fmt_growth(row[f"epk_{epk_type}_PCT"])}</td>')
            # Passengers (already in lakhs, rounded)
            html.append(f'<td>{fmt_pax(row.get("pax_CY", 0))}</td>')
            html.append(f'<td>{fmt_pax(row.get("pax_LY", 0))}</td>')
            html.append(f'<td class="{var_class(row.get("pax_VAR", 0))}">{fmt_pax(row.get("pax_VAR", 0))}</td>')
            html.append(f'<td class="{var_class(row.get("pax_PCT", 0))}">{fmt_growth(row.get("pax_PCT", 0))}</td>')
            html.append("</tr>")
        html.append("</table></div>")
        return "".join(html), merged
    # ==================== TAB 1 ====================
else:
    # Defaults so Schedules / Home / DOR do not recompute heavy filters
    depot = mhl = rtc = product = route = month = "ALL"
    passengers = "TOT"
    ac_type = "ALL"
    for_upto = "UPTO"
    net_gross = "Gross"
    service_col = next((col for col in ["SER_NO", "SERVICE_NO", "SERVICE", "SERVICE_NUMBER"] if col in df.columns), None)
    service_no = "ALL"
    prefix = "Gross"
    earn_tot, earn_fpd, earn_mhl = "GE_TOT", "GE_FPD", "GE_MHL"
    pax_col = "PSNGR_TOT"
    cy_data = df.iloc[0:0].copy()
    ly_data = df.iloc[0:0].copy()
    base_mask = pd.Series(True, index=df.index)
    selected_max_date = pd.NaT
    col_depot = col_mhl = col_rtc = col_product = col_route = col_month = col_date = None

if section == "Home":
    st.markdown(
        f"""<div class="title-bar" style="font-size:16px;padding:14px;">
        Welcome — Historical Analysis of Rangareddy Region<br>
        <span style="font-weight:600;font-size:13px;">Data updated on {DATA_UPDATED_ON}</span>
        </div>""",
        unsafe_allow_html=True,
    )
    st.markdown(
        """
        Use the **left menu** to open a report. Click the **«** control at the top of the sidebar to collapse or expand the menu.

        | Menu item | Description |
        |---|---|
        | **Schedules** | SCHs / Services / SCH KMs |
        | **Route Daywise** | Route × weekday EPK |
        | **ACT VS ACT** | Depot actuals CY vs LY (+ OR) |
        | **ACT vs ACT TRENDS** | Month-wise trends |
        | **Day wise** | Weekday performance |
        | **Product wise** | Product comparison |
        | **Service performance** | Service metrics |
        | **Service wise SROS** | SROS service-wise |
        | **Period Comparison** | Custom period compare |
        | **Trends from 2024** | Long-run trends |
        | **Task** | Daily depot performance |
        | **DOR** | Daily operating report |
        | **Monthly files** | Depot route / product trend / service / range / sector boards + Excel |
        """
    )

elif section == "DOR":
    st.markdown('<div class="title-bar">DOR — Daily Operating Report</div>', unsafe_allow_html=True)

    # ---- Formatters so render_act_table_with_or works ----
    def fmt(v):
        try:
            if v is None or (isinstance(v, float) and pd.isna(v)):
                return ""
            fv = float(v)
            if abs(fv) < 1e-12:
                return ""
            return f"{fv:,.2f}"
        except Exception:
            return ""

    def fmt_pax(v):
        try:
            if v is None or (isinstance(v, float) and pd.isna(v)):
                return ""
            return f"{int(round(float(v))):,}"
        except Exception:
            return ""

    def fmt_growth(v):
        try:
            if v is None or (isinstance(v, float) and pd.isna(v)):
                return ""
            fv = float(v)
            if abs(fv) < 1e-12:
                return ""
            return f"{fv:,.2f}%"
        except Exception:
            return ""

    def var_class(v):
        try:
            if v is None or (isinstance(v, float) and pd.isna(v)):
                return ""
            fv = float(v)
            if fv > 0:
                return "pos"
            if fv < 0:
                return "neg"
        except Exception:
            pass
        return ""

    globals()["fmt"] = fmt
    globals()["fmt_pax"] = fmt_pax
    globals()["fmt_growth"] = fmt_growth
    globals()["var_class"] = var_class

    # ---- Resolve columns from ser_wise ----
    _src = df_ser_wise if (df_ser_wise is not None and len(df_ser_wise)) else df
    if _src is None or len(_src) == 0:
        st.error("ser_wise.parquet not available for DOR.")
        st.stop()
    df = _src.copy()

    def _find_col_dor(frame, *cands):
        if frame is None or len(getattr(frame, "columns", [])) == 0:
            return None
        norm = {str(c).strip().lower().replace(" ", "").replace("_", ""): c for c in frame.columns}
        for cand in cands:
            k = cand.lower().replace(" ", "").replace("_", "")
            if k in norm:
                return norm[k]
        for c in frame.columns:
            cl = str(c).strip().lower().replace(" ", "").replace("_", "")
            for cand in cands:
                if cand.lower().replace(" ", "").replace("_", "") in cl:
                    return c
        return None

    col_depot = _find_col_dor(df, "DEPOT")
    col_mhl = _find_col_dor(df, "MHL_NMHL", "MHL/NMHL", "MHLNMHL")
    col_rtc = _find_col_dor(df, "RTC_HIRE", "RTC/HIRE", "RTCHIRE")
    col_product = _find_col_dor(df, "PRODUCT")
    col_route = None
    for rn in ("ROUTEE", "Routee", "routee", "ROUTE"):
        if rn in df.columns:
            col_route = rn
            break
    if col_route is None:
        col_route = _find_col_dor(df, "ROUTEE", "ROUTE")
    col_date = _find_col_dor(df, "Date", "DATE")
    if col_date is None:
        st.error("No Date column found – DOR needs a date field in ser_wise.parquet.")
        st.stop()
    if "Date" not in df.columns and col_date:
        df["Date"] = pd.to_datetime(df[col_date], errors="coerce")
        col_date = "Date"
    else:
        df["Date"] = pd.to_datetime(df[col_date], errors="coerce")
        col_date = "Date"

    for std, alts in {
        "Optd_KMs": ["OPD_KMS", "Optd_KMs", "optd_kms"],
        "GE_TOT": ["Gross Total", "GE_TOT"],
        "GE_FPD": ["Gross Fare Paid", "GE_FPD"],
        "GE_MHL": ["Gross MHL", "GE_MHL"],
        "NE_TOT": ["Net Total", "NE_TOT"],
        "NE_FPD": ["Net Fare Paid", "NE_FPD"],
        "NE_MHL": ["Net MHL", "NE_MHL"],
        "PSNGR_TOT": ["Passengers Total", "PSNGR_TOT"],
        "PSNGR_FPD": ["Passengers Fare Paid", "PSNGR_FPD"],
        "PSNGR_MHL": ["Passengers MHL", "PSNGR_MHL"],
        "DEPOT": ["DEPOT"],
    }.items():
        if std not in df.columns:
            src = next((a for a in alts if a in df.columns), None)
            if src is not None:
                if std.startswith(("GE_", "NE_", "Optd", "PSNGR")):
                    df[std] = pd.to_numeric(df[src], errors="coerce").fillna(0)
                else:
                    df[std] = df[src]
            else:
                df[std] = 0 if std.startswith(("GE_", "NE_", "Optd", "PSNGR")) else ""

    _all_dt = pd.to_datetime(df[col_date], errors="coerce")
    _data_min = _all_dt.min()
    _data_max = _all_dt.max()
    if pd.isna(_data_max):
        st.error("No valid dates in data.")
        st.stop()

    _default = _data_max.date() if hasattr(_data_max, "date") else pd.Timestamp(_data_max).date()
    _min_d = _data_min.date() if pd.notna(_data_min) and hasattr(_data_min, "date") else None
    _max_d = _data_max.date() if hasattr(_data_max, "date") else None

    # ---- DOR filters (non-depot cascading) ----
    st.markdown(
        "<div style='font-size:13px;font-weight:600;color:#334155;margin:8px 0 2px 0;'>DOR filters</div>",
        unsafe_allow_html=True,
    )
    fc1, fc2, fc3, fc4, fc5, fc6 = st.columns(6)
    with fc1:
        if col_mhl:
            mhl_opts = ["ALL"] + sorted(
                [str(x).strip() for x in df[col_mhl].dropna().unique()
                 if str(x).strip() and str(x).lower() != "nan"]
            )
        else:
            mhl_opts = ["ALL"]
        mhl = st.selectbox("MHL / NMHL", mhl_opts, index=0, key="dor_mhl")
    with fc2:
        rtc = st.selectbox("RTC / HIRE", ["ALL", "RTC", "HIRE"], index=0, key="dor_rtc")
    with fc3:
        if col_product:
            product_opts = ["ALL"] + sorted(
                [str(x).strip() for x in df[col_product].dropna().unique()
                 if str(x).strip() and str(x).lower() != "nan"]
            )
        else:
            product_opts = ["ALL"]
        product = st.selectbox("PRODUCT", product_opts, index=0, key="dor_product")
    with fc4:
        if col_route and col_route in df.columns:
            route_opts = ["ALL"] + sorted(
                [str(x).strip() for x in df[col_route].dropna().unique()
                 if str(x).strip() and str(x).lower() not in ("nan", "none", "")],
                key=lambda z: str(z).upper(),
            )
        else:
            route_opts = ["ALL"]
        route = st.selectbox("ROUTE", route_opts, index=0, key="dor_route")
    with fc5:
        passengers = st.selectbox("PASSENGERS", ["TOT", "FPD", "MHL"], index=0, key="dor_pax")
    with fc6:
        net_gross = st.selectbox("NET / GROSS (charts)", ["Gross", "Net"], index=0, key="dor_ng")

    depot = "ALL"

    # ---- Date controls ----
    st.markdown(
        "<div style='font-size:13px;font-weight:600;color:#334155;margin:8px 0 2px 0;'>DOR date filter</div>",
        unsafe_allow_html=True,
    )
    _dc1, _dc2, _dc3 = st.columns([1, 1, 2])
    with _dc1:
        dor_date = st.date_input(
            "As on date",
            value=_default,
            min_value=_min_d,
            max_value=_max_d,
            key="dor_as_on_date",
        )
    with _dc2:
        dor_period = st.selectbox(
            "Period (for tables)",
            ["For the day", "Upto the day", "Upto the month"],
            index=0,
            key="dor_period_mode",
            help=(
                "For the day: only the selected date. "
                "Upto the day: 1st of that month → selected date. "
                "Upto the month: 1 Apr of FY → selected date. "
                "Charts always show all three periods."
            ),
        )

    sel = pd.Timestamp(dor_date).normalize()
    ly_sel = (sel - pd.DateOffset(years=1)).normalize()

    def _windows(sel_ts, ly_ts):
        for_cy = (sel_ts, sel_ts)
        for_ly = (ly_ts, ly_ts)
        mtd_cy = (pd.Timestamp(year=sel_ts.year, month=sel_ts.month, day=1), sel_ts)
        mtd_ly = (pd.Timestamp(year=ly_ts.year, month=ly_ts.month, day=1), ly_ts)
        fy_year = sel_ts.year if sel_ts.month >= 4 else sel_ts.year - 1
        fy_start = pd.Timestamp(year=fy_year, month=4, day=1)
        fy_cy = (fy_start, sel_ts)
        fy_ly = (fy_start - pd.DateOffset(years=1), ly_ts)
        return {
            "For the day": (for_cy, for_ly),
            "Upto the day": (mtd_cy, mtd_ly),
            "Upto the month": (fy_cy, fy_ly),
        }

    WINS = _windows(sel, ly_sel)
    cy_start, cy_end = WINS[dor_period][0]
    ly_start, ly_end = WINS[dor_period][1]

    _no_dep = pd.Series(True, index=df.index)
    if mhl != "ALL" and col_mhl:
        _no_dep &= df[col_mhl].astype(str).str.strip().str.upper() == str(mhl).strip().upper()
    if route != "ALL" and col_route:
        _no_dep &= df[col_route].astype(str).str.strip().str.upper() == str(route).strip().upper()
    if product != "ALL" and col_product:
        _no_dep &= df[col_product].astype(str).str.strip().str.upper() == str(product).strip().upper()
    if rtc != "ALL" and col_rtc:
        _s = df[col_rtc].astype(str).str.strip().str.upper()
        if rtc == "RTC":
            _no_dep &= (_s == "RTC") | (_s == "R") | _s.str.startswith("RTC")
        elif rtc == "HIRE":
            _no_dep &= (
                (_s == "H")
                | _s.str.startswith("HIRE")
                | _s.str.contains(r"\bHIRE", regex=True, na=False)
            )

    _dt = pd.to_datetime(df[col_date], errors="coerce")

    def _slice(start, end):
        return df[_no_dep & (_dt >= start) & (_dt <= end)].copy()

    cy_data = _slice(cy_start, cy_end)
    ly_data = _slice(ly_start, ly_end)

    period_label = f"{dor_period}: {cy_start.strftime('%d-%b-%Y')} → {cy_end.strftime('%d-%b-%Y')}"
    st.caption(
        f"DOR tables | {period_label} | CY rows: {len(cy_data):,} | LY rows: {len(ly_data):,}"
    )

    pax_heading = {"FPD": "FPD PASSENGERS", "MHL": "MHL PASSENGERS"}.get(passengers, "TOTAL PASSENGERS")
    if passengers == "FPD":
        pax_col = "PSNGR_FPD"
    elif passengers == "MHL":
        pax_col = "PSNGR_MHL"
    else:
        pax_col = "PSNGR_TOT"

    orf_map, orf_by_prod, orf_err = load_orf_map()
    if orf_err:
        st.warning(f"ORF: {orf_err}")

    def build_dor_act_table(group_col, data_cy, data_ly, earn_tot_c, earn_fpd_c, earn_mhl_c, pax_c):
        """ACT VS ACT aggregation for DOR (kms/earn in lakhs after EPK)."""
        if len(data_cy) == 0 and len(data_ly) == 0:
            return pd.DataFrame()

        def agg_summary(data):
            if len(data) == 0:
                return pd.DataFrame(columns=[group_col, "kms", "earn_tot", "earn_fpd", "earn_mhl", "pax"])
            data = data.copy()
            for c in (earn_tot_c, earn_fpd_c, earn_mhl_c, "Optd_KMs", pax_c):
                if c not in data.columns:
                    data[c] = 0
            g = (
                data.groupby(group_col)
                .agg(
                    kms=("Optd_KMs", "sum"),
                    earn_tot=(earn_tot_c, "sum"),
                    earn_fpd=(earn_fpd_c, "sum"),
                    earn_mhl=(earn_mhl_c, "sum"),
                    pax=(pax_c, "sum"),
                )
                .reset_index()
            )
            return g

        cy_sum = agg_summary(data_cy)
        ly_sum = agg_summary(data_ly)
        merged = cy_sum.merge(ly_sum, on=group_col, how="outer", suffixes=("_CY", "_LY"))
        for col in [
            "kms_CY", "kms_LY", "earn_tot_CY", "earn_tot_LY",
            "earn_fpd_CY", "earn_fpd_LY", "earn_mhl_CY", "earn_mhl_LY",
            "pax_CY", "pax_LY",
        ]:
            if col not in merged.columns:
                merged[col] = 0.0
            merged[col] = pd.to_numeric(merged[col], errors="coerce").fillna(0.0)

        for side in ("CY", "LY"):
            k = merged[f"kms_{side}"]
            merged[f"epk_tot_{side}"] = np.where(k > 0, merged[f"earn_tot_{side}"] / k, np.nan)
            merged[f"epk_fpd_{side}"] = np.where(k > 0, merged[f"earn_fpd_{side}"] / k, np.nan)
            merged[f"epk_mhl_{side}"] = np.where(k > 0, merged[f"earn_mhl_{side}"] / k, np.nan)

        for col in ["kms_CY", "kms_LY", "earn_tot_CY", "earn_tot_LY",
                    "earn_fpd_CY", "earn_fpd_LY", "earn_mhl_CY", "earn_mhl_LY"]:
            merged[col] = merged[col] / 100000.0

        for base in ["kms", "earn_tot", "epk_tot", "epk_fpd", "epk_mhl"]:
            merged[f"{base}_VAR"] = merged[f"{base}_CY"] - merged[f"{base}_LY"]
            merged[f"{base}_PCT"] = np.where(
                pd.to_numeric(merged[f"{base}_LY"], errors="coerce").fillna(0) != 0,
                merged[f"{base}_VAR"] * 100 / merged[f"{base}_LY"],
                np.nan,
            )

        for col in ["pax_CY", "pax_LY"]:
            merged[col] = pd.to_numeric(merged[col], errors="coerce").fillna(0).round(0)
        merged["pax_VAR"] = merged["pax_CY"] - merged["pax_LY"]
        merged["pax_PCT"] = np.where(
            merged["pax_LY"] != 0, merged["pax_VAR"] * 100 / merged["pax_LY"], np.nan
        )

        merged[group_col] = merged[group_col].astype(str).str.strip()
        merged = merged[~merged[group_col].str.upper().isin(["", "NAN", "NONE", "NAT", "ALL"])]
        merged = merged.sort_values(group_col).reset_index(drop=True)

        t_kms_cy = float(data_cy["Optd_KMs"].sum()) / 100000 if len(data_cy) else 0.0
        t_kms_ly = float(data_ly["Optd_KMs"].sum()) / 100000 if len(data_ly) else 0.0
        t_earn_cy = float(data_cy[earn_tot_c].sum()) / 100000 if len(data_cy) and earn_tot_c in data_cy.columns else 0.0
        t_earn_ly = float(data_ly[earn_tot_c].sum()) / 100000 if len(data_ly) and earn_tot_c in data_ly.columns else 0.0
        t_earn_fpd_cy = float(data_cy[earn_fpd_c].sum()) if len(data_cy) and earn_fpd_c in data_cy.columns else 0.0
        t_earn_fpd_ly = float(data_ly[earn_fpd_c].sum()) if len(data_ly) and earn_fpd_c in data_ly.columns else 0.0
        t_earn_mhl_cy = float(data_cy[earn_mhl_c].sum()) if len(data_cy) and earn_mhl_c in data_cy.columns else 0.0
        t_earn_mhl_ly = float(data_ly[earn_mhl_c].sum()) if len(data_ly) and earn_mhl_c in data_ly.columns else 0.0
        kms_abs_cy = t_kms_cy * 100000
        kms_abs_ly = t_kms_ly * 100000
        t_epk_tot_cy = (t_earn_cy * 100000 / kms_abs_cy) if kms_abs_cy else np.nan
        t_epk_tot_ly = (t_earn_ly * 100000 / kms_abs_ly) if kms_abs_ly else np.nan
        t_epk_fpd_cy = (t_earn_fpd_cy / kms_abs_cy) if kms_abs_cy else np.nan
        t_epk_fpd_ly = (t_earn_fpd_ly / kms_abs_ly) if kms_abs_ly else np.nan
        t_epk_mhl_cy = (t_earn_mhl_cy / kms_abs_cy) if kms_abs_cy else np.nan
        t_epk_mhl_ly = (t_earn_mhl_ly / kms_abs_ly) if kms_abs_ly else np.nan
        t_pax_cy = float(data_cy[pax_c].sum()) if len(data_cy) and pax_c in data_cy.columns else 0.0
        t_pax_ly = float(data_ly[pax_c].sum()) if len(data_ly) and pax_c in data_ly.columns else 0.0

        def _var_pct(cy, ly):
            v = (cy - ly) if (pd.notna(cy) and pd.notna(ly)) else np.nan
            p = (v * 100 / ly) if (pd.notna(v) and pd.notna(ly) and ly != 0) else np.nan
            return v, p

        kms_v, kms_p = _var_pct(t_kms_cy, t_kms_ly)
        earn_v, earn_p = _var_pct(t_earn_cy, t_earn_ly)
        epk_tot_v, epk_tot_p = _var_pct(t_epk_tot_cy, t_epk_tot_ly)
        epk_fpd_v, epk_fpd_p = _var_pct(t_epk_fpd_cy, t_epk_fpd_ly)
        epk_mhl_v, epk_mhl_p = _var_pct(t_epk_mhl_cy, t_epk_mhl_ly)
        pax_v, pax_p = _var_pct(t_pax_cy, t_pax_ly)

        total = {
            group_col: "TOTAL",
            "kms_CY": t_kms_cy, "kms_LY": t_kms_ly, "kms_VAR": kms_v, "kms_PCT": kms_p,
            "earn_tot_CY": t_earn_cy, "earn_tot_LY": t_earn_ly, "earn_tot_VAR": earn_v, "earn_tot_PCT": earn_p,
            "epk_tot_CY": t_epk_tot_cy, "epk_tot_LY": t_epk_tot_ly, "epk_tot_VAR": epk_tot_v, "epk_tot_PCT": epk_tot_p,
            "epk_fpd_CY": t_epk_fpd_cy, "epk_fpd_LY": t_epk_fpd_ly, "epk_fpd_VAR": epk_fpd_v, "epk_fpd_PCT": epk_fpd_p,
            "epk_mhl_CY": t_epk_mhl_cy, "epk_mhl_LY": t_epk_mhl_ly, "epk_mhl_VAR": epk_mhl_v, "epk_mhl_PCT": epk_mhl_p,
            "pax_CY": round(t_pax_cy), "pax_LY": round(t_pax_ly),
            "pax_VAR": round(pax_v) if pd.notna(pax_v) else 0, "pax_PCT": pax_p,
        }
        merged = pd.concat([merged, pd.DataFrame([total])], ignore_index=True)
        return merged

    # ---- TABLE A: NET ----
    st.markdown(
        f'<div class="title-bar">TABLE A – NET | DOR | {dor_period} | As on {sel.strftime("%d-%b-%Y")}</div>',
        unsafe_allow_html=True,
    )
    df_net = build_dor_act_table(
        "DEPOT", cy_data, ly_data, "NE_TOT", "NE_FPD", "NE_MHL", pax_col
    )
    if df_net is None or len(df_net) == 0:
        st.warning("No data for NET table.")
        df_net = pd.DataFrame()
    else:
        df_net = add_or_columns_depot(
            df_net, cy_data, ly_data, "NE_TOT", "NE_FPD", "NE_MHL", orf_map,
            orf_by_prod=orf_by_prod, product_filter=product, depot_filter=depot,
        )
        try:
            fleet_map, _ferr = load_fleet_map()
            if not _ferr:
                df_net = add_avu_epb_columns(
                    df_net, cy_data, ly_data, fleet_map,
                    month_key=sel.strftime("%b-%Y"),
                    group_col="DEPOT", product_filter=product,
                )
        except Exception:
            pass
        st.markdown(
            render_act_table_with_or(df_net, "DEPOT", "NET", pax_heading),
            unsafe_allow_html=True,
        )

    # ---- TABLE B: GROSS ----
    st.markdown(
        f'<div class="title-bar">TABLE B – GROSS | DOR | {dor_period} | As on {sel.strftime("%d-%b-%Y")}</div>',
        unsafe_allow_html=True,
    )
    df_gr = build_dor_act_table(
        "DEPOT", cy_data, ly_data, "GE_TOT", "GE_FPD", "GE_MHL", pax_col
    )
    if df_gr is None or len(df_gr) == 0:
        st.warning("No data for GROSS table.")
        df_gr = pd.DataFrame()
    else:
        df_gr = add_or_columns_depot(
            df_gr, cy_data, ly_data, "GE_TOT", "GE_FPD", "GE_MHL", orf_map,
            orf_by_prod=orf_by_prod, product_filter=product, depot_filter=depot,
        )
        try:
            fleet_map, _ferr = load_fleet_map()
            if not _ferr:
                df_gr = add_avu_epb_columns(
                    df_gr, cy_data, ly_data, fleet_map,
                    month_key=sel.strftime("%b-%Y"),
                    group_col="DEPOT", product_filter=product,
                )
        except Exception:
            pass
        st.markdown(
            render_act_table_with_or(df_gr, "DEPOT", "GROSS", pax_heading),
            unsafe_allow_html=True,
        )

    _dn = df_net if df_net is not None else pd.DataFrame()
    _dg = df_gr if df_gr is not None else pd.DataFrame()
    if len(_dn) or len(_dg):
        _fname = f"DOR_NET_GROSS_{sel.strftime('%Y%m%d')}_{dor_period.replace(' ', '_')}.xlsx"
        st.download_button(
            "Download Excel – NET + GROSS",
            trends_dual_excel_bytes(
                _dn, _dg, pax_heading=pax_heading,
                report_title=f"DOR | {dor_period} | As on {sel.strftime('%d-%b-%Y')} | Depot={depot}",
            ),
            _fname,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="dl_dor_both",
        )

    # =============================================================================
    # 3 charts: For the day | Upto the day | Upto the month
    # =============================================================================
    try:
        import plotly.graph_objects as go
        from plotly.subplots import make_subplots

        if str(net_gross).upper() == "NET":
            _earn_col, _pfx = "NE_TOT", "Net"
        else:
            _earn_col, _pfx = "GE_TOT", "Gross"

        def _agg_kms_epk(data):
            if data is None or len(data) == 0 or "DEPOT" not in data.columns:
                return pd.DataFrame(columns=["DEPOT", "kms", "epk"])
            if "Optd_KMs" not in data.columns or _earn_col not in data.columns:
                return pd.DataFrame(columns=["DEPOT", "kms", "epk"])
            g = data.groupby("DEPOT", as_index=False).agg(
                Total_KMs=("Optd_KMs", "sum"),
                Total_Earn=(_earn_col, "sum"),
            )
            g["DEPOT"] = g["DEPOT"].astype(str).str.strip()
            g = g[~g["DEPOT"].str.upper().isin(["TOTAL", "REGION", "ALL", "NAN", ""])]
            g["kms"] = g["Total_KMs"] / 100000.0
            g["epk"] = np.where(g["Total_KMs"] > 0, g["Total_Earn"] / g["Total_KMs"], 0.0)
            return g[["DEPOT", "kms", "epk"]].sort_values("DEPOT").reset_index(drop=True)

        HIGH_SCALE = {"BHEL", "HYD1", "HYD2", "PKT"}

        def _left_range_for(deps, kms_vals):
            deps_u = {str(d).upper() for d in deps}
            mx = max(kms_vals) if kms_vals else 0.1
            if deps_u & HIGH_SCALE or mx > 0.75:
                top = max(1.2, (int(mx / 0.3) + 1) * 0.3)
                return [0, top], 0.3
            top = max(0.75, (int(mx / 0.15) + 1) * 0.15)
            return [0, top], 0.15

        st.markdown(
            f'<div style="text-align:center;margin:16px 0 8px 0;font-size:1.15rem;font-weight:800;color:#1e40af;">'
            f'DOR – KMs & {_pfx} EPK by Depot | As on {sel.strftime("%d-%b-%Y")}</div>',
            unsafe_allow_html=True,
        )

        chart_specs = [
            ("For the day", WINS["For the day"]),
            ("Upto the day", WINS["Upto the day"]),
            ("Upto the month", WINS["Upto the month"]),
        ]

        for title, ((c0, c1), (_l0, _l1)) in chart_specs:
            g = _agg_kms_epk(_slice(c0, c1))
            if len(g) == 0:
                st.warning(f"No data for chart: {title}")
                continue

            x_dep = g["DEPOT"].astype(str).tolist()
            kms_vals = pd.to_numeric(g["kms"], errors="coerce").fillna(0).tolist()
            epk_vals = pd.to_numeric(g["epk"], errors="coerce").fillna(0).tolist()
            left_range, left_dtick = _left_range_for(x_dep, kms_vals)
            epk_max = max(epk_vals) if epk_vals else 80
            right_top = max(80.0, epk_max * 1.15)

            win_txt = f"{c0.strftime('%d-%b-%Y')} → {c1.strftime('%d-%b-%Y')}"
            st.markdown(
                f'<div style="text-align:center;font-size:13px;font-weight:800;color:#1e40af;'
                f'background:#eff6ff;border:1px solid #bfdbfe;border-radius:6px;padding:6px;margin:10px 0 4px 0;">'
                f'{title} – KMs & {_pfx} EPK | {win_txt}</div>',
                unsafe_allow_html=True,
            )

            fig = make_subplots(specs=[[{"secondary_y": True}]])
            fig.add_trace(
                go.Bar(
                    name="KMs", x=x_dep, y=kms_vals, marker_color="#2563eb",
                    text=[f"{v:.2f}" for v in kms_vals], textposition="outside",
                    textfont=dict(size=10, color="#1e3a8a"), offsetgroup="a",
                ),
                secondary_y=False,
            )
            fig.add_trace(
                go.Bar(
                    name="E.P.K.", x=x_dep, y=epk_vals, marker_color="#15803d",
                    text=[f"{v:.2f}" for v in epk_vals], textposition="outside",
                    textfont=dict(size=10, color="#14532d"), offsetgroup="b",
                ),
                secondary_y=True,
            )
            fig.update_layout(
                barmode="group", height=380,
                margin=dict(l=50, r=50, t=40, b=50),
                legend=dict(orientation="h", y=1.12, x=0.5, xanchor="center"),
                xaxis=dict(title="Depot", tickfont=dict(color="#dc2626", size=12)),
                template="plotly_white", bargap=0.25,
            )
            fig.update_yaxes(
                title_text="KMs (lakhs)", range=left_range, dtick=left_dtick,
                secondary_y=False, color="#2563eb", title_font=dict(size=11),
                tickfont=dict(size=10),
            )
            fig.update_yaxes(
                title_text="EPK", range=[0, right_top],
                secondary_y=True, color="#15803d", title_font=dict(size=11),
                tickfont=dict(size=10), showgrid=False,
            )
            st.plotly_chart(fig, width="stretch")

    except Exception as _ce:
        st.caption(f"DOR charts: {_ce}")

elif section == "Monthly files":
    st.markdown(
        '<div class="title-bar">Monthly Files — Depot Route / Product / Service / Range / Sector boards</div>',
        unsafe_allow_html=True,
    )
    _mf = df_ser_wise if (df_ser_wise is not None and len(df_ser_wise)) else df
    if _mf is None or len(_mf) == 0:
        st.error("ser_wise.parquet not available.")
        st.stop()
    mf = _mf.copy()
    mf["Date"] = pd.to_datetime(mf["Date"], errors="coerce")
    mf = mf.dropna(subset=["Date"])
    _col_alts = {
        "GE_TOT": ["Gross Total", "GE_TOT"],
        "GE_FPD": ["Gross Fare Paid", "GE_FPD"],
        "GE_MHL": ["Gross MHL", "GE_MHL"],
        "NE_TOT": ["Net Total", "NE_TOT"],
        "NE_FPD": ["Net Fare Paid", "NE_FPD"],
        "NE_MHL": ["Net MHL", "NE_MHL"],
        "Optd_KMs": ["OPD_KMS", "Optd_KMs"],
        "NO_OF_SCHS": ["NO.OF SCHs", "NO_OF_SCHS", "NoOfSchedules"],
    }
    for c, alts in _col_alts.items():
        src = c if c in mf.columns else next((a for a in alts if a in mf.columns), None)
        if src is not None:
            s = mf[src]
            if isinstance(s, pd.DataFrame):
                s = s.iloc[:, 0]
            mf[c] = pd.to_numeric(s, errors="coerce").fillna(0)
        else:
            mf[c] = 0.0
    if "Month_Name" not in mf.columns and "Month Name" in mf.columns:
        mf["Month_Name"] = mf["Month Name"]
    if "RTC_HIRE" not in mf.columns and "RTC/HIRE" in mf.columns:
        mf["RTC_HIRE"] = mf["RTC/HIRE"]
    if "ROUTEE" not in mf.columns and "ROUTE" in mf.columns:
        mf["ROUTEE"] = mf["ROUTE"]
    if "PRODUCT" not in mf.columns:
        mf["PRODUCT"] = ""
    if "SER_NO" not in mf.columns:
        mf["SER_NO"] = ""
    if "TYPE" not in mf.columns and "TYPE OF SERV." in mf.columns:
        mf["TYPE"] = mf["TYPE OF SERV."]
    if "R_L" not in mf.columns:
        if "R/L" in mf.columns:
            mf["R_L"] = pd.to_numeric(mf["R/L"], errors="coerce").fillna(0)
        else:
            mf["R_L"] = 0.0
    if "SCH_KMS" not in mf.columns:
        if "DAY SCH_KMS" in mf.columns:
            mf["SCH_KMS"] = pd.to_numeric(mf["DAY SCH_KMS"], errors="coerce").fillna(0)
        elif "DAY_SCH_KMS" in mf.columns:
            mf["SCH_KMS"] = pd.to_numeric(mf["DAY_SCH_KMS"], errors="coerce").fillna(0)
        else:
            mf["SCH_KMS"] = 0.0
    if "SCH_DEP" not in mf.columns:
        if "SCH_DEP." in mf.columns:
            mf["SCH_DEP"] = mf["SCH_DEP."].astype(str)
        else:
            mf["SCH_DEP"] = ""
    if "NATURE" not in mf.columns:
        if "NATURE OF SERV." in mf.columns:
            mf["NATURE"] = mf["NATURE OF SERV."].astype(str)
        else:
            mf["NATURE"] = ""
    if "PRODUCT_NAME" not in mf.columns:
        if "PRODUCT NAME" in mf.columns:
            mf["PRODUCT_NAME"] = mf["PRODUCT NAME"].astype(str)
        elif "PRODUCT" in mf.columns:
            mf["PRODUCT_NAME"] = mf["PRODUCT"].astype(str)
        else:
            mf["PRODUCT_NAME"] = ""
    if "INTERSTATE" not in mf.columns:
        mf["INTERSTATE"] = ""
    if "MHL_NMHL" not in mf.columns and "MHL/NMHL" in mf.columns:
        mf["MHL_NMHL"] = mf["MHL/NMHL"]

    # Month options
    def _pm(m):
        try:
            return pd.to_datetime(str(m), format="%b-%Y")
        except Exception:
            return pd.to_datetime(str(m), errors="coerce")
    months = sorted([str(x).strip() for x in mf["Month_Name"].dropna().unique() if str(x).strip()], key=_pm, reverse=True)
    dep_opts = ["ALL"] + sorted({str(x).strip() for x in mf["DEPOT"].dropna().unique() if str(x).strip()})
    _route_col_mf = "ROUTEE" if "ROUTEE" in mf.columns else ("ROUTE" if "ROUTE" in mf.columns else None)
    if _route_col_mf:
        route_opts_mf = ["ALL"] + sorted(
            {str(x).strip() for x in mf[_route_col_mf].dropna().unique() if str(x).strip() and str(x).strip().lower() not in ("nan", "none", "")},
            key=lambda z: str(z).upper(),
        )
    else:
        route_opts_mf = ["ALL"]
    prod_opts_mf = ["ALL"] + sorted(
        {str(x).strip() for x in mf["PRODUCT"].dropna().unique() if str(x).strip() and str(x).strip().lower() not in ("nan", "none", "")},
        key=lambda z: str(z).upper(),
    )
    board_opts = [
        "1. Depot Route Wise",
        "2. Product Wise Trend",
        "6. Inter State Sector Wise",
        "7. Service Wise Performance",
        "8. Trend Service Wise (Gross)",
    ]
    # One filter row only (no global cascading row). Report board first.
    f1, f2, f3, f4, f5, f6 = st.columns(6)
    with f1:
        mf_board = st.selectbox("Report board", board_opts, key="mf_board")
    with f2:
        mf_month = st.selectbox("Month (FOR / Current)", months, index=0, key="mf_month")
    with f3:
        mf_depot = st.selectbox("Depot", dep_opts, index=0, key="mf_depot")
    with f4:
        mf_route = st.selectbox("Route", route_opts_mf, index=0, key="mf_route")
    with f5:
        mf_product = st.selectbox("Product", prod_opts_mf, index=0, key="mf_product")
    with f6:
        mf_ng = st.selectbox("Gross / Net", ["Gross", "Net"], index=0, key="mf_ng")

    # Parse selected month → FY window
    try:
        mon_name, yr = str(mf_month).split("-")[0], int(str(mf_month).split("-")[1])
    except Exception:
        mon_name, yr = "Jul", 2026
    cy_dt = _pm(mf_month)
    if pd.isna(cy_dt):
        cy_dt = pd.Timestamp(year=yr if yr > 100 else 2000 + yr, month=7, day=1)
    cy_year = int(cy_dt.year)
    cy_mon = int(cy_dt.month)
    fy_start_y = cy_year if cy_mon >= 4 else cy_year - 1
    fy_start = pd.Timestamp(year=fy_start_y, month=4, day=1)
    # month end of selected month
    cy_end = (cy_dt + pd.offsets.MonthEnd(0))
    ly_start = fy_start - pd.DateOffset(years=1)
    ly_end = cy_end - pd.DateOffset(years=1)
    ly_month_key = f"{mon_name}-{cy_year - 1}"

    base = mf.copy()
    if mf_depot != "ALL":
        base = base[base["DEPOT"].astype(str).str.strip().str.upper() == mf_depot.strip().upper()]

    # Period slices
    cm = base[base["Month_Name"].astype(str).str.strip() == str(mf_month).strip()].copy()  # current month FOR
    cy_um = base[(base["Date"] >= fy_start) & (base["Date"] <= cy_end)].copy()
    ly_um = base[(base["Date"] >= ly_start) & (base["Date"] <= ly_end)].copy()
    ly_cm = base[base["Month_Name"].astype(str).str.strip() == ly_month_key].copy()

    orf_map, orf_by_prod, orf_err = load_orf_map(r"D:\\dashboard\\ORF.xlsx")

    def _orf(dep, prod=None, side="cy"):
        d = str(dep).strip().upper()
        if prod and orf_by_prod:
            rec = orf_by_prod.get((d, str(prod).strip().upper()), {})
            v = rec.get(side, np.nan)
            if v is not None and not (isinstance(v, float) and np.isnan(v)):
                try:
                    return float(v)
                except Exception:
                    pass
        if orf_map and d in orf_map:
            try:
                return float(orf_map[d].get(side, np.nan))
            except Exception:
                return np.nan
        if orf_map and "REGION" in orf_map:
            try:
                return float(orf_map["REGION"].get(side, np.nan))
            except Exception:
                return np.nan
        return np.nan

    def _epk(earn, kms):
        try:
            e, k = float(earn), float(kms)
            return e / k if k > 0 else np.nan
        except Exception:
            return np.nan

    def _or(epk, orf):
        if pd.isna(epk) or pd.isna(orf) or not orf:
            return np.nan
        return epk * 10000 / orf

    def _f2(v):
        try:
            fv = float(v)
            if abs(fv) < 1e-12 or (isinstance(fv, float) and np.isnan(fv)):
                return ""
            return f"{fv:.2f}"
        except Exception:
            return ""

    def _f0(v):
        try:
            fv = float(v)
            if abs(fv) < 1e-12:
                return ""
            return f"{fv:.0f}"
        except Exception:
            return ""

    def _excel_bytes(df, title=""):
        bio = BytesIO()
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils.dataframe import dataframe_to_rows
            wb = Workbook()
            ws = wb.active
            ws.title = "Report"[:31]
            thin = Border(
                left=Side(style="thin", color="94A3B8"),
                right=Side(style="thin", color="94A3B8"),
                top=Side(style="thin", color="94A3B8"),
                bottom=Side(style="thin", color="94A3B8"),
            )
            hdr = PatternFill("solid", fgColor="1E3A8A")
            white = Font(bold=True, color="FFFFFF", size=10)
            center = Alignment(horizontal="center", vertical="center", wrap_text=True)
            r0 = 1
            if title:
                ws.cell(1, 1, title)
                ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(1, len(df.columns)))
                ws.cell(1, 1).font = Font(bold=True, size=12, color="1E3A8A")
                r0 = 2
            for j, col in enumerate(df.columns, 1):
                cell = ws.cell(r0, j, col)
                cell.fill = hdr
                cell.font = white
                cell.alignment = center
                cell.border = thin
            for i, row in enumerate(df.itertuples(index=False), r0 + 1):
                for j, val in enumerate(row, 1):
                    cell = ws.cell(i, j, None if (val is None or (isinstance(val, float) and np.isnan(val))) else val)
                    cell.alignment = center
                    cell.border = thin
            for col in ws.columns:
                letter = col[0].column_letter
                ws.column_dimensions[letter].width = min(16, max(8, max(len(str(c.value or "")) for c in col) + 2))
            wb.save(bio)
            bio.seek(0)
            return bio.getvalue()
        except Exception:
            bio = BytesIO()
            df.to_excel(bio, index=False)
            bio.seek(0)
            return bio.getvalue()

    def _th(text, bg="#1e3a8a", colspan=1, rowspan=1, color="#fff", top=None):
        sticky = f"position:sticky;top:{top}px;z-index:3;" if top is not None else ""
        return (
            f'<th colspan="{colspan}" rowspan="{rowspan}" '
            f'style="background:{bg};color:{color};padding:5px 4px;font-size:10px;'
            f'text-align:center;border:1px solid #94a3b8;{sticky}">{text}</th>'
        )

    def _cell(v, is_int=False, bg=None, row_hl=False):
        if v is None or v == "" or (isinstance(v, float) and (np.isnan(v) or abs(v) < 1e-12)):
            s = ""
        elif isinstance(v, (int, float)):
            try:
                fv = float(v)
                if is_int or abs(fv - round(fv)) < 1e-6:
                    s = f"{int(round(fv))}" if abs(fv) > 1e-12 else ""
                else:
                    s = f"{fv:.2f}"
            except Exception:
                s = str(v)
        else:
            s = str(v)
        style = "padding:3px 5px;text-align:center;border:1px solid #e2e8f0;font-size:11px;"
        if row_hl:
            style += "background:#fef08a !important;font-weight:700;"
        elif bg:
            style += f"background:{bg};"
        return f'<td style="{style}">{s}</td>'

    # Soft fills matching header groups (GROSS EPK / GROSS OR / NET EPK / NET OR)
    _BG_GEPK = "#f0fdf4"   # green tint
    _BG_GOR = "#eff6ff"    # blue tint
    _BG_NEPK = "#fffbeb"   # amber tint
    _BG_NOR = "#f5f3ff"    # violet tint
    _BG_OPTD = "#ecfdf5"   # teal tint

    def _render_board(title, thead_html, body_rows_html, df_for_excel, key_suffix):
        st.markdown(f'<div class="title-bar">{title}</div>', unsafe_allow_html=True)
        if not body_rows_html:
            st.warning("No rows for this board.")
            return
        html = [
            '<div class="table-scroll-fixable"><table class="excel-table" style="border-collapse:collapse;width:max-content;">',
            "<thead>", thead_html, "</thead><tbody>",
            body_rows_html, "</tbody></table></div>",
        ]
        st.markdown("".join(html), unsafe_allow_html=True)
        st.download_button(
            "Download Excel (as on screen)",
            _excel_bytes(df_for_excel, title),
            f"Monthly_{mf_month}_{key_suffix}.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key=f"mf_dl_{key_suffix}",
        )

    def _html_table(df, title, freeze_cols=2):
        """Flat fallback."""
        st.markdown(f'<div class="title-bar">{title}</div>', unsafe_allow_html=True)
        if df is None or len(df) == 0:
            st.warning("No rows for this board.")
            return
        cols = list(df.columns)
        html = ['<div class="table-scroll-fixable"><table class="excel-table" style="border-collapse:collapse;font-size:11px;">']
        html.append("<thead><tr>")
        for c in cols:
            html.append(_th(c, top=0))
        html.append("</tr></thead><tbody>")
        for _, row in df.iterrows():
            html.append("<tr>")
            for c in cols:
                html.append(_cell(row[c]))
            html.append("</tr>")
        html.append("</tbody></table></div>")
        st.markdown("".join(html), unsafe_allow_html=True)
        st.download_button(
            "Download Excel (as on screen)",
            _excel_bytes(df, title),
            f"Monthly_{mf_month}_{mf_board[:20].replace(' ', '_')}.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key=f"mf_dl_{mf_board[:12]}",
        )

    # ========== BOARD 1: Depot Route Wise ==========
    if mf_board.startswith("1."):
        # Filters from single top row: depot, route, product, gross/net
        b1_depot = mf_depot
        b1_route = mf_route
        b1_product = mf_product
        b1_ng = mf_ng
        _route_col = _route_col_mf

        # Board 1 now displays BOTH GROSS and NET EPK/OR.  Keep the existing
        # Gross/Net selector for backward-compatible UI state, but do not use it
        # to hide either metric group.
        b1_prefix = "GROSS + NET"
        if b1_ng == "Gross":
            _et, _em, _ef = "GE_TOT", "GE_MHL", "GE_FPD"
        else:
            _et, _em, _ef = "NE_TOT", "NE_MHL", "NE_FPD"

        # Period slices from full mf (month window only)
        _cm_full = mf[mf["Month_Name"].astype(str).str.strip() == str(mf_month).strip()].copy()
        _cy_full = mf[(mf["Date"] >= fy_start) & (mf["Date"] <= cy_end)].copy()
        _ly_full = mf[(mf["Date"] >= ly_start) & (mf["Date"] <= ly_end)].copy()

        def _b1_filter(data):
            out_d = data
            if b1_depot != "ALL" and "DEPOT" in out_d.columns:
                out_d = out_d[out_d["DEPOT"].astype(str).str.strip().str.upper() == str(b1_depot).strip().upper()]
            if b1_route != "ALL" and _route_col and _route_col in out_d.columns:
                out_d = out_d[out_d[_route_col].astype(str).str.strip().str.upper() == str(b1_route).strip().upper()]
            if b1_product != "ALL" and "PRODUCT" in out_d.columns:
                out_d = out_d[out_d["PRODUCT"].astype(str).str.strip().str.upper() == str(b1_product).strip().upper()]
            return out_d

        cm_b1 = _b1_filter(_cm_full)
        cy_um_b1 = _b1_filter(_cy_full)
        ly_um_b1 = _b1_filter(_ly_full)

        title = f"{b1_prefix} - DEPOT ROUTE WISE PERFORMANCE FOR AND UPTO {mf_month}"
        if b1_depot != "ALL":
            title += f" | Depot: {b1_depot}"
        if b1_route != "ALL":
            title += f" | Route: {b1_route}"
        if b1_product != "ALL":
            title += f" | Product: {b1_product}"

        keys = ["DEPOT", "ROUTEE", "PRODUCT", "RTC_HIRE"]

        def _agg_side(data):
            """Schs / Sch Kms: max per SER_NO then sum (avoid inflating daily rows).
            Ser: nunique SER_NO. Kms / earnings: sum as before.
            """
            if len(data) == 0:
                return pd.DataFrame(columns=keys + ["schs", "sers", "rl", "sch_kms", "kms", "gross_tot", "gross_mhl", "gross_fpd", "net_tot", "net_mhl", "net_fpd"])
            d = data.copy()
            if "SER_NO" not in d.columns:
                d["SER_NO"] = ""
            # Per service: schedule count = max(NO_OF_SCHS); sch kms = max(SCH_KMS) (not summed over days)
            ser_keys = keys + ["SER_NO"]
            per = (
                d.groupby(ser_keys, dropna=False)
                .agg(
                    schs=("NO_OF_SCHS", "max"),
                    sch_kms=("SCH_KMS", "max"),
                    rl=("R_L", "max"),
                    kms=("Optd_KMs", "sum"),
                    gross_tot=("GE_TOT", "sum"),
                    gross_mhl=("GE_MHL", "sum"),
                    gross_fpd=("GE_FPD", "sum"),
                    net_tot=("NE_TOT", "sum"),
                    net_mhl=("NE_MHL", "sum"),
                    net_fpd=("NE_FPD", "sum"),
                )
                .reset_index()
            )
            g = (
                per.groupby(keys, dropna=False)
                .agg(
                    schs=("schs", "sum"),
                    sers=("SER_NO", "nunique"),
                    rl=("rl", "max"),
                    sch_kms=("sch_kms", "sum"),
                    kms=("kms", "sum"),
                    gross_tot=("gross_tot", "sum"),
                    gross_mhl=("gross_mhl", "sum"),
                    gross_fpd=("gross_fpd", "sum"),
                    net_tot=("net_tot", "sum"),
                    net_mhl=("net_mhl", "sum"),
                    net_fpd=("net_fpd", "sum"),
                )
                .reset_index()
            )
            # If SCH_KMS blank but R/L and schs present → Sch Kms = R/L × No of Schs
            for i, row in g.iterrows():
                sk = float(row.get("sch_kms", 0) or 0)
                rl = float(row.get("rl", 0) or 0)
                sc = float(row.get("schs", 0) or 0)
                if sk == 0 and rl > 0 and sc > 0:
                    g.at[i, "sch_kms"] = rl * sc
            return g

        a_cm, a_cy, a_ly = _agg_side(cm_b1), _agg_side(cy_um_b1), _agg_side(ly_um_b1)
        merged = a_cm.merge(a_cy, on=keys, how="outer", suffixes=("_cm", "_cy"))
        merged = merged.merge(a_ly, on=keys, how="outer")
        for c in ["schs", "sers", "rl", "sch_kms", "kms", "earn_tot", "earn_mhl", "earn_fpd"]:
            if c in merged.columns:
                merged.rename(columns={c: f"{c}_ly"}, inplace=True)
        rows = []
        sno = 0
        for _, r in merged.fillna(0).iterrows():
            sno += 1
            kms_cm, kms_cy, kms_ly = float(r.get("kms_cm", 0)), float(r.get("kms_cy", 0)), float(r.get("kms_ly", 0))
            dep = str(r.get("DEPOT", ""))
            orf_c, orf_l = _orf(dep, side="cy"), _orf(dep, side="ly")
            def pack(earn_t, earn_m, earn_f, kms, orf):
                et, em, ef = _epk(earn_t, kms), _epk(earn_m, kms), _epk(earn_f, kms)
                return et, em, ef, _or(et, orf), _or(em, orf), _or(ef, orf)
            et_cm, em_cm, ef_cm, ot_cm, om_cm, of_cm = pack(r.get("gross_tot_cm", 0), r.get("gross_mhl_cm", 0), r.get("gross_fpd_cm", 0), kms_cm, orf_c)
            et_cy, em_cy, ef_cy, ot_cy, om_cy, of_cy = pack(r.get("gross_tot_cy", 0), r.get("gross_mhl_cy", 0), r.get("gross_fpd_cy", 0), kms_cy, orf_c)
            et_ly, em_ly, ef_ly, ot_ly, om_ly, of_ly = pack(r.get("gross_tot_ly", 0), r.get("gross_mhl_ly", 0), r.get("gross_fpd_ly", 0), kms_ly, orf_l)
            nt_cm, nm_cm, nf_cm, not_cm, nom_cm, nof_cm = pack(r.get("net_tot_cm", 0), r.get("net_mhl_cm", 0), r.get("net_fpd_cm", 0), kms_cm, orf_c)
            nt_cy, nm_cy, nf_cy, not_cy, nom_cy, nof_cy = pack(r.get("net_tot_cy", 0), r.get("net_mhl_cy", 0), r.get("net_fpd_cy", 0), kms_cy, orf_c)
            nt_ly, nm_ly, nf_ly, not_ly, nom_ly, nof_ly = pack(r.get("net_tot_ly", 0), r.get("net_mhl_ly", 0), r.get("net_fpd_ly", 0), kms_ly, orf_l)
            # NET is calculated from the same underlying period/KM rows, using
            # the NET earnings columns in the three period aggregations.
            # `_agg_side` stores the selected metric only, so for the combined
            # Board 1 view we calculate a second set directly from the filtered
            # source data below.
            rows.append({
                "SL.NO.": sno,
                "DEPOT": dep,
                "ROUTE": r.get("ROUTEE", ""),
                "Product": r.get("PRODUCT", ""),
                "RTC/HIRE": r.get("RTC_HIRE", ""),
                "No. of Schs": int(round(float(r.get("schs_cm", 0) or 0))) or "",
                "No. of Ser": int(round(float(r.get("sers_cm", 0) or 0))) or "",
                "R/L": _f0(r.get("rl_cm", 0)),
                "Sch Kms": _f0(r.get("sch_kms_cm", 0)),
                "CM TOT EPK": round(et_cm, 2) if pd.notna(et_cm) else None,
                "CM MHL EPK": round(em_cm, 2) if pd.notna(em_cm) else None,
                "CM FPD EPK": round(ef_cm, 2) if pd.notna(ef_cm) else None,
                "CY UM TOT EPK": round(et_cy, 2) if pd.notna(et_cy) else None,
                "CY UM MHL EPK": round(em_cy, 2) if pd.notna(em_cy) else None,
                "CY UM FPD EPK": round(ef_cy, 2) if pd.notna(ef_cy) else None,
                "LY UM TOT EPK": round(et_ly, 2) if pd.notna(et_ly) else None,
                "LY UM MHL EPK": round(em_ly, 2) if pd.notna(em_ly) else None,
                "LY UM FPD EPK": round(ef_ly, 2) if pd.notna(ef_ly) else None,
                "CM TOT OR": round(ot_cm, 0) if pd.notna(ot_cm) else None,
                "CM MHL OR": round(om_cm, 0) if pd.notna(om_cm) else None,
                "CM FPD OR": round(of_cm, 0) if pd.notna(of_cm) else None,
                "CY UM TOT OR": round(ot_cy, 0) if pd.notna(ot_cy) else None,
                "CY UM MHL OR": round(om_cy, 0) if pd.notna(om_cy) else None,
                "CY UM FPD OR": round(of_cy, 0) if pd.notna(of_cy) else None,
                "LY UM TOT OR": round(ot_ly, 0) if pd.notna(ot_ly) else None,
                "LY UM MHL OR": round(om_ly, 0) if pd.notna(om_ly) else None,
                "LY UM FPD OR": round(of_ly, 0) if pd.notna(of_ly) else None,
                "N CM TOT EPK": round(nt_cm, 2) if pd.notna(nt_cm) else None,
                "N CM MHL EPK": round(nm_cm, 2) if pd.notna(nm_cm) else None,
                "N CM FPD EPK": round(nf_cm, 2) if pd.notna(nf_cm) else None,
                "N CY UM TOT EPK": round(nt_cy, 2) if pd.notna(nt_cy) else None,
                "N CY UM MHL EPK": round(nm_cy, 2) if pd.notna(nm_cy) else None,
                "N CY UM FPD EPK": round(nf_cy, 2) if pd.notna(nf_cy) else None,
                "N LY UM TOT EPK": round(nt_ly, 2) if pd.notna(nt_ly) else None,
                "N LY UM MHL EPK": round(nm_ly, 2) if pd.notna(nm_ly) else None,
                "N LY UM FPD EPK": round(nf_ly, 2) if pd.notna(nf_ly) else None,
                "N CM TOT OR": round(not_cm, 0) if pd.notna(not_cm) else None,
                "N CM MHL OR": round(nom_cm, 0) if pd.notna(nom_cm) else None,
                "N CM FPD OR": round(nof_cm, 0) if pd.notna(nof_cm) else None,
                "N CY UM TOT OR": round(not_cy, 0) if pd.notna(not_cy) else None,
                "N CY UM MHL OR": round(nom_cy, 0) if pd.notna(nom_cy) else None,
                "N CY UM FPD OR": round(nof_cy, 0) if pd.notna(nof_cy) else None,
                "N LY UM TOT OR": round(not_ly, 0) if pd.notna(not_ly) else None,
                "N LY UM MHL OR": round(nom_ly, 0) if pd.notna(nom_ly) else None,
                "N LY UM FPD OR": round(nof_ly, 0) if pd.notna(nof_ly) else None,
            })
        out = pd.DataFrame(rows)

        def _b1_excel_bytes(df_out, report_title, prefix_label):
            """Excel matching on-screen multi-level headers (not flat raw column names)."""
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            wb = Workbook()
            ws = wb.active
            ws.title = "DepotRoute"[:31]
            thin = Border(
                left=Side(style="thin", color="94A3B8"),
                right=Side(style="thin", color="94A3B8"),
                top=Side(style="thin", color="94A3B8"),
                bottom=Side(style="thin", color="94A3B8"),
            )
            center = Alignment(horizontal="center", vertical="center", wrap_text=True)
            white = Font(bold=True, color="FFFFFF", size=9)
            dark = Font(bold=True, color="14532D", size=9)
            fills = {
                "left": PatternFill("solid", fgColor="1E3A8A"),
                "epk": PatternFill("solid", fgColor="DCFCE7"),
                "or": PatternFill("solid", fgColor="DBEAFE"),
                "cm": PatternFill("solid", fgColor="BBF7D0"),
                "cy": PatternFill("solid", fgColor="86EFAC"),
                "ly": PatternFill("solid", fgColor="BBF7D0"),
                "orc": PatternFill("solid", fgColor="BFDBFE"),
                "orcy": PatternFill("solid", fgColor="93C5FD"),
                "orly": PatternFill("solid", fgColor="BFDBFE"),
                "sub": PatternFill("solid", fgColor="FEF3C7"),
            }

            # Title — fixed 9 identity columns + 36 metric columns
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=45)
            tcell = ws.cell(1, 1, report_title)
            tcell.font = Font(bold=True, size=12, color="1E3A8A")
            tcell.alignment = Alignment(horizontal="center", vertical="center")

            fixed = ["SL.NO.", "DEPOT", "ROUTE", "Product", "RTC/HIRE", "No. of Schs", "No. of Ser", "R/L", "Sch Kms"]
            for c, h in enumerate(fixed, 1):
                cell = ws.cell(2, c, h)
                cell.fill = fills["left"]
                cell.font = white
                cell.alignment = center
                cell.border = thin
                ws.merge_cells(start_row=2, start_column=c, end_row=4, end_column=c)
                for rr in range(2, 5):
                    ws.cell(rr, c).fill = fills["left"]
                    ws.cell(rr, c).border = thin
                    ws.cell(rr, c).font = white
                    ws.cell(rr, c).alignment = center

            # Row 2: GROSS EPK | GROSS OR | NET EPK | NET OR
            groups = [
                (10, 18, "GROSS E P K", fills["epk"], "14532D"),
                (19, 27, "GROSS OR", fills["or"], "1E3A8A"),
                (28, 36, "NET E P K", PatternFill("solid", fgColor="FEF3C7"), "92400E"),
                (37, 45, "NET OR", PatternFill("solid", fgColor="EDE9FE"), "5B21B6"),
            ]
            for start, end, label, fill, font_color in groups:
                ws.merge_cells(start_row=2, start_column=start, end_row=2, end_column=end)
                c = ws.cell(2, start, label)
                c.fill = fill
                c.font = Font(bold=True, color=font_color, size=10)
                c.alignment = center
                for cc in range(start, end + 1):
                    ws.cell(2, cc).fill = fill
                    ws.cell(2, cc).border = thin

            # Row 3: Current Month / CY UM / LY UM × 4
            periods = [
                (10, "Current Month", fills["cm"]),
                (13, "CY Upto The Month", fills["cy"]),
                (16, "LY Upto The Month", fills["ly"]),
                (19, "Current Month", fills["orc"]),
                (22, "CY Upto The Month", fills["orcy"]),
                (25, "LY Upto The Month", fills["orly"]),
                (28, "Current Month", PatternFill("solid", fgColor="FDE68A")),
                (31, "CY Upto The Month", PatternFill("solid", fgColor="FBBF24")),
                (34, "LY Upto The Month", PatternFill("solid", fgColor="FEF3C7")),
                (37, "Current Month", PatternFill("solid", fgColor="C4B5FD")),
                (40, "CY Upto The Month", PatternFill("solid", fgColor="A78BFA")),
                (43, "LY Upto The Month", PatternFill("solid", fgColor="EDE9FE")),
            ]
            for start, lab, fl in periods:
                ws.merge_cells(start_row=3, start_column=start, end_row=3, end_column=start + 2)
                cell = ws.cell(3, start, lab)
                cell.fill = fl
                cell.font = dark
                cell.alignment = center
                for cc in range(start, start + 3):
                    ws.cell(3, cc).fill = fl
                    ws.cell(3, cc).border = thin

            # Row 4: TOT MHL FPD
            for start in [p[0] for p in periods]:
                for i, lab in enumerate(["TOT", "MHL", "FPD"]):
                    cell = ws.cell(4, start + i, lab)
                    cell.fill = fills["sub"]
                    cell.font = Font(bold=True, color="92400E", size=9)
                    cell.alignment = center
                    cell.border = thin

            data_cols = [
                "SL.NO.", "DEPOT", "ROUTE", "Product", "RTC/HIRE", "No. of Schs", "No. of Ser", "R/L", "Sch Kms",
                "CM TOT EPK", "CM MHL EPK", "CM FPD EPK", "CY UM TOT EPK", "CY UM MHL EPK", "CY UM FPD EPK",
                "LY UM TOT EPK", "LY UM MHL EPK", "LY UM FPD EPK",
                "CM TOT OR", "CM MHL OR", "CM FPD OR", "CY UM TOT OR", "CY UM MHL OR", "CY UM FPD OR",
                "LY UM TOT OR", "LY UM MHL OR", "LY UM FPD OR",
                "N CM TOT EPK", "N CM MHL EPK", "N CM FPD EPK", "N CY UM TOT EPK", "N CY UM MHL EPK", "N CY UM FPD EPK",
                "N LY UM TOT EPK", "N LY UM MHL EPK", "N LY UM FPD EPK",
                "N CM TOT OR", "N CM MHL OR", "N CM FPD OR", "N CY UM TOT OR", "N CY UM MHL OR", "N CY UM FPD OR",
                "N LY UM TOT OR", "N LY UM MHL OR", "N LY UM FPD OR",
            ]
            for ri, row in enumerate(df_out.itertuples(index=False), 5):
                rec = dict(zip(df_out.columns, row))
                for ci, col in enumerate(data_cols, 1):
                    val = rec.get(col, "")
                    if val is None or (isinstance(val, float) and (pd.isna(val) or abs(val) < 1e-12)):
                        val = None
                    elif isinstance(val, float) and abs(val - round(val)) < 1e-6 and col in (
                        "SL.NO.", "No. of Schs", "No. of Ser",
                        "CM TOT OR", "CM MHL OR", "CM FPD OR", "CY UM TOT OR", "CY UM MHL OR", "CY UM FPD OR",
                        "LY UM TOT OR", "LY UM MHL OR", "LY UM FPD OR",
                        "N CM TOT OR", "N CM MHL OR", "N CM FPD OR", "N CY UM TOT OR", "N CY UM MHL OR",
                        "N CY UM FPD OR", "N LY UM TOT OR", "N LY UM MHL OR", "N LY UM FPD OR",
                    ):
                        val = int(round(val))
                    cell = ws.cell(ri, ci, val if val != "" else None)
                    cell.alignment = center
                    cell.border = thin

            widths = {1: 6, 2: 10, 3: 10, 4: 10, 5: 9, 6: 8, 7: 8, 8: 8, 9: 10}
            for i in range(1, 46):
                ws.column_dimensions[get_column_letter(i)].width = widths.get(i, 9)
            ws.row_dimensions[1].height = 20
            ws.row_dimensions[2].height = 18
            ws.row_dimensions[3].height = 18
            ws.row_dimensions[4].height = 16
            ws.freeze_panes = "J5"
            bio = BytesIO()
            wb.save(bio)
            bio.seek(0)
            return bio.getvalue()

        # Multi-level headers: {GROSS|NET} EPK | {GROSS|NET} OR
        thead = (
            "<tr>"
            + _th("SL.<br>NO.", rowspan=3, top=0)
            + _th("DEPOT", rowspan=3, top=0)
            + _th("ROUTE", rowspan=3, top=0)
            + _th("Product", rowspan=3, top=0)
            + _th("RTC/<br>HIRE", rowspan=3, top=0)
            + _th("No.<br>of<br>Schs", rowspan=3, top=0)
            + _th("No.<br>of<br>Ser", rowspan=3, top=0)
            + _th("R/L", rowspan=3, top=0)
            + _th("Sch<br>Kms", rowspan=3, top=0)
            + _th("GROSS E P K", bg="#dcfce7", color="#14532d", colspan=9, top=0)
            + _th("GROSS OR", bg="#dbeafe", color="#1e3a8a", colspan=9, top=0)
            + _th("NET E P K", bg="#fef3c7", color="#92400e", colspan=9, top=0)
            + _th("NET OR", bg="#ede9fe", color="#5b21b6", colspan=9, top=0)
            + "</tr><tr>"
            + _th("Current Month", bg="#bbf7d0", color="#14532d", colspan=3, top=28)
            + _th("CY Upto The Month", bg="#86efac", color="#14532d", colspan=3, top=28)
            + _th("LY Upto The Month", bg="#bbf7d0", color="#14532d", colspan=3, top=28)
            + _th("Current Month", bg="#bfdbfe", color="#1e3a8a", colspan=3, top=28)
            + _th("CY Upto The Month", bg="#93c5fd", color="#1e3a8a", colspan=3, top=28)
            + _th("LY Upto The Month", bg="#bfdbfe", color="#1e3a8a", colspan=3, top=28)
            + _th("Current Month", bg="#fde68a", color="#92400e", colspan=3, top=28)
            + _th("CY Upto The Month", bg="#fbbf24", color="#92400e", colspan=3, top=28)
            + _th("LY Upto The Month", bg="#fef3c7", color="#92400e", colspan=3, top=28)
            + _th("Current Month", bg="#c4b5fd", color="#5b21b6", colspan=3, top=28)
            + _th("CY Upto The Month", bg="#a78bfa", color="#5b21b6", colspan=3, top=28)
            + _th("LY Upto The Month", bg="#ede9fe", color="#5b21b6", colspan=3, top=28)
            + "</tr><tr>"
        )
        # Row-3 sub-headers match parent group colours
        for bg, col in [
            ("#dcfce7", "#14532d"), ("#dcfce7", "#14532d"), ("#dcfce7", "#14532d"),  # GROSS EPK x3 periods
            ("#dbeafe", "#1e3a8a"), ("#dbeafe", "#1e3a8a"), ("#dbeafe", "#1e3a8a"),  # GROSS OR
            ("#fef3c7", "#92400e"), ("#fef3c7", "#92400e"), ("#fef3c7", "#92400e"),  # NET EPK
            ("#ede9fe", "#5b21b6"), ("#ede9fe", "#5b21b6"), ("#ede9fe", "#5b21b6"),  # NET OR
        ]:
            thead += _th("TOT", bg=bg, color=col, top=56) + _th("MHL", bg=bg, color=col, top=56) + _th("FPD", bg=bg, color=col, top=56)
        thead += "</tr>"
        data_cols = [
            "SL.NO.", "DEPOT", "ROUTE", "Product", "RTC/HIRE", "No. of Schs", "No. of Ser", "R/L", "Sch Kms",
            "CM TOT EPK", "CM MHL EPK", "CM FPD EPK", "CY UM TOT EPK", "CY UM MHL EPK", "CY UM FPD EPK",
            "LY UM TOT EPK", "LY UM MHL EPK", "LY UM FPD EPK",
            "CM TOT OR", "CM MHL OR", "CM FPD OR", "CY UM TOT OR", "CY UM MHL OR", "CY UM FPD OR",
            "LY UM TOT OR", "LY UM MHL OR", "LY UM FPD OR",
            "N CM TOT EPK", "N CM MHL EPK", "N CM FPD EPK", "N CY UM TOT EPK", "N CY UM MHL EPK", "N CY UM FPD EPK",
            "N LY UM TOT EPK", "N LY UM MHL EPK", "N LY UM FPD EPK",
            "N CM TOT OR", "N CM MHL OR", "N CM FPD OR", "N CY UM TOT OR", "N CY UM MHL OR", "N CY UM FPD OR",
            "N LY UM TOT OR", "N LY UM MHL OR", "N LY UM FPD OR",
        ]
        def _b1_col_bg(c):
            if c.startswith("N ") and "EPK" in c:
                return _BG_NEPK
            if c.startswith("N ") and "OR" in c:
                return _BG_NOR
            if "EPK" in c:
                return _BG_GEPK
            if "OR" in c:
                return _BG_GOR
            return None
        st.markdown(f'<div class="title-bar">{title}</div>', unsafe_allow_html=True)
        if out is None or len(out) == 0:
            st.warning("No rows for this board.")
            body = []
        else:
            _row_opts = ["(none)"] + [
                f"{int(r.get('SL.NO.', i+1))}. {r.get('DEPOT','')} / {r.get('ROUTE','')} / {r.get('Product','')}"
                for i, r in out.iterrows()
            ]
            b1_sel = st.selectbox("Highlight row", _row_opts, index=0, key="mf_b1_row_sel")
            body = []
            for i, r in out.iterrows():
                label = f"{int(r.get('SL.NO.', i+1))}. {r.get('DEPOT','')} / {r.get('ROUTE','')} / {r.get('Product','')}"
                hl = b1_sel == label
                tr_style = ' style="background:#fef08a;"' if hl else ""
                body.append(f"<tr{tr_style}>")
                for c in data_cols:
                    body.append(_cell(r.get(c), bg=_b1_col_bg(c), row_hl=hl))
                body.append("</tr>")
            html = [
                '<div class="table-scroll-fixable"><table class="excel-table" style="border-collapse:collapse;width:max-content;">',
                "<thead>", thead, "</thead><tbody>",
                "".join(body), "</tbody></table></div>",
            ]
            st.markdown("".join(html), unsafe_allow_html=True)
            st.download_button(
                "Download Excel (as on screen)",
                _b1_excel_bytes(out, title, b1_prefix),
                f"Monthly_{mf_month}_DepotRoute.xlsx",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="mf_dl_DepotRoute",
            )
        if orf_err:
            st.caption(f"OR blank without ORF.xlsx ({orf_err})")

    # ========== BOARD 2: Product Wise Trend ==========
    elif mf_board.startswith("2."):
        title = f"PRODUCT WISE TREND UPTO {mf_month}"
        # months in FY up to selected
        fy_months = []
        cur = fy_start
        while cur <= cy_end:
            fy_months.append(cur.strftime("%b-%Y"))
            cur += pd.DateOffset(months=1)
        # TYPE OF SERVICE ≈ PRODUCT_NAME or PRODUCT
        svc_col = "PRODUCT_NAME" if "PRODUCT_NAME" in base.columns else "PRODUCT"
        rows = []
        groups = base.groupby(["DEPOT", svc_col], dropna=False)
        for (dep, ptype), g in groups:
            if not str(ptype).strip() or str(ptype).lower() in ("nan", "none"):
                continue
            sch_for = int(round(g[g["Month_Name"].astype(str).str.strip() == str(mf_month).strip()]["NO_OF_SCHS"].sum())) if len(g) else 0
            sch_upto = int(round(g[(g["Date"] >= fy_start) & (g["Date"] <= cy_end)]["NO_OF_SCHS"].sum())) if len(g) else 0
            rec = {"DEPOT": dep, "TYPE OF SERVICE": ptype, "SCHs FOR": sch_for or "", "SCHs UPTO": sch_upto or ""}
            orf_c = _orf(dep, side="cy")
            orf_l = _orf(dep, side="ly")
            for mkey in fy_months:
                sub = g[g["Month_Name"].astype(str).str.strip() == mkey]
                kms = float(sub["Optd_KMs"].sum())
                ne = float(sub["NE_TOT"].sum())
                ge = float(sub["GE_TOT"].sum())
                nepk, gepk = _epk(ne, kms), _epk(ge, kms)
                lab = mkey.split("-")[0]  # Apr
                rec[f"{lab} NET EPK"] = round(nepk, 2) if pd.notna(nepk) else None
                rec[f"{lab} NET OR"] = round(_or(nepk, orf_c), 0) if pd.notna(nepk) else None
                rec[f"{lab} GROSS EPK"] = round(gepk, 2) if pd.notna(gepk) else None
                rec[f"{lab} GROSS OR"] = round(_or(gepk, orf_c), 0) if pd.notna(gepk) else None
            # CY UM / LY UM
            sub_cy = g[(g["Date"] >= fy_start) & (g["Date"] <= cy_end)]
            sub_ly = g[(g["Date"] >= ly_start) & (g["Date"] <= ly_end)]
            for label, sub, orf in (("CY UM", sub_cy, orf_c), ("LY UM", sub_ly, orf_l)):
                kms = float(sub["Optd_KMs"].sum())
                ne, ge = float(sub["NE_TOT"].sum()), float(sub["GE_TOT"].sum())
                nepk, gepk = _epk(ne, kms), _epk(ge, kms)
                rec[f"{label} NET EPK"] = round(nepk, 2) if pd.notna(nepk) else None
                rec[f"{label} NET OR"] = round(_or(nepk, orf), 0) if pd.notna(nepk) else None
                rec[f"{label} GROSS EPK"] = round(gepk, 2) if pd.notna(gepk) else None
                rec[f"{label} GROSS OR"] = round(_or(gepk, orf), 0) if pd.notna(gepk) else None
            rows.append(rec)
        out = pd.DataFrame(rows)
        if len(out):
            out = out.sort_values(["DEPOT", "TYPE OF SERVICE"])
        # Header: SCHs FOR/UPTO | each month NET/GROSS EPK OR | CY UM | LY UM
        mon_labs = [m.split("-")[0] for m in fy_months]
        thead = "<tr>" + _th("DEPOT", rowspan=3, top=0) + _th("TYPE OF SERVICE", rowspan=3, top=0)
        thead += _th("SCHs", bg="#1e40af", colspan=2, top=0)
        for lab in mon_labs:
            thead += _th(lab, bg="#0f766e", colspan=4, top=0)
        thead += _th("CY UM", bg="#7c3aed", colspan=4, top=0)
        thead += _th("LY UM", bg="#a21caf", colspan=4, top=0)
        thead += "</tr><tr>"
        thead += _th("FOR", bg="#3b82f6", top=28) + _th("UPTO", bg="#3b82f6", top=28)
        for _ in range(len(mon_labs) + 2):
            thead += _th("NET", bg="#dbeafe", color="#1e3a8a", colspan=2, top=28) + _th("GROSS", bg="#dcfce7", color="#14532d", colspan=2, top=28)
        thead += "</tr><tr>"
        thead += _th("", bg="#e2e8f0", color="#334155", top=56) + _th("", bg="#e2e8f0", color="#334155", top=56)
        for _ in range(len(mon_labs) + 2):
            for pair_bg in ("#eff6ff", "#eff6ff", "#f0fdf4", "#f0fdf4"):
                pass
            thead += _th("E.P.K", bg="#eff6ff", color="#1e3a8a", top=56) + _th("O.R", bg="#eff6ff", color="#1e3a8a", top=56)
            thead += _th("E.P.K", bg="#f0fdf4", color="#14532d", top=56) + _th("O.R", bg="#f0fdf4", color="#14532d", top=56)
        thead += "</tr>"
        body = []
        for _, r in out.iterrows():
            body.append("<tr>")
            body.append(_cell(r.get("DEPOT")))
            body.append(_cell(r.get("TYPE OF SERVICE")))
            body.append(_cell(r.get("SCHs FOR"), is_int=True))
            body.append(_cell(r.get("SCHs UPTO"), is_int=True))
            for lab in mon_labs:
                body.append(_cell(r.get(f"{lab} NET EPK")))
                body.append(_cell(r.get(f"{lab} NET OR"), is_int=True))
                body.append(_cell(r.get(f"{lab} GROSS EPK")))
                body.append(_cell(r.get(f"{lab} GROSS OR"), is_int=True))
            for label in ("CY UM", "LY UM"):
                body.append(_cell(r.get(f"{label} NET EPK")))
                body.append(_cell(r.get(f"{label} NET OR"), is_int=True))
                body.append(_cell(r.get(f"{label} GROSS EPK")))
                body.append(_cell(r.get(f"{label} GROSS OR"), is_int=True))
            body.append("</tr>")
        _render_board(title, thead, "".join(body), out, "ProductTrend")

    # ========== BOARD 6: Inter State Sector Wise ==========
    elif mf_board.startswith("6."):
        title = f"INTER STATE SECTOR WISE PERFORMANCE FOR & UPTO {mf_month} (WITH MYP-1)"
        # ROUTEE → SECTOR → STATE (only these routes)
        _sector_map_rows = [
            ("ADONI", "KURNOOL", "AP"), ("AMP", "VIJAYAWADA", "AP"), ("ATP", "KURNOOL", "AP"),
            ("AURAD", "BIDAR", "KA"), ("BLDL", "OTHER", "CS"), ("BNG", "BNG", "KA"),
            ("CHN", "ONGOLE", "TN"), ("CHINCHOLI", "OTHER", "KA"), ("CHRL", "GUNTUR", "AP"),
            ("DGLR", "OTHER", "KA"), ("ELR", "VIJAYAWADA", "AP"), ("GDV", "VIJAYAWADA", "AP"),
            ("GLB", "BIDAR", "KA"), ("GNT", "GUNTUR", "AP"), ("GURMITKAL", "YADGIR", "KA"),
            ("YDGR", "YADGIR", "KA"), ("JGDPR", "OTHER", "CS"), ("KDP", "KURNOOL", "AP"),
            ("KKD", "VIJAYAWADA", "AP"), ("KUNTA", "OTHER", "CS"), ("KRNL", "KURNOOL", "AP"),
            ("MCPTM", "VIJAYAWADA", "AP"), ("MTLM", "RAICHUR", "AP"), ("MRKP", "ONGOLE", "AP"),
            ("AMVT", "NAGPUR", "MH"), ("NGP-KRNL", "NAGPUR", "MH"), ("NDNL", "KURNOOL", "AP"),
            ("NLR", "ONGOLE", "AP"), ("ONG", "ONGOLE", "AP"), ("PUNE", "OTHER", "MH"),
            ("RCR", "RAICHUR", "KA"), ("SEDAM", "YADGIR", "KA"), ("SHRD", "OTHER", "MH"),
            ("SLPR", "OTHER", "MH"), ("SSLM", "SRISAILAM", "AP"), ("TDP", "KURNOOL", "AP"),
            ("TNL", "VIJAYAWADA", "AP"), ("TPT", "ONGOLE", "AP"), ("UDGIR", "BIDAR", "MH"),
            ("VJA", "VIJAYAWADA", "AP"), ("VSP", "VIJAYAWADA", "AP"), ("SUKMA", "OTHER", "CS"),
            ("NAGPUR", "NAGPUR", "MH"), ("BLRY", "RAICHUR", "KA"), ("BDR", "BIDAR", "KA"),
            ("GNPR", "BIDAR", "KA"), ("SINDHANUR", "RAICHUR", "KA"),
        ]
        _route_to_sec = {}
        for _rt, _sec, _st in _sector_map_rows:
            _k = str(_rt).strip().upper()
            if _k not in _route_to_sec:
                _route_to_sec[_k] = (str(_sec).strip().upper(), str(_st).strip().upper())

        def _norm_route6(v):
            s = str(v).strip().upper() if v is not None and not (isinstance(v, float) and pd.isna(v)) else ""
            if s.endswith(".0"):
                s = s[:-2]
            return s

        _route_col6 = "ROUTEE" if "ROUTEE" in mf.columns else ("ROUTE" if "ROUTE" in mf.columns else None)
        if not _route_col6:
            st.error("ROUTEE / ROUTE column not found.")
            st.stop()

        def _filter_sector_routes(data):
            if len(data) == 0:
                return data
            d = data.copy()
            d["_RK"] = d[_route_col6].map(_norm_route6)
            d = d[d["_RK"].isin(_route_to_sec.keys())]
            return d

        # Apply depot / product / gross-net context from top filters already in cm/cy_um/ly_um via base
        _cm6 = _filter_sector_routes(cm)
        _cy6 = _filter_sector_routes(cy_um)
        _ly6 = _filter_sector_routes(ly_um)
        if mf_depot != "ALL":
            for _name in ("_cm6", "_cy6", "_ly6"):
                pass
            if "DEPOT" in _cm6.columns:
                _cm6 = _cm6[_cm6["DEPOT"].astype(str).str.strip().str.upper() == str(mf_depot).strip().upper()]
            if "DEPOT" in _cy6.columns:
                _cy6 = _cy6[_cy6["DEPOT"].astype(str).str.strip().str.upper() == str(mf_depot).strip().upper()]
            if "DEPOT" in _ly6.columns:
                _ly6 = _ly6[_ly6["DEPOT"].astype(str).str.strip().str.upper() == str(mf_depot).strip().upper()]
        if mf_product != "ALL" and "PRODUCT" in mf.columns:
            if "PRODUCT" in _cm6.columns:
                _cm6 = _cm6[_cm6["PRODUCT"].astype(str).str.strip().str.upper() == str(mf_product).strip().upper()]
            if "PRODUCT" in _cy6.columns:
                _cy6 = _cy6[_cy6["PRODUCT"].astype(str).str.strip().str.upper() == str(mf_product).strip().upper()]
            if "PRODUCT" in _ly6.columns:
                _ly6 = _ly6[_ly6["PRODUCT"].astype(str).str.strip().str.upper() == str(mf_product).strip().upper()]

        if len(_cy6) == 0 and len(_cm6) == 0:
            st.warning("No rows for mapped inter-state routes with current filters.")
            st.caption("Mapped routes: " + ", ".join(sorted(_route_to_sec.keys())))
            st.stop()

        keys = ["ROUTEE", "DEPOT", "PRODUCT", "RTC_HIRE"]

        def agg6(data):
            """Schs / Sch Kms: max per service then sum (same fix as Board 1)."""
            if len(data) == 0:
                return pd.DataFrame()
            d = data.copy()
            if "ROUTEE" not in d.columns and _route_col6 != "ROUTEE":
                d["ROUTEE"] = d[_route_col6]
            d["ROUTEE"] = d["ROUTEE"].map(_norm_route6)
            if "SER_NO" not in d.columns:
                return d.groupby(keys, dropna=False).agg(
                    schs=("NO_OF_SCHS", "max"),
                    sers=("NO_OF_SCHS", "count"),
                    sch_kms=("SCH_KMS", "max"),
                    kms=("Optd_KMs", "sum"),
                    ge=("GE_TOT", "sum"),
                    ne=("NE_TOT", "sum"),
                ).reset_index()
            if "R_L" not in d.columns:
                d["R_L"] = 0
            ser_keys = keys + ["SER_NO"]
            per = d.groupby(ser_keys, dropna=False).agg(
                schs=("NO_OF_SCHS", "max"),
                sch_kms=("SCH_KMS", "max"),
                rl=("R_L", "max"),
                kms=("Optd_KMs", "sum"),
                ge=("GE_TOT", "sum"),
                ne=("NE_TOT", "sum"),
            ).reset_index()
            for i, row in per.iterrows():
                sk = float(row.get("sch_kms", 0) or 0)
                rl = float(row.get("rl", 0) or 0)
                sc = float(row.get("schs", 0) or 0)
                if sk == 0 and rl > 0 and sc > 0:
                    per.at[i, "sch_kms"] = rl * sc
            return per.groupby(keys, dropna=False).agg(
                schs=("schs", "sum"),
                sers=("SER_NO", "nunique"),
                sch_kms=("sch_kms", "sum"),
                kms=("kms", "sum"),
                ge=("ge", "sum"),
                ne=("ne", "sum"),
            ).reset_index()

        a_cm, a_cy, a_ly = agg6(_cm6), agg6(_cy6), agg6(_ly6)
        merged = a_cy.merge(a_cm, on=keys, how="outer", suffixes=("_cy", "_cm"))
        merged = merged.merge(a_ly, on=keys, how="outer")
        for c in ["schs", "sers", "sch_kms", "kms", "ge", "ne"]:
            if c in merged.columns:
                merged.rename(columns={c: f"{c}_ly"}, inplace=True)

        rows = []
        for _, r in merged.fillna(0).iterrows():
            dep = str(r.get("DEPOT", ""))
            route = _norm_route6(r.get("ROUTEE", ""))
            if route not in _route_to_sec:
                continue
            sector, state = _route_to_sec[route]
            orf_c, orf_l = _orf(dep, side="cy"), _orf(dep, side="ly")
            def side(ge, ne, kms, orf):
                g, n = _epk(ge, kms), _epk(ne, kms)
                return g, n, _or(g, orf), _or(n, orf)
            g_cm, n_cm, og_cm, on_cm = side(r.get("ge_cm", 0), r.get("ne_cm", 0), r.get("kms_cm", 0), orf_c)
            g_cy, n_cy, og_cy, on_cy = side(r.get("ge_cy", 0), r.get("ne_cy", 0), r.get("kms_cy", 0), orf_c)
            g_ly, n_ly, og_ly, on_ly = side(r.get("ge_ly", 0), r.get("ne_ly", 0), r.get("kms_ly", 0), orf_l)
            var_epk = (g_cy - g_ly) if pd.notna(g_cy) and pd.notna(g_ly) else np.nan
            var_or = (og_cy - og_ly) if pd.notna(og_cy) and pd.notna(og_ly) else np.nan
            rows.append({
                "SL NO.": len(rows) + 1,
                "SECTOR": sector,
                "STATE": state,
                "DEPOT": dep,
                "ROUTE": route,
                "LONG_TP.": r.get("PRODUCT", ""),
                "RTC/HIRE": r.get("RTC_HIRE", ""),
                "No. of Schs": int(round(float(r.get("schs_cy", 0) or 0))) or "",
                "No. of Ser": int(round(float(r.get("sers_cy", 0) or 0))) or "",
                "Sch Kms": _f0(r.get("sch_kms_cy", 0)),
                "GROSS EPK CM": round(g_cm, 2) if pd.notna(g_cm) else None,
                "GROSS EPK UM": round(g_cy, 2) if pd.notna(g_cy) else None,
                "GROSS EPK LYUM": round(g_ly, 2) if pd.notna(g_ly) else None,
                "NET EPK CM": round(n_cm, 2) if pd.notna(n_cm) else None,
                "NET EPK UM": round(n_cy, 2) if pd.notna(n_cy) else None,
                "NET EPK LYUM": round(n_ly, 2) if pd.notna(n_ly) else None,
                "GROSS OR CM": round(og_cm, 0) if pd.notna(og_cm) else None,
                "GROSS OR UM": round(og_cy, 0) if pd.notna(og_cy) else None,
                "GROSS OR LYUM": round(og_ly, 0) if pd.notna(og_ly) else None,
                "NET OR CM": round(on_cm, 0) if pd.notna(on_cm) else None,
                "NET OR UM": round(on_cy, 0) if pd.notna(on_cy) else None,
                "NET OR LYUM": round(on_ly, 0) if pd.notna(on_ly) else None,
                "VAR UM EPK": round(var_epk, 2) if pd.notna(var_epk) else None,
                "VAR UM OR": round(var_or, 0) if pd.notna(var_or) else None,
            })
        out = pd.DataFrame(rows)
        if len(out):
            out = out.sort_values(["SECTOR", "STATE", "DEPOT", "ROUTE"]).reset_index(drop=True)
            out["SL NO."] = range(1, len(out) + 1)
        thead = (
            "<tr>"
            + _th("SL<br>NO.", rowspan=2, top=0) + _th("SECTOR", rowspan=2, top=0)
            + _th("STATE", rowspan=2, top=0)
            + _th("DEPOT", rowspan=2, top=0) + _th("ROUTE", rowspan=2, top=0)
            + _th("LONG_TP.", rowspan=2, top=0) + _th("RTC/<br>HIRE", rowspan=2, top=0)
            + _th("No.<br>of<br>Schs", rowspan=2, top=0) + _th("No.<br>of<br>Ser", rowspan=2, top=0)
            + _th("Sch<br>Kms", rowspan=2, top=0)
            + _th("GROSS E P K", bg="#dcfce7", color="#14532d", colspan=3, top=0)
            + _th("NET E P K", bg="#fef3c7", color="#92400e", colspan=3, top=0)
            + _th("GROSS OR", bg="#dbeafe", color="#1e3a8a", colspan=3, top=0)
            + _th("NET OR", bg="#ede9fe", color="#5b21b6", colspan=3, top=0)
            + _th("VAR- GROSS", bg="#fee2e2", color="#991b1b", colspan=2, top=0)
            + "</tr><tr>"
        )
        for grp in range(4):
            thead += _th("CM", bg="#f1f5f9", color="#334155", top=28) + _th("UM", bg="#f1f5f9", color="#334155", top=28) + _th("LYUM", bg="#f1f5f9", color="#334155", top=28)
        thead += _th("UM EPK", bg="#fecaca", color="#991b1b", top=28) + _th("UM OR", bg="#fecaca", color="#991b1b", top=28)
        thead += "</tr>"
        body = []
        cols6 = [
            "SL NO.", "SECTOR", "STATE", "DEPOT", "ROUTE", "LONG_TP.", "RTC/HIRE",
            "No. of Schs", "No. of Ser", "Sch Kms",
            "GROSS EPK CM", "GROSS EPK UM", "GROSS EPK LYUM",
            "NET EPK CM", "NET EPK UM", "NET EPK LYUM",
            "GROSS OR CM", "GROSS OR UM", "GROSS OR LYUM",
            "NET OR CM", "NET OR UM", "NET OR LYUM",
            "VAR UM EPK", "VAR UM OR",
        ]
        for _, r in out.iterrows():
            body.append("<tr>")
            for c in cols6:
                body.append(_cell(r.get(c)))
            body.append("</tr>")
        _render_board(title, thead, "".join(body), out, "InterState")
        st.caption(f"Routes limited to sector map ({len(_route_to_sec)} codes) · Schs/Sch Kms = max per service then sum")

    # ========== BOARD 7: Service Wise Performance ==========
    # Master list = SMASTER only; table = identity + Optd + GROSS/NET EPK&OR (TOT/MHL/FPD × CM/UM/LUM)
    elif mf_board.startswith("7."):
        title = f"RR REGION SERVICE WISE PERFORMANCE FOR AND UPTO {mf_month}"
        if mf_depot != "ALL":
            title += f" | Depot: {mf_depot}"
        if mf_route != "ALL":
            title += f" | Route: {mf_route}"
        if mf_product != "ALL":
            title += f" | Product: {mf_product}"

        def _norm_svc7(v):
            if v is None or (isinstance(v, float) and pd.isna(v)):
                return ""
            s = str(v).strip()
            if s.lower() in ("nan", "none", ""):
                return ""
            if s.endswith(".0"):
                s = s[:-2]
            try:
                return str(int(float(s)))
            except Exception:
                return s

        def _mode_or_first(series):
            s = series.dropna().astype(str).str.strip()
            s = s[~s.str.lower().isin(["", "nan", "none"])]
            if len(s) == 0:
                return ""
            try:
                return s.mode().iloc[0]
            except Exception:
                return s.iloc[0]

        def _pick(*vals):
            for v in vals:
                if v is None or (isinstance(v, float) and pd.isna(v)):
                    continue
                s = str(v).strip()
                if s and s.lower() not in ("nan", "none", "0", "0.0"):
                    return v
            return ""

        # ---- Load SMASTER (required) ----
        _sros_p = Path(r"D:\Dashboard\SMASTER.parquet")
        if not _sros_p.exists():
            for alt in [
                Path(r"D:\dashboard\SMASTER.parquet"),
                Path(r"D:\MONTHLY\SMASTER.parquet"),
                Path("SMASTER.parquet"),
                Path(r"/home/workdir/attachments/SMASTER.parquet"),
            ]:
                if alt.exists():
                    _sros_p = alt
                    break
        if not _sros_p.exists():
            st.error("SMASTER.parquet not found — Board 7 lists services from SMASTER only.")
            st.stop()
        try:
            _sm_raw, _sm_err = load_smaster(str(_sros_p))
            if _sm_raw is None:
                raise RuntimeError(_sm_err or "Unable to load SMASTER.parquet")
            _sm_raw = _sm_raw.copy()
            _sm_raw.columns = [str(c).strip() for c in _sm_raw.columns]
        except Exception as _e:
            st.error(f"Could not read SMASTER: {_e}")
            st.stop()

        def _fc7(cands, frame=_sm_raw):
            n = {
                str(c).strip().lower().replace(" ", "").replace("_", "").replace("/", "").replace(".", ""): c
                for c in frame.columns
            }
            for cand in cands:
                k = cand.lower().replace(" ", "").replace("_", "").replace("/", "").replace(".", "")
                if k in n:
                    return n[k]
            for c in frame.columns:
                cl = str(c).strip().lower().replace(" ", "").replace("_", "").replace("/", "").replace(".", "")
                for cand in cands:
                    if cand.lower().replace(" ", "").replace("_", "").replace("/", "").replace(".", "") in cl:
                        return c
            return None

        c_svc = _fc7(["ServiceNo", "SER_NO", "SERVICE_NO", "SERVICENO"])
        c_dep = _fc7(["DEPOT"])
        c_prod = _fc7(["PRODUCT"])
        c_route = _fc7(["ROUTEE", "Routee", "ROUTE"])
        c_sch = _fc7(["NoOfSchedules", "NoOfSchedule", "NO_OF_SCHS"])
        c_kms = _fc7(["RevenueKms", "Revenue Kms", "SCH_KMS", "SchKms", "DAY_SCH_KMS"])
        c_mon = _fc7(["MONTH", "Month"])
        c_year = _fc7(["YEAR", "Year"])
        c_rtc = _fc7(["RTC/HIRE", "RTC_HIRE", "RTCHIRE"])
        c_rl = _fc7(["R/L", "R_L", "RL", "RouteLength"])
        c_schdep = _fc7(["SCH_DEP", "SCH_DEP.", "SchDep", "SCHDEP", "ScheduleDepot"])
        c_type = _fc7(["TYPE", "TYPE OF SERV.", "ServiceType", "TP"])
        # D/N from D.TYPE (SMASTER) — DO/SC/SO/NO etc.
        c_dtype = _fc7(["D.TYPE", "DTYPE", "D.Type", "DutyType", "DTYPE OF SERV"])
        c_nature = _fc7(["NATURE", "NATURE OF SERV.", "D/N", "DN", "DayNight"])
        if not c_svc:
            st.error(f"SMASTER missing ServiceNo. Columns: {list(_sm_raw.columns)}")
            st.stop()

        sm = _sm_raw.copy()
        sm["_SVC"] = sm[c_svc].map(_norm_svc7)
        sm["_DEP"] = sm[c_dep].astype(str).str.strip().str.upper() if c_dep else ""
        sm["_PROD"] = sm[c_prod].astype(str).str.strip() if c_prod else ""
        sm["_ROUTE"] = sm[c_route].astype(str).str.strip() if c_route else ""
        sm["_SCH"] = pd.to_numeric(sm[c_sch], errors="coerce").fillna(0) if c_sch else 0.0
        sm["_SCH_KMS"] = pd.to_numeric(sm[c_kms], errors="coerce").fillna(0) if c_kms else 0.0
        sm["_RL"] = pd.to_numeric(sm[c_rl], errors="coerce").fillna(0) if c_rl else 0.0
        sm["_RTC"] = sm[c_rtc].astype(str).str.strip() if c_rtc else ""
        sm["_SCH_DEP"] = sm[c_schdep].astype(str).str.strip() if c_schdep else ""
        sm["_TYPE"] = sm[c_type].astype(str).str.strip() if c_type else ""
        # Prefer D.TYPE for D/N column
        if c_dtype:
            sm["_DN"] = sm[c_dtype].astype(str).str.strip().str.upper()
            sm["_DN"] = sm["_DN"].replace({
                "DAY OUT": "DO", "DAYOUT": "DO", "DO": "DO",
                "NIGHT OUT": "NO", "NIGHTOUT": "NO", "NO": "NO",
                "SPECIAL": "SC", "SC": "SC", "SO": "SO",
            })
        elif c_nature:
            sm["_DN"] = sm[c_nature].astype(str).str.strip()
        else:
            sm["_DN"] = ""
        sm["_NATURE"] = sm["_DN"]
        if c_mon and c_year:
            def _mk7(row):
                try:
                    return f"{str(row[c_mon]).strip()[:3].title()}-{int(float(row[c_year]))}"
                except Exception:
                    return ""
            sm["_MK"] = sm.apply(_mk7, axis=1)
        else:
            sm["_MK"] = "ALL"

        if mf_depot not in ("ALL", "REGION", ""):
            sm = sm[sm["_DEP"] == str(mf_depot).strip().upper()]
        if mf_product != "ALL":
            sm = sm[sm["_PROD"].astype(str).str.strip().str.upper() == str(mf_product).strip().upper()]
        if mf_route != "ALL":
            sm = sm[sm["_ROUTE"].astype(str).str.strip().str.upper() == str(mf_route).strip().upper()]
        sm_mon = sm[sm["_MK"] == str(mf_month)].copy() if str(mf_month) else sm.copy()
        if len(sm_mon) == 0:
            st.warning(f"No SMASTER services for month {mf_month} with current Depot/Route/Product filters.")
            st.stop()

        # Service + Low EPK + Negative EPK filters
        svc_vals = {_norm_svc7(x) for x in sm_mon["_SVC"].dropna().unique()}
        svc_vals.discard("")
        def _svc_sort(z):
            try:
                return (0, int(z))
            except Exception:
                return (1, str(z))
        svc_opts7 = ["ALL"] + sorted(svc_vals, key=_svc_sort)
        low_opts = [
            "ALL",
            "TOT EPK < 40",
            "TOT EPK < 35",
            "TOT EPK < 30",
            "TOT EPK < 25",
            "FPD EPK < 25",
            "FPD EPK < 20",
        ]
        neg_opts = ["ALL", "NEG TOT EPK", "NEG FPD EPK"]
        f7a, f7b, f7c = st.columns(3)
        with f7a:
            mf_svc = st.selectbox("Select Service (SER NO)", svc_opts7, index=0, key="mf_b7_service")
        with f7b:
            mf_low = st.selectbox("Low EPK Ser", low_opts, index=0, key="mf_b7_low")
        with f7c:
            mf_neg = st.selectbox("Negative EPK Ser", neg_opts, index=0, key="mf_b7_neg")
        if mf_svc != "ALL":
            sm_mon = sm_mon[sm_mon["_SVC"].map(_norm_svc7) == str(mf_svc).strip()]
            title += f" | Service: {mf_svc}"
        if mf_low != "ALL":
            title += f" | {mf_low}"
        if mf_neg != "ALL":
            title += f" | {mf_neg}"

        master = (
            sm_mon.groupby(["_DEP", "_SVC"], dropna=False)
            .agg(
                route=("_ROUTE", _mode_or_first),
                prod=("_PROD", _mode_or_first),
                rtc=("_RTC", _mode_or_first),
                sch_dep=("_SCH_DEP", _mode_or_first),
                typ=("_TYPE", _mode_or_first),
                nature=("_DN", _mode_or_first),
                rl=("_RL", "max"),
                schs=("_SCH", "sum"),
                sch_kms=("_SCH_KMS", "sum"),
            )
            .reset_index()
            .rename(columns={"_DEP": "DEPOT", "_SVC": "SER_NO"})
        )
        master["sers"] = 1
        for i, row in master.iterrows():
            sk = float(row.get("sch_kms", 0) or 0)
            rl = float(row.get("rl", 0) or 0)
            sc = float(row.get("schs", 0) or 0)
            if sk == 0 and rl > 0 and sc > 0:
                master.at[i, "sch_kms"] = rl * sc

        # `master` is SMASTER-only for schedule counts/identity.  However,
        # Sch Dep, R/L and Type must come from the monthly service file.
        # Ensure SCH_DEP / R_L exist on monthly frame (fuzzy) for Board 7 meta
        def _ensure_sch_rl(frame):
            if frame is None or len(frame) == 0:
                return frame
            def nk(c):
                return str(c).strip().lower().replace("_", "").replace(" ", "").replace("/", "").replace(".", "")
            if "SCH_DEP" not in frame.columns:
                for c in frame.columns:
                    k = nk(c)
                    if ("sch" in k and "dep" in k) or k in ("schdep", "scheduledepot", "schdepot"):
                        frame = frame.copy()
                        frame["SCH_DEP"] = frame[c].astype(str)
                        break
            if "R_L" not in frame.columns:
                for c in frame.columns:
                    k = nk(c)
                    if k in ("rl", "routelength", "routelen") or (k.startswith("rl") and "epk" not in k and len(k) <= 6):
                        frame = frame.copy()
                        frame["R_L"] = pd.to_numeric(frame[c], errors="coerce").fillna(0.0)
                        break
            return frame
        _svc7 = _ensure_sch_rl(_svc7)

        _meta7 = _svc7[_svc7["Month_Name"].astype(str).str.strip() == str(mf_month).strip()].copy()
        if len(_meta7):
            _meta7["_SVC"] = _meta7["SER_NO"].map(_norm_svc7)
            _meta7["_DEP"] = _meta7["DEPOT"].astype(str).str.strip().str.upper()
            def _first_nonblank7(series):
                z = series.dropna().astype(str).str.strip()
                z = z[~z.str.lower().isin(["", "nan", "none"])]
                return z.mode().iloc[0] if len(z) else ""
            mcols7 = ["_DEP", "_SVC"]
            agg7_meta = {}
            if "SCH_DEP" in _meta7.columns:
                agg7_meta["sch_dep_m"] = ("SCH_DEP", _first_nonblank7)
            if "R_L" in _meta7.columns:
                agg7_meta["rl_m"] = ("R_L", "max")
            if "PRODUCT" in _meta7.columns:
                agg7_meta["typ_m"] = ("PRODUCT", _first_nonblank7)
            if agg7_meta:
                monthly_meta7 = _meta7.groupby(mcols7, dropna=False).agg(**agg7_meta).reset_index()
                master = master.merge(
                    monthly_meta7.rename(columns={"_DEP":"DEPOT","_SVC":"SER_NO"}),
                    on=["DEPOT","SER_NO"], how="left"
                )
                def _nonblank_series(s):
                    s = s.astype(str).str.strip()
                    return s.where(~s.str.lower().isin(["", "nan", "none", "nat", "<na>"]), other=pd.NA)

                if "sch_dep_m" in master.columns:
                    # Prefer monthly Sch Dep; fall back to SMASTER when blank
                    m = _nonblank_series(master["sch_dep_m"])
                    fb = master["sch_dep"].astype(str).str.strip() if "sch_dep" in master.columns else ""
                    master["sch_dep"] = m.fillna(fb).fillna("").astype(str).str.strip()
                if "rl_m" in master.columns:
                    # Prefer monthly R/L; fall back to SMASTER when 0/blank
                    mrl = pd.to_numeric(master["rl_m"], errors="coerce")
                    fbrl = pd.to_numeric(master.get("rl", 0), errors="coerce").fillna(0.0)
                    master["rl"] = mrl.where(mrl.fillna(0) != 0, fbrl).fillna(fbrl)
                if "typ_m" in master.columns:
                    m = _nonblank_series(master["typ_m"])
                    fb = master["typ"].astype(str).str.strip() if "typ" in master.columns else ""
                    master["typ"] = m.fillna(fb).fillna("").astype(str).str.strip()
                master = master.drop(columns=[c for c in ["sch_dep_m", "rl_m", "typ_m"] if c in master.columns])

        allowed = set(zip(master["DEPOT"].astype(str).str.upper(), master["SER_NO"].map(_norm_svc7)))

        def _b7_filter(data):
            out_d = data.copy() if len(data) else data
            if len(out_d) == 0:
                return out_d
            if mf_depot != "ALL" and "DEPOT" in out_d.columns:
                out_d = out_d[out_d["DEPOT"].astype(str).str.strip().str.upper() == str(mf_depot).strip().upper()]
            if mf_route != "ALL" and _route_col_mf and _route_col_mf in out_d.columns:
                out_d = out_d[out_d[_route_col_mf].astype(str).str.strip().str.upper() == str(mf_route).strip().upper()]
            if mf_product != "ALL" and "PRODUCT" in out_d.columns:
                out_d = out_d[out_d["PRODUCT"].astype(str).str.strip().str.upper() == str(mf_product).strip().upper()]
            if "SER_NO" in out_d.columns and "DEPOT" in out_d.columns:
                mask = out_d.apply(
                    lambda r: (str(r["DEPOT"]).strip().upper(), _norm_svc7(r["SER_NO"])) in allowed,
                    axis=1,
                )
                out_d = out_d[mask]
            if mf_svc != "ALL" and "SER_NO" in out_d.columns:
                out_d = out_d[out_d["SER_NO"].map(_norm_svc7) == str(mf_svc).strip()]
            return out_d

        # Performance metrics for Board 7 come ONLY from ser_monthly/ser_montly.
        # Service identity and displayed Ser No come ONLY from SMASTER `SER NO`.
        _svc7, _svc7_err = load_monthly_service_metrics()
        if _svc7_err:
            st.error(f"Could not read service-monthly parquet: {_svc7_err}")
            st.stop()
        if _svc7 is None or len(_svc7) == 0:
            st.error("Service-monthly parquet not found/empty. Checked ser_monthly.parquet and ser_montly.parquet.")
            st.stop()
        _svc7 = _svc7.copy()
        if "SER_NO" not in _svc7.columns or "DEPOT" not in _svc7.columns:
            st.error("Service-monthly parquet must contain DEPOT and SER_NO for SMASTER matching.")
            st.stop()
        _svc7["SER_NO"] = _svc7["SER_NO"].map(_norm_svc7)
        _svc7["DEPOT"] = _svc7["DEPOT"].astype(str).str.strip().str.upper()
        # Board 7 source rule:
        #   SMASTER -> Depot, Ser No, Route, RTC/HIRE, D/N, No.of Schs,
        #              No.of Ser, Sch.Kms
        #   ser_monthly -> Sch Dep, R/L, TYPE (= PRODUCT)
        # Monthly attributes are merged later by DEPOT+SER_NO and NEVER used
        # for schedule counts.
        # Metrics may contain services not present in the selected SMASTER month;
        # never let those create rows in Board 7.
        _svc7 = _svc7[_svc7.apply(
            lambda rr: (rr["DEPOT"], _norm_svc7(rr["SER_NO"])) in allowed, axis=1
        )].copy()
        _cm7 = _svc7[_svc7["Month_Name"].astype(str).str.strip() == str(mf_month).strip()].copy()
        # The monthly metrics parquet is month-grain data. It does not need
        # Date or Weekday; CY/LY are selected from Month_Name.
        _fy7_months = [x.strftime("%b-%Y") for x in pd.date_range(fy_start, cy_end, freq="MS")]
        _ly7_months = [(pd.to_datetime(x, format="%b-%Y") - pd.DateOffset(years=1)).strftime("%b-%Y") for x in _fy7_months]
        _cy7 = _svc7[_svc7["Month_Name"].astype(str).str.strip().isin(_fy7_months)].copy()
        _ly7 = _svc7[_svc7["Month_Name"].astype(str).str.strip().isin(_ly7_months)].copy()

        keys = ["DEPOT", "SER_NO"]

        def agg7_perf(data):
            if len(data) == 0:
                return pd.DataFrame(columns=keys + [
                    "kms", "ge_tot", "ge_mhl", "ge_fpd", "ne_tot", "ne_mhl", "ne_fpd", "days",
                ])
            d = data.copy()
            d["SER_NO"] = d["SER_NO"].map(_norm_svc7)
            d["DEPOT"] = d["DEPOT"].astype(str).str.strip().str.upper()
            rc = "ROUTEE" if "ROUTEE" in d.columns else ("ROUTE" if "ROUTE" in d.columns else None)
            if rc and "ROUTEE" not in d.columns:
                d["ROUTEE"] = d[rc]
            # IMPORTANT: this aggregation receives ONLY the monthly metrics source.
            # Identity/schedule columns are never requested here; they come from
            # `master`, which was built exclusively from SMASTER above.
            agg_map = {
                "kms": ("Optd_KMs", "sum"),
                "ge_tot": ("GE_TOT", "sum"),
                "ge_mhl": ("GE_MHL", "sum"),
                "ge_fpd": ("GE_FPD", "sum"),
                "ne_tot": ("NE_TOT", "sum"),
                "ne_mhl": ("NE_MHL", "sum"),
                "ne_fpd": ("NE_FPD", "sum"),
            }
            if "DAYS" in d.columns:
                agg_map["days"] = ("DAYS", "sum")
            elif "Date" in d.columns:
                agg_map["days"] = ("Date", "nunique")
            else:
                # Monthly parquet has one service/month record; this is only a
                # fallback for the display count and does not affect EPK/OR.
                agg_map["days"] = ("Month_Name", "nunique")
            return d.groupby(keys, dropna=False).agg(**agg_map).reset_index()

        a_cm, a_cy, a_ly = agg7_perf(_cm7), agg7_perf(_cy7), agg7_perf(_ly7)
        merged = master.merge(a_cy, on=keys, how="left")
        merged = merged.merge(a_cm, on=keys, how="left", suffixes=("_cy", "_cm"))
        merged = merged.merge(a_ly, on=keys, how="left")
        for c in ["kms", "ge_tot", "ge_mhl", "ge_fpd", "ne_tot", "ne_mhl", "ne_fpd", "days"]:
            if c in merged.columns:
                merged.rename(columns={c: f"{c}_ly"}, inplace=True)

        def _trio(et, em, ef, kms, orf):
            a, b, c = _epk(et, kms), _epk(em, kms), _epk(ef, kms)
            return a, b, c, _or(a, orf), _or(b, orf), _or(c, orf)

        def _r2(v):
            return round(v, 2) if pd.notna(v) else None

        def _r0(v):
            return round(v, 0) if pd.notna(v) else None

        rows = []
        for _, r in merged.fillna(0).iterrows():
            dep = str(r.get("DEPOT", "")).strip().upper()
            ser = _norm_svc7(r.get("SER_NO", ""))
            if not ser:
                continue
            orf_c, orf_l = _orf(dep, side="cy"), _orf(dep, side="ly")
            k_cm = float(r.get("kms_cm", 0) or 0)
            k_cy = float(r.get("kms_cy", 0) or 0)
            k_ly = float(r.get("kms_ly", 0) or 0)
            gt_cm, gm_cm, gf_cm, ot_cm, om_cm, of_cm = _trio(
                r.get("ge_tot_cm", 0), r.get("ge_mhl_cm", 0), r.get("ge_fpd_cm", 0), k_cm, orf_c)
            gt_um, gm_um, gf_um, ot_um, om_um, of_um = _trio(
                r.get("ge_tot_cy", 0), r.get("ge_mhl_cy", 0), r.get("ge_fpd_cy", 0), k_cy, orf_c)
            gt_lm, gm_lm, gf_lm, ot_lm, om_lm, of_lm = _trio(
                r.get("ge_tot_ly", 0), r.get("ge_mhl_ly", 0), r.get("ge_fpd_ly", 0), k_ly, orf_l)
            nt_cm, nm_cm, nf_cm, ont_cm, onm_cm, onf_cm = _trio(
                r.get("ne_tot_cm", 0), r.get("ne_mhl_cm", 0), r.get("ne_fpd_cm", 0), k_cm, orf_c)
            nt_um, nm_um, nf_um, ont_um, onm_um, onf_um = _trio(
                r.get("ne_tot_cy", 0), r.get("ne_mhl_cy", 0), r.get("ne_fpd_cy", 0), k_cy, orf_c)
            nt_lm, nm_lm, nf_lm, ont_lm, onm_lm, onf_lm = _trio(
                r.get("ne_tot_ly", 0), r.get("ne_mhl_ly", 0), r.get("ne_fpd_ly", 0), k_ly, orf_l)

            # Low EPK filter — based on UM (upto month) gross metrics
            if mf_low != "ALL":
                if mf_low == "TOT EPK < 40" and not (pd.notna(gt_um) and gt_um < 40):
                    continue
                if mf_low == "TOT EPK < 35" and not (pd.notna(gt_um) and gt_um < 35):
                    continue
                if mf_low == "TOT EPK < 30" and not (pd.notna(gt_um) and gt_um < 30):
                    continue
                if mf_low == "TOT EPK < 25" and not (pd.notna(gt_um) and gt_um < 25):
                    continue
                if mf_low == "FPD EPK < 25" and not (pd.notna(gf_um) and gf_um < 25):
                    continue
                if mf_low == "FPD EPK < 20" and not (pd.notna(gf_um) and gf_um < 20):
                    continue

            # Negative EPK: UM vs LUM (UM < LUM → decline)
            neg_tot = pd.notna(gt_um) and pd.notna(gt_lm) and (gt_um - gt_lm) < 0
            neg_fpd = pd.notna(gf_um) and pd.notna(gf_lm) and (gf_um - gf_lm) < 0
            if mf_neg == "NEG TOT EPK" and not neg_tot:
                continue
            if mf_neg == "NEG FPD EPK" and not neg_fpd:
                continue

            schs_v = float(r.get("schs", 0) or 0)
            sch_kms_v = float(r.get("sch_kms", 0) or 0)
            rl_v = float(r.get("rl", 0) or 0) or float(r.get("rl_p_cy", 0) or 0) or float(r.get("rl_p_cm", 0) or 0)
            if sch_kms_v == 0 and rl_v > 0 and schs_v > 0:
                sch_kms_v = rl_v * schs_v

            rows.append({
                "SL NO": len(rows) + 1,
                "Depot": dep,
                "Ser No": ser,
                "Sch Dep": _pick(r.get("sch_dep")),
                "Route": _pick(r.get("route")),
                "RTC/HIRE": _pick(r.get("rtc")),
                "R/L": _f0(rl_v),
                "Type": _pick(r.get("typ")),
                "D/N": _pick(r.get("nature")),
                "No.of Schs": int(round(schs_v)) if schs_v else "",
                "No.of Sers": 1,
                "Sch. Kms": _f0(sch_kms_v),
                "Optd CM": int(r.get("days_cm", 0) or 0) or "",
                "Optd UM": int(r.get("days_cy", 0) or 0) or "",
                # GROSS EPK
                "G EPK CM TOT": _r2(gt_cm), "G EPK CM MHL": _r2(gm_cm), "G EPK CM FPD": _r2(gf_cm),
                "G EPK UM TOT": _r2(gt_um), "G EPK UM MHL": _r2(gm_um), "G EPK UM FPD": _r2(gf_um),
                "G EPK LUM TOT": _r2(gt_lm), "G EPK LUM MHL": _r2(gm_lm), "G EPK LUM FPD": _r2(gf_lm),
                # GROSS OR
                "G OR CM TOT": _r0(ot_cm), "G OR CM MHL": _r0(om_cm), "G OR CM FPD": _r0(of_cm),
                "G OR UM TOT": _r0(ot_um), "G OR UM MHL": _r0(om_um), "G OR UM FPD": _r0(of_um),
                "G OR LUM TOT": _r0(ot_lm), "G OR LUM MHL": _r0(om_lm), "G OR LUM FPD": _r0(of_lm),
                # NET EPK
                "N EPK CM TOT": _r2(nt_cm), "N EPK CM MHL": _r2(nm_cm), "N EPK CM FPD": _r2(nf_cm),
                "N EPK UM TOT": _r2(nt_um), "N EPK UM MHL": _r2(nm_um), "N EPK UM FPD": _r2(nf_um),
                "N EPK LUM TOT": _r2(nt_lm), "N EPK LUM MHL": _r2(nm_lm), "N EPK LUM FPD": _r2(nf_lm),
                # NET OR
                "N OR CM TOT": _r0(ont_cm), "N OR CM MHL": _r0(onm_cm), "N OR CM FPD": _r0(onf_cm),
                "N OR UM TOT": _r0(ont_um), "N OR UM MHL": _r0(onm_um), "N OR UM FPD": _r0(onf_um),
                "N OR LUM TOT": _r0(ont_lm), "N OR LUM MHL": _r0(onm_lm), "N OR LUM FPD": _r0(onf_lm),
                # flags for highlight: UM < LUM
                "_neg_g_tot": 1 if (pd.notna(gt_um) and pd.notna(gt_lm) and gt_um < gt_lm) else 0,
                "_neg_g_mhl": 1 if (pd.notna(gm_um) and pd.notna(gm_lm) and gm_um < gm_lm) else 0,
                "_neg_g_fpd": 1 if (pd.notna(gf_um) and pd.notna(gf_lm) and gf_um < gf_lm) else 0,
                "_neg_n_tot": 1 if (pd.notna(nt_um) and pd.notna(nt_lm) and nt_um < nt_lm) else 0,
                "_neg_n_mhl": 1 if (pd.notna(nm_um) and pd.notna(nm_lm) and nm_um < nm_lm) else 0,
                "_neg_n_fpd": 1 if (pd.notna(nf_um) and pd.notna(nf_lm) and nf_um < nf_lm) else 0,
                "_neg_go_tot": 1 if (pd.notna(ot_um) and pd.notna(ot_lm) and ot_um < ot_lm) else 0,
                "_neg_go_mhl": 1 if (pd.notna(om_um) and pd.notna(om_lm) and om_um < om_lm) else 0,
                "_neg_go_fpd": 1 if (pd.notna(of_um) and pd.notna(of_lm) and of_um < of_lm) else 0,
                "_neg_no_tot": 1 if (pd.notna(ont_um) and pd.notna(ont_lm) and ont_um < ont_lm) else 0,
                "_neg_no_mhl": 1 if (pd.notna(onm_um) and pd.notna(onm_lm) and onm_um < onm_lm) else 0,
                "_neg_no_fpd": 1 if (pd.notna(onf_um) and pd.notna(onf_lm) and onf_um < onf_lm) else 0,
            })

        out = pd.DataFrame(rows)
        if len(out):
            def _ser_ord(v):
                s = str(v).strip()
                try:
                    return (0, int(s))
                except Exception:
                    return (1, s)
            out["_ord"] = out["Ser No"].map(_ser_ord)
            out = out.sort_values(["Depot", "_ord"]).drop(columns=["_ord"]).reset_index(drop=True)
            out["SL NO"] = range(1, len(out) + 1)

        def _cell_neg(v, is_neg=False, is_int=False, bg=None, row_hl=False):
            if v is None or v == "" or (isinstance(v, float) and (pd.isna(v) or abs(v) < 1e-12)):
                s = ""
            elif isinstance(v, (int, float)):
                try:
                    fv = float(v)
                    if is_int or abs(fv - round(fv)) < 1e-6:
                        s = f"{int(round(fv))}" if abs(fv) > 1e-12 else ""
                    else:
                        s = f"{fv:.2f}"
                except Exception:
                    s = str(v)
            else:
                s = str(v)
            style = "padding:3px 5px;text-align:center;border:1px solid #e2e8f0;font-size:11px;"
            if row_hl:
                style += "background:#fef08a !important;font-weight:700;"
            elif is_neg and s:
                style += "background:#fee2e2;color:#b91c1c;font-weight:700;"
            elif bg:
                style += f"background:{bg};"
            return f'<td style="{style}">{s}</td>'

        # Freeze left columns through Sch Kms (12 cols, TP removed)
        def _fth(text, left, width, rowspan=3, top=0, bg="#1e3a8a", color="#fff"):
            return (
                f'<th rowspan="{rowspan}" style="position:sticky;left:{left}px;top:{top}px;z-index:6;'
                f'background:{bg};color:{color};padding:5px 4px;font-size:10px;text-align:center;'
                f'border:1px solid #94a3b8;min-width:{width}px;max-width:{width}px;">{text}</th>'
            )
        # widths: SL NO 36, Depot 64, Ser 56, SchDep 56, Route 64, RTC 56, R/L 44, Type 56, D/N 44, Schs 48, Sers 48, SchKms 56
        _fw = [36, 64, 56, 56, 64, 56, 44, 56, 44, 48, 48, 56]
        _fl = [0]
        for w in _fw[:-1]:
            _fl.append(_fl[-1] + w)
        _flabels = [
            "SL<br>NO", "Depot", "Ser No", "Sch<br>Dep", "Route", "RTC/<br>HIRE",
            "R/L", "Type", "D/N", "No.of<br>Schs", "No.of<br>Sers", "Sch.<br>Kms",
        ]
        # Header layout (matches screenshot):
        # Row1: identity(rowspan3) | Optd(2) | GROSS EPK(9) | GROSS OR(9) | NET EPK(9) | NET OR(9)
        # Row2: Optd CM/UM (rowspan2 — no TOT/MHL/FPD) | CM/UM/LUM (colspan3 each) × 4 groups
        # Row3: TOT MHL FPD under each CM/UM/LUM only (starts under GROSS EPK)
        thead = "<tr>"
        for lab, left, width in zip(_flabels, _fl, _fw):
            thead += _fth(lab, left, width, rowspan=3, top=0)
        thead += (
            _th("Optd", bg="#0f766e", color="#fff", colspan=2, top=0)
            + _th("GROSS EPK", bg="#dcfce7", color="#14532d", colspan=9, top=0)
            + _th("GROSS OR", bg="#dbeafe", color="#1e3a8a", colspan=9, top=0)
            + _th("NET EPK", bg="#fef3c7", color="#92400e", colspan=9, top=0)
            + _th("NET OR", bg="#ede9fe", color="#5b21b6", colspan=9, top=0)
            + "</tr><tr>"
        )
        # Optd CM / UM — rowspan 2 so row 3 has no sub-headers under Optd
        thead += (
            _th("CM", bg="#fef08a", color="#854d0e", rowspan=2, top=28)
            + _th("UM", bg="#fef08a", color="#854d0e", rowspan=2, top=28)
        )
        # Period labels under GROSS/NET groups only
        period_hdrs = [
            ("CM", "#86efac", "#14532d"), ("UM", "#4ade80", "#14532d"), ("LUM", "#bbf7d0", "#14532d"),
            ("CM", "#93c5fd", "#1e3a8a"), ("UM", "#60a5fa", "#1e3a8a"), ("LUM", "#bfdbfe", "#1e3a8a"),
            ("CM", "#fde68a", "#92400e"), ("UM", "#fbbf24", "#92400e"), ("LUM", "#fef3c7", "#92400e"),
            ("CM", "#c4b5fd", "#5b21b6"), ("UM", "#a78bfa", "#5b21b6"), ("LUM", "#ede9fe", "#5b21b6"),
        ]
        for lab, bg, col in period_hdrs:
            thead += _th(lab, bg=bg, color=col, colspan=3, top=28)
        thead += "</tr><tr>"
        # TOT MHL FPD only under GROSS EPK / GROSS OR / NET EPK / NET OR (12 periods × 3)
        for _ in range(12):
            thead += (
                _th("TOT", bg="#f1f5f9", color="#334155", top=56)
                + _th("MHL", bg="#f1f5f9", color="#334155", top=56)
                + _th("FPD", bg="#f1f5f9", color="#334155", top=56)
            )
        thead += "</tr>"

        um_epk_cols = {
            "G EPK UM TOT": "_neg_g_tot", "G EPK UM MHL": "_neg_g_mhl", "G EPK UM FPD": "_neg_g_fpd",
            "N EPK UM TOT": "_neg_n_tot", "N EPK UM MHL": "_neg_n_mhl", "N EPK UM FPD": "_neg_n_fpd",
        }
        um_or_cols = {
            "G OR UM TOT": "_neg_go_tot", "G OR UM MHL": "_neg_go_mhl", "G OR UM FPD": "_neg_go_fpd",
            "N OR UM TOT": "_neg_no_tot", "N OR UM MHL": "_neg_no_mhl", "N OR UM FPD": "_neg_no_fpd",
        }
        data_cols = [
            "SL NO", "Depot", "Ser No", "Sch Dep", "Route", "RTC/HIRE", "R/L", "Type", "D/N",
            "No.of Schs", "No.of Sers", "Sch. Kms", "Optd CM", "Optd UM",
            "G EPK CM TOT", "G EPK CM MHL", "G EPK CM FPD",
            "G EPK UM TOT", "G EPK UM MHL", "G EPK UM FPD",
            "G EPK LUM TOT", "G EPK LUM MHL", "G EPK LUM FPD",
            "G OR CM TOT", "G OR CM MHL", "G OR CM FPD",
            "G OR UM TOT", "G OR UM MHL", "G OR UM FPD",
            "G OR LUM TOT", "G OR LUM MHL", "G OR LUM FPD",
            "N EPK CM TOT", "N EPK CM MHL", "N EPK CM FPD",
            "N EPK UM TOT", "N EPK UM MHL", "N EPK UM FPD",
            "N EPK LUM TOT", "N EPK LUM MHL", "N EPK LUM FPD",
            "N OR CM TOT", "N OR CM MHL", "N OR CM FPD",
            "N OR UM TOT", "N OR UM MHL", "N OR UM FPD",
            "N OR LUM TOT", "N OR LUM MHL", "N OR LUM FPD",
        ]
        freeze_cols = {
            "SL NO", "Depot", "Ser No", "Sch Dep", "Route", "RTC/HIRE", "R/L", "Type", "D/N",
            "No.of Schs", "No.of Sers", "Sch. Kms",
        }
        def _b7_col_bg(c):
            if c.startswith("G EPK"):
                return _BG_GEPK
            if c.startswith("G OR"):
                return _BG_GOR
            if c.startswith("N EPK"):
                return _BG_NEPK
            if c.startswith("N OR"):
                return _BG_NOR
            if c in ("Optd CM", "Optd UM"):
                return _BG_OPTD
            return None

        _row_opts7 = ["(none)"] + [
            f"{int(r.get('SL NO', i+1))}. {r.get('Depot','')} / {r.get('Ser No','')} / {r.get('Route','')}"
            for i, r in out.iterrows()
        ]
        b7_sel = st.selectbox("Highlight service row", _row_opts7, index=0, key="mf_b7_row_sel")

        body = []
        for i, r in out.iterrows():
            label = f"{int(r.get('SL NO', i+1))}. {r.get('Depot','')} / {r.get('Ser No','')} / {r.get('Route','')}"
            hl = b7_sel == label
            tr_style = ' style="background:#fef08a;"' if hl else ""
            body.append(f"<tr{tr_style}>")
            for ci, c in enumerate(data_cols):
                is_neg = False
                if c in um_epk_cols:
                    is_neg = bool(r.get(um_epk_cols[c], 0))
                elif c in um_or_cols:
                    is_neg = bool(r.get(um_or_cols[c], 0))
                is_int = c in ("SL NO", "No.of Schs", "No.of Sers", "Optd CM", "Optd UM") or " OR " in c or c.startswith("G OR") or c.startswith("N OR")
                col_bg = _b7_col_bg(c)
                if c in freeze_cols:
                    left = _fl[ci]
                    width = _fw[ci]
                    v = r.get(c)
                    if v is None or v == "" or (isinstance(v, float) and (pd.isna(v) or abs(v) < 1e-12)):
                        s = ""
                    elif isinstance(v, (int, float)):
                        fv = float(v)
                        s = f"{int(round(fv))}" if is_int or abs(fv - round(fv)) < 1e-6 else f"{fv:.2f}"
                        if abs(float(v)) < 1e-12:
                            s = ""
                    else:
                        s = str(v)
                    bg = "#fef08a" if hl else "#f8fafc"
                    body.append(
                        f'<td style="position:sticky;left:{left}px;z-index:2;background:{bg};'
                        f'padding:3px 5px;text-align:center;border:1px solid #e2e8f0;font-size:11px;'
                        f'min-width:{width}px;max-width:{width}px;{"font-weight:700;" if hl else ""}">{s}</td>'
                    )
                else:
                    # group color under full columns; red still wins for negative
                    body.append(_cell_neg(r.get(c), is_neg=is_neg, is_int=is_int, bg=None if is_neg else col_bg, row_hl=hl))
            body.append("</tr>")

        def _b7_excel(df_out, report_title):
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            wb = Workbook()
            ws = wb.active
            ws.title = "ServiceWise"[:31]
            thin = Border(
                left=Side(style="thin", color="94A3B8"), right=Side(style="thin", color="94A3B8"),
                top=Side(style="thin", color="94A3B8"), bottom=Side(style="thin", color="94A3B8"),
            )
            center = Alignment(horizontal="center", vertical="center", wrap_text=True)
            white = Font(bold=True, color="FFFFFF", size=9)
            red_fill = PatternFill("solid", fgColor="FEE2E2")
            red_font = Font(color="B91C1C", bold=True, size=9)
            left_fill = PatternFill("solid", fgColor="1E3A8A")
            fills_top = {
                "optd": PatternFill("solid", fgColor="0F766E"),
                "gepk": PatternFill("solid", fgColor="DCFCE7"),
                "gor": PatternFill("solid", fgColor="DBEAFE"),
                "nepk": PatternFill("solid", fgColor="FEF3C7"),
                "nor": PatternFill("solid", fgColor="EDE9FE"),
            }
            sub = PatternFill("solid", fgColor="F1F5F9")
            dark = Font(bold=True, color="14532D", size=9)

            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=51)
            t = ws.cell(1, 1, report_title)
            t.font = Font(bold=True, size=12, color="1E3A8A")
            t.alignment = Alignment(horizontal="center")

            fixed = ["SL NO", "Depot", "Ser No", "Sch Dep", "Route", "RTC/HIRE", "R/L", "Type", "D/N",
                     "No.of Schs", "No.of Sers", "Sch. Kms"]
            for c, h in enumerate(fixed, 1):
                cell = ws.cell(2, c, h)
                cell.fill = left_fill
                cell.font = white
                cell.alignment = center
                cell.border = thin
                ws.merge_cells(start_row=2, start_column=c, end_row=4, end_column=c)
                for rr in range(2, 5):
                    ws.cell(rr, c).fill = left_fill
                    ws.cell(rr, c).border = thin
                    ws.cell(rr, c).font = white
                    ws.cell(rr, c).alignment = center

            # Row 2 groups (12 fixed cols → metrics start at col 13)
            groups = [
                (13, 14, "Optd", fills_top["optd"]),
                (15, 23, "GROSS EPK", fills_top["gepk"]),
                (24, 32, "GROSS OR", fills_top["gor"]),
                (33, 41, "NET EPK", fills_top["nepk"]),
                (42, 50, "NET OR", fills_top["nor"]),
            ]
            for a, b, lab, fl in groups:
                ws.merge_cells(start_row=2, start_column=a, end_row=2, end_column=b)
                cell = ws.cell(2, a, lab)
                cell.fill = fl
                cell.font = Font(bold=True, color="14532D" if "EPK" in lab or lab == "Optd" else "1E3A8A", size=10)
                if "NET" in lab:
                    cell.font = Font(bold=True, color="92400E" if "EPK" in lab else "5B21B6", size=10)
                cell.alignment = center
                for cc in range(a, b + 1):
                    ws.cell(2, cc).fill = fl
                    ws.cell(2, cc).border = thin

            # Row 3–4: Optd CM / UM (no TOT/MHL/FPD under them — rowspan 2)
            for cc, lab in ((13, "CM"), (14, "UM")):
                ws.merge_cells(start_row=3, start_column=cc, end_row=4, end_column=cc)
                cell = ws.cell(3, cc, lab)
                cell.fill = PatternFill("solid", fgColor="FEF08A")
                cell.font = Font(bold=True, color="854D0E", size=9)
                cell.alignment = center
                cell.border = thin
                ws.cell(4, cc).fill = PatternFill("solid", fgColor="FEF08A")
                ws.cell(4, cc).border = thin
            period_starts = [15, 18, 21, 24, 27, 30, 33, 36, 39, 42, 45, 48]
            period_labs = ["CM", "UM", "LUM"] * 4
            period_colors = (
                ["86EFAC", "4ADE80", "BBF7D0"]
                + ["93C5FD", "60A5FA", "BFDBFE"]
                + ["FDE68A", "FBBF24", "FEF3C7"]
                + ["C4B5FD", "A78BFA", "EDE9FE"]
            )
            for start, lab, col in zip(period_starts, period_labs, period_colors):
                ws.merge_cells(start_row=3, start_column=start, end_row=3, end_column=start + 2)
                cell = ws.cell(3, start, lab)
                cell.fill = PatternFill("solid", fgColor=col)
                cell.font = dark
                cell.alignment = center
                for cc in range(start, start + 3):
                    ws.cell(3, cc).fill = PatternFill("solid", fgColor=col)
                    ws.cell(3, cc).border = thin

            # Row 4 TOT MHL FPD under GROSS/NET only (not under Optd)
            for start in period_starts:
                for i, lab in enumerate(["TOT", "MHL", "FPD"]):
                    cell = ws.cell(4, start + i, lab)
                    cell.fill = sub
                    cell.font = Font(bold=True, color="334155", size=8)
                    cell.alignment = center
                    cell.border = thin

            export_cols = data_cols
            neg_map = {**um_epk_cols, **um_or_cols}
            for ri, row in enumerate(df_out.itertuples(index=False), 5):
                rec = dict(zip(df_out.columns, row))
                for ci, col in enumerate(export_cols, 1):
                    val = rec.get(col, "")
                    if val is None or (isinstance(val, float) and (pd.isna(val) or abs(val) < 1e-12)):
                        val = None
                    cell = ws.cell(ri, ci, val if val != "" else None)
                    cell.alignment = center
                    cell.border = thin
                    if col in neg_map and rec.get(neg_map[col], 0):
                        cell.fill = red_fill
                        cell.font = red_font

            for i in range(1, 51):
                ws.column_dimensions[get_column_letter(i)].width = 8 if i > 12 else 9
            ws.column_dimensions["A"].width = 6
            ws.column_dimensions["B"].width = 10
            ws.column_dimensions["C"].width = 9
            ws.freeze_panes = "M5"  # freeze through Sch. Kms (col 12)
            bio = BytesIO()
            wb.save(bio)
            bio.seek(0)
            return bio.getvalue()

        st.markdown(f'<div class="title-bar">{title}</div>', unsafe_allow_html=True)
        if not body:
            st.warning("No services match the selected filters.")
        else:
            html = [
                '<div class="table-scroll-fixable"><table class="excel-table" style="border-collapse:collapse;width:max-content;">',
                "<thead>", thead, "</thead><tbody>",
                "".join(body), "</tbody></table></div>",
            ]
            st.markdown("".join(html), unsafe_allow_html=True)
            st.download_button(
                "Download Excel (as on screen)",
                _b7_excel(out, title),
                f"Monthly_{mf_month}_ServiceWise.xlsx",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="mf_dl_ServiceWise",
            )

        # ========== Range-wise EPK services by PRODUCT (reacts to Gross/Net + TOT/MHL/FPD) ==========
        st.markdown("---")
        _basis = "GROSS" if str(mf_ng).upper().startswith("G") else "NET"
        r1, r2 = st.columns([1, 3])
        with r1:
            mf_rng_metric = st.selectbox(
                "Range metric (EPK)",
                ["TOT", "MHL", "FPD"],
                index=0,
                key="mf_b7_range_metric",
            )
        rng_title = f"{_basis} RANGE WISE {mf_rng_metric} EPK SERVICES UPTO {mf_month} (by Product)"
        st.markdown(f'<div class="title-bar">{rng_title}</div>', unsafe_allow_html=True)

        # Build service-level UM EPK for the selected basis + metric, restricted to SMASTER services
        if _basis == "GROSS":
            _earn_map = {"TOT": "ge_tot_cy", "MHL": "ge_mhl_cy", "FPD": "ge_fpd_cy"}
        else:
            _earn_map = {"TOT": "ne_tot_cy", "MHL": "ne_mhl_cy", "FPD": "ne_fpd_cy"}
        _ecol = _earn_map.get(mf_rng_metric, "ge_tot_cy")

        # Use merged frame (SMASTER master + CY metrics) before low/neg row filters
        _rng_src = merged.copy() if "merged" in dir() else pd.DataFrame()
        # Prefer building from a_cy + master for full SMASTER list
        try:
            _rng_base = master.merge(a_cy, on=["DEPOT", "SER_NO"], how="left")
        except Exception:
            _rng_base = pd.DataFrame()
        ranges = [
            ("01-15", 0, 15),
            ("15-20", 15, 20),
            ("21-30", 21, 30),
            ("31-40", 31, 40),
            ("41-50", 41, 50),
            (">51", 51, 1e9),
        ]
        products = sorted(
            {
                str(x).strip()
                for x in master.get("prod", pd.Series(dtype=str)).dropna().unique()
                if str(x).strip() and str(x).strip().lower() not in ("nan", "none", "")
            },
            key=lambda z: str(z).upper(),
        )
        # Also include products from performance if master prod blank
        if not products and len(_rng_base):
            products = sorted(
                {
                    str(x).strip()
                    for x in _rng_base.get("prod", pd.Series(dtype=str)).dropna().unique()
                    if str(x).strip() and str(x).strip().lower() not in ("nan", "none", "")
                },
                key=lambda z: str(z).upper(),
            )
        if len(_rng_base) and products:
            _rb = _rng_base.copy()
            _rb["SER_NO"] = _rb["SER_NO"].map(_norm_svc7)
            _rb["DEPOT"] = _rb["DEPOT"].astype(str).str.strip().str.upper()
            _rb["PRODUCT"] = _rb["prod"].astype(str).str.strip()
            _kms = pd.to_numeric(_rb.get("kms", 0), errors="coerce").fillna(0)
            _earn = pd.to_numeric(_rb.get(_ecol.replace("_cy", ""), _rb.get(_ecol, 0)), errors="coerce")
            # a_cy columns may already be without _cy suffix after merge
            if _ecol not in _rb.columns:
                # after merge suffixes: kms from a_cy may be plain "kms"
                alt = {
                    "ge_tot_cy": "ge_tot", "ge_mhl_cy": "ge_mhl", "ge_fpd_cy": "ge_fpd",
                    "ne_tot_cy": "ne_tot", "ne_mhl_cy": "ne_mhl", "ne_fpd_cy": "ne_fpd",
                }.get(_ecol, _ecol)
                _earn = pd.to_numeric(_rb.get(alt, 0), errors="coerce")
            _rb["_EPK"] = np.where(_kms > 0, _earn.fillna(0) / _kms, np.nan)
            depots_r = sorted(_rb["DEPOT"].dropna().unique())
            rng_rows = []
            for dep in depots_r:
                gd = _rb[_rb["DEPOT"] == dep]
                for rlab, lo, hi in ranges:
                    rec = {"Depot": dep, "EPK RANGE": rlab}
                    total = 0
                    for p in products:
                        if rlab == ">51":
                            sub = gd[(gd["PRODUCT"].str.upper() == p.upper()) & (gd["_EPK"] >= 51)]
                        else:
                            sub = gd[
                                (gd["PRODUCT"].str.upper() == p.upper())
                                & (gd["_EPK"] >= lo)
                                & (gd["_EPK"] < hi)
                            ]
                        cnt = int(sub["SER_NO"].nunique())
                        rec[p] = cnt if cnt else ""
                        total += cnt
                    rec["Grand Total"] = total if total else ""
                    rng_rows.append(rec)
                tot = {"Depot": dep, "EPK RANGE": "Total"}
                grand = 0
                for p in products:
                    cnt = int(gd[gd["PRODUCT"].str.upper() == p.upper()]["SER_NO"].nunique())
                    tot[p] = cnt if cnt else ""
                    grand += cnt
                tot["Grand Total"] = grand if grand else ""
                rng_rows.append(tot)
            out_r = pd.DataFrame(rng_rows)
            cols_r = ["Depot", "EPK RANGE"] + products + ["Grand Total"]
            out_r = out_r.reindex(columns=[c for c in cols_r if c in out_r.columns])
            thead_r = (
                "<tr>"
                + _th("Depot", rowspan=2, top=0)
                + _th("EPK<br>RANGE", rowspan=2, top=0)
                + _th(f"{_basis} · {mf_rng_metric} EPK by Product", bg="#1e40af", colspan=len(products), top=0)
                + _th("Grand<br>Total", rowspan=2, bg="#0f172a", top=0)
                + "</tr><tr>"
            )
            for p in products:
                thead_r += _th(p, bg="#3b82f6", top=28)
            thead_r += "</tr>"
            body_r = []
            for _, rr in out_r.iterrows():
                is_tot = str(rr.get("EPK RANGE", "")) == "Total"
                style = ' style="font-weight:700;background:#e2efda;"' if is_tot else ""
                body_r.append(f"<tr{style}>")
                body_r.append(_cell(rr.get("Depot")))
                body_r.append(_cell(rr.get("EPK RANGE")))
                for p in products:
                    body_r.append(_cell(rr.get(p), is_int=True))
                body_r.append(_cell(rr.get("Grand Total"), is_int=True))
                body_r.append("</tr>")
            if body_r:
                html_r = [
                    '<div class="table-scroll-fixable"><table class="excel-table" style="border-collapse:collapse;width:max-content;">',
                    "<thead>", thead_r, "</thead><tbody>",
                    "".join(body_r), "</tbody></table></div>",
                ]
                st.markdown("".join(html_r), unsafe_allow_html=True)
                st.download_button(
                    "Download Range-wise Excel",
                    _excel_bytes(out_r, rng_title),
                    f"Monthly_{mf_month}_RangeEPK_Product.xlsx",
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="mf_dl_RangeEPK_Product",
                )
            else:
                st.warning("No range-wise rows.")
        else:
            st.caption("Range-wise table: no product list from SMASTER.")

        st.caption(
            f"SMASTER services only · {len(out)} row(s) · D/N from D.TYPE · "
            f"Columns frozen through Sch. Kms · Low/Neg filters on UM vs LUM · Red = UM below LUM"
        )

    # ========== BOARD 8: Trend Service Wise Gross ==========
    elif mf_board.startswith("8."):
        title = f"TREND- GROSS DEPOT SERVICE WISE PERFORMANCE FOR AND UPTO {mf_month}"
        if mf_depot != "ALL":
            title += f" | Depot: {mf_depot}"
        if mf_product != "ALL":
            title += f" | Product: {mf_product}"

        # ===== Board 8 master: SMASTER supplies service identity =====
        _sm8_path = _resolve_smaster_path()
        if not _sm8_path.exists():
            st.error("SMASTER.parquet not found — Board 8 requires SMASTER for Ser No/service identity.")
            st.stop()
        _sm8_raw, _sm8_err = load_smaster(str(_sm8_path))
        if _sm8_raw is None or len(_sm8_raw) == 0:
            st.error(f"Could not read SMASTER for Board 8: {_sm8_err or 'empty file'}")
            st.stop()
        _sm8_raw = _sm8_raw.copy()
        _sm8_raw.columns = [str(c).strip() for c in _sm8_raw.columns]
        def _fc8(cands):
            norm = {str(c).strip().lower().replace(" ", "").replace("_", "").replace("/", "").replace(".", ""): c for c in _sm8_raw.columns}
            for cand in cands:
                k = cand.lower().replace(" ", "").replace("_", "").replace("/", "").replace(".", "")
                if k in norm:
                    return norm[k]
            return None
        def _norm8(v):
            if v is None or (isinstance(v, float) and pd.isna(v)):
                return ""
            z = str(v).strip()
            if z.lower() in ("", "nan", "none"):
                return ""
            try:
                return str(int(float(z)))
            except Exception:
                return z
        def _mode8(series):
            z = series.dropna().astype(str).str.strip()
            z = z[~z.str.lower().isin(["", "nan", "none"])]
            if len(z) == 0:
                return ""
            try:
                return z.mode().iloc[0]
            except Exception:
                return z.iloc[0]
        c8_svc = _fc8(["ServiceNo", "SER_NO", "SERVICE_NO", "SERVICENO"])
        c8_dep = _fc8(["DEPOT"])
        c8_prod = _fc8(["PRODUCT"])
        c8_route = _fc8(["ROUTEE", "ROUTE"])
        c8_sch = _fc8(["NoOfSchedules", "NoOfSchedule", "NO_OF_SCHS"])
        c8_kms = _fc8(["RevenueKms", "Revenue Kms", "SCH_KMS", "SchKms", "DAY_SCH_KMS"])
        c8_rtc = _fc8(["RTC/HIRE", "RTC_HIRE", "RTCHIRE"])
        c8_rl = _fc8(["R/L", "R_L", "RL", "RouteLength"])
        c8_schdep = _fc8(["SCH_DEP", "SCH_DEP.", "SchDep", "SCHDEP", "ScheduleDepot"])
        c8_type = _fc8(["TYPE", "TYPE OF SERV.", "ServiceType", "TP"])
        c8_dtype = _fc8(["D.TYPE", "DTYPE", "DutyType"])
        if not c8_svc or not c8_dep:
            st.error("SMASTER must contain DEPOT and SER NO/ServiceNo for Board 8.")
            st.stop()
        sm8 = _sm8_raw.copy()
        sm8["_SVC"] = sm8[c8_svc].map(_norm8)
        sm8["_DEP"] = sm8[c8_dep].astype(str).str.strip().str.upper()
        sm8["_PROD"] = sm8[c8_prod].astype(str).str.strip() if c8_prod else ""
        sm8["_ROUTE"] = sm8[c8_route].astype(str).str.strip() if c8_route else ""
        sm8["_SCH"] = pd.to_numeric(sm8[c8_sch], errors="coerce").fillna(0) if c8_sch else 0.0
        sm8["_SCH_KMS"] = pd.to_numeric(sm8[c8_kms], errors="coerce").fillna(0) if c8_kms else 0.0
        sm8["_RL"] = pd.to_numeric(sm8[c8_rl], errors="coerce").fillna(0) if c8_rl else 0.0
        sm8["_RTC"] = sm8[c8_rtc].astype(str).str.strip() if c8_rtc else ""
        sm8["_SCH_DEP"] = sm8[c8_schdep].astype(str).str.strip() if c8_schdep else ""
        sm8["_TYPE"] = sm8[c8_type].astype(str).str.strip() if c8_type else ""
        sm8["_DN"] = sm8[c8_dtype].astype(str).str.strip().str.upper() if c8_dtype else ""
        if mf_depot != "ALL":
            sm8 = sm8[sm8["_DEP"] == str(mf_depot).strip().upper()]
        if mf_product != "ALL":
            sm8 = sm8[sm8["_PROD"].astype(str).str.strip().str.upper() == str(mf_product).strip().upper()]
        if mf_route != "ALL":
            sm8 = sm8[sm8["_ROUTE"].astype(str).str.strip().str.upper() == str(mf_route).strip().upper()]
        master8 = (sm8.groupby(["_DEP", "_SVC"], dropna=False).agg(
            route=("_ROUTE", _mode8), prod=("_PROD", _mode8), rtc=("_RTC", _mode8),
            sch_dep=("_SCH_DEP", _mode8), typ=("_TYPE", _mode8), nature=("_DN", _mode8),
            rl=("_RL", "max"), schs=("_SCH", "sum"), sch_kms=("_SCH_KMS", "sum")
        ).reset_index().rename(columns={"_DEP": "DEPOT", "_SVC": "SER_NO"}))
        for ii, rr in master8.iterrows():
            sk, rl, sc = float(rr.get("sch_kms", 0) or 0), float(rr.get("rl", 0) or 0), float(rr.get("schs", 0) or 0)
            if sk == 0 and rl > 0 and sc > 0:
                master8.at[ii, "sch_kms"] = rl * sc
        allowed8 = set(zip(master8["DEPOT"].astype(str).str.upper(), master8["SER_NO"].map(_norm8)))

        # ===== Board 8 metrics: dedicated service-monthly parquet =====
        _svc8, _svc8_err = load_monthly_service_metrics()
        if _svc8_err:
            st.error(f"Could not read service-monthly parquet: {_svc8_err}")
            st.stop()
        if _svc8 is None or len(_svc8) == 0:
            st.error("Service-monthly parquet not found/empty. Checked ser_monthly.parquet and ser_montly.parquet.")
            st.stop()
        if "SER_NO" not in _svc8.columns or "DEPOT" not in _svc8.columns:
            st.error("Service-monthly parquet must contain DEPOT and SER_NO for SMASTER matching.")
            st.stop()
        _svc8 = _svc8.copy()
        _svc8["SER_NO"] = _svc8["SER_NO"].map(_norm8)
        _svc8["DEPOT"] = _svc8["DEPOT"].astype(str).str.strip().str.upper()
        _svc8 = _svc8[_svc8.apply(lambda rr: (rr["DEPOT"], _norm8(rr["SER_NO"])) in allowed8, axis=1)].copy()

        # Board 8 source rule:
        #   SMASTER -> Depot, Ser No, Route, RTC/HIRE, D/N,
        #              No.of Schs, No.of Ser, Sch.Kms
        #   ser_monthly -> Sch Dep, R/L, Type (= PRODUCT)
        def _ensure_sch_rl8(frame):
            if frame is None or len(frame) == 0:
                return frame
            def nk(c):
                return str(c).strip().lower().replace("_", "").replace(" ", "").replace("/", "").replace(".", "")
            if "SCH_DEP" not in frame.columns:
                for c in frame.columns:
                    k = nk(c)
                    if ("sch" in k and "dep" in k) or k in ("schdep", "scheduledepot", "schdepot"):
                        frame = frame.copy()
                        frame["SCH_DEP"] = frame[c].astype(str)
                        break
            if "R_L" not in frame.columns:
                for c in frame.columns:
                    k = nk(c)
                    if k in ("rl", "routelength", "routelen") or (k.startswith("rl") and "epk" not in k and len(k) <= 6):
                        frame = frame.copy()
                        frame["R_L"] = pd.to_numeric(frame[c], errors="coerce").fillna(0.0)
                        break
            return frame
        _svc8 = _ensure_sch_rl8(_svc8)

        _meta8 = _svc8[_svc8["Month_Name"].astype(str).str.strip() == str(mf_month).strip()].copy()
        if len(_meta8):
            _meta8["_SVC"] = _meta8["SER_NO"].map(_norm8)
            def _first_nonblank8(series):
                z = series.dropna().astype(str).str.strip()
                z = z[~z.str.lower().isin(["", "nan", "none"])]
                return z.mode().iloc[0] if len(z) else ""
            agg8_meta = {}
            if "SCH_DEP" in _meta8.columns:
                agg8_meta["sch_dep_m"] = ("SCH_DEP", _first_nonblank8)
            if "R_L" in _meta8.columns:
                agg8_meta["rl_m"] = ("R_L", "max")
            if "PRODUCT" in _meta8.columns:
                agg8_meta["typ_m"] = ("PRODUCT", _first_nonblank8)
            if agg8_meta:
                mm8 = _meta8.groupby(["DEPOT","_SVC"], dropna=False).agg(**agg8_meta).reset_index()
                master8 = master8.merge(
                    mm8.rename(columns={"_SVC":"SER_NO"}), on=["DEPOT","SER_NO"], how="left"
                )
                def _nonblank8(s):
                    s = s.astype(str).str.strip()
                    return s.where(~s.str.lower().isin(["", "nan", "none", "nat", "<na>"]), other=pd.NA)

                if "sch_dep_m" in master8.columns:
                    m = _nonblank8(master8["sch_dep_m"])
                    fb = master8["sch_dep"].astype(str).str.strip() if "sch_dep" in master8.columns else ""
                    master8["sch_dep"] = m.fillna(fb).fillna("").astype(str).str.strip()
                if "rl_m" in master8.columns:
                    mrl = pd.to_numeric(master8["rl_m"], errors="coerce")
                    fbrl = pd.to_numeric(master8.get("rl", 0), errors="coerce").fillna(0.0)
                    master8["rl"] = mrl.where(mrl.fillna(0) != 0, fbrl).fillna(fbrl)
                if "typ_m" in master8.columns:
                    m = _nonblank8(master8["typ_m"])
                    fb = master8["typ"].astype(str).str.strip() if "typ" in master8.columns else ""
                    master8["typ"] = m.fillna(fb).fillna("").astype(str).str.strip()
                master8 = master8.drop(columns=[c for c in ["sch_dep_m", "rl_m", "typ_m"] if c in master8.columns])
        # Recalculate Sch Kms only if SMASTER Sch Kms is blank, using the
        # monthly R/L as requested.
        for ii, rr in master8.iterrows():
            sk = float(rr.get("sch_kms", 0) or 0)
            rl = float(rr.get("rl", 0) or 0)
            sc = float(rr.get("schs", 0) or 0)
            if sk == 0 and rl > 0 and sc > 0:
                master8.at[ii, "sch_kms"] = rl * sc

        fy_months = []
        cur = fy_start
        while cur <= cy_end:
            fy_months.append(cur.strftime("%b-%Y"))
            cur += pd.DateOffset(months=1)
        mon_labs = [m.split("-")[0] for m in fy_months]
        keys = ["DEPOT", "SER_NO"]

        # Scope data once (vectorized — avoid per-service filters)
        _b8 = _svc8.copy()
        if mf_depot != "ALL" and "DEPOT" in _b8.columns:
            _b8 = _b8[_b8["DEPOT"].astype(str).str.strip().str.upper() == str(mf_depot).strip().upper()]
        if mf_product != "ALL" and "PRODUCT" in _b8.columns:
            _b8 = _b8[_b8["PRODUCT"].astype(str).str.strip().str.upper() == str(mf_product).strip().upper()]
        if mf_route != "ALL" and _route_col_mf and _route_col_mf in _b8.columns:
            _b8 = _b8[_b8[_route_col_mf].astype(str).str.strip().str.upper() == str(mf_route).strip().upper()]

        _cy8 = _svc8.copy()
        _ly8 = _svc8.copy()
        if mf_depot != "ALL":
            if "DEPOT" in _cy8.columns:
                _cy8 = _cy8[_cy8["DEPOT"].astype(str).str.strip().str.upper() == str(mf_depot).strip().upper()]
            if "DEPOT" in _ly8.columns:
                _ly8 = _ly8[_ly8["DEPOT"].astype(str).str.strip().str.upper() == str(mf_depot).strip().upper()]
        if mf_product != "ALL":
            if "PRODUCT" in _cy8.columns:
                _cy8 = _cy8[_cy8["PRODUCT"].astype(str).str.strip().str.upper() == str(mf_product).strip().upper()]
            if "PRODUCT" in _ly8.columns:
                _ly8 = _ly8[_ly8["PRODUCT"].astype(str).str.strip().str.upper() == str(mf_product).strip().upper()]

        if "Date" in _cy8.columns:
            _cy8 = _cy8[(_cy8["Date"] >= fy_start) & (_cy8["Date"] <= cy_end)].copy()
            _ly8 = _ly8[(_ly8["Date"] >= ly_start) & (_ly8["Date"] <= ly_end)].copy()
        elif "Month_Name" in _cy8.columns:
            _cy8 = _cy8[_cy8["Month_Name"].astype(str).str.strip().isin(fy_months)].copy()
            _ly_months = [(pd.to_datetime(x, format="%b-%Y") - pd.DateOffset(years=1)).strftime("%b-%Y") for x in fy_months]
            _ly8 = _ly8[_ly8["Month_Name"].astype(str).str.strip().isin(_ly_months)].copy()

        if len(_cy8) == 0 and len(_b8) == 0:
            st.warning("No data for selected filters.")
            st.stop()

        # Identity + sch counts (max per service for schs/sch_kms)
        def _mode_first(s):
            s = s.dropna().astype(str).str.strip()
            s = s[~s.str.lower().isin(["", "nan", "none"])]
            if len(s) == 0:
                return ""
            try:
                return s.mode().iloc[0]
            except Exception:
                return s.iloc[0]

        # Identity and schedule information are exclusively from SMASTER.
        meta = master8.copy()
        meta["sers"] = 1

        # Pre-aggregate monthly metrics: groupby DEPOT, SER_NO, Month_Name once
        _b8m = _b8.copy()
        if "Month_Name" not in _b8m.columns:
            st.error("Month_Name missing — cannot build trend.")
            st.stop()
        mon_g = (
            _b8m.groupby(["DEPOT", "SER_NO", "Month_Name"], dropna=False)
            .agg(kms=("Optd_KMs", "sum"), ge=("GE_TOT", "sum"), ne=("NE_TOT", "sum"))
            .reset_index()
        )
        mon_g["Month_Name"] = mon_g["Month_Name"].astype(str).str.strip()
        mon_g["g_epk"] = np.where(mon_g["kms"] > 0, mon_g["ge"] / mon_g["kms"], np.nan)
        mon_g["n_epk"] = np.where(mon_g["kms"] > 0, mon_g["ne"] / mon_g["kms"], np.nan)

        # Pivot GROSS and NET EPK by month label (Jan, Feb, ...)
        mon_g["_lab"] = mon_g["Month_Name"].map(lambda x: str(x).split("-")[0] if x else "")
        mon_g = mon_g[mon_g["Month_Name"].isin(fy_months)]
        g_epk_piv = mon_g.pivot_table(index=["DEPOT", "SER_NO"], columns="_lab", values="g_epk", aggfunc="first")
        n_epk_piv = mon_g.pivot_table(index=["DEPOT", "SER_NO"], columns="_lab", values="n_epk", aggfunc="first")
        for lab in mon_labs:
            if lab not in g_epk_piv.columns:
                g_epk_piv[lab] = np.nan
            if lab not in n_epk_piv.columns:
                n_epk_piv[lab] = np.nan
        g_epk_piv = g_epk_piv.reindex(columns=mon_labs)
        n_epk_piv = n_epk_piv.reindex(columns=mon_labs)

        # UM / LY UM aggregates
        def _um_agg(frame):
            if len(frame) == 0:
                return pd.DataFrame(columns=keys + ["kms", "ge", "ne", "g_epk", "n_epk"]).set_index(keys)
            g = frame.groupby(keys, dropna=False).agg(
                kms=("Optd_KMs", "sum"), ge=("GE_TOT", "sum"), ne=("NE_TOT", "sum")
            )
            g["g_epk"] = np.where(g["kms"] > 0, g["ge"] / g["kms"], np.nan)
            g["n_epk"] = np.where(g["kms"] > 0, g["ne"] / g["kms"], np.nan)
            return g

        cy_agg = _um_agg(_cy8)
        ly_agg = _um_agg(_ly8)

        # ORF cache by depot
        _orf_cache_c = {}
        _orf_cache_l = {}
        def _orf_c(dep):
            dep = str(dep)
            if dep not in _orf_cache_c:
                _orf_cache_c[dep] = _orf(dep, side="cy")
            return _orf_cache_c[dep]
        def _orf_l(dep):
            dep = str(dep)
            if dep not in _orf_cache_l:
                _orf_cache_l[dep] = _orf(dep, side="ly")
            return _orf_cache_l[dep]

        rows = []
        for _, m in meta.iterrows():
            dep, ser = m["DEPOT"], m["SER_NO"]
            orf_c = _orf_c(dep)
            orf_l = _orf_l(dep)
            rec = {
                "SL NO": len(rows) + 1,
                "Depot": dep,
                "Ser No": ser,
                "Sch Dep": m.get("sch_dep", ""),
                "Route": m.get("route", ""),
                "RTC/HIRE": m.get("rtc", ""),
                "Type": m.get("typ", ""),
                "D/N": m.get("nature", ""),
                "No. of Schs": int(round(float(m.get("schs", 0) or 0))) or "",
                "No. of Ser": 1,
                "Sch. Kms": _f0(m.get("sch_kms", 0)),
            }
            key = (dep, ser)
            for lab in mon_labs:
                g_epk = n_epk = np.nan
                if key in g_epk_piv.index and lab in g_epk_piv.columns:
                    try:
                        g_epk = g_epk_piv.loc[key, lab]
                    except Exception:
                        pass
                if key in n_epk_piv.index and lab in n_epk_piv.columns:
                    try:
                        n_epk = n_epk_piv.loc[key, lab]
                    except Exception:
                        pass
                rec[f"{lab} GROSS EPK"] = round(float(g_epk), 2) if pd.notna(g_epk) else None
                rec[f"{lab} GROSS OR"] = round(_or(g_epk, orf_c), 0) if pd.notna(g_epk) else None
                rec[f"{lab} NET EPK"] = round(float(n_epk), 2) if pd.notna(n_epk) else None
                rec[f"{lab} NET OR"] = round(_or(n_epk, orf_c), 0) if pd.notna(n_epk) else None
            for label, agg, orf in (("UM", cy_agg, orf_c), ("LY UM", ly_agg, orf_l)):
                g_epk = n_epk = np.nan
                if key in agg.index:
                    try:
                        g_epk = agg.loc[key, "g_epk"]
                        n_epk = agg.loc[key, "n_epk"]
                    except Exception:
                        pass
                rec[f"{label} GROSS EPK"] = round(float(g_epk), 2) if pd.notna(g_epk) else None
                rec[f"{label} GROSS OR"] = round(_or(g_epk, orf), 0) if pd.notna(g_epk) else None
                rec[f"{label} NET EPK"] = round(float(n_epk), 2) if pd.notna(n_epk) else None
                rec[f"{label} NET OR"] = round(_or(n_epk, orf), 0) if pd.notna(n_epk) else None
            rows.append(rec)

        out = pd.DataFrame(rows)
        if len(out):
            out = out.sort_values(["Depot", "Ser No"]).reset_index(drop=True)
            out["SL NO"] = range(1, len(out) + 1)
        thead = (
            "<tr>"
            + _th("SL<br>NO", rowspan=2, top=0) + _th("Depot", rowspan=2, top=0)
            + _th("Ser No", rowspan=2, top=0) + _th("Sch<br>Dep", rowspan=2, top=0)
            + _th("Route", rowspan=2, top=0) + _th("RTC/<br>HIRE", rowspan=2, top=0)
            + _th("Type", rowspan=2, top=0)
            + _th("D/N", rowspan=2, top=0) + _th("No.<br>of<br>Schs", rowspan=2, top=0)
            + _th("No.<br>of<br>Ser", rowspan=2, top=0) + _th("Sch.<br>Kms", rowspan=2, top=0)
            + _th("GROSS EPK", bg="#dcfce7", color="#14532d", colspan=len(mon_labs) + 2, top=0)
            + _th("GROSS OR", bg="#dbeafe", color="#1e3a8a", colspan=len(mon_labs) + 2, top=0)
            + _th("NET EPK", bg="#fef3c7", color="#92400e", colspan=len(mon_labs) + 2, top=0)
            + _th("NET OR", bg="#ede9fe", color="#5b21b6", colspan=len(mon_labs) + 2, top=0)
            + "</tr><tr>"
        )
        for lab in mon_labs:
            thead += _th(lab.upper()[:3], bg="#bbf7d0", color="#14532d", top=28)
        thead += _th("UM", bg="#86efac", color="#14532d", top=28) + _th("LY UM", bg="#bbf7d0", color="#14532d", top=28)
        for lab in mon_labs:
            thead += _th(lab.upper()[:3], bg="#bfdbfe", color="#1e3a8a", top=28)
        thead += _th("UM", bg="#93c5fd", color="#1e3a8a", top=28) + _th("LY UM", bg="#bfdbfe", color="#1e3a8a", top=28)
        for lab in mon_labs:
            thead += _th(lab.upper()[:3], bg="#fef3c7", color="#92400e", top=28)
        thead += _th("UM", bg="#fbbf24", color="#92400e", top=28) + _th("LY UM", bg="#fef3c7", color="#92400e", top=28)
        for lab in mon_labs:
            thead += _th(lab.upper()[:3], bg="#ddd6fe", color="#5b21b6", top=28)
        thead += _th("UM", bg="#c4b5fd", color="#5b21b6", top=28) + _th("LY UM", bg="#ddd6fe", color="#5b21b6", top=28)
        thead += "</tr>"
        _row_opts8 = ["(none)"] + [
            f"{int(r.get('SL NO', i+1))}. {r.get('Depot','')} / {r.get('Ser No','')}"
            for i, r in out.iterrows()
        ]
        b8_sel = st.selectbox("Highlight service row", _row_opts8, index=0, key="mf_b8_row_sel")
        body = []
        for i, r in out.iterrows():
            label = f"{int(r.get('SL NO', i+1))}. {r.get('Depot','')} / {r.get('Ser No','')}"
            hl = b8_sel == label
            tr_style = ' style="background:#fef08a;"' if hl else ""
            body.append(f"<tr{tr_style}>")
            for c in ["SL NO", "Depot", "Ser No", "Sch Dep", "Route", "RTC/HIRE", "Type", "D/N", "No. of Schs", "No. of Ser", "Sch. Kms"]:
                body.append(_cell(r.get(c), row_hl=hl))
            for lab in mon_labs:
                body.append(_cell(r.get(f"{lab} GROSS EPK"), bg=_BG_GEPK, row_hl=hl))
            body.append(_cell(r.get("UM GROSS EPK"), bg=_BG_GEPK, row_hl=hl))
            body.append(_cell(r.get("LY UM GROSS EPK"), bg=_BG_GEPK, row_hl=hl))
            for lab in mon_labs:
                body.append(_cell(r.get(f"{lab} GROSS OR"), is_int=True, bg=_BG_GOR, row_hl=hl))
            body.append(_cell(r.get("UM GROSS OR"), is_int=True, bg=_BG_GOR, row_hl=hl))
            body.append(_cell(r.get("LY UM GROSS OR"), is_int=True, bg=_BG_GOR, row_hl=hl))
            for lab in mon_labs:
                body.append(_cell(r.get(f"{lab} NET EPK"), bg=_BG_NEPK, row_hl=hl))
            body.append(_cell(r.get("UM NET EPK"), bg=_BG_NEPK, row_hl=hl))
            body.append(_cell(r.get("LY UM NET EPK"), bg=_BG_NEPK, row_hl=hl))
            for lab in mon_labs:
                body.append(_cell(r.get(f"{lab} NET OR"), is_int=True, bg=_BG_NOR, row_hl=hl))
            body.append(_cell(r.get("UM NET OR"), is_int=True, bg=_BG_NOR, row_hl=hl))
            body.append(_cell(r.get("LY UM NET OR"), is_int=True, bg=_BG_NOR, row_hl=hl))
            body.append("</tr>")
        _render_board(title, thead, "".join(body), out, "TrendService")
        st.caption(f"Vectorized trend · {len(out)} services · months={', '.join(mon_labs)}")


    else:
        st.info("Select a report board.")

elif section == "Route Day-wise":
    title = f"Route WISE, Day Wise {prefix} EPK ({for_upto}) - {month}"
    if depot != "ALL":
        title += f" of {depot} Depot"
    st.markdown(f'<div class="title-bar">{title}</div>', unsafe_allow_html=True)
    st.caption(f"CY rows: {len(cy_data):,} | LY rows: {len(ly_data):,}")
    if len(cy_data) == 0:
        st.warning("No data for selected filters.")
    else:
        all_parts = []
        for label, earn_col in [("TOT", earn_tot), ("FPD", earn_fpd), ("MHL", earn_mhl)]:
            cy = weighted_epk(cy_data, earn_col)
            ly = weighted_epk(ly_data, earn_col)
            all_idx = cy.index.union(ly.index) if len(ly) > 0 else cy.index
            cy = cy.reindex(all_idx)
            ly = ly.reindex(all_idx) if len(ly) > 0 else pd.DataFrame(np.nan, index=all_idx, columns=day_order + ["UPTO"], dtype=float)
            block = pd.DataFrame(index=all_idx)
            for d in day_order:
                block[f"{label}_{d}"] = pd.to_numeric(cy[d] if d in cy.columns else np.nan, errors="coerce")
            block[f"{label}_CY"] = pd.to_numeric(cy["UPTO"] if "UPTO" in cy.columns else np.nan, errors="coerce")
            block[f"{label}_LY"] = pd.to_numeric(ly["UPTO"] if "UPTO" in ly.columns else np.nan, errors="coerce")
            block[f"{label}_Var"] = (block[f"{label}_CY"] - block[f"{label}_LY"]).astype(float).round(2)
            all_parts.append(block)
        result = pd.concat(all_parts, axis=1).reset_index()
        result = result.rename(columns={"ROUTEE": "ROUTE"})
        result = result.sort_values(["DEPOT", "ROUTE", "PRODUCT"]).reset_index(drop=True)
        html = ['<div class="table-scroll"><table class="excel-table">']
        html.append("<tr>")
        html.append('<th class="header-left" rowspan="2">S.No</th>')
        html.append('<th class="header-left" rowspan="2">DEPOT</th>')
        html.append('<th class="header-left" rowspan="2">ROUTE</th>')
        html.append('<th class="header-left" rowspan="2">PRODUCT</th>')
        html.append(f'<th class="header-tot" colspan="10">{prefix} TOT. E.P.K (in Ps/kms.)</th>')
        html.append(f'<th class="header-fpd" colspan="10">{prefix} FPD. E.P.K (in Ps/kms.)</th>')
        html.append(f'<th class="header-mhl" colspan="10">{prefix} MHL. E.P.K (in Ps/kms.)</th>')
        html.append("</tr><tr>")
        for _ in range(3):
            for d in day_short:
                html.append(f'<th class="header-sub">{d}</th>')
            html.append(f'<th class="header-sub">{for_upto} CY</th>')
            html.append(f'<th class="header-sub">{for_upto} LY</th>')
            html.append('<th class="header-sub">Var</th>')
        html.append("</tr>")
        html.append("</thead><tbody>")
        for sno, (_, row) in enumerate(result.iterrows(), 1):
            html.append("<tr>")
            html.append(f'<td>{sno}</td>')
            html.append(f'<td>{row["DEPOT"]}</td>')
            html.append(f'<td>{row["ROUTE"]}</td>')
            html.append(f'<td>{row["PRODUCT"]}</td>')
            for label in ["TOT", "FPD", "MHL"]:
                # Peak / Slack among Mon-Sun for this metric on this row
                day_vals = {}
                for d in day_order:
                    v = row.get(f"{label}_{d}")
                    try:
                        fv = float(v) if pd.notna(v) else None
                    except Exception:
                        fv = None
                    if fv is not None and fv != 0:
                        day_vals[d] = fv
                peak_d = max(day_vals, key=day_vals.get) if day_vals else None
                slack_d = min(day_vals, key=day_vals.get) if day_vals else None
                for d in day_order:
                    val = row.get(f"{label}_{d}")
                    style = ""
                    if peak_d and d == peak_d:
                        style = 'background-color:#c6efce; color:#006100; font-weight:600;'
                    elif slack_d and d == slack_d:
                        style = 'background-color:#ffc7ce; color:#9c0006; font-weight:600;'
                    html.append(f'<td style="{style}">{fmt(val)}</td>')
                html.append(f'<td>{fmt(row.get(f"{label}_CY"))}</td>')
                html.append(f'<td>{fmt(row.get(f"{label}_LY"))}</td>')
                var_val = row.get(f"{label}_Var")
                html.append(f'<td class="{var_class(var_val)}">{fmt(var_val)}</td>')
            html.append("</tr>")
        html.append("</tbody></table></div>")
        st.markdown("".join(html), unsafe_allow_html=True)
        st.markdown("""
<span style="background:#c6efce; padding:2px 8px;">Peak day</span> = Highest EPK among Mon–Sun &nbsp;
<span style="background:#ffc7ce; padding:2px 8px;">Slack day</span> = Lowest EPK among Mon–Sun
&nbsp;(per row, for each of TOT / FPD / MHL)
""", unsafe_allow_html=True)
        st.download_button(
            "Download Excel",
            excel_with_title(
                result,
                "Route_Daywise",
                report_title=f"Route Day-wise | Depot={depot} | Month={month} | {for_upto} | {net_gross}",
            ),
            f"Route_Daywise_{month}_{depot}.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="dl1",
        )

# ==================== TAB 2 ====================
elif section == "ACT VS ACT":
    pax_heading = {"FPD": "FPD PASSENGERS", "MHL": "MHL PASSENGERS"}.get(passengers, "TOTAL PASSENGERS")
    # Tables ignore depot filter — show all depots (other cascading filters still apply)
    try:
        _no_dep = pd.Series(True, index=df.index)
        if mhl != "ALL" and col_mhl:
            _no_dep &= df[col_mhl].astype(str).str.strip().str.upper() == str(mhl).strip().upper()
        if route != "ALL" and col_route:
            _no_dep &= df[col_route].astype(str).str.strip().str.upper() == str(route).strip().upper()
        if product != "ALL" and col_product:
            _no_dep &= df[col_product].astype(str).str.strip().str.upper() == str(product).strip().upper()
        if rtc != "ALL" and col_rtc:
            _no_dep &= _rtc_match_mask(df[col_rtc], rtc)
        if ac_type != "ALL" and col_product:
            _ac_mask_act = df[col_product].map(_is_ac_product)
            if ac_type == "AC":
                _no_dep &= _ac_mask_act
            else:
                _no_dep &= ~_ac_mask_act
        _mcol2 = col_month if col_month else ("Month_Name" if "Month_Name" in df.columns else None)
        _dcol2 = "Date" if "Date" in df.columns else (col_date if col_date else None)
        # CY max date from selected month (available data)
        if _mcol2 and _dcol2:
            _cy_max = pd.to_datetime(
                df.loc[df[_mcol2].astype(str).str.strip() == str(month).strip(), _dcol2], errors="coerce"
            ).max()
        elif _dcol2:
            _cy_max = pd.to_datetime(df[_dcol2], errors="coerce").max()
        else:
            _cy_max = pd.NaT
        _ly_max = _cy_max - pd.DateOffset(years=1) if pd.notna(_cy_max) else pd.NaT
        if for_upto == "FOR":
            _cy_m = _no_dep & (df[_mcol2].astype(str).str.strip() == str(month).strip()) if _mcol2 else _no_dep
            try:
                mon_name, yr = str(month).split("-")[0], int(str(month).split("-")[1])
                cands = [f"{mon_name}-{yr-1}", f"{mon_name}-{str(yr-1)[-2:]}"]
                _ly_m = pd.Series(False, index=df.index)
                if _mcol2:
                    for c in cands:
                        _ly_m = _ly_m | (df[_mcol2].astype(str).str.strip() == c)
                _ly_m = _no_dep & _ly_m
            except Exception:
                _ly_m = pd.Series(False, index=df.index)
        else:
            # UPTO: from Apr FY start to max available date
            if pd.notna(_cy_max):
                _fy_y = _cy_max.year if _cy_max.month >= 4 else _cy_max.year - 1
                _fy_start = pd.Timestamp(year=_fy_y, month=4, day=1)
                _dt = pd.to_datetime(df[_dcol2], errors="coerce") if _dcol2 else None
                if _dt is not None:
                    _cy_m = _no_dep & (_dt >= _fy_start) & (_dt <= _cy_max)
                    _ly_m = _no_dep & (_dt >= (_fy_start - pd.DateOffset(years=1))) & (_dt <= _ly_max)
                else:
                    _cy_m, _ly_m = _no_dep.copy(), pd.Series(False, index=df.index)
            else:
                _cy_m, _ly_m = _no_dep.copy(), pd.Series(False, index=df.index)
        # Cap FOR to available day-of-month
        if _dcol2 and pd.notna(_cy_max):
            _dt2 = pd.to_datetime(df[_dcol2], errors="coerce")
            _cy_m = _cy_m & (_dt2 <= _cy_max)
            if pd.notna(_ly_max):
                _ly_m = _ly_m & (_dt2 <= _ly_max)
        cy_data = df[_cy_m].copy()
        ly_data = df[_ly_m].copy()
    except Exception as _act_e:
        st.caption(f"ACT VS ACT data rebuild: {_act_e}")

    orf_map, orf_by_prod, orf_err = load_orf_map(r"D:\dashboard\ORF.xlsx")
    if orf_map:
        _bhel = orf_map.get("BHEL", {})
        st.caption(
            f"ORF loaded: {len(orf_map)} depots (from PRODUCT=TOTAL) | "
            f"BHEL CY ORF={_bhel.get('cy', 'n/a')} | example OR if EPK=47.64: "
            f"{(47.64*10000/_bhel['cy']) if _bhel.get('cy') else 'n/a'}"
        )
    if orf_err:
        st.warning(f"ORF: {orf_err}")

    # Heading helper: Actual Vs Actuals {FOR/UPTO} the month of {month} (NET/GROSS) [last date if current month]
    def _act_heading(ng_label):
        fu = str(for_upto).upper()
        date_br = ""
        try:
            if pd.notna(selected_max_date):
                now = pd.Timestamp.now()
                if selected_max_date.year == now.year and selected_max_date.month == now.month:
                    date_br = f" ({selected_max_date.strftime('%d-%m-%Y')})"
                elif fu == "FOR" and selected_max_date.day < 28:
                    # partial month even if not "today's" calendar month
                    date_br = f" ({selected_max_date.strftime('%d-%m-%Y')})"
        except Exception:
            date_br = ""
        return f"Actual Vs Actuals {fu} the month of {month}{date_br} ({ng_label})"

    # Fleet map for AVU / EPB
    try:
        _fleet_map, _fleet_err = load_fleet_map()
    except Exception:
        _fleet_map, _fleet_err = {}, "fleet load failed"
    if _fleet_err:
        st.caption(f"Fleet: {_fleet_err}")

    # ---- TABLE A: NET ----
    st.markdown(f'<div class="title-bar">{_act_heading("NET")}</div>', unsafe_allow_html=True)
    html_net, df_net = build_act_vs_act_table(
        group_col="DEPOT",
        data_cy=cy_data,
        data_ly=ly_data,
        cy_label="CY",
        ly_label="LY",
    )
    if html_net is None or df_net is None or len(df_net) == 0:
        st.warning("No data for NET table.")
    else:
        _et, _ef, _em, _px = earn_tot, earn_fpd, earn_mhl, prefix
        earn_tot, earn_fpd, earn_mhl, prefix = "NE_TOT", "NE_FPD", "NE_MHL", "Net"
        html_net, df_net = build_act_vs_act_table(group_col="DEPOT")
        if df_net is not None and len(df_net) > 0:
            df_net = add_or_columns_depot(df_net, cy_data, ly_data, "NE_TOT", "NE_FPD", "NE_MHL", orf_map, orf_by_prod=orf_by_prod, product_filter=product, depot_filter=depot)
            df_net = add_avu_epb_columns(df_net, cy_data, ly_data, _fleet_map, month, group_col="DEPOT", product_filter=product)
            st.markdown(render_act_table_with_or(df_net, "DEPOT", "NET", pax_heading), unsafe_allow_html=True)
            st.caption("AVU = KMs ÷ (Fleet × Days) · EPB = Earnings ÷ (Fleet × Days) · Fleet from FLEET.parquet · Days = unique dates in period")
        earn_tot, earn_fpd, earn_mhl, prefix = _et, _ef, _em, _px

    # ---- TABLE B: GROSS ----
    st.markdown(f'<div class="title-bar">{_act_heading("GROSS")}</div>', unsafe_allow_html=True)
    _et, _ef, _em, _px = earn_tot, earn_fpd, earn_mhl, prefix
    earn_tot, earn_fpd, earn_mhl, prefix = "GE_TOT", "GE_FPD", "GE_MHL", "Gross"
    html_gr, df_gr = build_act_vs_act_table(group_col="DEPOT")
    if df_gr is None or len(df_gr) == 0:
        st.warning("No data for GROSS table.")
        df_gr = pd.DataFrame()
    else:
        df_gr = add_or_columns_depot(df_gr, cy_data, ly_data, "GE_TOT", "GE_FPD", "GE_MHL", orf_map, orf_by_prod=orf_by_prod, product_filter=product, depot_filter=depot)
        df_gr = add_avu_epb_columns(df_gr, cy_data, ly_data, _fleet_map, month, group_col="DEPOT", product_filter=product)
        st.markdown(render_act_table_with_or(df_gr, "DEPOT", "GROSS", pax_heading), unsafe_allow_html=True)
    earn_tot, earn_fpd, earn_mhl, prefix = _et, _ef, _em, _px

    # Combined formatted Excel (NET + GROSS single sheet)
    _dn = df_net if df_net is not None else pd.DataFrame()
    _dg = df_gr if df_gr is not None else pd.DataFrame()
    if len(_dn) or len(_dg):
        st.download_button(
            "Download Excel – NET + GROSS",
            trends_dual_excel_bytes(
                _dn, _dg, pax_heading=pax_heading,
                report_title=f"ACT VS ACT | Depot={depot} | Month={month} | {for_upto}",
            ),
            f"ACT_VS_ACT_NET_GROSS_{month}.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="dl2_both",
        )

    # ---- 5 charts: CY vs LY by Depot (ignore depot filter; other filters apply) ----
    try:
        import plotly.graph_objects as go
        # Build data WITHOUT depot restriction so all depots always appear
        _cm = pd.Series(True, index=df.index)
        if product != "ALL" and "PRODUCT" in df.columns:
            _cm &= df["PRODUCT"].astype(str).str.strip().str.upper() == str(product).strip().upper()
        if route != "ALL" and "ROUTEE" in df.columns:
            _cm &= df["ROUTEE"].astype(str).str.strip().str.upper() == str(route).strip().upper()
        if mhl != "ALL" and "MHL_NMHL" in df.columns:
            _cm &= df["MHL_NMHL"].astype(str).str.strip().str.upper() == str(mhl).strip().upper()
        if rtc != "ALL" and col_rtc and col_rtc in df.columns:
            _cm &= _rtc_match_mask(df[col_rtc], rtc)
        elif rtc != "ALL" and "RTC_HIRE" in df.columns:
            _cm &= _rtc_match_mask(df["RTC_HIRE"], rtc)
        if ac_type != "ALL" and col_product and col_product in df.columns:
            _ac_ch = df[col_product].map(_is_ac_product)
            if ac_type == "AC":
                _cm &= _ac_ch
            else:
                _cm &= ~_ac_ch
        elif ac_type != "ALL" and "PRODUCT" in df.columns:
            _ac_ch = df["PRODUCT"].map(_is_ac_product)
            if ac_type == "AC":
                _cm &= _ac_ch
            else:
                _cm &= ~_ac_ch
        def _months_upto_local(end_month_str, start_month=4):
            try:
                parts = str(end_month_str).replace(" ", "").split("-")
                mon_map = {"Jan":1,"Feb":2,"Mar":3,"Apr":4,"May":5,"Jun":6,"Jul":7,"Aug":8,"Sep":9,"Oct":10,"Nov":11,"Dec":12}
                em = mon_map.get(parts[0][:3], 4)
                ey = int(parts[1]) if len(parts[1]) == 4 else int("20" + parts[1])
                start_y = ey if em >= start_month else ey - 1
                abbr = {1:"Jan",2:"Feb",3:"Mar",4:"Apr",5:"May",6:"Jun",7:"Jul",8:"Aug",9:"Sep",10:"Oct",11:"Nov",12:"Dec"}
                result, y, m = [], start_y, start_month
                while True:
                    result.append(f"{abbr[m]}-{y}")
                    if y == ey and m == em:
                        break
                    m += 1
                    if m > 12:
                        m, y = 1, y + 1
                    if y > ey + 1:
                        break
                return result
            except Exception:
                return [end_month_str]

        if for_upto == "FOR":
            _cy_ch = df[_cm & (df["Month_Name"].astype(str).str.strip() == str(month).strip())].copy()
            try:
                _parts = month.split("-")
                _ly_m = f"{_parts[0]}-{int(_parts[1]) - 1}"
            except Exception:
                _ly_m = None
            _ly_ch = df[_cm & (df["Month_Name"].astype(str).str.strip() == str(_ly_m).strip())].copy() if _ly_m else pd.DataFrame()
        else:
            _cy_months = _months_upto_local(month)
            _cy_ch = df[_cm & df["Month_Name"].isin(_cy_months)].copy()
            try:
                _parts = month.split("-")
                _ly_m = f"{_parts[0]}-{int(_parts[1]) - 1}"
                _ly_months = _months_upto_local(_ly_m)
            except Exception:
                _ly_months = []
            _ly_ch = df[_cm & df["Month_Name"].isin(_ly_months)].copy() if _ly_months else pd.DataFrame()

        # Cap LY to same calendar day as available CY data
        # e.g. CY max 22-08-2026 → LY only up to 22-08-2025 (not full LY month)
        _dcol_ch = "Date" if "Date" in df.columns else next(
            (c for c in ["DATE", "TravelDate", "TripDate"] if c in df.columns), None
        )
        if _dcol_ch and len(_cy_ch) > 0:
            _cy_max_dt = pd.to_datetime(_cy_ch[_dcol_ch], errors="coerce").max()
            if pd.notna(_cy_max_dt):
                _cy_ch = _cy_ch[pd.to_datetime(_cy_ch[_dcol_ch], errors="coerce") <= _cy_max_dt].copy()
                _ly_max_dt = _cy_max_dt - pd.DateOffset(years=1)
                if len(_ly_ch) > 0:
                    _ly_ch = _ly_ch[pd.to_datetime(_ly_ch[_dcol_ch], errors="coerce") <= _ly_max_dt].copy()

        # Use NET or GROSS earnings columns from filter
        if str(net_gross).upper() == "NET":
            _et, _ef, _em, _pfx = "NE_TOT", "NE_FPD", "NE_MHL", "Net"
        else:
            _et, _ef, _em, _pfx = "GE_TOT", "GE_FPD", "GE_MHL", "Gross"

        def _agg_dep(data, earn_col):
            if data is None or len(data) == 0 or "DEPOT" not in data.columns:
                return pd.DataFrame(columns=["DEPOT", "kms", "earn", "pax"])
            g = data.groupby("DEPOT").agg(
                kms=("Optd_KMs", "sum"),
                earn=(earn_col, "sum"),
                pax=(pax_col, "sum"),
            ).reset_index()
            g["epk"] = np.where(g["kms"] > 0, g["earn"] / g["kms"], np.nan)
            g["kms"] = g["kms"] / 100000.0
            g["earn"] = g["earn"] / 100000.0
            return g

        cy_g = _agg_dep(_cy_ch, _et)
        ly_g = _agg_dep(_ly_ch, _et)
        plot_df = cy_g.merge(ly_g, on="DEPOT", how="outer", suffixes=("_CY", "_LY")).fillna(0)

        # ORF / OR (TOT) per depot from PRODUCT=TOTAL
        try:
            _orf_map, _, _ = load_orf_map(r"D:\dashboard\ORF.xlsx")
        except Exception:
            _orf_map = {}
        def _orf_val(dep, which="cy"):
            d = str(dep).strip().upper()
            if d in ("REGION", "TOTAL", "ALL"):
                d = "REGION"
            rec = _orf_map.get(d, {}) if isinstance(_orf_map, dict) else {}
            return float(rec.get(which) or 0) or np.nan
        plot_df["orf_CY"] = plot_df["DEPOT"].map(lambda d: _orf_val(d, "cy"))
        plot_df["orf_LY"] = plot_df["DEPOT"].map(lambda d: _orf_val(d, "ly"))
        plot_df["or_CY"] = np.where(
            pd.to_numeric(plot_df["orf_CY"], errors="coerce").fillna(0) != 0,
            pd.to_numeric(plot_df["epk_CY"], errors="coerce") * 10000 / plot_df["orf_CY"],
            np.nan,
        )
        plot_df["or_LY"] = np.where(
            pd.to_numeric(plot_df["orf_LY"], errors="coerce").fillna(0) != 0,
            pd.to_numeric(plot_df["epk_LY"], errors="coerce") * 10000 / plot_df["orf_LY"],
            np.nan,
        )

        # Drop TOTAL if any; keep REGION at end
        plot_df = plot_df[~plot_df["DEPOT"].astype(str).str.upper().isin(["TOTAL", ""])].copy()
        plot_df = plot_df.sort_values("DEPOT").reset_index(drop=True)
        if (plot_df["DEPOT"].astype(str).str.upper() == "REGION").any():
            _reg = plot_df[plot_df["DEPOT"].astype(str).str.upper() == "REGION"]
            _oth = plot_df[plot_df["DEPOT"].astype(str).str.upper() != "REGION"]
            plot_df = pd.concat([_oth, _reg], ignore_index=True)

        x_dep = plot_df["DEPOT"].astype(str).tolist()

        def _s(col):
            if col not in plot_df.columns:
                return [0.0] * len(plot_df)
            return pd.to_numeric(plot_df[col], errors="coerce").fillna(0).tolist()

        charts_spec = [
            ("Kilometers", "kms_CY", "kms_LY", "KMs (lakhs)"),
            (f"{_pfx} Earnings", "earn_CY", "earn_LY", f"{_pfx} Earnings (lakhs)"),
            (f"{_pfx} EPK (TOT)", "epk_CY", "epk_LY", f"{_pfx} EPK"),
            (f"{_pfx} OR (TOT)", "or_CY", "or_LY", f"{_pfx} OR"),
            ("Passengers", "pax_CY", "pax_LY", "Passengers"),
        ]

        st.markdown(
            f'<div style="text-align:center;margin:16px 0 8px 0;font-size:1.15rem;font-weight:800;color:#1e40af;">'
            f'ACT vs ACT Charts – {_pfx} | {for_upto} | {month}</div>',
            unsafe_allow_html=True,
        )
        for i in range(0, len(charts_spec), 2):
            cols = st.columns(2)
            for j, col in enumerate(cols):
                if i + j >= len(charts_spec):
                    break
                title, cy_c, ly_c, ylab = charts_spec[i + j]
                cy_vals = _s(cy_c)
                ly_vals = _s(ly_c)
                with col:
                    st.markdown(
                        f'<div style="text-align:center;font-size:13px;font-weight:800;color:#1e40af;'
                        f'background:#eff6ff;border:1px solid #bfdbfe;border-radius:6px;padding:4px 6px;margin-bottom:4px;">'
                        f'{title} – CY vs LY by Depot</div>',
                        unsafe_allow_html=True,
                    )
                    fig = go.Figure()
                    fig.add_trace(go.Bar(
                        name="CY", x=x_dep, y=cy_vals, marker_color="#2563eb",
                        text=[f"{v:.2f}" if abs(v) < 1000 else f"{v:,.0f}" for v in cy_vals],
                        textposition="outside", textfont=dict(size=9, color="#1e3a8a"),
                    ))
                    fig.add_trace(go.Bar(
                        name="LY", x=x_dep, y=ly_vals, marker_color="#15803d",
                        text=[f"{v:.2f}" if abs(v) < 1000 else f"{v:,.0f}" for v in ly_vals],
                        textposition="outside", textfont=dict(size=9, color="#14532d"),
                    ))
                    ymax = max([0.1] + [abs(v) for v in cy_vals + ly_vals]) * 1.25
                    fig.update_layout(
                        barmode="group", height=340,
                        margin=dict(l=50, r=20, t=30, b=50),
                        legend=dict(orientation="h", y=1.12, x=0.5, xanchor="center"),
                        xaxis=dict(tickfont=dict(color="#dc2626", size=11), title="Depot"),
                        yaxis=dict(title=ylab, range=[0, ymax]),
                        template="plotly_white", bargap=0.25,
                    )
                    st.plotly_chart(fig, width="stretch")

        # ---- 6th chart: REGION – each metric own scale (so Earnings does not dwarf others) ----
        try:
            from plotly.subplots import make_subplots
            _reg_src = plot_df[~plot_df["DEPOT"].astype(str).str.upper().isin(["REGION", "TOTAL"])].copy()
            if len(_reg_src) == 0:
                _reg_src = plot_df.copy()

            def _sum(col):
                if col not in _reg_src.columns:
                    return 0.0
                return float(pd.to_numeric(_reg_src[col], errors="coerce").fillna(0).sum())

            def _wavg(val_col, w_col):
                if val_col not in _reg_src.columns or w_col not in _reg_src.columns:
                    return 0.0
                w = pd.to_numeric(_reg_src[w_col], errors="coerce").fillna(0)
                v = pd.to_numeric(_reg_src[val_col], errors="coerce").fillna(0)
                return float((v * w).sum() / w.sum()) if w.sum() else 0.0

            kms_cy, kms_ly = _sum("kms_CY"), _sum("kms_LY")
            earn_cy, earn_ly = _sum("earn_CY"), _sum("earn_LY")
            pax_cy, pax_ly = _sum("pax_CY") / 100000.0, _sum("pax_LY") / 100000.0
            epk_cy = (earn_cy / kms_cy) if kms_cy else 0.0
            epk_ly = (earn_ly / kms_ly) if kms_ly else 0.0
            or_cy = _wavg("or_CY", "kms_CY")
            or_ly = _wavg("or_LY", "kms_LY")

            pax_lbl = {"FPD": "FPD Passengers (lakhs)", "MHL": "MHL Passengers (lakhs)"}.get(
                str(passengers).upper() if passengers is not None else "", "TOT Passengers (lakhs)"
            )
            if str(passengers).upper() not in ("FPD", "MHL"):
                pax_lbl = "TOT Passengers (lakhs)"

            metrics = [
                ("KMs", kms_cy, kms_ly),
                (f"{_pfx} Earnings", earn_cy, earn_ly),
                (f"{_pfx} EPK", epk_cy, epk_ly),
                (f"{_pfx} OR", or_cy, or_ly),
                (pax_lbl, pax_cy, pax_ly),
            ]

            st.markdown(
                f'<div style="text-align:center;font-size:14px;font-weight:800;color:#1e40af;'
                f'background:#eff6ff;border:1px solid #bfdbfe;border-radius:6px;padding:6px;margin:12px 0 6px 0;">'
                f'Region – CY vs LY | {_pfx} | {for_upto} | {month}</div>',
                unsafe_allow_html=True,
            )

            fig_r = make_subplots(rows=1, cols=5, shared_yaxes=False,
                                  subplot_titles=[m[0] for m in metrics],
                                  horizontal_spacing=0.06)
            for i, (lab, cy_v, ly_v) in enumerate(metrics, start=1):
                fig_r.add_trace(
                    go.Bar(
                        name="CY", x=["CY"], y=[cy_v], marker_color="#2563eb",
                        text=[f"{cy_v:.2f}" if abs(cy_v) < 1000 else f"{cy_v:,.0f}"],
                        textposition="outside", textfont=dict(size=11, color="#1e3a8a"),
                        showlegend=(i == 1),
                    ),
                    row=1, col=i,
                )
                fig_r.add_trace(
                    go.Bar(
                        name="LY", x=["LY"], y=[ly_v], marker_color="#15803d",
                        text=[f"{ly_v:.2f}" if abs(ly_v) < 1000 else f"{ly_v:,.0f}"],
                        textposition="outside", textfont=dict(size=11, color="#14532d"),
                        showlegend=(i == 1),
                    ),
                    row=1, col=i,
                )
                top = max(abs(cy_v), abs(ly_v), 0.1) * 1.35
                fig_r.update_yaxes(
                    range=[0, top], visible=False, showticklabels=False,
                    showgrid=False, zeroline=False, row=1, col=i,
                )
                fig_r.update_xaxes(
                    showticklabels=False, showgrid=False, title=None, row=1, col=i,
                )

            fig_r.update_layout(
                barmode="group", height=320,
                margin=dict(l=10, r=10, t=50, b=20),
                legend=dict(orientation="h", y=1.18, x=0.5, xanchor="center"),
                template="plotly_white", bargap=0.25,
            )
            # subplot title style
            for ann in fig_r["layout"]["annotations"]:
                ann["font"] = dict(size=12, color="#dc2626")
            st.plotly_chart(fig_r, width="stretch")
        except Exception as _re:
            st.caption(f"Region chart: {_re}")

    except Exception as _ace:
        st.caption(f"ACT vs ACT charts: {_ace}")


elif section == "Product wise":
    pax_heading = {"FPD": "FPD PASSENGERS", "MHL": "MHL PASSENGERS"}.get(passengers, "TOTAL PASSENGERS")
    orf_map, orf_by_prod, orf_err = load_orf_map(r"D:\dashboard\ORF.xlsx")
    if orf_err:
        st.warning(f"ORF: {orf_err}")

    # Schedules from SMASTER.parquet (shared for both tables)
    sch_map = {}
    try:
        sros_path = Path(r"D:\Dashboard\SMASTER.parquet")
        if not sros_path.exists():
            for alt in [Path(r"D:\dashboard\SMASTER.parquet"), Path(r"D:\MONTHLY\SMASTER.parquet"), Path("SMASTER.parquet"), Path(r"/home/workdir/attachments/SMASTER.parquet")]:
                if alt.exists():
                    sros_path = alt
                    break
        if sros_path.exists():
            sch_map = _load_sros_sch_map_pw(str(sros_path), str(month), str(depot))
    except Exception as _e:
        st.caption(f"SROS schedules: {_e}")

    def _attach_sch(df):
        df = df.copy()
        df["sch_CY"] = df["PRODUCT"].map(lambda p: sch_map.get(str(p).strip(), (0.0, 0.0))[0] if str(p).upper() != "TOTAL" else 0.0)
        df["sch_LY"] = df["PRODUCT"].map(lambda p: sch_map.get(str(p).strip(), (0.0, 0.0))[1] if str(p).upper() != "TOTAL" else 0.0)
        if (df["PRODUCT"].astype(str).str.upper() == "TOTAL").any():
            det = df[df["PRODUCT"].astype(str).str.upper() != "TOTAL"]
            df.loc[df["PRODUCT"].astype(str).str.upper() == "TOTAL", "sch_CY"] = det["sch_CY"].sum()
            df.loc[df["PRODUCT"].astype(str).str.upper() == "TOTAL", "sch_LY"] = det["sch_LY"].sum()
        df["sch_VAR"] = df["sch_CY"] - df["sch_LY"]
        return df

    # ---- TABLE A NET ----
    st.markdown(f'<div class="title-bar">TABLE A – NET | Product Wise | {for_upto} | {month}</div>', unsafe_allow_html=True)
    _et, _ef, _em, _px = earn_tot, earn_fpd, earn_mhl, prefix
    earn_tot, earn_fpd, earn_mhl, prefix = "NE_TOT", "NE_FPD", "NE_MHL", "Net"
    _, df_net = build_act_vs_act_table(group_col="PRODUCT")
    earn_tot, earn_fpd, earn_mhl, prefix = _et, _ef, _em, _px
    if df_net is None or len(df_net) == 0:
        st.warning("No data for NET product table.")
    else:
        df_net = _attach_sch(df_net)
        df_net = add_or_columns_product(df_net, orf_map, orf_by_prod=orf_by_prod, depot_filter=depot)
        st.markdown(render_product_table_with_or(df_net, "NET", pax_heading), unsafe_allow_html=True)

    # ---- TABLE B GROSS ----
    st.markdown('<div style="height:24px;clear:both;"></div>', unsafe_allow_html=True)
    st.markdown(f'<div class="title-bar" style="margin-top:12px;">TABLE B – GROSS | Product Wise | {for_upto} | {month}</div>', unsafe_allow_html=True)
    earn_tot, earn_fpd, earn_mhl, prefix = "GE_TOT", "GE_FPD", "GE_MHL", "Gross"
    _, df_gr = build_act_vs_act_table(group_col="PRODUCT")
    earn_tot, earn_fpd, earn_mhl, prefix = _et, _ef, _em, _px
    if df_gr is None or len(df_gr) == 0:
        st.warning("No data for GROSS product table.")
        df_gr = pd.DataFrame()
    else:
        df_gr = _attach_sch(df_gr)
        df_gr = add_or_columns_product(df_gr, orf_map, orf_by_prod=orf_by_prod, depot_filter=depot)
        st.markdown(render_product_table_with_or(df_gr, "GROSS", pax_heading), unsafe_allow_html=True)

    _dn = df_net if df_net is not None else pd.DataFrame()
    _dg = df_gr if df_gr is not None else pd.DataFrame()
    if len(_dn) or len(_dg):
        st.download_button(
            "Download Excel – NET + GROSS",
            trends_dual_excel_bytes(
                _dn, _dg, pax_heading=pax_heading,
                report_title=f"Product Wise | Depot={depot} | Month={month} | {for_upto}",
            ),
            f"Product_NET_GROSS_{month}.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="dl3_both",
        )

# ==================== TAB 4 ====================
elif section == "Day wise":
    prep = "upto" if for_upto == "UPTO" else "for"

    # Construct the dynamic title
    title5 = f"'AVERAGE' Weekday Wise Actual vs Actuals"
    if depot != "ALL":
        title5 += f" of {depot} Depot"
    else:
        title5 += " of ALL Depots"

    title5 += f" {prep} the month of {month}"

    st.markdown(f'<div class="title-bar">{title5}</div>', unsafe_allow_html=True)
    if len(cy_data) == 0 and for_upto != "UPTO":
        st.warning("No data for selected filters.")
    else:
        from calendar import monthcalendar
        day_order = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
        mon_map = {
            "Jan": 1, "Feb": 2, "Mar": 3, "Apr": 4, "May": 5, "Jun": 6,
            "Jul": 7, "Aug": 8, "Sep": 9, "Oct": 10, "Nov": 11, "Dec": 12
        }
        def parse_month(mstr):
            try:
                parts = mstr.replace(" ", "").split("-")
                return mon_map.get(parts[0][:3], 1), int(parts[1])
            except:
                return 1, 2026
        def count_weekdays_in_month(month_str):
            """Full calendar month (fallback only)."""
            try:
                m, y = parse_month(month_str)
                cal = monthcalendar(y, m)
                counts = {d: 0 for d in day_order}
                for week in cal:
                    for idx, day in enumerate(week):
                        if day != 0:
                            counts[day_order[idx]] += 1
                return counts
            except:
                return {d: 4 for d in day_order}

        def count_weekdays_from_data(data, fallback_month=None):
            """Count weekdays from first to last available Date in data (not full month)."""
            try:
                if data is None or len(data) == 0:
                    return count_weekdays_in_month(fallback_month) if fallback_month else {d: 4 for d in day_order}
                date_col = next((c for c in ["Date", "DATE", "TravelDate", "TripDate"] if c in data.columns), None)
                if date_col is None:
                    return count_weekdays_in_month(fallback_month) if fallback_month else {d: 4 for d in day_order}
                dts = pd.to_datetime(data[date_col], errors="coerce").dropna()
                if len(dts) == 0:
                    return count_weekdays_in_month(fallback_month) if fallback_month else {d: 4 for d in day_order}
                start = dts.min().normalize()
                end = dts.max().normalize()
                counts = {d: 0 for d in day_order}
                for d in pd.date_range(start, end, freq="D"):
                    counts[day_order[d.weekday()]] += 1
                return counts
            except Exception:
                return count_weekdays_in_month(fallback_month) if fallback_month else {d: 4 for d in day_order}

        def count_weekdays_upto(end_month_str, start_month=4):
            try:
                em, ey = parse_month(end_month_str)
                start_y = ey if em >= start_month else ey - 1
                counts = {d: 0 for d in day_order}
                y, m = start_y, start_month
                while True:
                    cal = monthcalendar(y, m)
                    for week in cal:
                        for idx, day in enumerate(week):
                            if day != 0:
                                counts[day_order[idx]] += 1
                    if y == ey and m == em:
                        break
                    m += 1
                    if m > 12:
                        m = 1
                        y += 1
                    if y > ey + 1:
                        break
                return counts
            except:
                return {d: 4 for d in day_order}
        def get_months_upto(end_month_str, start_month=4):
            try:
                em, ey = parse_month(end_month_str)
                start_y = ey if em >= start_month else ey - 1
                result = []
                y, m = start_y, start_month
                abbr = {1:"Jan",2:"Feb",3:"Mar",4:"Apr",5:"May",6:"Jun",
                        7:"Jul",8:"Aug",9:"Sep",10:"Oct",11:"Nov",12:"Dec"}
                while True:
                    result.append(f"{abbr[m]}-{y}")
                    if y == ey and m == em:
                        break
                    m += 1
                    if m > 12:
                        m = 1
                        y += 1
                    if y > ey + 1:
                        break
                return result
            except:
                return [end_month_str]
        try:
            parts = month.split("-")
            _ly_month = f"{parts[0]}-{int(parts[1]) - 1}"
        except:
            _ly_month = None
        if for_upto == "UPTO":
            cy_months = get_months_upto(month)
            ly_months = get_months_upto(_ly_month) if _ly_month else []
            cy_use = df[base_mask & df["Month_Name"].isin(cy_months)].copy()
            ly_use = df[base_mask & df["Month_Name"].isin(ly_months)].copy() if ly_months else pd.DataFrame()
        else:
            cy_use = cy_data
            ly_use = ly_data
        # Weekday counts from available data only (first→last Date present), not full calendar month
        cy_counts = count_weekdays_from_data(cy_use, fallback_month=month)
        ly_counts = count_weekdays_from_data(ly_use, fallback_month=_ly_month) if _ly_month else {d: 4 for d in day_order}
        def weekday_excel_style(data, earn_tot_col, earn_fpd_col, earn_mhl_col, wd_counts):
            if len(data) == 0:
                return pd.DataFrame()
            data = data.copy()
            # Ensure Weekday column exists (Mon, Tue, ...)
            if "Weekday" not in data.columns:
                date_col = next((c for c in ["Date", "DATE", "TravelDate", "TripDate"] if c in data.columns), None)
                if date_col is None:
                    return pd.DataFrame()
                data["_dt"] = pd.to_datetime(data[date_col], errors="coerce")
                data = data.dropna(subset=["_dt"])
                data["Weekday"] = data["_dt"].dt.strftime("%a")  # Mon, Tue, Wed...
            else:
                # normalize existing weekday values
                data["Weekday"] = data["Weekday"].astype(str).str.strip().str[:3].str.title()
                data["Weekday"] = data["Weekday"].replace({
                    "Mon": "Mon", "Tue": "Tue", "Wed": "Wed", "Thu": "Thu",
                    "Fri": "Fri", "Sat": "Sat", "Sun": "Sun",
                    "Monday": "Mon", "Tuesday": "Tue", "Wednesday": "Wed",
                    "Thursday": "Thu", "Friday": "Fri", "Saturday": "Sat", "Sunday": "Sun",
                })
            if len(data) == 0 or "Weekday" not in data.columns:
                return pd.DataFrame()
            g = data.groupby("Weekday").agg(
                Total_KMs=("Optd_KMs", "sum"),
                Total_Earn_TOT=(earn_tot_col, "sum"),
                Total_Earn_FPD=(earn_fpd_col, "sum"),
                Total_Earn_MHL=(earn_mhl_col, "sum"),
                Total_Pax=(pax_col, "sum"),
            ).reset_index()
            g["Days_In_Month"] = g["Weekday"].map(wd_counts).fillna(4)
            g["Avg_KMs"] = (g["Total_KMs"] / g["Days_In_Month"]) / 100000
            g["Avg_Earn_TOT"] = (g["Total_Earn_TOT"] / g["Days_In_Month"]) / 100000
            g["Avg_Pax"] = (g["Total_Pax"] / g["Days_In_Month"]).round(0)
            g["EPK_TOT"] = np.where(g["Total_KMs"] > 0, g["Total_Earn_TOT"] / g["Total_KMs"], np.nan)
            g["EPK_FPD"] = np.where(g["Total_KMs"] > 0, g["Total_Earn_FPD"] / g["Total_KMs"], np.nan)
            g["EPK_MHL"] = np.where(g["Total_KMs"] > 0, g["Total_Earn_MHL"] / g["Total_KMs"], np.nan)
            g["Weekday"] = pd.Categorical(g["Weekday"], categories=day_order, ordered=True)
            return g.sort_values("Weekday").reset_index(drop=True)
        cy_wd = weekday_excel_style(cy_use, earn_tot, earn_fpd, earn_mhl, cy_counts)
        ly_wd = weekday_excel_style(ly_use, earn_tot, earn_fpd, earn_mhl, ly_counts)
        if len(cy_wd) == 0:
            st.warning("No data for selected filters.")
        else:
            merged = cy_wd.merge(ly_wd, on="Weekday", how="outer", suffixes=("_CY", "_LY"))
            merged = merged.sort_values("Weekday").reset_index(drop=True)
            for metric in ["Avg_KMs", "Avg_Earn_TOT", "EPK_TOT", "EPK_FPD", "EPK_MHL", "Avg_Pax"]:
                merged[f"{metric}_VAR"] = merged.get(f"{metric}_CY", 0).fillna(0) - merged.get(f"{metric}_LY", 0).fillna(0)
            # TOTAL = average of Mon-Sun
            total_row = {"Weekday": "TOTAL"}
            for metric in ["Avg_KMs", "Avg_Earn_TOT", "EPK_TOT", "EPK_FPD", "EPK_MHL", "Avg_Pax"]:
                total_row[f"{metric}_CY"] = merged[f"{metric}_CY"].mean()
                total_row[f"{metric}_LY"] = merged[f"{metric}_LY"].mean()
                total_row[f"{metric}_VAR"] = (total_row[f"{metric}_CY"] or 0) - (total_row[f"{metric}_LY"] or 0)
            if len(cy_use) > 0 and cy_use["Optd_KMs"].sum() > 0:
                total_row["EPK_TOT_CY"] = cy_use[earn_tot].sum() / cy_use["Optd_KMs"].sum()
                total_row["EPK_FPD_CY"] = cy_use[earn_fpd].sum() / cy_use["Optd_KMs"].sum()
                total_row["EPK_MHL_CY"] = cy_use[earn_mhl].sum() / cy_use["Optd_KMs"].sum()
            if len(ly_use) > 0 and ly_use["Optd_KMs"].sum() > 0:
                total_row["EPK_TOT_LY"] = ly_use[earn_tot].sum() / ly_use["Optd_KMs"].sum()
                total_row["EPK_FPD_LY"] = ly_use[earn_fpd].sum() / ly_use["Optd_KMs"].sum()
                total_row["EPK_MHL_LY"] = ly_use[earn_mhl].sum() / ly_use["Optd_KMs"].sum()
            for m in ["EPK_TOT", "EPK_FPD", "EPK_MHL"]:
                total_row[f"{m}_VAR"] = (total_row.get(f"{m}_CY") or 0) - (total_row.get(f"{m}_LY") or 0)
            merged = pd.concat([merged, pd.DataFrame([total_row])], ignore_index=True)
            # Colour scale helper for CY values (High=Green → Low=Red)
            def cy_color(val, series):
                if pd.isna(val):
                    return ""
                s = series.dropna()
                if len(s) < 2:
                    return ""
                mn, mx = s.min(), s.max()
                if mx == mn:
                    return "background-color:#ffff99;"
                # 0 = low (red), 1 = high (green)
                ratio = (float(val) - mn) / (mx - mn)
                # interpolate red -> yellow -> green
                if ratio >= 0.6:
                    return "background-color:#c6efce; color:#006100;" # high green
                elif ratio >= 0.3:
                    return "background-color:#ffeb9c; color:#9c5700;" # mid yellow
                else:
                    return "background-color:#ffc7ce; color:#9c0006;" # low red
            # Precompute series for each CY metric (exclude TOTAL)
            non_tot = merged[merged["Weekday"] != "TOTAL"]
            cy_series = {
                "Avg_KMs": non_tot["Avg_KMs_CY"] if "Avg_KMs_CY" in non_tot else pd.Series(),
                "Avg_Earn_TOT": non_tot["Avg_Earn_TOT_CY"] if "Avg_Earn_TOT_CY" in non_tot else pd.Series(),
                "EPK_TOT": non_tot["EPK_TOT_CY"] if "EPK_TOT_CY" in non_tot else pd.Series(),
                "EPK_FPD": non_tot["EPK_FPD_CY"] if "EPK_FPD_CY" in non_tot else pd.Series(),
                "EPK_MHL": non_tot["EPK_MHL_CY"] if "EPK_MHL_CY" in non_tot else pd.Series(),
                "Avg_Pax": non_tot["Avg_Pax_CY"] if "Avg_Pax_CY" in non_tot else pd.Series(),
            }
            # Charts
            import plotly.graph_objects as go
            from plotly.subplots import make_subplots
            chart_data = merged[merged["Weekday"] != "TOTAL"].copy()
            fig = make_subplots(rows=1, cols=5,
                subplot_titles=("KILOMETERS (in lks.)", f"{prefix} TOT. E.P.K",
                                f"{prefix} FPD. E.P.K", f"{prefix} MHL. E.P.K", {"FPD": "FPD PASSENGERS", "MHL": "MHL PASSENGERS"}.get(passengers, "TOTAL PASSENGERS")),
                horizontal_spacing=0.04)
            for metric, col in [("Avg_KMs",1),("EPK_TOT",2),("EPK_FPD",3),("EPK_MHL",4),("Avg_Pax",5)]:
                fig.add_trace(go.Bar(name="CY", x=chart_data["Weekday"], y=chart_data[f"{metric}_CY"],
                    marker_color="#4472C4", showlegend=(col==1),
                    text=[f"{v:.2f}" if pd.notna(v) else "" for v in chart_data[f"{metric}_CY"]],
                    textposition="outside", textfont_size=9), row=1, col=col)
                fig.add_trace(go.Bar(name="LY", x=chart_data["Weekday"], y=chart_data[f"{metric}_LY"],
                    marker_color="#ED7D31", showlegend=(col==1),
                    text=[f"{v:.2f}" if pd.notna(v) else "" for v in chart_data[f"{metric}_LY"]],
                    textposition="outside", textfont_size=9), row=1, col=col)
            fig.update_layout(barmode="group", height=340, margin=dict(l=20,r=20,t=80,b=20),
                legend=dict(orientation="h", y=1.15, font=dict(color="white")),
                plot_bgcolor="white", paper_bgcolor="#1e3a5f",
                font=dict(color="white", size=10),
                title_font=dict(color="white"))
            fig.update_xaxes(tickfont=dict(size=9, color="white"))
            fig.update_yaxes(tickfont=dict(size=9, color="white"), gridcolor="#555")
            for ann in fig["layout"]["annotations"]:
                ann["font"] = dict(size=11, color="#FFD700", family="Segoe UI")
                try:
                    ann["y"] = float(ann["y"]) + 0.08 if ann["y"] is not None else 1.08
                except Exception:
                    pass
            st.plotly_chart(fig, width="stretch")
            # Table - only CY, LY, VAR (no % Growth)
            html = ['<div class="table-scroll"><table class="excel-table">']
            html.append('<tr>')
            html.append('<th rowspan="2" style="background:#c000c0; color:white;">DAY</th>')
            html.append('<th colspan="3" style="background:#c000c0; color:white;">KILOMETERS (in lks.)</th>')
            html.append(f'<th colspan="3" style="background:#c000c0; color:white;">{prefix} EARNINGS (Rs.in lks.)</th>')
            html.append(f'<th colspan="3" style="background:#c000c0; color:white;">{prefix} TOT. E.P.K</th>')
            html.append(f'<th colspan="3" style="background:#c000c0; color:white;">{prefix} FPD. E.P.K</th>')
            html.append(f'<th colspan="3" style="background:#c000c0; color:white;">{prefix} MHL. E.P.K</th>')
            pax_heading = {"FPD": "FPD PASSENGERS", "MHL": "MHL PASSENGERS"}.get(passengers, "TOTAL PASSENGERS")
            html.append(f'<th colspan="3" style="background:#c000c0; color:white;">{pax_heading}</th>')
            html.append('</tr><tr>')
            for _ in range(6):
                for h in ["CY", "LY", "VAR"]:
                    html.append(f'<th style="background:#ff66ff; color:black; font-size:10px;">{h}</th>')
            html.append('</tr>')
            for _, row in merged.iterrows():
                is_total = str(row["Weekday"]) == "TOTAL"
                style = 'font-weight:bold; background:#e2efda;' if is_total else ''
                html.append(f'<tr style="{style}">')
                html.append(f'<td><b>{row["Weekday"]}</b></td>')
                for metric in ["Avg_KMs", "Avg_Earn_TOT", "EPK_TOT", "EPK_FPD", "EPK_MHL", "Avg_Pax"]:
                    cy_val = row.get(f"{metric}_CY")
                    ly_val = row.get(f"{metric}_LY")
                    var_val = row.get(f"{metric}_VAR")
                    show = fmt_pax if metric == "Avg_Pax" else fmt
                    if is_total:
                        html.append(f'<td>{show(cy_val)}</td>')
                    else:
                        html.append(f'<td style="{cy_color(cy_val, cy_series[metric])}">{show(cy_val)}</td>')
                    html.append(f'<td>{show(ly_val)}</td>')
                    html.append(f'<td>{show(var_val)}</td>')
                html.append('</tr>')
            html.append('</table></div>')
            st.markdown("".join(html), unsafe_allow_html=True)
            st.markdown("""
**designed by:**
&nbsp;&nbsp;|&nbsp;&nbsp; kiran kumar
""", unsafe_allow_html=True)

            # ---- Per-depot dual-axis charts (3 per row) + Region chart ----
            try:
                import plotly.graph_objects as go
                from plotly.subplots import make_subplots

                day_ord = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
                day_upper = ["MON", "TUE", "WED", "THU", "FRI", "SAT", "SUN"]

                # Source: all depots for selected month/filters (ignore single-depot pick for grid)
                _src = cy_use.copy()
                if "DEPOT" not in _src.columns or _src["DEPOT"].nunique() < 2:
                    _src = df.copy()
                    if product != "ALL" and "PRODUCT" in _src.columns:
                        _src = _src[_src["PRODUCT"].astype(str).str.strip() == str(product).strip()]
                    if route != "ALL" and "ROUTEE" in _src.columns:
                        _src = _src[_src["ROUTEE"].astype(str).str.strip() == str(route).strip()]
                    if mhl != "ALL" and "MHL_NMHL" in _src.columns:
                        _src = _src[_src["MHL_NMHL"].astype(str).str.strip() == str(mhl).strip()]
                    if rtc != "ALL" and "RTC_HIRE" in _src.columns:
                        _src = _src[_src["RTC_HIRE"].astype(str).str.strip() == str(rtc).strip()]
                    if for_upto == "FOR" and "Month_Name" in _src.columns:
                        _src = _src[_src["Month_Name"].astype(str).str.strip() == str(month).strip()]
                    elif "Month_Name" in _src.columns:
                        try:
                            cy_months = get_months_upto(month)
                            _src = _src[_src["Month_Name"].isin(cy_months)]
                        except Exception:
                            _src = _src[_src["Month_Name"].astype(str).str.strip() == str(month).strip()]

                if "Weekday" not in _src.columns:
                    _src["_dt"] = pd.to_datetime(_src["Date"], errors="coerce")
                    _src = _src.dropna(subset=["_dt"])
                    _src["Weekday"] = _src["_dt"].dt.strftime("%a")
                else:
                    _src["Weekday"] = _src["Weekday"].astype(str).str.strip().str[:3].str.title()

                def _wd_metrics(data):
                    if len(data) == 0:
                        return {d: (0.0, 0.0) for d in day_ord}
                    g = data.groupby("Weekday").agg(
                        Total_KMs=("Optd_KMs", "sum"),
                        Total_Earn=(earn_tot, "sum"),
                    ).reset_index()
                    g["Weekday"] = g["Weekday"].astype(str).str[:3].str.title()
                    g["Days"] = g["Weekday"].map(cy_counts).fillna(4)
                    g["Vol"] = (g["Total_KMs"] / g["Days"]) / 100000.0
                    g["EPK"] = np.where(g["Total_KMs"] > 0, g["Total_Earn"] / g["Total_KMs"], 0.0)
                    out = {}
                    for d in day_ord:
                        row = g[g["Weekday"] == d]
                        if len(row):
                            out[d] = (float(row["Vol"].iloc[0]), float(row["EPK"].iloc[0]))
                        else:
                            out[d] = (0.0, 0.0)
                    return out

                def _chart_heading(depot_label):
                    # Day-wise KMs & Gross E.P.K. of HYD2 Depot FOR Aug-2026
                    return f"Day-wise KMs & {prefix} E.P.K. of {depot_label} {for_upto} {month}"

                def _make_fig(metrics, left_range, left_dtick, left_tickvals=None):
                    vol = [metrics[d][0] for d in day_ord]
                    epk = [metrics[d][1] for d in day_ord]
                    fig = make_subplots(specs=[[{"secondary_y": True}]])
                    fig.add_trace(
                        go.Bar(
                            name="Volume", x=day_upper, y=vol, marker_color="#2563eb",
                            text=[f"{v:.2f}" for v in vol], textposition="outside",
                            textfont=dict(size=10, color="#1e3a8a"), offsetgroup="a",
                            showlegend=False,
                        ),
                        secondary_y=False,
                    )
                    fig.add_trace(
                        go.Bar(
                            name="E.P.K.", x=day_upper, y=epk, marker_color="#15803d",
                            text=[f"{v:.2f}" for v in epk], textposition="outside",
                            textfont=dict(size=10, color="#14532d"), offsetgroup="b",
                            showlegend=False,
                        ),
                        secondary_y=True,
                    )
                    epk_max = max(epk) if epk else 80
                    right_top = max(80.0, epk_max * 1.15)
                    fig.update_layout(
                        # no plotly title — heading rendered above via HTML (full text visible)
                        barmode="group", height=320,
                        margin=dict(l=50, r=50, t=20, b=30),
                        template="plotly_white", bargap=0.2,
                        xaxis=dict(tickfont=dict(color="#dc2626", size=11)),
                    )
                    _yaxis_left = dict(
                        title_text="KMs", range=left_range,
                        secondary_y=False, color="#2563eb", title_font=dict(size=10),
                        tickfont=dict(size=9),
                    )
                    if left_tickvals is not None:
                        _yaxis_left["tickvals"] = left_tickvals
                        _yaxis_left["tickmode"] = "array"
                    else:
                        _yaxis_left["dtick"] = left_dtick
                    fig.update_yaxes(**_yaxis_left)
                    fig.update_yaxes(
                        title_text="EPK", range=[0, right_top],
                        secondary_y=True, color="#15803d", title_font=dict(size=10),
                        tickfont=dict(size=9), showgrid=False,
                    )
                    return fig, vol, epk

                def _summary_html(vol, epk):
                    # 2 Peak days (highest EPK) + 2 Slack days (lowest EPK)
                    valid = [(i, v) for i, v in enumerate(epk) if v is not None and not (isinstance(v, float) and v != v)]
                    # sort by EPK descending for peaks, ascending for slack
                    by_high = sorted(valid, key=lambda x: x[1], reverse=True)
                    by_low = sorted(valid, key=lambda x: x[1])
                    peak_set = set(i for i, _ in by_high[:2]) if by_high else set()
                    # exclude peak days from slack so no overlap
                    slack_cands = [(i, v) for i, v in by_low if i not in peak_set]
                    slack_set = set(i for i, _ in slack_cands[:2]) if slack_cands else set()

                    html_t = ['<table style="margin:4px auto 16px auto;border-collapse:collapse;font-size:12px;width:100%;">']
                    html_t.append('<tr><th style="background:#1e3a8a;color:#fff;padding:5px 6px;">Day</th>')
                    for d in day_upper:
                        html_t.append(f'<th style="background:#1e3a8a;color:#fff;padding:5px 6px;">{d}</th>')
                    html_t.append('</tr><tr><td style="background:#2563eb;color:#fff;padding:5px 6px;font-weight:700;">Volume</td>')
                    for v in vol:
                        html_t.append(f'<td style="border:1px solid #cbd5e1;padding:5px 6px;text-align:center;">{v:.2f}</td>')
                    html_t.append('</tr><tr><td style="background:#15803d;color:#fff;padding:5px 6px;font-weight:700;">E.P.K.</td>')
                    for i, v in enumerate(epk):
                        if i in peak_set:
                            sty = "background:#c6efce;color:#006100;font-weight:700;border:1px solid #86efac;"
                        elif i in slack_set:
                            sty = "background:#ffc7ce;color:#9c0006;font-weight:700;border:1px solid #fca5a5;"
                        else:
                            sty = "border:1px solid #cbd5e1;"
                        html_t.append(f'<td style="{sty}padding:5px 6px;text-align:center;">{v:.2f}</td>')
                    html_t.append('</tr></table>')
                    return "".join(html_t)

                depots_list = sorted(
                    [str(x).strip() for x in _src["DEPOT"].dropna().unique()
                     if str(x).strip() and str(x).upper() not in ("NAN", "REGION", "TOTAL", "ALL")]
                )
                # Left axis scales:
                #   BHEL, HYD1, HYD2 (+ Region): 0.3, 0.6, 0.9, 1.2
                #   PKT, TNDR, VKB (+ others): existing 0.15 steps to 0.75
                HIGH_SCALE_DEPOTS = {"BHEL", "HYD1", "HYD2", "PKT"}
                chart_items = []
                for dep in depots_list:
                    sub = _src[_src["DEPOT"].astype(str).str.strip().str.upper() == dep.upper()]
                    if dep.upper() in HIGH_SCALE_DEPOTS:
                        left_range, left_dtick = [0, 1.2], 0.3
                    else:
                        left_range, left_dtick = [0, 0.75], 0.15  # PKT, TNDR, VKB, PRG, etc.
                    chart_items.append((_chart_heading(f"{dep} Depot"), _wd_metrics(sub), left_range, left_dtick, None))
                # Region same high scale as BHEL/HYD1/HYD2
                chart_items.append((_chart_heading("Rangareddy Region"), _wd_metrics(_src), [0, 6], 1.5, [1.5, 3, 3.5, 4, 4.5, 6]))

                # 2 charts per row for better visibility — full heading above each chart
                for i in range(0, len(chart_items), 2):
                    cols = st.columns(2)
                    for j, col in enumerate(cols):
                        if i + j < len(chart_items):
                            title, mets, lr, dt, tickvals = chart_items[i + j]
                            with col:
                                st.markdown(
                                    f'<div style="text-align:center;font-size:13px;font-weight:800;color:#1e40af;'
                                    f'line-height:1.35;margin:4px 2px 2px 2px;padding:4px 6px;'
                                    f'background:#eff6ff;border-radius:6px;border:1px solid #bfdbfe;">'
                                    f'{title}</div>',
                                    unsafe_allow_html=True,
                                )
                                fig, vol, epk = _make_fig(mets, lr, dt, left_tickvals=tickvals)
                                st.plotly_chart(fig, width="stretch")
                                st.markdown(_summary_html(vol, epk), unsafe_allow_html=True)
            except Exception as _ce:
                st.caption(f"Day-wise chart: {_ce}")

            st.caption(f"Mode: {for_upto} | Weekday counts CY: {cy_counts}")
            st.download_button(
                "Download Excel",
                excel_with_title(
                    merged,
                    "WeekDay",
                    report_title=f"Week Day | Depot={depot} | Month={month} | {for_upto} | {net_gross}",
                ),
                f"WeekDay_{month}_{depot}.xlsx",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="dl5",
            )

# ==================== TAB 5 ====================

elif section == "ACT vs ACT TRENDS":
    prep_text = "upto" if for_upto == "UPTO" else "for"
    loc_text = f"{depot} Depot" if depot != "ALL" else "ALL Depots"
    _date_br = ""
    try:
        if pd.notna(selected_max_date):
            _now = pd.Timestamp.now()
            if selected_max_date.year == _now.year and selected_max_date.month == _now.month:
                _date_br = f" ({selected_max_date.strftime('%d-%m-%Y')})"
    except Exception:
        pass
    _ng = str(net_gross).upper()
    title_tab5 = f"Actual Vs Actuals {str(for_upto).upper()} the month of {month}{_date_br} ({_ng}) — Trends · {loc_text}"
    st.markdown(f'<div class="title-bar">{title_tab5}</div>', unsafe_allow_html=True)

    m_base = df[base_mask].copy() if base_mask is not None else df.copy()
    if len(m_base) == 0:
        st.warning("No data found for selected filters.")
    else:
        date_col = next((col for col in ["Date", "Month_Name", "Month", "MONTH"] if col in m_base.columns), None)
        if date_col is None:
            st.error("Could not find a valid Date or Month column.")
        else:
            m_base = m_base.copy()
            m_base["_parsed_date"] = pd.to_datetime(m_base[date_col], errors="coerce")
            m_base = m_base.dropna(subset=["_parsed_date"]).copy()
            m_base["_month_short"] = m_base["_parsed_date"].dt.strftime("%b")
            m_base["_fy"] = np.where(
                m_base["_parsed_date"].dt.month >= 4,
                m_base["_parsed_date"].dt.year,
                m_base["_parsed_date"].dt.year - 1,
            )

            available_fys = sorted(m_base["_fy"].unique(), reverse=True)
            fy_options = [f"{fy}-{str(fy + 1)[-2:]}" for fy in available_fys]

            # Service filter options
            _svc_col = next((c for c in ["SER_NO", "SERVICE_NO", "SERVICE", "ServiceNo"] if c in m_base.columns), None)
            svc_opts = ["ALL"]
            if _svc_col:
                def _ns(v):
                    try:
                        return str(int(float(v)))
                    except Exception:
                        return str(v).strip()
                svc_opts += sorted(
                    {_ns(x) for x in m_base[_svc_col].dropna().unique() if str(x).strip() and str(x).lower() != "nan"},
                    key=lambda z: (0, int(z)) if str(z).isdigit() else (1, str(z)),
                )

            col_fy, col_svc, _sp = st.columns([1, 1, 3])
            with col_fy:
                selected_fy_str = st.selectbox(
                    "Select Financial Year",
                    options=fy_options if fy_options else ["2025-26"],
                    index=0,
                    key="fy_select_trends",
                )
                selected_fy = int(str(selected_fy_str).split("-")[0])
            with col_svc:
                tr_service = st.selectbox("SERVICE NO", svc_opts, index=0, key="trends_service_no")

            if _svc_col and tr_service != "ALL":
                m_base = m_base[m_base[_svc_col].map(lambda v: str(int(float(v))) if str(v).replace(".", "", 1).isdigit() else str(v).strip()) == str(tr_service)]

            cy_tr = m_base[m_base["_fy"] == selected_fy].copy()
            ly_tr = m_base[m_base["_fy"] == (selected_fy - 1)].copy()
            cy_yr = str(selected_fy)[-2:]
            next_yr = str(selected_fy + 1)[-2:]
            fy_months = ["Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec", "Jan", "Feb", "Mar"]
            month_label_map = {
                "Apr": f"Apr-{cy_yr}", "May": f"May-{cy_yr}", "Jun": f"Jun-{cy_yr}", "Jul": f"Jul-{cy_yr}",
                "Aug": f"Aug-{cy_yr}", "Sep": f"Sep-{cy_yr}", "Oct": f"Oct-{cy_yr}", "Nov": f"Nov-{cy_yr}",
                "Dec": f"Dec-{cy_yr}", "Jan": f"Jan-{next_yr}", "Feb": f"Feb-{next_yr}", "Mar": f"Mar-{next_yr}",
            }

            def _agg_side(data, e_tot, e_fpd, e_mhl):
                if data is None or len(data) == 0:
                    return pd.DataFrame(index=fy_months, columns=["kms", "earn_tot", "earn_fpd", "earn_mhl", "pax"])
                return data.groupby("_month_short").agg(
                    kms=("Optd_KMs", "sum"),
                    earn_tot=(e_tot, "sum"),
                    earn_fpd=(e_fpd, "sum"),
                    earn_mhl=(e_mhl, "sum"),
                    pax=(pax_col, "sum"),
                ).reindex(fy_months)

            def _build_month_df(cy_agg, ly_agg):
                valid = cy_agg[cy_agg["kms"].notna() & (cy_agg["kms"] > 0)].index.tolist()
                active = fy_months[: max(fy_months.index(m) for m in valid) + 1] if valid else fy_months
                rows = []
                for m in active:
                    cy_kms = float(cy_agg.loc[m, "kms"] or 0)
                    ly_kms = float(ly_agg.loc[m, "kms"] or 0) if m in ly_agg.index else 0.0
                    cy_earn = float(cy_agg.loc[m, "earn_tot"] or 0)
                    ly_earn = float(ly_agg.loc[m, "earn_tot"] or 0) if m in ly_agg.index else 0.0
                    cy_fpd = float(cy_agg.loc[m, "earn_fpd"] or 0)
                    ly_fpd = float(ly_agg.loc[m, "earn_fpd"] or 0) if m in ly_agg.index else 0.0
                    cy_mhl = float(cy_agg.loc[m, "earn_mhl"] or 0)
                    ly_mhl = float(ly_agg.loc[m, "earn_mhl"] or 0) if m in ly_agg.index else 0.0
                    cy_pax = float(cy_agg.loc[m, "pax"] or 0)
                    ly_pax = float(ly_agg.loc[m, "pax"] or 0) if m in ly_agg.index else 0.0
                    rec = {
                        "Month": month_label_map.get(m, m),
                        "kms_CY": cy_kms / 100000.0,
                        "kms_LY": ly_kms / 100000.0,
                        "earn_tot_CY": cy_earn / 100000.0,
                        "earn_tot_LY": ly_earn / 100000.0,
                        "epk_tot_CY": (cy_earn / cy_kms) if cy_kms > 0 else np.nan,
                        "epk_tot_LY": (ly_earn / ly_kms) if ly_kms > 0 else np.nan,
                        "epk_fpd_CY": (cy_fpd / cy_kms) if cy_kms > 0 else np.nan,
                        "epk_fpd_LY": (ly_fpd / ly_kms) if ly_kms > 0 else np.nan,
                        "epk_mhl_CY": (cy_mhl / cy_kms) if cy_kms > 0 else np.nan,
                        "epk_mhl_LY": (ly_mhl / ly_kms) if ly_kms > 0 else np.nan,
                        "pax_CY": cy_pax,
                        "pax_LY": ly_pax,
                    }
                    for base in ["kms", "earn_tot", "epk_tot", "epk_fpd", "epk_mhl", "pax"]:
                        rec[f"{base}_VAR"] = (rec[f"{base}_CY"] - rec[f"{base}_LY"]) if pd.notna(rec[f"{base}_CY"]) and pd.notna(rec[f"{base}_LY"]) else np.nan
                        rec[f"{base}_PCT"] = (rec[f"{base}_VAR"] * 100 / rec[f"{base}_LY"]) if pd.notna(rec[f"{base}_LY"]) and rec[f"{base}_LY"] not in (0, 0.0) else np.nan
                    rows.append(rec)

                # ---- UPTO total row (sum of active months) ----
                if rows:
                    tot_c_kms = tot_l_kms = 0.0
                    tot_c_earn = tot_l_earn = 0.0
                    tot_c_fpd = tot_l_fpd = 0.0
                    tot_c_mhl = tot_l_mhl = 0.0
                    tot_c_pax = tot_l_pax = 0.0
                    for m in active:
                        tot_c_kms += float(cy_agg.loc[m, "kms"] or 0) if m in cy_agg.index else 0.0
                        tot_l_kms += float(ly_agg.loc[m, "kms"] or 0) if m in ly_agg.index else 0.0
                        tot_c_earn += float(cy_agg.loc[m, "earn_tot"] or 0) if m in cy_agg.index else 0.0
                        tot_l_earn += float(ly_agg.loc[m, "earn_tot"] or 0) if m in ly_agg.index else 0.0
                        tot_c_fpd += float(cy_agg.loc[m, "earn_fpd"] or 0) if m in cy_agg.index else 0.0
                        tot_l_fpd += float(ly_agg.loc[m, "earn_fpd"] or 0) if m in ly_agg.index else 0.0
                        tot_c_mhl += float(cy_agg.loc[m, "earn_mhl"] or 0) if m in cy_agg.index else 0.0
                        tot_l_mhl += float(ly_agg.loc[m, "earn_mhl"] or 0) if m in ly_agg.index else 0.0
                        tot_c_pax += float(cy_agg.loc[m, "pax"] or 0) if m in cy_agg.index else 0.0
                        tot_l_pax += float(ly_agg.loc[m, "pax"] or 0) if m in ly_agg.index else 0.0
                    last_lab = month_label_map.get(active[-1], active[-1])
                    upto = {
                        "Month": f"UPTO {last_lab}",
                        "kms_CY": tot_c_kms / 100000.0,
                        "kms_LY": tot_l_kms / 100000.0,
                        "earn_tot_CY": tot_c_earn / 100000.0,
                        "earn_tot_LY": tot_l_earn / 100000.0,
                        "epk_tot_CY": (tot_c_earn / tot_c_kms) if tot_c_kms > 0 else np.nan,
                        "epk_tot_LY": (tot_l_earn / tot_l_kms) if tot_l_kms > 0 else np.nan,
                        "epk_fpd_CY": (tot_c_fpd / tot_c_kms) if tot_c_kms > 0 else np.nan,
                        "epk_fpd_LY": (tot_l_fpd / tot_l_kms) if tot_l_kms > 0 else np.nan,
                        "epk_mhl_CY": (tot_c_mhl / tot_c_kms) if tot_c_kms > 0 else np.nan,
                        "epk_mhl_LY": (tot_l_mhl / tot_l_kms) if tot_l_kms > 0 else np.nan,
                        "pax_CY": tot_c_pax,
                        "pax_LY": tot_l_pax,
                    }
                    for base in ["kms", "earn_tot", "epk_tot", "epk_fpd", "epk_mhl", "pax"]:
                        upto[f"{base}_VAR"] = (upto[f"{base}_CY"] - upto[f"{base}_LY"]) if pd.notna(upto[f"{base}_CY"]) and pd.notna(upto[f"{base}_LY"]) else np.nan
                        upto[f"{base}_PCT"] = (upto[f"{base}_VAR"] * 100 / upto[f"{base}_LY"]) if pd.notna(upto[f"{base}_LY"]) and upto[f"{base}_LY"] not in (0, 0.0) else np.nan
                    rows.append(upto)
                return pd.DataFrame(rows), active

            def _add_or_month(df_m):
                """Apply same ORF as ACT (depot/product filters) to every month row."""
                if df_m is None or len(df_m) == 0:
                    return df_m
                orf_map, orf_by_prod, _err = load_orf_map(r"D:\dashboard\ORF.xlsx")
                # Fake a one-row depot frame to reuse add_or_columns_depot logic per EPK
                # Simpler: get single ORF rates for current depot/product selection
                depot_orf = orf_map if isinstance(orf_map, dict) else {}
                by_prod = orf_by_prod if isinstance(orf_by_prod, dict) else {}
                prod_sel = str(product).strip().upper() if product else "ALL"
                dep_sel = str(depot).strip().upper() if depot else "ALL"
                prod_is_all = prod_sel in ("ALL", "NONE", "", "NAN")
                dep_is_all = dep_sel in ("ALL", "NONE", "", "NAN")

                def lookup(dep_key, prod_key, side):
                    d, p = str(dep_key).upper(), str(prod_key).upper()
                    rec = by_prod.get((d, p), {})
                    val = rec.get(side, np.nan)
                    if val is not None and not (isinstance(val, float) and np.isnan(val)):
                        return float(val)
                    if p == "TOTAL":
                        val2 = depot_orf.get(d, {}).get(side, np.nan)
                        if val2 is not None and not (isinstance(val2, float) and np.isnan(val2)):
                            return float(val2)
                    return np.nan

                def orf_side(side):
                    if dep_is_all and prod_is_all:
                        return lookup("REGION", "TOTAL", side)
                    if dep_is_all and not prod_is_all:
                        return lookup("REGION", prod_sel, side)
                    if not dep_is_all and prod_is_all:
                        return lookup(dep_sel, "TOTAL", side)
                    return lookup(dep_sel, prod_sel, side)

                oc, ol = orf_side("cy"), orf_side("ly")
                df_m = df_m.copy()
                for kind in ["tot", "fpd", "mhl"]:
                    df_m[f"or_{kind}_CY"] = np.where(oc and oc != 0, (pd.to_numeric(df_m[f"epk_{kind}_CY"], errors="coerce") * 10000) / oc, np.nan)
                    df_m[f"or_{kind}_LY"] = np.where(ol and ol != 0, (pd.to_numeric(df_m[f"epk_{kind}_LY"], errors="coerce") * 10000) / ol, np.nan)
                    df_m[f"or_{kind}_VAR"] = df_m[f"or_{kind}_CY"] - df_m[f"or_{kind}_LY"]
                    df_m[f"or_{kind}_PCT"] = np.where(
                        pd.to_numeric(df_m[f"or_{kind}_LY"], errors="coerce").fillna(0) != 0,
                        df_m[f"or_{kind}_VAR"] * 100 / df_m[f"or_{kind}_LY"],
                        np.nan,
                    )
                return df_m

            def _render_trends_table(df_m, prefix_label, pax_heading):
                html = [
                    '<div class="table-scroll-fixable">',
                    '<table class="excel-table" style="border-collapse:separate;border-spacing:0;width:max-content;"><thead><tr>',
                ]
                html.append('<th rowspan="2" style="position:sticky;left:0;top:0;z-index:4;background:#0369a1;color:#fff;min-width:48px;">S.No</th>')
                html.append('<th rowspan="2" style="position:sticky;left:48px;top:0;z-index:4;background:#0369a1;color:#fff;min-width:90px;">MONTH</th>')
                groups = [
                    ("KILOMETERS (in lks.)", "#c2410c"),
                    (f"{prefix_label} EARNINGS (in lks.)", "#047857"),
                    (f"{prefix_label} TOT EPK", "#6b21a8"),
                    (f"{prefix_label} FPD EPK", "#15803d"),
                    (f"{prefix_label} MHL EPK", "#1d4ed8"),
                    ("TOT OR", "#7c3aed"),
                    ("FPD OR", "#0d9488"),
                    ("MHL OR", "#2563eb"),
                    (pax_heading, "#0369a1"),
                ]
                for title, bg in groups:
                    html.append(f'<th colspan="4" style="position:sticky;top:0;z-index:3;background:{bg};color:#fff;">{title}</th>')
                html.append("</tr><tr>")
                for _ in range(9):
                    for s in ("CY", "LY", "VAR", "% ▲/▼"):
                        html.append(f'<th style="position:sticky;top:32px;z-index:3;background:#f1f5f9;font-size:10px;">{s}</th>')
                html.append("</tr></thead><tbody>")
                for i, row in df_m.iterrows():
                    is_tot = "UPTO" in str(row.get("Month", "")).upper() or str(row.get("Month", "")).upper() == "TOTAL"
                    bg = "#e2efda" if is_tot else "#fff"
                    lbg = "#e2efda" if is_tot else "#f8fafc"
                    html.append(f'<tr style="background:{bg};{"font-weight:700;" if is_tot else ""}">')
                    html.append(f'<td style="position:sticky;left:0;z-index:2;background:{lbg};">{"" if is_tot else (i + 1)}</td>')
                    html.append(f'<td style="position:sticky;left:48px;z-index:2;background:{lbg};">{row["Month"]}</td>')
                    for metric in ["kms", "earn_tot"]:
                        html.append(f'<td>{fmt(row[f"{metric}_CY"])}</td>')
                        html.append(f'<td>{fmt(row[f"{metric}_LY"])}</td>')
                        html.append(f'<td class="{var_class(row[f"{metric}_VAR"])}">{fmt(row[f"{metric}_VAR"])}</td>')
                        html.append(f'<td class="{var_class(row[f"{metric}_PCT"])}">{fmt_growth(row[f"{metric}_PCT"])}</td>')
                    for epk in ["tot", "fpd", "mhl"]:
                        html.append(f'<td>{fmt(row[f"epk_{epk}_CY"])}</td>')
                        html.append(f'<td>{fmt(row[f"epk_{epk}_LY"])}</td>')
                        html.append(f'<td class="{var_class(row[f"epk_{epk}_VAR"])}">{fmt(row[f"epk_{epk}_VAR"])}</td>')
                        html.append(f'<td class="{var_class(row[f"epk_{epk}_PCT"])}">{fmt_growth(row[f"epk_{epk}_PCT"])}</td>')
                    for ot in ["tot", "fpd", "mhl"]:
                        html.append(f'<td>{fmt(row.get(f"or_{ot}_CY"))}</td>')
                        html.append(f'<td>{fmt(row.get(f"or_{ot}_LY"))}</td>')
                        html.append(f'<td class="{var_class(row.get(f"or_{ot}_VAR"))}">{fmt(row.get(f"or_{ot}_VAR"))}</td>')
                        html.append(f'<td class="{var_class(row.get(f"or_{ot}_PCT"))}">{fmt_growth(row.get(f"or_{ot}_PCT"))}</td>')
                    html.append(f'<td>{fmt_pax(row.get("pax_CY", 0))}</td>')
                    html.append(f'<td>{fmt_pax(row.get("pax_LY", 0))}</td>')
                    html.append(f'<td class="{var_class(row.get("pax_VAR", 0))}">{fmt_pax(row.get("pax_VAR", 0))}</td>')
                    html.append(f'<td class="{var_class(row.get("pax_PCT", 0))}">{fmt_growth(row.get("pax_PCT", 0))}</td>')
                    html.append("</tr>")
                html.append("</tbody></table></div>")
                return "".join(html)

            pax_heading = {"FPD": "FPD PASSENGERS", "MHL": "MHL PASSENGERS"}.get(passengers, "TOTAL PASSENGERS")

            # Charts react to NET/GROSS filter; exclude UPTO from charts
            if str(net_gross).strip().lower().startswith("net"):
                _et, _ef, _em = "NE_TOT", "NE_FPD", "NE_MHL"
                _chart_prefix = "NET"
            else:
                _et, _ef, _em = "GE_TOT", "GE_FPD", "GE_MHL"
                _chart_prefix = "GROSS"
            cy_ch = _agg_side(cy_tr, _et, _ef, _em)
            ly_ch = _agg_side(ly_tr, _et, _ef, _em)
            df_chart, _active = _build_month_df(cy_ch, ly_ch)
            # drop UPTO row from charts
            if len(df_chart) > 0:
                df_chart_plot = df_chart[~df_chart["Month"].astype(str).str.upper().str.startswith("UPTO")].copy()
            else:
                df_chart_plot = df_chart

            if len(df_chart_plot) > 0:
                import plotly.graph_objects as go
                def _combo(dfc, cy_col, ly_col, title, color_cy, color_ly, is_pax=False):
                    fig = go.Figure()
                    txt = "%{text:,.0f}" if is_pax else "%{text:.2f}"
                    fig.add_trace(go.Bar(name="CY", x=dfc["Month"], y=dfc[cy_col], marker_color=color_cy,
                                         text=dfc[cy_col], textposition="outside", texttemplate=txt, textfont_size=9))
                    fig.add_trace(go.Bar(name="LY", x=dfc["Month"], y=dfc[ly_col], marker_color=color_ly,
                                         text=dfc[ly_col], textposition="outside", texttemplate=txt, textfont_size=9))
                    fig.update_layout(
                        barmode="group",
                        height=280,
                        margin=dict(l=4, r=4, t=40, b=30),
                        title=dict(text=f"<b>{title}</b>", x=0.5, font=dict(size=11)),
                        legend=dict(orientation="h", y=1.15, font=dict(size=9)),
                        xaxis=dict(tickangle=-45, tickfont=dict(size=9)),
                        yaxis=dict(showticklabels=False, title=""),
                        template="plotly_white",
                    )
                    return fig
                # Single row, 5 charts side by side (narrower)
                ch1, ch2, ch3, ch4, ch5 = st.columns(5)
                with ch1:
                    st.plotly_chart(_combo(df_chart_plot, "kms_CY", "kms_LY", f"KMs ({_chart_prefix})", "#0284c7", "#93c5fd"), width="stretch")
                with ch2:
                    st.plotly_chart(_combo(df_chart_plot, "epk_tot_CY", "epk_tot_LY", f"{_chart_prefix} TOT EPK", "#6b21a8", "#c4b5fd"), width="stretch")
                with ch3:
                    st.plotly_chart(_combo(df_chart_plot, "epk_fpd_CY", "epk_fpd_LY", f"{_chart_prefix} FPD EPK", "#15803d", "#86efac"), width="stretch")
                with ch4:
                    st.plotly_chart(_combo(df_chart_plot, "epk_mhl_CY", "epk_mhl_LY", f"{_chart_prefix} MHL EPK", "#1d4ed8", "#93c5fd"), width="stretch")
                with ch5:
                    st.plotly_chart(_combo(df_chart_plot, "pax_CY", "pax_LY", pax_heading, "#db2777", "#f9a8d4", is_pax=True), width="stretch")

            # TABLE A NET
            st.markdown(
                f'<div class="title-bar">Actual Vs Actuals {str(for_upto).upper()} the month of {month}{_date_br} (NET) — FY {selected_fy_str}</div>',
                unsafe_allow_html=True,
            )
            cy_n = _agg_side(cy_tr, "NE_TOT", "NE_FPD", "NE_MHL")
            ly_n = _agg_side(ly_tr, "NE_TOT", "NE_FPD", "NE_MHL")
            df_net, _ = _build_month_df(cy_n, ly_n)
            df_net = _add_or_month(df_net)
            if len(df_net):
                st.markdown(_render_trends_table(df_net, "NET", pax_heading), unsafe_allow_html=True)

            # TABLE B GROSS
            st.markdown(
                f'<div class="title-bar">Actual Vs Actuals {str(for_upto).upper()} the month of {month}{_date_br} (GROSS) — FY {selected_fy_str}</div>',
                unsafe_allow_html=True,
            )
            cy_g = _agg_side(cy_tr, "GE_TOT", "GE_FPD", "GE_MHL")
            ly_g = _agg_side(ly_tr, "GE_TOT", "GE_FPD", "GE_MHL")
            df_gr, _ = _build_month_df(cy_g, ly_g)
            df_gr = _add_or_month(df_gr)
            if len(df_gr):
                st.markdown(_render_trends_table(df_gr, "GROSS", pax_heading), unsafe_allow_html=True)

            # One Excel file with both NET and GROSS sheets, on-screen style headings
            if (df_net is not None and len(df_net)) or (df_gr is not None and len(df_gr)):
                st.download_button(
                    "Download Excel – NET + GROSS",
                    trends_dual_excel_bytes(
                        df_net if df_net is not None else pd.DataFrame(),
                        df_gr if df_gr is not None else pd.DataFrame(),
                        pax_heading=pax_heading,
                        report_title=f"ACT vs ACT TRENDS | FY={selected_fy_str} | Depot={depot} | Product={product} | {for_upto}",
                    ),
                    f"ACT_vs_ACT_TRENDS_{selected_fy_str}.xlsx",
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="dl_tr_both",
                )

# ==================== TAB 6 ====================
elif section == "Trends from 2024":
    prep_text = "upto" if for_upto == "UPTO" else "for"
    loc_text = f"{depot} Depot" if depot != "ALL" else "ALL Depots"
    title_tab5 = f"Month Wise Actual vs Actuals of {loc_text} {prep_text} the month of {month}"
    st.markdown(f'<div class="title-bar">{title_tab5}</div>', unsafe_allow_html=True)

    # Filter data matching base filters (Depot, Route, Product, MHL/NMHL, RTC/HIRE)
    m_base = df[base_mask].copy()

    if len(m_base) == 0:
        st.warning("No data found for selected filters.")
    else:
        # Dynamically determine months present in the filtered dataset in chronological order
        unique_months = m_base["Month_Name"].dropna().unique()
        def parse_m_sort(m_str):
            try:
                return pd.to_datetime(m_str, format="%b-%y")
            except Exception:
                try:
                    return pd.to_datetime(m_str, format="%b-%Y")
                except Exception:
                    return pd.to_datetime(m_str, errors="coerce")

        months_order = sorted([str(m).strip() for m in unique_months if str(m).strip()], key=parse_m_sort)

        if not months_order:
            st.warning("No valid month data available.")
        else:
            # 1. Aggregate Month-wise values for CY
            m_cy_agg = m_base.groupby("Month_Name").agg(
                kms=("Optd_KMs", "sum"),
                earn_tot=(earn_tot, "sum"),
                earn_fpd=(earn_fpd, "sum"),
                earn_mhl=(earn_mhl, "sum"),
                pax=(pax_col, "sum")
            ).reindex(months_order).fillna(0)

            # Calculated Monthly Metrics
            m_chart = pd.DataFrame(index=months_order)
            m_chart["Month_Group"] = m_chart.index
            m_chart["Kms_Lakhs"] = (m_cy_agg["kms"] / 100000).round(2)
            m_chart["Earn_Lakhs"] = (m_cy_agg["earn_tot"] / 100000).round(2)
            m_chart["Pax_Lakhs"] = (m_cy_agg["pax"] / 100000).round(0).astype(int)
            m_chart["Tot_EPK"] = np.where(m_cy_agg["kms"] > 0, (m_cy_agg["earn_tot"] / m_cy_agg["kms"]).round(2), 0.0)
            m_chart["FPD_EPK"] = np.where(m_cy_agg["kms"] > 0, (m_cy_agg["earn_fpd"] / m_cy_agg["kms"]).round(2), 0.0)
            m_chart["MHL_EPK"] = np.where(m_cy_agg["kms"] > 0, (m_cy_agg["earn_mhl"] / m_cy_agg["kms"]).round(2), 0.0)

            # 2. Render 5 Top Bar Charts
            def create_card_chart(df_chart, y_col, chart_title, bar_color="#1d4ed8", is_pax=False):
                fig = px.bar(
                    df_chart,
                    x="Month_Group",
                    y=y_col,
                    title=f"<b>{chart_title}</b>",
                    text=y_col,
                    color_discrete_sequence=[bar_color]
                )
                text_fmt = '%{text:.0f}' if is_pax else '%{text:.2f}'
                fig.update_traces(
                    texttemplate=text_fmt,
                    textposition='outside',
                    textangle=-90
                )
                fig.update_layout(
                    height=250,
                    margin=dict(l=5, r=5, t=30, b=5),
                    xaxis_title="",
                    yaxis_title="",
                    xaxis=dict(tickangle=-90, showgrid=False),
                    yaxis=dict(showgrid=True, showticklabels=False),
                    template="plotly_white",
                    title_font_size=11,
                    title_x=0.5
                )
                return fig

            ch1, ch2, ch3, ch4, ch5 = st.columns(5)
            with ch1:
                st.plotly_chart(create_card_chart(m_chart, "Kms_Lakhs", "KILOMETERS (in lks.)", "#0284c7"), width="stretch")
            with ch2:
                st.plotly_chart(create_card_chart(m_chart, "Tot_EPK", f"{prefix} TOT. E.P.K (in Ps/kms.)", "#2563eb"), width="stretch")
            with ch3:
                st.plotly_chart(create_card_chart(m_chart, "FPD_EPK", f"{prefix} FPD. E.P.K (in Ps/kms.)", "#16a34a"), width="stretch")
            with ch4:
                st.plotly_chart(create_card_chart(m_chart, "MHL_EPK", f"{prefix} MHL. E.P.K (in Ps/kms.)", "#9333ea"), width="stretch")
            with ch5:
                pax_title = {"FPD": "FPD PASSENGERS (in lks.)", "MHL": "MHL PASSENGERS (in lks.)"}.get(passengers, "TOTAL PASSENGERS (in lks.)")
                st.plotly_chart(create_card_chart(m_chart, "Pax_Lakhs", pax_title, "#db2777", is_pax=True), width="stretch")

            # 3. Cumulative CY & LY Totals using Global Filtering Rules
            cy_kms_tot = cy_data["Optd_KMs"].sum() / 100000 if len(cy_data) else 0
            ly_kms_tot = ly_data["Optd_KMs"].sum() / 100000 if len(ly_data) else 0

            cy_earn_tot = cy_data[earn_tot].sum() / 100000 if len(cy_data) else 0
            ly_earn_tot = ly_data[earn_tot].sum() / 100000 if len(ly_data) else 0

            cy_epk_tot = (cy_data[earn_tot].sum() / cy_data["Optd_KMs"].sum()) if len(cy_data) and cy_data["Optd_KMs"].sum() > 0 else np.nan
            ly_epk_tot = (ly_data[earn_tot].sum() / ly_data["Optd_KMs"].sum()) if len(ly_data) and ly_data["Optd_KMs"].sum() > 0 else np.nan

            cy_epk_fpd = (cy_data[earn_fpd].sum() / cy_data["Optd_KMs"].sum()) if len(cy_data) and cy_data["Optd_KMs"].sum() > 0 else np.nan
            ly_epk_fpd = (ly_data[earn_fpd].sum() / ly_data["Optd_KMs"].sum()) if len(ly_data) and ly_data["Optd_KMs"].sum() > 0 else np.nan

            cy_epk_mhl = (cy_data[earn_mhl].sum() / cy_data["Optd_KMs"].sum()) if len(cy_data) and cy_data["Optd_KMs"].sum() > 0 else np.nan
            ly_epk_mhl = (ly_data[earn_mhl].sum() / ly_data["Optd_KMs"].sum()) if len(ly_data) and ly_data["Optd_KMs"].sum() > 0 else np.nan

            cy_pax_tot = round(cy_data[pax_col].sum() / 100000) if len(cy_data) else 0
            ly_pax_tot = round(ly_data[pax_col].sum() / 100000) if len(ly_data) else 0

            
            # Construct Month-Wise Dashboard Rows
            # Schedules / Services / Sch KMs from SROS (react to depot / product filter)
            sch_by_m = {m: 0.0 for m in months_order}
            svc_by_m = {m: 0.0 for m in months_order}
            schkms_by_m = {m: 0.0 for m in months_order}
            _sros_dbg = ""
            try:
                sros_path = Path(r"D:\Dashboard\SMASTER.parquet")
                if not sros_path.exists():
                    for alt in [Path(r"D:\dashboard\SMASTER.parquet"), Path(r"D:\MONTHLY\SMASTER.parquet"), Path("SMASTER.parquet"), Path(r"/home/workdir/attachments/SMASTER.parquet")]:
                        if alt.exists():
                            sros_path = alt
                            break
                if not sros_path.exists():
                    st.warning(r"SMASTER.parquet not found (checked D:\Dashboard and D:\MONTHLY)")
                else:
                    sros, _sros_err = load_smaster(str(sros_path))
                    if sros is None:
                        raise RuntimeError(_sros_err or "Unable to load SMASTER.parquet")
                    sros = sros.copy()
                    sros.columns = [str(c).strip() for c in sros.columns]
                    def _fc(cands):
                        n = {str(c).strip().lower().replace(" ", "").replace("_", ""): c for c in sros.columns}
                        for cand in cands:
                            k = cand.lower().replace(" ", "").replace("_", "")
                            if k in n:
                                return n[k]
                        for c in sros.columns:
                            cl = str(c).strip().lower().replace(" ", "").replace("_", "")
                            for cand in cands:
                                if cand.lower().replace(" ", "").replace("_", "") in cl:
                                    return c
                        return None
                    c_sch = _fc(["NoOfSchedules", "NoOfSchedule", "NOSCH", "SCHEDULES"])
                    c_svc = _fc(["ServiceNo", "SER_NO", "SERVICE_NO", "SERVICENO"])
                    c_kms = _fc(["RevenueKms", "SCHKMS", "SchKms", "Revenue_Kms", "SCH KMS", "ScheduledKms"])
                    c_mon = _fc(["MONTH", "Month", "MON"])
                    c_year = _fc(["YEAR", "Year", "YR"])
                    c_dep = _fc(["DEPOT", "Depot"])
                    c_prod = _fc(["PRODUCT", "Product"])
                    ss = sros.copy()
                    if c_dep and str(depot).upper() not in ("ALL", "REGION", ""):
                        ss = ss[ss[c_dep].astype(str).str.strip().str.upper() == str(depot).strip().upper()]
                    if c_prod and str(product).upper() not in ("ALL", ""):
                        ss = ss[ss[c_prod].astype(str).str.strip().str.upper() == str(product).strip().upper()]

                    mon_map_num = {
                        "1": "Jan", "2": "Feb", "3": "Mar", "4": "Apr", "5": "May", "6": "Jun",
                        "7": "Jul", "8": "Aug", "9": "Sep", "10": "Oct", "11": "Nov", "12": "Dec",
                        "01": "Jan", "02": "Feb", "03": "Mar", "04": "Apr", "05": "May", "06": "Jun",
                        "07": "Jul", "08": "Aug", "09": "Sep",
                        "JAN": "Jan", "FEB": "Feb", "MAR": "Mar", "APR": "Apr", "MAY": "May", "JUN": "Jun",
                        "JUL": "Jul", "AUG": "Aug", "SEP": "Sep", "OCT": "Oct", "NOV": "Nov", "DEC": "Dec",
                        "JANUARY": "Jan", "FEBRUARY": "Feb", "MARCH": "Mar", "APRIL": "Apr",
                        "JUNE": "Jun", "JULY": "Jul", "AUGUST": "Aug", "SEPTEMBER": "Sep",
                        "OCTOBER": "Oct", "NOVEMBER": "Nov", "DECEMBER": "Dec",
                    }

                    def _norm_month_key(mon_val, year_val=None):
                        """Return set of possible keys: Jul-26, Jul-2026, Jul-26 etc."""
                        keys = set()
                        try:
                            s = str(mon_val).strip()
                            # already like Jul-26 / Jul-2026
                            if "-" in s and any(c.isalpha() for c in s):
                                parts = s.replace(" ", "").split("-")
                                mon = parts[0][:3].title()
                                yr = parts[1] if len(parts) > 1 else ""
                                if yr:
                                    keys.add(f"{mon}-{yr}")
                                    if len(yr) == 4:
                                        keys.add(f"{mon}-{yr[-2:]}")
                                    elif len(yr) == 2:
                                        keys.add(f"{mon}-20{yr}")
                                return keys
                            # numeric month or name
                            mon = mon_map_num.get(s.upper(), mon_map_num.get(s.lstrip("0"), s[:3].title()))
                            yr = None
                            if year_val is not None and str(year_val).strip() not in ("", "nan", "None"):
                                yr = int(float(year_val))
                            if yr:
                                keys.add(f"{mon}-{yr}")
                                keys.add(f"{mon}-{str(yr)[-2:]}")
                            else:
                                keys.add(mon)
                        except Exception:
                            pass
                        return keys

                    def _norm_target(mstr):
                        keys = set()
                        s = str(mstr).strip().replace(" ", "")
                        keys.add(s)
                        try:
                            dt = pd.to_datetime(s, format="%b-%y", errors="coerce")
                            if pd.isna(dt):
                                dt = pd.to_datetime(s, format="%b-%Y", errors="coerce")
                            if pd.isna(dt):
                                dt = pd.to_datetime(s, errors="coerce")
                            if not pd.isna(dt):
                                keys.add(dt.strftime("%b-%y"))
                                keys.add(dt.strftime("%b-%Y"))
                                keys.add(f"{dt.strftime('%b')}-{dt.year}")
                                keys.add(f"{dt.strftime('%b')}-{str(dt.year)[-2:]}")
                        except Exception:
                            pass
                        if "-" in s:
                            a, b = s.split("-", 1)
                            mon = a[:3].title()
                            keys.add(f"{mon}-{b}")
                            if len(b) == 4:
                                keys.add(f"{mon}-{b[-2:]}")
                            elif len(b) == 2:
                                keys.add(f"{mon}-20{b}")
                        return keys

                    # Build key -> list of row indices
                    key_to_idx = {}
                    for idx, row in ss.iterrows():
                        mon_v = row[c_mon] if c_mon else None
                        yr_v = row[c_year] if c_year else None
                        for k in _norm_month_key(mon_v, yr_v):
                            key_to_idx.setdefault(k, []).append(idx)

                    matched_any = False
                    for m in months_order:
                        targets = _norm_target(m)
                        idxs = set()
                        for t in targets:
                            idxs.update(key_to_idx.get(t, []))
                        if not idxs:
                            # try partial: month name only + year from m
                            continue
                        matched_any = True
                        sub = ss.loc[sorted(idxs)]
                        if c_sch:
                            sch_by_m[m] = float(pd.to_numeric(sub[c_sch], errors="coerce").fillna(0).sum())
                        if c_svc:
                            svc_by_m[m] = float(sub[c_svc].nunique())
                        if c_kms:
                            schkms_by_m[m] = float(pd.to_numeric(sub[c_kms], errors="coerce").fillna(0).sum()) / 100000.0

                    sample_keys = list(key_to_idx.keys())[:8]
                    _sros_dbg = (
                        f"SROS ok | cols sch={c_sch}, svc={c_svc}, kms={c_kms}, mon={c_mon}, yr={c_year} | "
                        f"rows={len(ss)} | sample keys={sample_keys} | months={months_order[:3]} | matched={matched_any}"
                    )
                    st.caption(_sros_dbg)
            except Exception as _se:
                st.warning(f"SROS load error: {_se}")

            # ORF for OR rows (same rules as ACT)
            try:
                orf_map, orf_by_prod, _oe = load_orf_map(r"D:\dashboard\ORF.xlsx")
            except Exception:
                orf_map, orf_by_prod = {}, {}
            depot_orf = orf_map if isinstance(orf_map, dict) else {}
            by_prod = orf_by_prod if isinstance(orf_by_prod, dict) else {}
            prod_sel = str(product).strip().upper() if product else "ALL"
            dep_sel = str(depot).strip().upper() if depot else "ALL"
            prod_is_all = prod_sel in ("ALL", "NONE", "", "NAN")
            dep_is_all = dep_sel in ("ALL", "NONE", "", "NAN")

            def _orf_lookup(dep_key, prod_key, side="cy"):
                d, p = str(dep_key).upper(), str(prod_key).upper()
                rec = by_prod.get((d, p), {})
                val = rec.get(side, np.nan)
                if val is not None and not (isinstance(val, float) and np.isnan(val)):
                    return float(val)
                if p == "TOTAL":
                    val2 = depot_orf.get(d, {}).get(side, np.nan)
                    if val2 is not None and not (isinstance(val2, float) and np.isnan(val2)):
                        return float(val2)
                return np.nan

            def _orf_rate(side="cy"):
                if dep_is_all and prod_is_all:
                    return _orf_lookup("REGION", "TOTAL", side)
                if dep_is_all and not prod_is_all:
                    return _orf_lookup("REGION", prod_sel, side)
                if not dep_is_all and prod_is_all:
                    return _orf_lookup(dep_sel, "TOTAL", side)
                return _orf_lookup(dep_sel, prod_sel, side)

            orf_cy = _orf_rate("cy")
            orf_ly = _orf_rate("ly")

            def _or_series(epk_series):
                out = {}
                for m in months_order:
                    v = float(epk_series.get(m, 0) or 0)
                    out[m] = (v * 10000 / orf_cy) if orf_cy and orf_cy != 0 else np.nan
                return out

            or_tot_m = _or_series(m_chart["Tot_EPK"])
            or_fpd_m = _or_series(m_chart["FPD_EPK"])
            or_mhl_m = _or_series(m_chart["MHL_EPK"])
            cy_or_tot = (cy_epk_tot * 10000 / orf_cy) if orf_cy and orf_cy != 0 else np.nan
            ly_or_tot = (ly_epk_tot * 10000 / orf_ly) if orf_ly and orf_ly != 0 else np.nan
            cy_or_fpd = (cy_epk_fpd * 10000 / orf_cy) if orf_cy and orf_cy != 0 else np.nan
            ly_or_fpd = (ly_epk_fpd * 10000 / orf_ly) if orf_ly and orf_ly != 0 else np.nan
            cy_or_mhl = (cy_epk_mhl * 10000 / orf_cy) if orf_cy and orf_cy != 0 else np.nan
            ly_or_mhl = (ly_epk_mhl * 10000 / orf_ly) if orf_ly and orf_ly != 0 else np.nan

            sch_cy = sum(sch_by_m.values())
            svc_cy = sum(svc_by_m.values())
            schkms_cy = sum(schkms_by_m.values())

            fleet_by_m = {m: 0.0 for m in months_order}
            _fm, _fe = load_fleet_map()
            if not _fe and _fm:
                for m in months_order:
                    keys_try = [m]
                    try:
                        dt = pd.to_datetime(m, errors="coerce")
                        if pd.notna(dt):
                            keys_try += [dt.strftime("%b-%Y"), dt.strftime("%b-%y")]
                    except Exception:
                        pass
                    val = 0.0
                    for k in keys_try:
                        if str(depot).upper() not in ("ALL", "REGION") and str(product).upper() not in ("ALL", ""):
                            val = _fm.get("by_dpm", {}).get((str(depot).upper(), str(product).upper(), k), 0) or val
                        if not val and str(depot).upper() not in ("ALL", "REGION"):
                            val = _fm.get("by_dm", {}).get((str(depot).upper(), k), 0) or val
                        if not val:
                            val = _fm.get("by_m", {}).get(k, 0) or val
                    fleet_by_m[m] = float(val or 0)
            fleet_cy = sum(fleet_by_m.values())
            rows_data = [
                ("FLEET", pd.Series(fleet_by_m), fleet_cy, np.nan, "sch"),
                ("NO OF SCHEDULES", pd.Series(sch_by_m), sch_cy, np.nan, "sch"),
                ("NO OF SERVICES", pd.Series(svc_by_m), svc_cy, np.nan, "sch"),
                ("SCH KMS (in lks.)", pd.Series(schkms_by_m), schkms_cy, np.nan, False),
                ("KILOMETERS (in lks.)", m_chart["Kms_Lakhs"], cy_kms_tot, ly_kms_tot, False),
                (f"{prefix} EARNINGS (in lks.)", m_chart["Earn_Lakhs"], cy_earn_tot, ly_earn_tot, False),
                (f"{prefix} TOT. E.P.K (in Ps/kms.)", m_chart["Tot_EPK"], cy_epk_tot, ly_epk_tot, False),
                (f"{prefix} FPD. E.P.K (in Ps/kms.)", m_chart["FPD_EPK"], cy_epk_fpd, ly_epk_fpd, False),
                (f"{prefix} MHL. E.P.K (in Ps/kms.)", m_chart["MHL_EPK"], cy_epk_mhl, ly_epk_mhl, False),
                ("TOT OR", pd.Series(or_tot_m), cy_or_tot, ly_or_tot, False),
                ("FPD OR", pd.Series(or_fpd_m), cy_or_fpd, ly_or_fpd, False),
                ("MHL OR", pd.Series(or_mhl_m), cy_or_mhl, ly_or_mhl, False),
                (pax_title, m_chart["Pax_Lakhs"], cy_pax_tot, ly_pax_tot, True),
            ]

            # Line chart – Total EPK (reacts to filters via m_chart)
            try:
                import plotly.graph_objects as go
                fig_line = go.Figure()
                fig_line.add_trace(go.Scatter(
                    x=list(months_order),
                    y=[float(m_chart["Tot_EPK"].get(m, 0) or 0) for m in months_order],
                    mode="lines+markers+text",
                    name=f"{prefix} TOT EPK",
                    line=dict(color="#6b21a8", width=3),
                    marker=dict(size=8),
                    text=[f"{float(m_chart['Tot_EPK'].get(m, 0) or 0):.2f}" for m in months_order],
                    textposition="top center",
                ))
                fig_line.update_layout(
                    title=dict(text=f"<b>{prefix} TOTAL EPK – Trends from 2024</b>", x=0.5),
                    height=320,
                    margin=dict(l=40, r=20, t=50, b=40),
                    xaxis_title="Month",
                    yaxis_title="TOT EPK",
                    template="plotly_white",
                )
                st.plotly_chart(fig_line, width="stretch")
            except Exception as _le:
                st.caption(f"Line chart: {_le}")

            # Coloured frozen table
            html = [
                '<div style="max-height:70vh;overflow:auto;border:1px solid #cbd5e1;">',
                '<table class="excel-table" style="border-collapse:separate;border-spacing:0;width:max-content;">',
            ]
            html.append("<tr>")
            html.append('<th style="position:sticky;left:0;top:0;z-index:6;background:#0369a1;color:#fff;min-width:48px;">S.No</th>')
            html.append('<th style="position:sticky;left:48px;top:0;z-index:6;background:#0369a1;color:#fff;min-width:180px;text-align:left;">PERFORMANCE PARAMETER</th>')
            for m_name in months_order:
                html.append(f'<th style="position:sticky;top:0;z-index:5;background:#0f172a;color:#fff;">{m_name}</th>')
            html.append(f'<th style="position:sticky;top:0;z-index:5;background:#6b21a8;color:#fff;">{for_upto} CY</th>')
            html.append(f'<th style="position:sticky;top:0;z-index:5;background:#6b21a8;color:#fff;">{for_upto} LY</th>')
            html.append('<th style="position:sticky;top:0;z-index:5;background:#0d9488;color:#fff;">VAR</th>')
            html.append('<th style="position:sticky;top:0;z-index:5;background:#0d9488;color:#fff;">% ⬆/⬇</th>')
            html.append("</tr>")

            export_rows = []
            sno = 0
            for param_label, month_vals, cy_val, ly_val, is_pax in rows_data:
                sno += 1
                var_val = cy_val - ly_val if pd.notna(cy_val) and pd.notna(ly_val) else np.nan
                pct_val = (var_val * 100 / ly_val) if (pd.notna(ly_val) and ly_val != 0) else np.nan
                row_dict = {"S.No": sno, "Parameter": param_label}
                html.append("<tr>")
                html.append(f'<td style="position:sticky;left:0;z-index:3;background:#f8fafc;">{sno}</td>')
                html.append(f'<td style="position:sticky;left:48px;z-index:3;background:#f8fafc;font-weight:700;text-align:left;">{param_label}</td>')
                for m_name in months_order:
                    v = month_vals.get(m_name, 0) if hasattr(month_vals, "get") else (month_vals[m_name] if m_name in month_vals.index else 0)
                    if is_pax is True:
                        cell_fmt = fmt_pax(v)
                    elif is_pax == "sch":
                        try:
                            cell_fmt = f"{int(round(float(v))):,}" if float(v) else ""
                        except Exception:
                            cell_fmt = ""
                    else:
                        cell_fmt = fmt(v)
                    html.append(f'<td>{cell_fmt}</td>')
                    row_dict[m_name] = v
                if is_pax is True:
                    c_str, l_str, v_str = fmt_pax(cy_val), fmt_pax(ly_val), fmt_pax(var_val)
                elif is_pax == "sch":
                    def _sch(x):
                        try:
                            return f"{int(round(float(x))):,}" if pd.notna(x) and float(x) else ""
                        except Exception:
                            return ""
                    c_str, l_str, v_str = _sch(cy_val), _sch(ly_val), _sch(var_val)
                else:
                    c_str, l_str, v_str = fmt(cy_val), fmt(ly_val), fmt(var_val)
                p_str = fmt_growth(pct_val)
                html.append(f'<td style="font-weight:700; background:#f1f5f9;">{c_str}</td>')
                html.append(f'<td style="font-weight:700; background:#f1f5f9;">{l_str}</td>')
                html.append(f'<td class="{var_class(var_val)}">{v_str}</td>')
                html.append(f'<td class="{var_class(pct_val)}">{p_str}</td>')
                html.append("</tr>")
                row_dict["CY"] = cy_val
                row_dict["LY"] = ly_val
                row_dict["VAR"] = var_val
                row_dict["PCT"] = pct_val
                export_rows.append(row_dict)

            html.append("</table></div>")
            st.markdown("".join(html), unsafe_allow_html=True)

            export_df = pd.DataFrame(export_rows)
            st.download_button(
                "Download Excel",
                excel_with_title(
                    export_df,
                    "Trends_2024",
                    report_title=f"Trends from 2024 | Depot={depot} | Product={product} | {for_upto} | {net_gross}",
                ),
                f"Trends_from_2024_{month}_{depot}.xlsx",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key=f"dl_trends2024_{month}_{depot}",
            )

# ==================== TAB 7 ====================
elif section == "Service performance":
    title7 = f"Service WISE, Day Wise {prefix} EPK ({for_upto}) - {month}"
    if depot != "ALL":
        title7 += f" of {depot} Depot"
    st.markdown(f'<div class="title-bar">{title7}</div>', unsafe_allow_html=True)

    if not service_col:
        st.warning("Service column not found in dataset.")
    else:
        # Simple service filter from performance data (SER_NO) only
        svc_vals = set()
        if len(cy_data) and service_col in cy_data.columns:
            svc_vals.update(cy_data[service_col].dropna().unique().tolist())
        if len(ly_data) and service_col in ly_data.columns:
            svc_vals.update(ly_data[service_col].dropna().unique().tolist())

        def _svc_key(v):
            s = str(v).strip()
            if s.endswith(".0"):
                s = s[:-2]
            try:
                return (0, int(float(s)))
            except Exception:
                return (1, s)

        svc_opts = ["ALL"] + sorted(
            [x for x in svc_vals if str(x).strip() and str(x).lower() != "nan"],
            key=_svc_key,
        )
        st.markdown("**Service filter** (uses depot / product / route filters above)")
        _c1, _c2 = st.columns([1, 3])
        with _c1:
            service_no = st.selectbox("SERVICE NO", svc_opts, index=0, key="tab7_service")

        if service_no == "ALL":
            cy_svc = cy_data.copy()
            ly_svc = ly_data.copy()
        else:
            # match as string and numeric
            def _match(data, svc):
                if len(data) == 0:
                    return data.copy()
                s = data[service_col].astype(str).str.strip().str.replace(r"\.0$", "", regex=True)
                target = str(svc).strip()
                if target.endswith(".0"):
                    target = target[:-2]
                m = data[s == target]
                if len(m) == 0:
                    try:
                        tnum = float(target)
                        m = data[pd.to_numeric(data[service_col], errors="coerce") == tnum]
                    except Exception:
                        pass
                return m.copy()
            cy_svc = _match(cy_data, service_no)
            ly_svc = _match(ly_data, service_no)

        st.caption(f"CY rows: {len(cy_svc):,} | LY rows: {len(ly_svc):,} | service={service_no} | col={service_col}")
        if len(cy_svc) == 0 and len(ly_svc) == 0:
            st.warning("No data for selected filters.")
        else:
            all_parts = []
            for label, earn_col in [("TOT", earn_tot), ("FPD", earn_fpd), ("MHL", earn_mhl)]:
                cy = weighted_epk(cy_svc, earn_col, group_cols=["DEPOT", service_col, "ROUTEE", "PRODUCT"])
                ly = weighted_epk(ly_svc, earn_col, group_cols=["DEPOT", service_col, "ROUTEE", "PRODUCT"])
                all_idx = cy.index.union(ly.index) if len(ly) > 0 else cy.index
                cy = cy.reindex(all_idx)
                ly = ly.reindex(all_idx) if len(ly) > 0 else pd.DataFrame(np.nan, index=all_idx, columns=day_order + ["UPTO"], dtype=float)
                block = pd.DataFrame(index=all_idx)
                for d in day_order:
                    block[f"{label}_{d}"] = pd.to_numeric(cy[d] if d in cy.columns else np.nan, errors="coerce")
                block[f"{label}_CY"] = pd.to_numeric(cy["UPTO"] if "UPTO" in cy.columns else np.nan, errors="coerce")
                block[f"{label}_LY"] = pd.to_numeric(ly["UPTO"] if "UPTO" in ly.columns else np.nan, errors="coerce")
                block[f"{label}_Var"] = (block[f"{label}_CY"] - block[f"{label}_LY"]).astype(float).round(2)
                all_parts.append(block)
            result7 = pd.concat(all_parts, axis=1).reset_index()
            result7 = result7.rename(columns={"ROUTEE": "ROUTE", service_col: "SERVICE NO"})
            result7 = result7.sort_values(["DEPOT", "SERVICE NO", "ROUTE", "PRODUCT"]).reset_index(drop=True)
            # Freeze: same pattern as Schedules Table 3 (inline sticky only)
            html = ['<div class="op-wrap"><table class="op-table" style="border-collapse:separate;border-spacing:0;width:max-content;"><thead>']
            html.append("<tr>")
            html.append('<th rowspan="2" style="position:sticky;left:0;top:0;z-index:6;background:#0f172a;color:white;padding:8px 6px;font-size:12px;min-width:70px;">DEPOT</th>')
            html.append('<th rowspan="2" style="position:sticky;left:70px;top:0;z-index:6;background:#0f172a;color:white;padding:8px 6px;font-size:12px;min-width:90px;">SERVICE NO</th>')
            html.append('<th rowspan="2" style="position:sticky;left:160px;top:0;z-index:6;background:#0f172a;color:white;padding:8px 6px;font-size:12px;min-width:80px;">ROUTE</th>')
            html.append('<th rowspan="2" style="position:sticky;left:240px;top:0;z-index:6;background:#0f172a;color:white;padding:8px 6px;font-size:12px;min-width:80px;">PRODUCT</th>')
            html.append(f'<th colspan="10" style="position:sticky;top:0;z-index:4;background:#6b21a8;color:white;padding:8px;">{prefix} TOT. E.P.K</th>')
            html.append(f'<th colspan="10" style="position:sticky;top:0;z-index:4;background:#15803d;color:white;padding:8px;">{prefix} FPD. E.P.K</th>')
            html.append(f'<th colspan="10" style="position:sticky;top:0;z-index:4;background:#1d4ed8;color:white;padding:8px;">{prefix} MHL. E.P.K</th>')
            html.append("</tr><tr>")
            for _ in range(3):
                for d in day_short:
                    html.append(f'<th style="position:sticky;top:36px;z-index:4;background:#f1f5f9;color:#0f172a;padding:6px;font-size:11px;">{d}</th>')
                html.append(f'<th style="position:sticky;top:36px;z-index:4;background:#f1f5f9;color:#0f172a;padding:6px;font-size:11px;">{for_upto} CY</th>')
                html.append(f'<th style="position:sticky;top:36px;z-index:4;background:#f1f5f9;color:#0f172a;padding:6px;font-size:11px;">{for_upto} LY</th>')
                html.append('<th style="position:sticky;top:36px;z-index:4;background:#f1f5f9;color:#0f172a;padding:6px;font-size:11px;">Var</th>')
            html.append("</tr></thead><tbody>")
            sticky_base = "position:sticky;z-index:2;padding:6px;font-size:12px;"
            for _, row in result7.iterrows():
                html.append("<tr>")
                html.append(f'<td style="{sticky_base}left:0;background:#e0f2fe;font-weight:600;min-width:70px;">{row["DEPOT"]}</td>')
                html.append(f'<td style="{sticky_base}left:70px;background:#f0f9ff;min-width:90px;">{row["SERVICE NO"]}</td>')
                html.append(f'<td style="{sticky_base}left:160px;background:#f8fafc;min-width:80px;">{row["ROUTE"]}</td>')
                html.append(f'<td style="{sticky_base}left:240px;background:#f8fafc;min-width:80px;">{row["PRODUCT"]}</td>')
                for label in ["TOT", "FPD", "MHL"]:
                    # Peak / Slack among Mon-Sun for this metric on this row
                    day_vals = {}
                    for d in day_order:
                        v = row.get(f"{label}_{d}")
                        try:
                            fv = float(v) if pd.notna(v) else None
                        except Exception:
                            fv = None
                        if fv is not None and fv != 0:
                            day_vals[d] = fv
                    peak_d = max(day_vals, key=day_vals.get) if day_vals else None
                    slack_d = min(day_vals, key=day_vals.get) if day_vals else None
                    for d in day_order:
                        val = row.get(f"{label}_{d}")
                        style = ""
                        if peak_d and d == peak_d:
                            style = 'background-color:#c6efce; color:#006100; font-weight:600;'
                        elif slack_d and d == slack_d:
                            style = 'background-color:#ffc7ce; color:#9c0006; font-weight:600;'
                        html.append(f'<td style="{style}">{fmt(val)}</td>')
                    html.append(f'<td>{fmt(row.get(f"{label}_CY"))}</td>')
                    html.append(f'<td>{fmt(row.get(f"{label}_LY"))}</td>')
                    var_val = row.get(f"{label}_Var")
                    html.append(f'<td class="{var_class(var_val)}">{fmt(var_val)}</td>')
                html.append("</tr>")
            html.append("</tbody></table></div>")
            st.markdown("".join(html), unsafe_allow_html=True)
            st.markdown(
                '<div style="clear:both;margin-top:10px;">'
                '<span style="background:#c6efce;padding:2px 8px;">Peak day</span> = Highest EPK among Mon–Sun &nbsp;'
                '<span style="background:#ffc7ce;padding:2px 8px;">Slack day</span> = Lowest EPK among Mon–Sun'
                '&nbsp;(per row, for each of TOT / FPD / MHL)</div>',
                unsafe_allow_html=True,
            )
            st.download_button(
                "Download Excel",
                excel_with_title(
                    result7,
                    "Service",
                    report_title=f"Service Performance | Depot={depot} | Month={month} | {for_upto} | {net_gross}",
                ),
                f"Service_Performance_{month}_{depot}.xlsx",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="dl7",
            )

# ==================== TAB 8 ====================
elif section == "Period Comparison":
    st.markdown('<div class="title-bar">Date-Range Set vs Date-Range Set Variance Analysis</div>', unsafe_allow_html=True)
    
    # Base dataset filtered by cascading filters (excluding month/upto date limitations)
    filtered_df = df[base_mask].copy()
    
    if filtered_df.empty or "Date" not in filtered_df.columns:
        st.warning("No data available for date range selection under current filters.")
    else:
        min_date = filtered_df["Date"].min().date()
        max_date = filtered_df["Date"].max().date()
        
        c_set1, c_set2 = st.columns(2)
        with c_set1:
            st.markdown("##### Period Set 1 (Target Period)")
            set1_range = st.date_input(
                "Select Date Range for Set 1",
                value=(min_date, max_date),
                min_value=min_date,
                max_value=max_date,
                key="tab8_set1_dates"
            )
        with c_set2:
            st.markdown("##### Period Set 2 (Comparison Period)")
            set2_range = st.date_input(
                "Select Date Range for Set 2",
                value=(min_date, max_date),
                min_value=min_date,
                max_value=max_date,
                key="tab8_set2_dates"
            )
            
        if isinstance(set1_range, (list, tuple)) and len(set1_range) == 2 and isinstance(set2_range, (list, tuple)) and len(set2_range) == 2:
            s1_start, s1_end = pd.Timestamp(set1_range[0]), pd.Timestamp(set1_range[1])
            s2_start, s2_end = pd.Timestamp(set2_range[0]), pd.Timestamp(set2_range[1])
            
            data_set1 = filtered_df[(filtered_df["Date"] >= s1_start) & (filtered_df["Date"] <= s1_end)]
            data_set2 = filtered_df[(filtered_df["Date"] >= s2_start) & (filtered_df["Date"] <= s2_end)]
            
            p1_label = f"{s1_start.strftime('%d-%b-%Y')} to {s1_end.strftime('%d-%b-%Y')}"
            p2_label = f"{s2_start.strftime('%d-%b-%Y')} to {s2_end.strftime('%d-%b-%Y')}"
            
            st.caption(f"Comparing **{p1_label}** vs **{p2_label}**")
            
            html_str, merged_df = build_act_vs_act_table(
                group_col="DEPOT",
                data_cy=data_set1,
                data_ly=data_set2,
                cy_label="SET 1",
                ly_label="SET 2"
            )
            
            if html_str is None:
                st.warning("No data matching the selected date ranges.")
            else:
                st.markdown(html_str, unsafe_allow_html=True)
                st.download_button(
                    "Download CSV",
                    merged_df.to_csv(index=False).encode("utf-8"),
                    f"Period_Set_Comparison_{depot}.csv",
                    "text/csv",
                    key="dl8"
                )
        else:
            st.info("Please select valid start and end dates for both Period Set 1 and Period Set 2.")






# ==================== TAB: Service-wise (SROS) ====================
elif section == "Service-wise (SROS)":
    title_sw = f"Service-wise Performance ({net_gross}) ({for_upto}) - Month: {month}"
    if depot != "ALL":
        title_sw += f" | Depot: {depot}"
    st.markdown(f'<div class="title-bar">{title_sw}</div>', unsafe_allow_html=True)

    st.markdown("**Filters** (cascading from top: Depot / Product / Route / Month)")
    fpd_opts = ["ALL", "FPD EPK < 20", "FPD EPK < 25", "FPD EPK < 30", "FPD EPK < 35", "FPD EPK < 40"]
    fpd_filter = "ALL"

    # --- Load SMASTER services from parquet ---
    sros_path = Path(r"D:\Dashboard\SMASTER.parquet")
    if not sros_path.exists():
        for alt in [Path(r"D:\dashboard\SMASTER.parquet"), Path(r"D:\MONTHLY\SMASTER.parquet"), Path("SMASTER.parquet"), Path(r"/home/workdir/attachments/SMASTER.parquet")]:
            if alt.exists():
                sros_path = alt
                break
    if not sros_path.exists():
        st.error(f"SMASTER.parquet not found: {sros_path}")
    else:
        sros = _load_sros_services_full(str(sros_path), str(depot), str(month))
        if len(sros) == 0:
            st.warning("No services found in SROS/SMASTER for current Depot.")
        else:
            # Apply cascading filters (Product / Route) on SROS list
            if product != "ALL" and "_PROD" in sros.columns:
                sros = sros[sros["_PROD"].astype(str).str.strip().str.upper() == str(product).strip().upper()]
            if route != "ALL" and "_ROUTE" in sros.columns:
                sros = sros[sros["_ROUTE"].astype(str).str.strip().str.upper() == str(route).strip().upper()]

            # SERVICE NO filter (from SROS list after depot/product/route)
            svc_list = sorted(
                {str(x) for x in sros["_SVC"].dropna().unique() if str(x).strip() and str(x).lower() != "nan"},
                key=lambda z: (0, int(z)) if str(z).isdigit() else (1, str(z)),
            )
            _sw1, _sw2, _sw3 = st.columns([1, 1, 2])
            with _sw1:
                fpd_filter = st.selectbox("FPD EPK Filter", fpd_opts, index=0, key="sw_fpd_filter")
            with _sw2:
                sw_svc = st.selectbox("SERVICE NO", ["ALL"] + svc_list, index=0, key="sw_service_no")
            if sw_svc != "ALL":
                sros = sros[sros["_SVC"].astype(str) == str(sw_svc)]

            # CY schedules for month
            s_cy = sros[sros["_MK"] == str(month)] if str(month) else sros
            ly_key = ""
            try:
                parts = str(month).split("-")
                ly_key = f"{parts[0]}-{int(parts[1]) - 1}"
            except Exception:
                pass
            s_ly = sros[sros["_MK"] == ly_key] if ly_key else sros.iloc[0:0]

            # Aggregate SROS by service + product + route + depot
            def agg_sros(frame):
                if len(frame) == 0:
                    return pd.DataFrame(columns=["_SVC", "_DEP", "_PROD", "_ROUTE", "SCH"])
                g = frame.groupby(["_SVC", "_DEP", "_PROD", "_ROUTE"], dropna=False).agg(SCH=("_SCH", "sum")).reset_index()
                return g

            cy_sch = agg_sros(s_cy)
            ly_sch = agg_sros(s_ly)
            sch = cy_sch.merge(ly_sch, on=["_SVC", "_DEP", "_PROD", "_ROUTE"], how="outer", suffixes=("_CY", "_LY")).fillna(0)
            sch = sch.rename(columns={"SCH_CY": "sch_CY", "SCH_LY": "sch_LY"})
            if "sch_CY" not in sch.columns:
                sch["sch_CY"] = 0
            if "sch_LY" not in sch.columns:
                sch["sch_LY"] = 0

            # Performance from parquet (already filtered cy_data / ly_data by global filters)
            svc_col = service_col if service_col else next(
                (c for c in ["SER_NO", "SERVICE_NO", "SERVICE"] if c in df.columns), None
            )
            if not svc_col:
                st.error("SER_NO not found in performance data.")
            else:
                def norm_svc(v):
                    try:
                        return str(int(float(v)))
                    except Exception:
                        return str(v).strip()

                def perf_agg(data):
                    if len(data) == 0:
                        return pd.DataFrame(columns=["_SVC", "_DEP", "_PROD", "_ROUTE", "kms", "earn_tot", "earn_fpd", "earn_mhl", "pax"])
                    d = data.copy()
                    d["_SVC"] = d[svc_col].map(norm_svc)
                    d["_DEP"] = d["DEPOT"].astype(str).str.strip().str.upper() if "DEPOT" in d.columns else ""
                    d["_PROD"] = d["PRODUCT"].astype(str).str.strip() if "PRODUCT" in d.columns else ""
                    rc = "ROUTEE" if "ROUTEE" in d.columns else ("ROUTE" if "ROUTE" in d.columns else None)
                    d["_ROUTE"] = d[rc].astype(str).str.strip() if rc else ""
                    g = d.groupby(["_SVC", "_DEP", "_PROD", "_ROUTE"], dropna=False).agg(
                        kms=("Optd_KMs", "sum"),
                        earn_tot=(earn_tot, "sum"),
                        earn_fpd=(earn_fpd, "sum"),
                        earn_mhl=(earn_mhl, "sum"),
                        pax=(pax_col, "sum"),
                    ).reset_index()
                    return g

                cy_p = perf_agg(cy_data)
                ly_p = perf_agg(ly_data)
                perf = cy_p.merge(ly_p, on=["_SVC", "_DEP", "_PROD", "_ROUTE"], how="outer", suffixes=("_CY", "_LY")).fillna(0)

                # Master list from SROS (services), left-join performance
                master = sch.merge(perf, on=["_SVC", "_DEP", "_PROD", "_ROUTE"], how="left").fillna(0)

                # Metrics in same units as ACT VS ACT (kms/earn already summed; convert to lakhs like build)
                for side in ["CY", "LY"]:
                    master[f"kms_{side}"] = master[f"kms_{side}"] / 100000.0
                    master[f"earn_tot_{side}"] = master[f"earn_tot_{side}"] / 100000.0
                    # FPD/MHL earnings stay for EPK only
                for side in ["CY", "LY"]:
                    master[f"epk_tot_{side}"] = np.where(
                        master[f"kms_{side}"] > 0,
                        (master[f"earn_tot_{side}"] * 100000) / (master[f"kms_{side}"] * 100000),
                        np.nan,
                    )
                    # Recalculate EPK from original units: earn/kms before /100000
                # Better recompute EPK from pre-lakh values - we already divided. Fix:
                # Reload ratios: epk = earn_raw / kms_raw = (earn_lakh*1e5)/(kms_lakh*1e5) = earn_lakh/kms_lakh
                    master[f"epk_tot_{side}"] = np.where(
                        master[f"kms_{side}"] > 0,
                        master[f"earn_tot_{side}"] / master[f"kms_{side}"],
                        np.nan,
                    )
                # Need fpd/mhl earn in same scale for epk
                for side in ["CY", "LY"]:
                    master[f"epk_fpd_{side}"] = np.where(
                        master[f"kms_{side}"] > 0,
                        (master[f"earn_fpd_{side}"] / 100000.0) / master[f"kms_{side}"],
                        np.nan,
                    )
                    master[f"epk_mhl_{side}"] = np.where(
                        master[f"kms_{side}"] > 0,
                        (master[f"earn_mhl_{side}"] / 100000.0) / master[f"kms_{side}"],
                        np.nan,
                    )
                    master[f"pax_{side}"] = master.get(f"pax_{side}", 0)

                # VAR / PCT
                for base in ["sch", "kms", "earn_tot", "epk_tot", "epk_fpd", "epk_mhl", "pax"]:
                    master[f"{base}_VAR"] = master[f"{base}_CY"] - master[f"{base}_LY"]
                    master[f"{base}_PCT"] = np.where(
                        master[f"{base}_LY"] != 0,
                        master[f"{base}_VAR"] * 100 / master[f"{base}_LY"],
                        np.nan,
                    )

                # FPD EPK filter
                thresh = None
                if fpd_filter and fpd_filter != "ALL":
                    try:
                        thresh = float(fpd_filter.split("<")[-1].strip())
                    except Exception:
                        thresh = None
                if thresh is not None:
                    master = master[master["epk_fpd_CY"].fillna(999) < thresh].copy()

                master = master.sort_values(["_DEP", "_SVC", "_PROD", "_ROUTE"]).reset_index(drop=True)

                if len(master) == 0:
                    st.warning("No rows after filters.")
                else:
                    html = ['<div class="table-scroll"><table class="excel-table"><thead>']
                    html.append("<tr>")
                    html.append('<th class="header-left" rowspan="2">S.No</th>')
                    html.append('<th class="header-left" rowspan="2">Depot</th>')
                    html.append('<th class="header-left" rowspan="2">service no</th>')
                    html.append('<th class="header-left" rowspan="2">PRODUCT</th>')
                    html.append('<th class="header-left" rowspan="2">ROUTE</th>')
                    html.append('<th colspan="3" style="background:#b45309;color:white;">NO OF SCHEDULES</th>')
                    html.append('<th class="header-km" colspan="4">KILOMETERS (IN LKS.)</th>')
                    html.append(f'<th class="header-earn" colspan="4">{prefix.upper()} EARNINGS (IN LKS.)</th>')
                    html.append(f'<th class="header-tot" colspan="4">{prefix.upper()} TOT EPK</th>')
                    html.append(f'<th class="header-fpd" colspan="4">{prefix.upper()} FPD EPK</th>')
                    html.append(f'<th class="header-mhl" colspan="4">{prefix.upper()} MHL EPK</th>')
                    pax_heading = {"FPD": "FPD PASSENGERS", "MHL": "MHL PASSENGERS"}.get(passengers, "TOTAL PASSENGERS")
                    html.append(f'<th class="header-left" colspan="4">{pax_heading}</th>')
                    html.append("</tr><tr>")
                    for _i_sh, _ in enumerate(range(7)):
                        if _i_sh == 0:
                            html.append('<th class="header-sub">CY</th>')
                            html.append('<th class="header-sub">LY</th>')
                            html.append('<th class="header-sub">VAR</th>')
                            continue

                        html.append('<th class="header-sub">CY</th>')
                        html.append('<th class="header-sub">LY</th>')
                        html.append('<th class="header-sub">VAR</th>')
                        html.append('<th class="header-sub">% ▲/▼</th>')
                    html.append("</tr></thead><tbody>")
                    for sno, row in master.iterrows():
                        html.append("<tr>")
                        html.append(f"<td>{sno + 1}</td>")
                        html.append(f'<td>{row["_DEP"]}</td>')
                        html.append(f'<td>{row["_SVC"]}</td>')
                        html.append(f'<td>{row["_PROD"]}</td>')
                        html.append(f'<td>{row["_ROUTE"]}</td>')
                        # sch (whole numbers, no % column)
                        def _fmt_sch(v):
                            try:
                                if pd.isna(v) or float(v) == 0:
                                    return ""
                                return f"{int(round(float(v))):,}"
                            except Exception:
                                return ""
                        html.append(f'<td>{_fmt_sch(row["sch_CY"])}</td>')
                        html.append(f'<td>{_fmt_sch(row["sch_LY"])}</td>')
                        html.append(f'<td class="{var_class(row["sch_VAR"])}">{_fmt_sch(row["sch_VAR"])}</td>')
                        for metric in ["kms", "earn_tot"]:
                            html.append(f'<td>{fmt(row[f"{metric}_CY"])}</td>')
                            html.append(f'<td>{fmt(row[f"{metric}_LY"])}</td>')
                            html.append(f'<td class="{var_class(row[f"{metric}_VAR"])}">{fmt(row[f"{metric}_VAR"])}</td>')
                            html.append(f'<td class="{var_class(row[f"{metric}_PCT"])}">{fmt_growth(row[f"{metric}_PCT"])}</td>')
                        for epk in ["tot", "fpd", "mhl"]:
                            html.append(f'<td>{fmt(row[f"epk_{epk}_CY"])}</td>')
                            html.append(f'<td>{fmt(row[f"epk_{epk}_LY"])}</td>')
                            html.append(f'<td class="{var_class(row[f"epk_{epk}_VAR"])}">{fmt(row[f"epk_{epk}_VAR"])}</td>')
                            html.append(f'<td class="{var_class(row[f"epk_{epk}_PCT"])}">{fmt_growth(row[f"epk_{epk}_PCT"])}</td>')
                        html.append(f'<td>{fmt_pax(row.get("pax_CY", 0))}</td>')
                        html.append(f'<td>{fmt_pax(row.get("pax_LY", 0))}</td>')
                        html.append(f'<td class="{var_class(row.get("pax_VAR", 0))}">{fmt_pax(row.get("pax_VAR", 0))}</td>')
                        html.append(f'<td class="{var_class(row.get("pax_PCT", 0))}">{fmt_growth(row.get("pax_PCT", 0))}</td>')
                        html.append("</tr>")
                    html.append("</tbody></table></div>")
                    st.markdown("".join(html), unsafe_allow_html=True)
                    st.caption(f"Rows: {len(master):,} | FPD filter: {fpd_filter} | Services from SROS/SMASTER · Performance from parquet")
                    out_df = master.rename(columns={
                        "_SVC": "service_no", "_DEP": "Depot", "_PROD": "PRODUCT", "_ROUTE": "ROUTE"
                    })
                    st.download_button(
                        "Download Excel",
                        excel_with_title(
                            out_df,
                            "Service_wise",
                            report_title=f"Service-wise (SROS) | Depot={depot} | Month={month} | {for_upto} | {net_gross} | {fpd_filter}",
                        ),
                        f"Service_wise_SROS_{month}_{depot}.xlsx",
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key="dl_sw_sros",
                    )

# ==================== TAB 9: Schedules (SMASTER) ====================

elif section == "Task":
    st.markdown(f'<div class="title-bar">Daily Depot Performance – {month}</div>', unsafe_allow_html=True)
    # Daily Depot Performance (moved from ACT VS ACT)


    # Build calendar month date range from month filter (e.g. Apr-2026 → 2026-04-01 to 2026-04-30)
    mon_map = {
        "Jan": 1, "Feb": 2, "Mar": 3, "Apr": 4, "May": 5, "Jun": 6,
        "Jul": 7, "Aug": 8, "Sep": 9, "Oct": 10, "Nov": 11, "Dec": 12
    }
    try:
        parts = month.replace(" ", "").split("-")
        m_num = mon_map.get(parts[0][:3], 1)
        y_num = int(parts[1]) if len(parts[1]) == 4 else int("20" + parts[1])
        month_start = pd.Timestamp(year=y_num, month=m_num, day=1)
        # last day of month
        if m_num == 12:
            month_end = pd.Timestamp(year=y_num, month=12, day=31)
        else:
            month_end = pd.Timestamp(year=y_num, month=m_num + 1, day=1) - pd.Timedelta(days=1)
    except Exception:
        month_start = selected_max_date.replace(day=1) if not pd.isna(selected_max_date) else None
        month_end = selected_max_date if not pd.isna(selected_max_date) else None

    if month_start is not None:
        daily_mask = base_mask & (df["Date"] >= month_start) & (df["Date"] <= month_end)
        daily_df = df[daily_mask].copy()
        if len(daily_df) == 0:
            st.warning("No daily data for selected month and filters.")
        else:
            g = daily_df.groupby(["DEPOT", "Date"]).agg(
                Kilometers=("Optd_KMs", "sum"),
                Earnings=(earn_tot, "sum"),
                Earn_FPD=(earn_fpd, "sum"),
                Earn_MHL=(earn_mhl, "sum"),
                Passengers=(pax_col, "sum"),
            ).reset_index()
            g["EPK_TOT"] = np.where(g["Kilometers"] > 0, g["Earnings"] / g["Kilometers"], np.nan)
            g["EPK_FPD"] = np.where(g["Kilometers"] > 0, g["Earn_FPD"] / g["Kilometers"], np.nan)
            g["EPK_MHL"] = np.where(g["Kilometers"] > 0, g["Earn_MHL"] / g["Kilometers"], np.nan)
            g = g.sort_values(["DEPOT", "Date"]).reset_index(drop=True)
            g["_dt"] = pd.to_datetime(g["Date"])
            g["Weekday"] = g["_dt"].dt.strftime("%a")  # Mon, Tue, ...
            g["Date"] = g["_dt"].dt.strftime("%d-%m-%Y")
            g["Kilometers"] = (g["Kilometers"] / 100000).round(2)
            g["Earnings"] = (g["Earnings"] / 100000).round(2)
            g["EPK_TOT"] = g["EPK_TOT"].round(2)
            g["EPK_FPD"] = g["EPK_FPD"].round(2)
            g["EPK_MHL"] = g["EPK_MHL"].round(2)
            g["Passengers"] = g["Passengers"].round(0)

            # REGION total when ALL/REGION selected
            if str(depot).upper() in ("ALL", "REGION"):
                reg_raw = daily_df.groupby("Date").agg(
                    Kilometers=("Optd_KMs", "sum"),
                    Earnings=(earn_tot, "sum"),
                    Earn_FPD=(earn_fpd, "sum"),
                    Earn_MHL=(earn_mhl, "sum"),
                    Passengers=(pax_col, "sum"),
                ).reset_index()
                reg_raw["DEPOT"] = "REGION"
                reg_raw["EPK_TOT"] = np.where(reg_raw["Kilometers"] > 0, reg_raw["Earnings"] / reg_raw["Kilometers"], np.nan)
                reg_raw["EPK_FPD"] = np.where(reg_raw["Kilometers"] > 0, reg_raw["Earn_FPD"] / reg_raw["Kilometers"], np.nan)
                reg_raw["EPK_MHL"] = np.where(reg_raw["Kilometers"] > 0, reg_raw["Earn_MHL"] / reg_raw["Kilometers"], np.nan)
                reg_raw["_dt"] = pd.to_datetime(reg_raw["Date"])
                reg_raw["Weekday"] = reg_raw["_dt"].dt.strftime("%a")
                reg_raw["Date"] = reg_raw["_dt"].dt.strftime("%d-%m-%Y")
                reg_raw["Kilometers"] = (reg_raw["Kilometers"] / 100000).round(2)
                reg_raw["Earnings"] = (reg_raw["Earnings"] / 100000).round(2)
                reg_raw["EPK_TOT"] = reg_raw["EPK_TOT"].round(2)
                reg_raw["EPK_FPD"] = reg_raw["EPK_FPD"].round(2)
                reg_raw["EPK_MHL"] = reg_raw["EPK_MHL"].round(2)
                reg_raw["Passengers"] = reg_raw["Passengers"].round(0)
                g = pd.concat([g, reg_raw[[c for c in g.columns if c in reg_raw.columns]]], ignore_index=True)
            if str(depot).upper() == "REGION":
                g = g[g["DEPOT"].astype(str).str.upper() == "REGION"].copy()

            # ORF / OR columns
            try:
                orf_map, orf_by_prod, _oe = load_orf_map(r"D:\dashboard\ORF.xlsx")
            except Exception:
                orf_map, orf_by_prod = {}, {}
            def _orf_dep(d):
                d = str(d).strip().upper()
                if d in ("REGION", "TOTAL", "ALL"):
                    d = "REGION"
                rec = orf_map.get(d, {}) if isinstance(orf_map, dict) else {}
                return float(rec.get("cy") or 0) or np.nan
            g["ORF"] = g["DEPOT"].map(_orf_dep)
            for epk_c, or_c in [("EPK_TOT", "OR_TOT"), ("EPK_FPD", "OR_FPD"), ("EPK_MHL", "OR_MHL")]:
                g[or_c] = np.where(
                    pd.to_numeric(g["ORF"], errors="coerce").fillna(0) != 0,
                    pd.to_numeric(g[epk_c], errors="coerce") * 10000 / g["ORF"],
                    np.nan,
                ).round(2)

            # TOTAL last row
            if len(g) > 0:
                tot = {
                    "DEPOT": "TOTAL", "Date": "", "Weekday": "",
                    "Kilometers": g["Kilometers"].sum(),
                    "Earnings": g["Earnings"].sum(),
                    "Passengers": g["Passengers"].sum(),
                    "EPK_TOT": round(g["Earnings"].sum() / g["Kilometers"].sum(), 2) if g["Kilometers"].sum() else np.nan,
                    "EPK_FPD": np.nan, "EPK_MHL": np.nan,
                    "OR_TOT": g["OR_TOT"].mean() if "OR_TOT" in g.columns else np.nan,
                    "OR_FPD": g["OR_FPD"].mean() if "OR_FPD" in g.columns else np.nan,
                    "OR_MHL": g["OR_MHL"].mean() if "OR_MHL" in g.columns else np.nan,
                }
                g = pd.concat([g, pd.DataFrame([tot])], ignore_index=True)

            pax_heading = {"FPD": "FPD PASSENGERS", "MHL": "MHL PASSENGERS"}.get(passengers, "TOTAL PASSENGERS")

            # Compact table: ~header + 31 rows visible per viewport
            html_d = ['<div class="table-scroll" style="max-height: 75vh;"><table class="excel-table" style="font-size:10px;">']
            html_d.append("<tr>")
            html_d.append('<th class="header-left" style="padding:2px 3px; font-size:9px;">DEPOT</th>')
            html_d.append('<th class="header-left" style="padding:2px 3px; font-size:9px;">DATE</th>')
            html_d.append('<th class="header-left" style="padding:2px 3px; font-size:9px;">WEEKDAY</th>')
            html_d.append('<th class="header-km" style="padding:2px 3px; font-size:9px;">KMs (lks.)</th>')
            html_d.append(f'<th class="header-earn" style="padding:2px 3px; font-size:9px;">{prefix} EARN (lks.)</th>')
            html_d.append(f'<th class="header-tot" style="padding:2px 3px; font-size:9px;">{prefix} TOT EPK</th>')
            html_d.append(f'<th class="header-fpd" style="padding:2px 3px; font-size:9px;">{prefix} FPD EPK</th>')
            html_d.append(f'<th class="header-mhl" style="padding:2px 3px; font-size:9px;">{prefix} MHL EPK</th>')
            html_d.append('<th style="background:#7c3aed;color:#fff;padding:2px 3px;font-size:9px;">TOT OR</th>')
            html_d.append('<th style="background:#0d9488;color:#fff;padding:2px 3px;font-size:9px;">FPD OR</th>')
            html_d.append('<th style="background:#2563eb;color:#fff;padding:2px 3px;font-size:9px;">MHL OR</th>')
            html_d.append(f'<th class="header-left" style="padding:2px 3px; font-size:9px;">{pax_heading}</th>')
            html_d.append("</tr>")
            # Peak / Slack by TOT EPK within each depot
            peak_dates = set()
            slack_dates = set()
            for dep, grp in g.groupby("DEPOT"):
                vals = grp[["Date", "EPK_TOT"]].dropna()
                vals = vals[vals["EPK_TOT"] != 0]
                if len(vals) == 0:
                    continue
                # top 2 peak, bottom 2 slack
                sorted_v = vals.sort_values("EPK_TOT", ascending=False)
                for d in sorted_v.head(2)["Date"]:
                    peak_dates.add((dep, d))
                for d in sorted_v.tail(2)["Date"]:
                    slack_dates.add((dep, d))
            # avoid same day marked both if few days
            slack_dates = slack_dates - peak_dates

            for _, r in g.iterrows():
                key = (r["DEPOT"], r["Date"])
                row_style = ""
                date_style = "font-weight:700;"
                if key in peak_dates:
                    row_style = "background-color:#c6efce;"
                    date_style = "font-weight:700; color:#006100;"
                elif key in slack_dates:
                    row_style = "background-color:#ffc7ce;"
                    date_style = "font-weight:700; color:#9c0006;"
                html_d.append(f'<tr style="{row_style}">')
                html_d.append(f'<td style="font-weight:700; padding:1px 3px;">{r["DEPOT"]}</td>')
                html_d.append(f'<td style="{date_style} padding:1px 3px;">{r["Date"]}</td>')
                html_d.append(f'<td style="padding:1px 3px;">{r["Weekday"]}</td>')
                html_d.append(f'<td style="padding:1px 3px;">{fmt(r["Kilometers"])}</td>')
                html_d.append(f'<td style="padding:1px 3px;">{fmt(r["Earnings"])}</td>')
                html_d.append(f'<td style="font-weight:700; padding:1px 3px;">{fmt(r["EPK_TOT"])}</td>')
                html_d.append(f'<td style="padding:1px 3px;">{fmt(r["EPK_FPD"])}</td>')
                html_d.append(f'<td style="padding:1px 3px;">{fmt(r["EPK_MHL"])}</td>')
                html_d.append(f'<td style="padding:1px 3px;">{fmt(r.get("OR_TOT", np.nan))}</td>')
                html_d.append(f'<td style="padding:1px 3px;">{fmt(r.get("OR_FPD", np.nan))}</td>')
                html_d.append(f'<td style="padding:1px 3px;">{fmt(r.get("OR_MHL", np.nan))}</td>')
                html_d.append(f'<td style="padding:1px 3px;">{fmt_pax(r["Passengers"])}</td>')
                html_d.append("</tr>")
            html_d.append("</table></div>")
            st.markdown("""
<span style="background:#c6efce; padding:2px 8px; font-weight:700;">Peak</span> = Top 2 TOT EPK days per depot &nbsp;
<span style="background:#ffc7ce; padding:2px 8px; font-weight:700;">Slack</span> = Bottom 2 TOT EPK days per depot
""", unsafe_allow_html=True)
            st.markdown("".join(html_d), unsafe_allow_html=True)
            st.caption(f"Date range: {month_start.strftime('%d-%m-%Y')} to {month_end.strftime('%d-%m-%Y')} | Rows: {len(g):,}")
            st.download_button(
                "Download Daily CSV",
                g.to_csv(index=False).encode("utf-8"),
                f"Daily_Depot_{month}.csv",
                "text/csv",
                key="dl2_daily"
            )


# ==================== TAB 3 ====================




elif section == "Schedules":
    st.markdown('<div class="title-bar">Schedules – SCHs / SERVICES / SCH KMS</div>', unsafe_allow_html=True)

    SCHEDULE_PARQUET = r"D:\Dashboard\SMASTER.parquet"

    sched_raw, sched_err = load_smaster(SCHEDULE_PARQUET)
    if sched_err:
        st.warning(sched_err)
    elif sched_raw is None or len(sched_raw) == 0:
        st.warning("SMASTER empty")
    else:
        def find_col(cands):
            def normkey(s):
                return str(s).strip().lower().replace(" ", "").replace("_", "").replace("/", "").replace(".", "")
            norm = {normkey(c): c for c in sched_raw.columns}
            for cand in cands:
                k = normkey(cand)
                if k in norm:
                    return norm[k]
            for c in sched_raw.columns:
                cl = normkey(c)
                for cand in cands:
                    if normkey(cand) in cl or cl in normkey(cand):
                        return c
            # last resort: strip-only match on display name
            strip_map = {str(c).strip().upper(): c for c in sched_raw.columns}
            for cand in cands:
                if cand.strip().upper() in strip_map:
                    return strip_map[cand.strip().upper()]
            return None

        col_depot = find_col(["DEPOT"])
        col_product = find_col(["PRODUCT"])
        col_mhl = find_col(["MHL/NMHL", "MHL_NMHL", "MHLNMHL"])
        col_month = find_col(["MONTH", "Month"])
        col_year = find_col(["YEAR", "Year"])
        col_rtc = find_col(["RTC/HIRE", "RTC_HIRE"])
        col_sch = find_col(["NoOfSchedules", "NoOfSchedule"])
        col_svc = find_col(["ServiceNo", "SERVICENO"])
        col_kms = find_col(["RevenueKms", "Revenue Kms"])
        col_dtype = find_col(["D.TYPE", "DTYPE", "DutyType", "OtService", "ServiceType"])

        required = [col_depot, col_product, col_rtc, col_sch, col_svc, col_kms]
        if any(x is None for x in required):
            st.error("Missing required columns")
            st.write(list(sched_raw.columns))
        else:
            sdf = sched_raw.copy()
            if col_month and col_year:
                def make_mon(row):
                    try:
                        m = str(row[col_month]).strip()[:3].title()
                        y = int(float(row[col_year]))
                        return f"{m}-{y}"
                    except Exception:
                        return str(row[col_month])
                sdf["_MonthKey"] = sdf.apply(make_mon, axis=1)
            elif col_month:
                sdf["_MonthKey"] = sdf[col_month].astype(str).str.strip()
            else:
                sdf["_MonthKey"] = "ALL"

            def norm_rtc(v):
                s = str(v).strip().upper()
                if s in ("R", "RTC"):
                    return "RTC"
                if s in ("H", "HIRE"):
                    return "HIRE"
                return s

            sdf["_RTC"] = sdf[col_rtc].map(norm_rtc)
            sdf["_SCH"] = pd.to_numeric(sdf[col_sch], errors="coerce").fillna(0)
            sdf["_KMS"] = pd.to_numeric(sdf[col_kms], errors="coerce").fillna(0)
            sdf["_DEPOT"] = sdf[col_depot].astype(str).str.strip()
            sdf["_PRODUCT"] = sdf[col_product].astype(str).str.strip()
            sdf["_MHL"] = sdf[col_mhl].astype(str).str.strip() if col_mhl else "ALL"
            if col_dtype:
                sdf["_DTYPE"] = sdf[col_dtype].astype(str).str.strip().str.upper()
            else:
                sdf["_DTYPE"] = ""


            # ---- Shared filters (Table A uses only these) ----
            fc1, fc2, fc3 = st.columns(3)
            with fc1:
                d_opts = ["ALL", "REGION"] + sorted([str(x).strip() for x in sdf["_DEPOT"].dropna().unique() if str(x).strip() and str(x).strip().lower() != "nan"])
                f_depot = st.selectbox("DEPOT", d_opts, key="sch_depot")
            with fc2:
                mon_opts = sorted(
                    [x for x in sdf["_MonthKey"].dropna().unique() if x and x != "ALL"],
                    key=lambda m: pd.to_datetime(str(m), format="%b-%Y", errors="coerce"),
                    reverse=True,
                ) or ["ALL"]
                f_month = st.selectbox("MONTH", mon_opts, key="sch_month")
            with fc3:
                f_compare = st.selectbox("COMPARE WITH", ["LAST YEAR", "PREVIOUS MONTH"], key="sch_compare")

            # Base for Table A: only depot + month (no product/mhl/route filter)
            base_a = sdf.copy()
            if f_depot not in ("ALL", "REGION"):
                base_a = base_a[base_a["_DEPOT"] == f_depot]

            cy_sdf = base_a[base_a["_MonthKey"] == f_month].copy() if f_month != "ALL" else base_a.copy()
            ly_sdf = base_a.iloc[0:0].copy()
            ly_key = ""
            try:
                parts = str(f_month).split("-")
                mon_abbr, yr = parts[0], int(parts[1])
                mon_map = {"Jan":1,"Feb":2,"Mar":3,"Apr":4,"May":5,"Jun":6,"Jul":7,"Aug":8,"Sep":9,"Oct":10,"Nov":11,"Dec":12}
                inv_map = {v: k for k, v in mon_map.items()}
                if f_compare == "LAST YEAR":
                    ly_key = f"{mon_abbr}-{yr-1}"
                else:
                    mnum = mon_map.get(mon_abbr[:3], 1)
                    if mnum == 1:
                        pm, py = 12, yr - 1
                    else:
                        pm, py = mnum - 1, yr
                    ly_key = f"{inv_map[pm]}-{py}"
                ly_sdf = base_a[base_a["_MonthKey"] == ly_key].copy()
            except Exception:
                pass

            st.caption(f"Filters: Depot={f_depot} | Month={f_month} | Compare={f_compare} ({ly_key or 'N/A'})")


            # ========== TABLE 1: Depot-wise full summary (DO/SC/SO/NO) ==========
            st.markdown('<hr style="margin:4px 0;border:none;border-top:1px solid #e2e8f0;">', unsafe_allow_html=True)
            st.markdown(f"#### Depot Wise Summary of SCHs and Services — {f_month}")

            col_route = find_col(["ROUTEE", "Routee", "routee"]) or find_col(["ROUTEE"])
            col_dtype = find_col(["D.TYPE", "DTYPE", "D.Type", "DutyType"])
            col_cond_mus = find_col(["COND MUSTERS", "CONDMUSTERS", "COND_MUSTERS", "ConductorMuster", "CND MUSTER", "COND MUSTER"])
            col_dri_mus = find_col(["DRI MUS", "DRIMUS", "DRI_MUS", "DriverMuster", "DRV MUSTER", "DRI MUSTERS", "DRI MUSTER"])
            # CND OT / DRV OT often have leading spaces in Excel headers e.g. " CND OT"
            col_cnd_ot = find_col(["CND OT", "CNDOT", "CND_OT", "ConductorOT", "CNDOT"])
            col_drv_ot = find_col(["DRV OT", "DRVOT", "DRV_OT", "DriverOT", "DRVOT"])
            if col_cnd_ot is None:
                for c in sched_raw.columns:
                    if str(c).strip().upper().replace(" ", "") in ("CNDOT", "CNDOUT"):
                        col_cnd_ot = c
                        break
            if col_drv_ot is None:
                for c in sched_raw.columns:
                    if str(c).strip().upper().replace(" ", "") in ("DRVOT", "DRVOUT"):
                        col_drv_ot = c
                        break
            col_muster3 = find_col(["3 DAYS MUSTER", "3DAYSMUSTER", "Muster3"])

            # ---- Table B extra filters: MHL/NMHL, ROUTE, PRODUCT ----
            st.markdown("##### Table B filters")
            bf1, bf2, bf3 = st.columns(3)
            with bf1:
                b_mhl_opts = ["ALL"] + sorted([x for x in sdf["_MHL"].dropna().unique() if x and str(x).lower() != "nan"])
                b_mhl = st.selectbox("MHL / NMHL", b_mhl_opts, key="sch_b_mhl")
            with bf2:
                col_route = find_col(["ROUTEE", "Routee", "routee"]) or find_col(["ROUTEE"])
                if col_route:
                    sdf["_ROUTE"] = sdf[col_route].astype(str).str.strip()
                    # month-scoped route list
                    route_base = sdf[sdf["_MonthKey"] == f_month] if f_month != "ALL" else sdf
                    b_route_opts = ["ALL"] + sorted([x for x in route_base["_ROUTE"].dropna().unique() if str(x).strip() and str(x).strip().lower() != "nan"])
                else:
                    b_route_opts = ["ALL"]
                b_route = st.selectbox("ROUTE", b_route_opts, key="sch_b_route")
            with bf3:
                prod_base = sdf[sdf["_MonthKey"] == f_month] if f_month != "ALL" else sdf
                if b_mhl != "ALL":
                    prod_base = prod_base[prod_base["_MHL"] == b_mhl]
                if col_route and b_route != "ALL":
                    prod_base = prod_base[prod_base.get("_ROUTE", prod_base.columns[0]) == b_route] if "_ROUTE" in prod_base.columns else prod_base
                b_prod_opts = ["ALL"] + sorted([x for x in prod_base["_PRODUCT"].dropna().unique() if x and str(x).lower() != "nan"])
                b_product = st.selectbox("PRODUCT", b_prod_opts, key="sch_b_product")

            # Source data for depot summary: month + shared depot + table B filters
            dep_src = sdf[sdf["_MonthKey"] == f_month].copy() if f_month != "ALL" else sdf.copy()
            if f_depot not in ("ALL", "REGION"):
                dep_src = dep_src[dep_src["_DEPOT"] == f_depot]
            if b_mhl != "ALL":
                dep_src = dep_src[dep_src["_MHL"] == b_mhl]
            if col_route and b_route != "ALL" and "_ROUTE" in dep_src.columns:
                dep_src = dep_src[dep_src["_ROUTE"] == b_route]
            if b_product != "ALL":
                dep_src = dep_src[dep_src["_PRODUCT"] == b_product]

            if len(dep_src) == 0:
                st.warning("No depot-wise data.")
            else:
                # Normalize D.TYPE to DO/SC/SO/NO
                if col_dtype:
                    dep_src = dep_src.copy()
                    dep_src["_DTYPE"] = dep_src[col_dtype].astype(str).str.strip().str.upper()
                    # map common variants
                    dep_src["_DTYPE"] = dep_src["_DTYPE"].replace({
                        "D.O": "DO", "D/O": "DO", "DO.": "DO",
                        "S.C": "SC", "S/C": "SC",
                        "S.O": "SO", "S/O": "SO",
                        "N.O": "NO", "N/O": "NO",
                    })
                else:
                    dep_src["_DTYPE"] = "OTHER"

                def parse_time_to_minutes(v):
                    """Convert HH:MM / H:MM / datetime.time / timedelta to total minutes."""
                    if v is None or (isinstance(v, float) and pd.isna(v)):
                        return 0.0
                    if isinstance(v, (int, float)):
                        # already numeric - treat as hours if small, or minutes if large
                        return float(v)
                    # datetime.time
                    try:
                        import datetime as _dt
                        if isinstance(v, _dt.time):
                            return v.hour * 60 + v.minute + v.second / 60.0
                        if isinstance(v, _dt.timedelta):
                            return v.total_seconds() / 60.0
                        if isinstance(v, _dt.datetime):
                            return v.hour * 60 + v.minute
                    except Exception:
                        pass
                    s = str(v).strip()
                    if not s or s.lower() in ("nan", "nat", "none", "null"):
                        return 0.0
                    # formats: 38:42, 38:42:00, 1:05
                    if ":" in s:
                        parts = s.split(":")
                        try:
                            h = int(parts[0])
                            m = int(parts[1]) if len(parts) > 1 else 0
                            sec = int(parts[2]) if len(parts) > 2 else 0
                            return h * 60 + m + sec / 60.0
                        except Exception:
                            return 0.0
                    try:
                        return float(s)
                    except Exception:
                        return 0.0

                def minutes_to_hhmm(mins):
                    if mins is None or mins == 0:
                        return ""
                    mins = int(round(float(mins)))
                    h, m = divmod(abs(mins), 60)
                    sign = "-" if mins < 0 else ""
                    return f"{sign}{h}:{m:02d}"

                for c, alias in [
                    (col_cond_mus, "_COND_MUS"),
                    (col_dri_mus, "_DRI_MUS"),
                ]:
                    if c:
                        dep_src[alias] = pd.to_numeric(dep_src[c], errors="coerce").fillna(0)
                    else:
                        dep_src[alias] = 0
                # OT columns are time (HH:MM) → minutes
                if col_cnd_ot:
                    dep_src["_CND_OT"] = dep_src[col_cnd_ot].map(parse_time_to_minutes)
                else:
                    dep_src["_CND_OT"] = 0
                if col_drv_ot:
                    dep_src["_DRV_OT"] = dep_src[col_drv_ot].map(parse_time_to_minutes)
                else:
                    dep_src["_DRV_OT"] = 0

                DTYPES = ["DO", "SC", "SO", "NO"]

                def side_metrics(sub):
                    schs = sub["_SCH"].sum()
                    ser = sub[col_svc].nunique()
                    kms = sub["_KMS"].sum()
                    cnd_req = sub["_COND_MUS"].sum() * 1.3
                    drv_req = sub["_DRI_MUS"].sum() * 1.3
                    cnd_ot = sub["_CND_OT"].sum()
                    drv_ot = sub["_DRV_OT"].sum()
                    return {
                        "SCHS": schs, "SER": ser, "SCH_KMS": kms,
                        "CND_REQ": cnd_req, "DRV_REQ": drv_req,
                        "CND_OT": cnd_ot, "DRV_OT": drv_ot,
                    }

                def build_side_row(dep, side, base):
                    """Build one REGION/RTC_HIRE row. TOTAL SER/SCHS/KMS = RTC+HIRE (Excel style)."""
                    rec = {"REGION": dep, "RTC_HIRE": side}
                    if side == "TOTAL":
                        m_rtc = side_metrics(base[base["_RTC"] == "RTC"])
                        m_hire = side_metrics(base[base["_RTC"] == "HIRE"])
                        # Additive totals matching Excel
                        rec["T_SCHS"] = m_rtc["SCHS"] + m_hire["SCHS"]
                        rec["T_SER"] = m_rtc["SER"] + m_hire["SER"]
                        rec["T_KMS"] = m_rtc["SCH_KMS"] + m_hire["SCH_KMS"]
                        rec["T_CND_REQ"] = m_rtc["CND_REQ"] + m_hire["CND_REQ"]
                        rec["T_DRV_REQ"] = m_rtc["DRV_REQ"] + m_hire["DRV_REQ"]
                        for dt in DTYPES:
                            sub_r = base[(base["_RTC"] == "RTC") & (base["_DTYPE"] == dt)]
                            sub_h = base[(base["_RTC"] == "HIRE") & (base["_DTYPE"] == dt)]
                            mr, mh = side_metrics(sub_r), side_metrics(sub_h)
                            rec[f"{dt}_SCHS"] = mr["SCHS"] + mh["SCHS"]
                            rec[f"{dt}_SER"] = mr["SER"] + mh["SER"]
                            rec[f"{dt}_KMS"] = mr["SCH_KMS"] + mh["SCH_KMS"]
                            rec[f"{dt}_CND_REQ"] = mr["CND_REQ"] + mh["CND_REQ"]
                            rec[f"{dt}_DRV_REQ"] = mr["DRV_REQ"] + mh["DRV_REQ"]
                            rec[f"{dt}_CND_OT"] = mr["CND_OT"] + mh["CND_OT"]
                            rec[f"{dt}_DRV_OT"] = mr["DRV_OT"] + mh["DRV_OT"]
                    else:
                        sub = base[base["_RTC"] == side]
                        tot_m = side_metrics(sub)
                        rec["T_SCHS"] = tot_m["SCHS"]
                        rec["T_SER"] = tot_m["SER"]
                        rec["T_KMS"] = tot_m["SCH_KMS"]
                        rec["T_CND_REQ"] = tot_m["CND_REQ"]
                        rec["T_DRV_REQ"] = tot_m["DRV_REQ"]
                        for dt in DTYPES:
                            dsub = sub[sub["_DTYPE"] == dt]
                            m = side_metrics(dsub)
                            rec[f"{dt}_SCHS"] = m["SCHS"]
                            rec[f"{dt}_SER"] = m["SER"]
                            rec[f"{dt}_KMS"] = m["SCH_KMS"]
                            rec[f"{dt}_CND_REQ"] = m["CND_REQ"]
                            rec[f"{dt}_DRV_REQ"] = m["DRV_REQ"]
                            rec[f"{dt}_CND_OT"] = m["CND_OT"]
                            rec[f"{dt}_DRV_OT"] = m["DRV_OT"]
                    return rec

                def fmt_num(v, is_time=False):
                    try:
                        fv = float(v)
                    except Exception:
                        return ""
                    if fv == 0:
                        return ""
                    if is_time:
                        # if stored as hours decimal or minutes - show as number for now
                        return f"{fv:,.0f}" if abs(fv - round(fv)) < 1e-6 else f"{fv:,.2f}"
                    return f"{fv:,.0f}" if abs(fv - round(fv)) < 1e-6 else f"{fv:,.1f}"

                rows_out = []
                depots = sorted([str(d).strip() for d in dep_src["_DEPOT"].dropna().unique() if str(d).strip() and str(d).strip().lower() != "nan"])
                for dep in depots:
                    dgrp = dep_src[dep_src["_DEPOT"] == dep]
                    for side in ["RTC", "HIRE", "TOTAL"]:
                        rows_out.append(build_side_row(dep, side, dgrp))

                # REGION totals = SUM of depot rows (Excel: REGION SER = sum of depot SERs)
                import copy
                depot_rows = [r for r in rows_out if r["REGION"] != "REGION"]
                for side in ["RTC", "HIRE", "TOTAL"]:
                    side_rows = [r for r in depot_rows if r["RTC_HIRE"] == side]
                    rec = {"REGION": "REGION", "RTC_HIRE": side}
                    if not side_rows:
                        rows_out.append(build_side_row("REGION", side, dep_src))
                        continue
                    # sum all numeric fields
                    keys = [k for k in side_rows[0].keys() if k not in ("REGION", "RTC_HIRE")]
                    for k in keys:
                        rec[k] = sum(float(r.get(k, 0) or 0) for r in side_rows)
                    rows_out.append(rec)

                dsum = pd.DataFrame(rows_out)

                # HTML - compact
                html2 = ['<div class="table-scroll"><table class="excel-table" style="table-layout:auto;font-size:9px;">']
                # header row 1
                html2.append("<tr>")
                html2.append('<th rowspan="2" style="background:#0f172a;color:white;padding:6px;">REGION</th>')
                html2.append('<th rowspan="2" style="background:#0f172a;color:white;padding:6px;">RTC/HIRE</th>')
                html2.append('<th colspan="3" style="background:#b91c1c;color:white;padding:6px;">DO</th>')
                html2.append('<th colspan="7" style="background:#7c3aed;color:white;padding:6px;">SC</th>')
                html2.append('<th colspan="5" style="background:#15803d;color:white;padding:6px;">SO</th>')
                html2.append('<th colspan="5" style="background:#0369a1;color:white;padding:6px;">NO</th>')
                html2.append('<th colspan="3" style="background:#c2410c;color:white;padding:6px;">TOTAL</th>')
                html2.append('<th colspan="2" style="background:#166534;color:white;padding:6px;">REQMT CREW</th>')
                html2.append("</tr>")
                # header row 2
                html2.append("<tr>")
                # DO: SCHS SER SCH KMS
                for h in ["SCHS", "SER", "SCH KMS"]:
                    html2.append(f'<th class="header-sub">{h}</th>')
                # SC: SCHS SER SCH KMS CND REQ DRV REQ CNDOT DRVOT
                for h in ["SCHS", "SER", "SCH KMS", "CND REQ", "DRV REQ", "CNDOT", "DRVOT"]:
                    html2.append(f'<th class="header-sub">{h}</th>')
                # SO
                for h in ["SCHS", "SER", "SCH KMS", "CND REQ", "DRV REQ"]:
                    html2.append(f'<th class="header-sub">{h}</th>')
                # NO
                for h in ["SCHS", "SER", "SCH KMS", "CND REQ", "DRV REQ"]:
                    html2.append(f'<th class="header-sub">{h}</th>')
                # TOTAL
                for h in ["SCHS", "SER", "SCH KMS"]:
                    html2.append(f'<th class="header-sub">{h}</th>')
                # REQMT
                for h in ["COND", "DRI"]:
                    html2.append(f'<th class="header-sub">{h}</th>')
                html2.append("</tr>")

                for _, r in dsum.iterrows():
                    is_tot = str(r["RTC_HIRE"]) == "TOTAL" or str(r["REGION"]) == "REGION"
                    style = "font-weight:bold;background:#e2efda;" if is_tot else ""
                    html2.append(f'<tr style="{style}">')
                    html2.append(f'<td>{r["REGION"]}</td>')
                    html2.append(f'<td>{r["RTC_HIRE"]}</td>')
                    def td(v, bg, is_ot=False):
                        val = minutes_to_hhmm(v) if is_ot else fmt_num(v)
                        return f'<td style="background:{bg};padding:4px;font-size:11px;">{val}</td>'
                    # DO
                    for k in ["DO_SCHS", "DO_SER", "DO_KMS"]:
                        html2.append(td(r.get(k, 0), "#fef2f2"))
                    # SC
                    for k in ["SC_SCHS", "SC_SER", "SC_KMS"]:
                        html2.append(td(r.get(k, 0), "#f5f3ff"))
                    for k in ["SC_CND_REQ", "SC_DRV_REQ"]:
                        html2.append(td(r.get(k, 0), "#eff6ff"))
                    html2.append(td(r.get("SC_CND_OT", 0), "#fff7ed", is_ot=True))
                    html2.append(td(r.get("SC_DRV_OT", 0), "#fff7ed", is_ot=True))
                    # SO
                    for k in ["SO_SCHS", "SO_SER", "SO_KMS"]:
                        html2.append(td(r.get(k, 0), "#f0fdf4"))
                    for k in ["SO_CND_REQ", "SO_DRV_REQ"]:
                        html2.append(td(r.get(k, 0), "#eff6ff"))
                    # NO
                    for k in ["NO_SCHS", "NO_SER", "NO_KMS"]:
                        html2.append(td(r.get(k, 0), "#fefce8"))
                    for k in ["NO_CND_REQ", "NO_DRV_REQ"]:
                        html2.append(td(r.get(k, 0), "#eff6ff"))
                    # TOTAL
                    for k in ["T_SCHS", "T_SER", "T_KMS"]:
                        html2.append(td(r.get(k, 0), "#e2e8f0"))
                    # REQMT crew
                    html2.append(td(r.get("T_CND_REQ", 0), "#dbeafe"))
                    html2.append(td(r.get("T_DRV_REQ", 0), "#dbeafe"))
                    html2.append("</tr>")
                html2.append("</table></div>")
                st.markdown("".join(html2), unsafe_allow_html=True)
                st.download_button(
                    "Download Depot-wise CSV",
                    dsum.to_csv(index=False).encode("utf-8"),
                    f"Schedules_DepotDetail_{f_month}.csv",
                    "text/csv",
                    key="dl9b",
                )


            st.markdown('<hr style="margin:4px 0;border:none;border-top:1px solid #e2e8f0;">', unsafe_allow_html=True)
            # ========== TABLE 2: Product-wise Summary ==========
            st.markdown("#### Product-wise Summary")

            # Labels for comparison columns
            if f_compare == "PREVIOUS MONTH":
                h_cy, h_ly, h_var = "CM", "PM", "VAR"
            else:
                h_cy, h_ly, h_var = "CY", "LY", "VAR"

            # Ensure OT / muster cols exist on cy_sdf / ly_sdf
            def ensure_ot_cols(data):
                data = data.copy()
                def parse_time_to_minutes(v):
                    if v is None or (isinstance(v, float) and pd.isna(v)):
                        return 0.0
                    if isinstance(v, (int, float)):
                        return float(v)
                    try:
                        import datetime as _dt
                        if isinstance(v, _dt.time):
                            return v.hour * 60 + v.minute + v.second / 60.0
                        if isinstance(v, _dt.timedelta):
                            return v.total_seconds() / 60.0
                        if isinstance(v, _dt.datetime):
                            return v.hour * 60 + v.minute
                    except Exception:
                        pass
                    s = str(v).strip()
                    if not s or s.lower() in ("nan", "nat", "none", "null"):
                        return 0.0
                    if ":" in s:
                        parts = s.split(":")
                        try:
                            h = int(parts[0]); m = int(parts[1]) if len(parts) > 1 else 0
                            return h * 60 + m
                        except Exception:
                            return 0.0
                    try:
                        return float(s)
                    except Exception:
                        return 0.0

                c_cond = find_col(["COND MUSTERS", "CONDMUSTERS", "COND MUSTER"])
                c_dri = find_col(["DRI MUS", "DRIMUS", "DRI MUSTER", "DRI MUSTERS"])
                c_cnd_ot = find_col(["CND OT", "CNDOT"])
                c_drv_ot = find_col(["DRV OT", "DRVOT"])
                if c_cnd_ot is None:
                    for c in data.columns:
                        if str(c).strip().upper().replace(" ", "") == "CNDOT":
                            c_cnd_ot = c; break
                if c_drv_ot is None:
                    for c in data.columns:
                        if str(c).strip().upper().replace(" ", "") == "DRVOT":
                            c_drv_ot = c; break
                data["_COND_MUS"] = pd.to_numeric(data[c_cond], errors="coerce").fillna(0) if c_cond else 0
                data["_DRI_MUS"] = pd.to_numeric(data[c_dri], errors="coerce").fillna(0) if c_dri else 0
                data["_CND_OT"] = data[c_cnd_ot].map(parse_time_to_minutes) if c_cnd_ot else 0
                data["_DRV_OT"] = data[c_drv_ot].map(parse_time_to_minutes) if c_drv_ot else 0
                return data

            cy_p = ensure_ot_cols(cy_sdf)
            ly_p = ensure_ot_cols(ly_sdf) if len(ly_sdf) else ly_sdf


            PRODUCT_ORDER = [
                "AC-SLP", "AC-HBD", "GRD+", "RJD", "e-GRD",
                "N-HBD H", "N-HBD R", "SLX", "DLX",
                "EXP-R", "ME", "EXP-H", "PVG-R", "PVG-H",
            ]

            def agg_product(data):
                metric_keys = []
                for side in ["RTC", "HIRE"]:
                    metric_keys += [f"SCH_{side}", f"SVC_{side}", f"KMS_{side}",
                                   f"CREW_COND_{side}", f"CREW_DRI_{side}",
                                   f"OT_COND_{side}", f"OT_DRI_{side}"]
                metric_keys += ["SCH_TOTAL", "SVC_TOTAL", "KMS_TOTAL"]
                base_cols = ["DEPOT", "PRODUCT", "MHL_NMHL"]
                if data is None or len(data) == 0:
                    return pd.DataFrame(columns=base_cols + metric_keys)
                rows = []
                if f_depot == "REGION":
                    group_keys = ["_PRODUCT", "_MHL"]
                else:
                    group_keys = ["_DEPOT", "_PRODUCT", "_MHL"]
                for keys, grp in data.groupby(group_keys):
                    if f_depot == "REGION":
                        if isinstance(keys, tuple):
                            prod, mhl = keys[0], keys[1]
                        else:
                            prod, mhl = keys, ""
                        dep = "REGION"
                    else:
                        dep, prod, mhl = keys
                    rec = {"DEPOT": dep, "PRODUCT": prod, "MHL_NMHL": mhl}
                    for side in ["RTC", "HIRE"]:
                        sub = grp[grp["_RTC"] == side]
                        rec[f"SCH_{side}"] = sub["_SCH"].sum()
                        rec[f"SVC_{side}"] = sub[col_svc].nunique()
                        rec[f"KMS_{side}"] = sub["_KMS"].sum()
                        rec[f"CREW_COND_{side}"] = sub["_COND_MUS"].sum() * 1.3
                        rec[f"CREW_DRI_{side}"] = sub["_DRI_MUS"].sum() * 1.3
                        rec[f"OT_COND_{side}"] = sub["_CND_OT"].sum()
                        rec[f"OT_DRI_{side}"] = sub["_DRV_OT"].sum()
                    rec["SCH_TOTAL"] = rec["SCH_RTC"] + rec["SCH_HIRE"]
                    rec["SVC_TOTAL"] = rec["SVC_RTC"] + rec["SVC_HIRE"]
                    rec["KMS_TOTAL"] = rec["KMS_RTC"] + rec["KMS_HIRE"]
                    rec["FLEET"] = 0.0
                    rows.append(rec)
                return pd.DataFrame(rows)

            def sort_products(df):
                if len(df) == 0:
                    return df
                order_map = {p: i for i, p in enumerate(PRODUCT_ORDER)}
                df = df.copy()
                df["_ord"] = df["PRODUCT"].map(lambda x: order_map.get(str(x).strip(), 500))
                df = df.sort_values(["DEPOT", "_ord", "PRODUCT"]).drop(columns=["_ord"])
                return df.reset_index(drop=True)


            cy_agg = agg_product(cy_p)
            ly_agg = agg_product(ly_p)

            _fmap, _ferr = load_fleet_map()

            def _month_key_list(mon_key):
                keys = []
                if mon_key:
                    keys.append(str(mon_key).strip())
                try:
                    dt = pd.to_datetime(str(mon_key), errors="coerce")
                    if pd.notna(dt):
                        keys += [
                            dt.strftime("%b-%Y"),
                            dt.strftime("%b-%y"),
                            f"{dt.strftime('%b')}-{dt.year}",
                        ]
                        prev = dt - pd.DateOffset(months=1)
                        keys += [prev.strftime("%b-%Y"), prev.strftime("%b-%y")]
                except Exception:
                    pass
                seen, out = set(), []
                for k in keys:
                    if k and k not in seen:
                        seen.add(k)
                        out.append(k)
                return out

            def _fleet_one(dep, prod, mon_key):
                """Per product fleet. REGION = sum across depots. Fallback previous month."""
                if _ferr or not _fmap:
                    return 0.0
                dep = str(dep).strip().upper()
                prod = str(prod).strip().upper()
                if prod in ("TOTAL", "MHL", "NMHL", ""):
                    return 0.0
                keys_try = _month_key_list(mon_key)
                by_dpm = _fmap.get("by_dpm", {})
                for k in keys_try:
                    if dep in ("REGION", "ALL", ""):
                        val = 0.0
                        for (d, pr, mk), v in by_dpm.items():
                            if str(pr).upper() == prod and str(mk) == k and str(d).upper() not in ("REGION", "TOTAL", ""):
                                val += float(v or 0)
                        if val:
                            return val
                    else:
                        val = by_dpm.get((dep, prod, k), 0) or 0
                        if val:
                            return float(val)
                return 0.0

            def _attach_fleet(adf, mon_key):
                if adf is None or len(adf) == 0:
                    return adf
                adf = adf.copy()
                if "FLEET" not in adf.columns:
                    adf["FLEET"] = 0.0
                for idx, row in adf.iterrows():
                    adf.at[idx, "FLEET"] = _fleet_one(
                        row.get("DEPOT", ""),
                        row.get("PRODUCT", ""),
                        mon_key,
                    )
                return adf

            cy_agg = _attach_fleet(cy_agg, f_month)
            ly_agg = _attach_fleet(ly_agg, ly_key if ly_key else "")
            metrics = [
                "FLEET",
                "SCH_RTC", "SCH_HIRE", "SCH_TOTAL",
                "SVC_RTC", "SVC_HIRE", "SVC_TOTAL",
                "KMS_RTC", "KMS_HIRE", "KMS_TOTAL",
                "CREW_COND_RTC", "CREW_DRI_RTC", "CREW_COND_HIRE",
                "OT_COND_RTC", "OT_DRI_RTC", "OT_COND_HIRE",
            ]

            def minutes_to_hhmm(mins):
                if mins is None:
                    return ""
                try:
                    mins = float(mins)
                except Exception:
                    return ""
                if mins == 0:
                    return ""
                mins = int(round(mins))
                h, m = divmod(abs(mins), 60)
                sign = "-" if mins < 0 else ""
                return f"{sign}{h}:{m:02d}"

            if len(cy_agg) == 0 and len(ly_agg) == 0:
                st.warning("No product-wise data for selection.")
            else:
                cy_agg = sort_products(cy_agg)
                ly_agg = sort_products(ly_agg)
                merged = cy_agg.merge(ly_agg, on=["DEPOT", "PRODUCT", "MHL_NMHL"], how="outer", suffixes=("_CY", "_LY"))
                if "MHL_NMHL" not in merged.columns:
                    merged["MHL_NMHL"] = ""
                merged["MHL_NMHL"] = merged["MHL_NMHL"].fillna("")
                for m in metrics:
                    for s in ("_CY", "_LY"):
                        c = f"{m}{s}"
                        if c not in merged.columns:
                            merged[c] = 0
                        merged[c] = pd.to_numeric(merged[c], errors="coerce").fillna(0)
                    merged[f"{m}_VAR"] = merged[f"{m}_CY"] - merged[f"{m}_LY"]
                merged = sort_products(merged)

                extra_rows = []
                dep_label = "REGION" if f_depot in ("ALL", "REGION") else (str(merged["DEPOT"].iloc[0]) if len(merged) else str(f_depot))
                # Only detail product rows (exclude any prior totals)
                detail = merged[~merged["PRODUCT"].astype(str).str.upper().isin(["TOTAL", "MHL", "NMHL"])].copy()
                # TOTAL = all products
                tot = {"DEPOT": dep_label, "PRODUCT": "TOTAL", "MHL_NMHL": ""}
                for m in metrics:
                    tot[f"{m}_CY"] = detail[f"{m}_CY"].sum()
                    tot[f"{m}_LY"] = detail[f"{m}_LY"].sum()
                    tot[f"{m}_VAR"] = tot[f"{m}_CY"] - tot[f"{m}_LY"]
                extra_rows.append(tot)
                # MHL = sum where MHL_NMHL is exactly MHL (not NMHL)
                # NMHL = sum where MHL_NMHL is exactly NMHL
                for label in ["MHL", "NMHL"]:
                    mask = detail["MHL_NMHL"].astype(str).str.strip().str.upper() == label
                    sub = detail[mask]
                    row = {"DEPOT": dep_label, "PRODUCT": label, "MHL_NMHL": label}
                    for m in metrics:
                        row[f"{m}_CY"] = sub[f"{m}_CY"].sum() if len(sub) else 0
                        row[f"{m}_LY"] = sub[f"{m}_LY"].sum() if len(sub) else 0
                        row[f"{m}_VAR"] = row[f"{m}_CY"] - row[f"{m}_LY"]
                    extra_rows.append(row)
                merged = pd.concat([detail, pd.DataFrame(extra_rows)], ignore_index=True)

                def cell_num(v, is_var=False, is_ot=False):
                    try:
                        fv = float(v)
                    except Exception:
                        return "<td style='padding:4px 5px;font-size:12px;'></td>"
                    if fv == 0:
                        return "<td style='padding:4px 5px;font-size:12px;'></td>"
                    if is_ot:
                        s = minutes_to_hhmm(fv)
                    else:
                        s = f"{fv:,.0f}" if abs(fv - round(fv)) < 1e-6 else f"{fv:,.1f}"
                    if is_var:
                        cls = "pos" if fv > 0 else "neg"
                        return f"<td class='{cls}' style='padding:4px 5px;font-size:12px;'>{s}</td>"
                    return f"<td style='padding:4px 5px;font-size:12px;'>{s}</td>"

                # No freeze - normal scroll, larger font & row height
                
                # ---- Build clean 3-row header + body ----
                # Column plan (after DEPOT, PRODUCT):
                # SCH: RTC×3, HIRE×3, TOTAL×3 = 9
                # SVC: 9
                # KMS: 9
                # CREW: RTC COND×3, RTC DRI×3, HIRE COND×3, HIRE DRI×3 = 12
                # OT: same 12
                # Total data cols = 9+9+9+12+12 = 51

                def th(text, colspan=1, rowspan=1, bg="#334155", extra=""):
                    return (
                        f'<th colspan="{colspan}" rowspan="{rowspan}" '
                        f'style="background:{bg};color:#fff;padding:6px 4px;font-size:11px;'
                        f'text-align:center;border:1px solid #94a3b8;{extra}">{text}</th>'
                    )

                def th_sub(text):
                    return (
                        f'<th style="background:#e2e8f0;color:#0f172a;padding:5px 4px;font-size:10px;'
                        f'text-align:center;border:1px solid #94a3b8;">{text}</th>'
                    )

                html = ['<div class="freeze-wrap"><table class="freeze-table"><thead>']

                # Row 1 – group titles
                html.append("<tr>")
                html.append(th("DEPOT", rowspan=3, bg="#0369a1", extra="position:sticky;left:0;z-index:6;min-width:70px;"))
                html.append(th("PRODUCT", rowspan=3, bg="#0369a1", extra="position:sticky;left:70px;z-index:6;min-width:80px;"))
                html.append(th("MHL/NMHL", rowspan=3, bg="#0369a1", extra="position:sticky;left:150px;z-index:6;min-width:60px;"))
                html.append(th("FLEET", colspan=3, bg="#0f766e"))
                html.append(th("SCHs", colspan=9, bg="#b91c1c"))
                html.append(th("SERVICES", colspan=9, bg="#a21caf"))
                html.append(th("SCH KMS", colspan=9, bg="#15803d"))
                html.append(th("CREW REQUIREMENT", colspan=9, bg="#1d4ed8"))
                html.append(th("SCH OVER TIME", colspan=9, bg="#c2410c"))
                html.append("</tr>")

                # Row 2 – RTC / HIRE / TOTAL (and COND/DRI under crew & OT)
                html.append("<tr>")
                html.append(th("TOTAL", colspan=3, bg="#0f766e"))  # FLEET
                for _ in range(3):  # SCH, SVC, KMS
                    html.append(th("RTC", colspan=3, bg="#475569"))
                    html.append(th("HIRE", colspan=3, bg="#475569"))
                    html.append(th("TOTAL", colspan=3, bg="#475569"))
                # CREW: RTC-COND, RTC-DRI, HIRE-COND (no HIRE DRI)
                html.append(th("RTC COND", colspan=3, bg="#1e40af"))
                html.append(th("RTC DRI", colspan=3, bg="#1e40af"))
                html.append(th("HIRE COND", colspan=3, bg="#1e40af"))
                # OT: RTC-COND, RTC-DRI, HIRE-COND (no HIRE DRI)
                html.append(th("RTC COND", colspan=3, bg="#9a3412"))
                html.append(th("RTC DRI", colspan=3, bg="#9a3412"))
                html.append(th("HIRE COND", colspan=3, bg="#9a3412"))
                html.append("</tr>")

                # Row 3 – CM/PM/VAR or CY/LY/VAR
                html.append("<tr>")
                for _ in range(16):  # 1 fleet + 9 (sch/svc/kms) + 3 crew + 3 ot
                    for h in (h_cy, h_ly, h_var):
                        html.append(th_sub(h))
                html.append("</tr>")
                html.append("</thead><tbody>")

                ot_metrics = {"OT_COND_RTC", "OT_DRI_RTC", "OT_COND_HIRE"}
                for _, row in merged.iterrows():
                    prod = str(row.get("PRODUCT", ""))
                    is_tot = prod in ("TOTAL", "MHL", "NMHL")
                    style = "font-weight:bold;background:#e2efda;" if is_tot else ""
                    html.append(f'<tr style="{style}">')
                    html.append(
                        f'<td style="position:sticky;left:0;z-index:2;background:#e0f2fe;'
                        f'padding:5px 6px;font-size:12px;font-weight:600;min-width:70px;">{row["DEPOT"]}</td>'
                    )
                    html.append(
                        f'<td style="position:sticky;left:70px;z-index:2;background:#f0f9ff;'
                        f'padding:5px 6px;font-size:12px;min-width:80px;">{row["PRODUCT"]}</td>'
                    )
                    html.append(
                        f'<td style="position:sticky;left:150px;z-index:2;background:#f8fafc;'
                        f'padding:5px 6px;font-size:11px;min-width:60px;">{row.get("MHL_NMHL", "")}</td>'
                    )
                    for m in metrics:
                        is_ot = m in ot_metrics
                        if m.startswith("SCH"):
                            bg = "#fef2f2"
                        elif m.startswith("SVC"):
                            bg = "#f5f3ff"
                        elif m.startswith("KMS"):
                            bg = "#f0fdf4"
                        elif m.startswith("CREW"):
                            bg = "#eff6ff"
                        elif m.startswith("OT"):
                            bg = "#fff7ed"
                        else:
                            bg = "#fff"
                        def cell_bg(v, is_var=False, is_ot=False, bg="#fff"):
                            try:
                                fv = float(v)
                            except Exception:
                                return f"<td style='padding:4px 5px;font-size:12px;background:{bg};'></td>"
                            if fv == 0:
                                return f"<td style='padding:4px 5px;font-size:12px;background:{bg};'></td>"
                            if is_ot:
                                s = minutes_to_hhmm(fv)
                            else:
                                s = f"{fv:,.0f}" if abs(fv - round(fv)) < 1e-6 else f"{fv:,.1f}"
                            if is_var:
                                cls = "pos" if fv > 0 else "neg"
                                return f"<td class='{cls}' style='padding:4px 5px;font-size:12px;background:{bg};'>{s}</td>"
                            return f"<td style='padding:4px 5px;font-size:12px;background:{bg};'>{s}</td>"
                        html.append(cell_bg(row[f"{m}_CY"], is_ot=is_ot, bg=bg))
                        html.append(cell_bg(row[f"{m}_LY"], is_ot=is_ot, bg=bg))
                        html.append(cell_bg(row[f"{m}_VAR"], is_var=True, is_ot=is_ot, bg=bg))
                    html.append("</tr>")
                html.append("</tbody></table></div>")

                st.markdown("".join(html), unsafe_allow_html=True)
                st.caption(
                    f"Headings: {h_cy}/{h_ly}/{h_var} | Crew REQ = MUSTERS×1.3 | OT = sum of CND OT / DRV OT (HH:MM)"
                )
                st.download_button(
                    "Download Product-wise CSV",
                    merged.to_csv(index=False).encode("utf-8"),
                    f"Schedules_Product_{f_month}.csv",
                    "text/csv",
                    key="dl9a",
                )


            # ========== TABLE 3: Summary of Operation ==========
            st.markdown('<hr style="margin:4px 0;border:none;border-top:1px solid #e2e8f0;">', unsafe_allow_html=True)
            st.markdown(f"#### Summary of Operation — {f_month}")

            # Reuse Table B filtered source (dep_src) if available, else rebuild
            try:
                op_src = dep_src.copy()
            except NameError:
                op_src = sdf[sdf["_MonthKey"] == f_month].copy() if f_month != "ALL" else sdf.copy()
                if f_depot not in ("ALL", "REGION"):
                    op_src = op_src[op_src["_DEPOT"] == f_depot]

            col_route = find_col(["ROUTEE", "Routee", "routee"]) or find_col(["ROUTEE"])
            if col_route and "_ROUTE" not in op_src.columns:
                op_src["_ROUTE"] = op_src[col_route].astype(str).str.strip()
            elif "_ROUTE" not in op_src.columns:
                op_src["_ROUTE"] = ""

            _dtype_col = find_col(["D.TYPE", "DTYPE", "D.Type", "DutyType"])
            if _dtype_col:
                op_src["_DTYPE"] = op_src[_dtype_col].astype(str).str.strip().str.upper()
                op_src["_DTYPE"] = op_src["_DTYPE"].replace({
                    "D.O": "DO", "D/O": "DO", "DO.": "DO",
                    "S.C": "SC", "S/C": "SC",
                    "S.O": "SO", "S/O": "SO",
                    "N.O": "NO", "N/O": "NO",
                })
            else:
                op_src["_DTYPE"] = ""

            DTYPES3 = ["DO", "NO", "SC", "SO"]

            # Crew + OT columns for Table C
            def _parse_ot_min(v):
                if v is None or (isinstance(v, float) and pd.isna(v)):
                    return 0.0
                if isinstance(v, (int, float)):
                    return float(v)
                try:
                    import datetime as _dt
                    if isinstance(v, _dt.time):
                        return v.hour * 60 + v.minute
                    if isinstance(v, _dt.timedelta):
                        return v.total_seconds() / 60.0
                    if isinstance(v, _dt.datetime):
                        return v.hour * 60 + v.minute
                except Exception:
                    pass
                s = str(v).strip()
                if not s or s.lower() in ("nan", "nat", "none", "null"):
                    return 0.0
                if ":" in s:
                    parts = s.split(":")
                    try:
                        return int(parts[0]) * 60 + int(parts[1])
                    except Exception:
                        return 0.0
                try:
                    return float(s)
                except Exception:
                    return 0.0

            def _min_to_hhmm(mins):
                try:
                    mins = int(round(float(mins)))
                except Exception:
                    return ""
                if mins == 0:
                    return ""
                h, m = divmod(abs(mins), 60)
                return f"{'-' if mins < 0 else ''}{h}:{m:02d}"

            c_cond = find_col(["COND MUSTERS", "CONDMUSTERS", "COND MUSTER"])
            c_dri = find_col(["DRI MUS", "DRIMUS", "DRI MUSTER", "DRI MUSTERS"])
            c_cnd_ot = find_col(["CND OT", "CNDOT"])
            c_drv_ot = find_col(["DRV OT", "DRVOT"])
            if c_cnd_ot is None:
                for c in op_src.columns:
                    if str(c).strip().upper().replace(" ", "") == "CNDOT":
                        c_cnd_ot = c
                        break
            if c_drv_ot is None:
                for c in op_src.columns:
                    if str(c).strip().upper().replace(" ", "") == "DRVOT":
                        c_drv_ot = c
                        break
            op_src = op_src.copy()
            op_src["_COND_MUS"] = pd.to_numeric(op_src[c_cond], errors="coerce").fillna(0) if c_cond else 0
            op_src["_DRI_MUS"] = pd.to_numeric(op_src[c_dri], errors="coerce").fillna(0) if c_dri else 0
            op_src["_CND_OT"] = op_src[c_cnd_ot].map(_parse_ot_min) if c_cnd_ot else 0
            op_src["_DRV_OT"] = op_src[c_drv_ot].map(_parse_ot_min) if c_drv_ot else 0

            if len(op_src) == 0:
                st.warning("No data for Summary of Operation.")
            else:
                rows3 = []
                sno = 1
                group_cols = ["_DEPOT", "_ROUTE", "_PRODUCT"]
                for keys, grp in op_src.groupby(group_cols, dropna=False):
                    dep, route, prod = keys
                    svcs = sorted({str(x).strip() for x in grp[col_svc].dropna().unique() if str(x).strip() and str(x).lower() != "nan"})
                    svc_str = ",".join(svcs) + ("," if svcs else "")
                    rec = {
                        "SNo": sno,
                        "DEPOT": dep,
                        "ROUTE": route,
                        "PRODUCT": prod,
                        "SERVICE_NUMBERS": svc_str,
                    }
                    for dt in DTYPES3:
                        sub = grp[grp["_DTYPE"] == dt]
                        rec[f"SCH_{dt}"] = sub["_SCH"].sum()
                        rec[f"SER_{dt}"] = sub[col_svc].nunique()
                        rec[f"KMS_{dt}"] = sub["_KMS"].sum()
                    rec["SCH_TOTAL"] = sum(rec[f"SCH_{dt}"] for dt in DTYPES3)
                    rec["SER_TOTAL"] = sum(rec[f"SER_{dt}"] for dt in DTYPES3)
                    rec["KMS_TOTAL"] = sum(rec[f"KMS_{dt}"] for dt in DTYPES3)
                    # Crew requirement (musters x 1.3) and Schedule OT
                    rec["CREW_COND"] = grp["_COND_MUS"].sum() * 1.3
                    rec["CREW_DRI"] = grp["_DRI_MUS"].sum() * 1.3
                    rec["OT_COND"] = grp["_CND_OT"].sum()
                    rec["OT_DRI"] = grp["_DRV_OT"].sum()
                    rows3.append(rec)
                    sno += 1

                df3 = pd.DataFrame(rows3)

                # TOTAL row
                if len(df3) > 0:
                    tot = {"SNo": "", "DEPOT": "TOTAL", "ROUTE": "", "PRODUCT": "", "SERVICE_NUMBERS": ""}
                    for dt in DTYPES3:
                        tot[f"SCH_{dt}"] = df3[f"SCH_{dt}"].sum()
                        tot[f"SER_{dt}"] = df3[f"SER_{dt}"].sum()
                        tot[f"KMS_{dt}"] = df3[f"KMS_{dt}"].sum()
                    tot["SCH_TOTAL"] = df3["SCH_TOTAL"].sum()
                    tot["SER_TOTAL"] = df3["SER_TOTAL"].sum()
                    tot["KMS_TOTAL"] = df3["KMS_TOTAL"].sum()
                    tot["CREW_COND"] = df3["CREW_COND"].sum()
                    tot["CREW_DRI"] = df3["CREW_DRI"].sum()
                    tot["OT_COND"] = df3["OT_COND"].sum()
                    tot["OT_DRI"] = df3["OT_DRI"].sum()
                    df3 = pd.concat([df3, pd.DataFrame([tot])], ignore_index=True)

                def n3(v):
                    try:
                        fv = float(v)
                        return "" if fv == 0 else (f"{fv:,.0f}" if abs(fv - round(fv)) < 1e-6 else f"{fv:,.1f}")
                    except Exception:
                        return ""

                html3 = ['<div class="op-wrap"><table class="op-table"><thead>']
                html3.append("<tr>")
                html3.append('<th rowspan="2" style="position:sticky;left:0;top:0;z-index:3;background:#0f172a;color:white;padding:8px 6px;font-size:13px;min-width:40px;">S.No</th>')
                html3.append('<th rowspan="2" style="position:sticky;left:40px;top:0;z-index:3;background:#0f172a;color:white;padding:8px 6px;font-size:13px;min-width:60px;">DEPOT</th>')
                html3.append('<th rowspan="2" style="position:sticky;left:100px;top:0;z-index:3;background:#0f172a;color:white;padding:8px 6px;font-size:13px;min-width:80px;">ROUTE NO.</th>')
                html3.append('<th rowspan="2" style="position:sticky;left:180px;top:0;z-index:3;background:#0f172a;color:white;padding:8px 6px;font-size:13px;min-width:70px;">PRODUCT</th>')
                html3.append('<th rowspan="2" style="position:sticky;left:250px;top:0;z-index:3;background:#0f172a;color:white;padding:8px 6px;font-size:13px;min-width:190px;width:190px;">SERVICE NUMBERS</th>')
                html3.append('<th colspan="5" style="position:sticky;top:0;z-index:4;background:#b91c1c;color:white;padding:8px;font-size:13px;">NO OF SCHEDULES</th>')
                html3.append('<th colspan="5" style="position:sticky;top:0;z-index:4;background:#7c3aed;color:white;padding:8px;font-size:13px;">SERVICES</th>')
                html3.append('<th colspan="5" style="position:sticky;top:0;z-index:4;background:#15803d;color:white;padding:8px;font-size:13px;">SCHEDULE KILOMETERS</th>')
                html3.append('<th colspan="2" style="position:sticky;top:0;z-index:4;background:#1d4ed8;color:white;padding:8px;font-size:13px;">CREW REQUIREMENT</th>')
                html3.append('<th colspan="2" style="position:sticky;top:0;z-index:4;background:#c2410c;color:white;padding:8px;font-size:13px;">SCH OVER TIME</th>')
                html3.append("</tr><tr>")
                for bg in ("#fecaca", "#ddd6fe", "#bbf7d0"):
                    for h in ["DO", "NO", "SC", "SO", "TOTAL"]:
                        html3.append(f'<th style="position:sticky;top:36px;z-index:4;background:{bg};color:#0f172a;padding:6px;font-size:12px;">{h}</th>')
                for h in ["COND", "DRI"]:
                    html3.append(f'<th style="position:sticky;top:36px;z-index:4;background:#dbeafe;color:#0f172a;padding:6px;font-size:12px;">{h}</th>')
                for h in ["COND", "DRI"]:
                    html3.append(f'<th style="position:sticky;top:36px;z-index:4;background:#ffedd5;color:#0f172a;padding:6px;font-size:12px;">{h}</th>')
                html3.append("</tr></thead><tbody>")

                sticky_base = 'position:sticky;z-index:2;padding:6px;font-size:13px;'
                lefts = [0, 45, 110, 200, 280]  # approx sticky offsets for 5 cols
                for _, r in df3.iterrows():
                    is_tot = str(r["DEPOT"]) == "TOTAL"
                    style = "font-weight:bold;background:#e2efda;" if is_tot else ""
                    html3.append(f'<tr style="{style}">')
                    bg0 = "#e2efda" if is_tot else "#e0f2fe"
                    html3.append(f'<td style="{sticky_base}left:0;background:{bg0};min-width:40px;">{r["SNo"]}</td>')
                    html3.append(f'<td style="{sticky_base}left:40px;background:{bg0};min-width:60px;">{r["DEPOT"]}</td>')
                    html3.append(f'<td style="{sticky_base}left:100px;background:{bg0};min-width:80px;">{r["ROUTE"]}</td>')
                    html3.append(f'<td style="{sticky_base}left:180px;background:{bg0};min-width:70px;">{r["PRODUCT"]}</td>')
                    html3.append(f'<td style="{sticky_base}left:250px;background:{bg0};min-width:70px;max-width:90px;text-align:left;white-space:normal;font-size:12px;">{r["SERVICE_NUMBERS"]}</td>')
                    for dt in DTYPES3:
                        html3.append(f'<td style="padding:6px;font-size:13px;background:#fef2f2;">{n3(r[f"SCH_{dt}"])}</td>')
                    html3.append(f'<td style="padding:6px;font-size:13px;background:#fee2e2;font-weight:600;">{n3(r["SCH_TOTAL"])}</td>')
                    for dt in DTYPES3:
                        html3.append(f'<td style="padding:6px;font-size:13px;background:#f5f3ff;">{n3(r[f"SER_{dt}"])}</td>')
                    html3.append(f'<td style="padding:6px;font-size:13px;background:#ede9fe;font-weight:600;">{n3(r["SER_TOTAL"])}</td>')
                    for dt in DTYPES3:
                        html3.append(f'<td style="padding:6px;font-size:13px;background:#f0fdf4;">{n3(r[f"KMS_{dt}"])}</td>')
                    html3.append(f'<td style="padding:6px;font-size:13px;background:#dcfce7;font-weight:600;">{n3(r["KMS_TOTAL"])}</td>')
                    html3.append(f'<td style="padding:6px;font-size:13px;background:#eff6ff;">{n3(r.get("CREW_COND", 0))}</td>')
                    html3.append(f'<td style="padding:6px;font-size:13px;background:#eff6ff;">{n3(r.get("CREW_DRI", 0))}</td>')
                    html3.append(f'<td style="padding:6px;font-size:13px;background:#fff7ed;">{_min_to_hhmm(r.get("OT_COND", 0))}</td>')
                    html3.append(f'<td style="padding:6px;font-size:13px;background:#fff7ed;">{_min_to_hhmm(r.get("OT_DRI", 0))}</td>')
                    html3.append("</tr>")
                html3.append("</tbody></table></div>")
                st.markdown("".join(html3), unsafe_allow_html=True)

                def summary_operation_excel(mdf, mon):
                    """Excel matching on-screen Summary of Operation board."""
                    from openpyxl import Workbook
                    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                    from openpyxl.utils import get_column_letter
                    wb = Workbook()
                    ws = wb.active
                    ws.title = "Summary of Operation"
                    thin = Border(
                        left=Side(style="thin", color="94A3B8"),
                        right=Side(style="thin", color="94A3B8"),
                        top=Side(style="thin", color="94A3B8"),
                        bottom=Side(style="thin", color="94A3B8"),
                    )
                    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    left_al = Alignment(horizontal="left", vertical="center", wrap_text=True)
                    white = Font(bold=True, color="FFFFFF", size=11)
                    dark = Font(bold=True, color="0F172A", size=10)
                    fills = {
                        "left": PatternFill("solid", fgColor="0F172A"),
                        "sch": PatternFill("solid", fgColor="B91C1C"),
                        "ser": PatternFill("solid", fgColor="7C3AED"),
                        "kms": PatternFill("solid", fgColor="15803D"),
                        "crew": PatternFill("solid", fgColor="1D4ED8"),
                        "ot": PatternFill("solid", fgColor="C2410C"),
                        "sch_sub": PatternFill("solid", fgColor="FECACA"),
                        "ser_sub": PatternFill("solid", fgColor="DDD6FE"),
                        "kms_sub": PatternFill("solid", fgColor="BBF7D0"),
                        "crew_sub": PatternFill("solid", fgColor="DBEAFE"),
                        "ot_sub": PatternFill("solid", fgColor="FFEDD5"),
                        "tot": PatternFill("solid", fgColor="E2EFDA"),
                        "sch_bg": PatternFill("solid", fgColor="FEF2F2"),
                        "ser_bg": PatternFill("solid", fgColor="F5F3FF"),
                        "kms_bg": PatternFill("solid", fgColor="F0FDF4"),
                        "crew_bg": PatternFill("solid", fgColor="EFF6FF"),
                        "ot_bg": PatternFill("solid", fgColor="FFF7ED"),
                    }

                    # Title
                    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=24)
                    tcell = ws.cell(1, 1, f"Summary of Operation | {mon}")
                    tcell.font = Font(bold=True, size=13, color="1E3A8A")
                    tcell.alignment = Alignment(horizontal="left", vertical="center")
                    ws.row_dimensions[1].height = 22

                    def th(r, c, text, fill, rowspan=1, colspan=1, font=None):
                        cell = ws.cell(r, c, text)
                        cell.fill = fill
                        cell.font = font or white
                        cell.alignment = center
                        cell.border = thin
                        if rowspan > 1 or colspan > 1:
                            ws.merge_cells(
                                start_row=r, start_column=c,
                                end_row=r + rowspan - 1, end_column=c + colspan - 1,
                            )
                        for rr in range(r, r + rowspan):
                            for cc in range(c, c + colspan):
                                ws.cell(rr, cc).fill = fill
                                ws.cell(rr, cc).border = thin
                                ws.cell(rr, cc).alignment = center

                    # Header row 2 – groups
                    th(2, 1, "S.No", fills["left"], rowspan=2)
                    th(2, 2, "DEPOT", fills["left"], rowspan=2)
                    th(2, 3, "ROUTE NO.", fills["left"], rowspan=2)
                    th(2, 4, "PRODUCT", fills["left"], rowspan=2)
                    th(2, 5, "SERVICE NUMBERS", fills["left"], rowspan=2)
                    th(2, 6, "NO OF SCHEDULES", fills["sch"], colspan=5)
                    th(2, 11, "SERVICES", fills["ser"], colspan=5)
                    th(2, 16, "SCHEDULE KILOMETERS", fills["kms"], colspan=5)
                    th(2, 21, "CREW REQUIREMENT", fills["crew"], colspan=2)
                    th(2, 23, "SCH OVER TIME", fills["ot"], colspan=2)

                    # Header row 3 – sub
                    for i, h in enumerate(["DO", "NO", "SC", "SO", "TOTAL"]):
                        th(3, 6 + i, h, fills["sch_sub"], font=dark)
                    for i, h in enumerate(["DO", "NO", "SC", "SO", "TOTAL"]):
                        th(3, 11 + i, h, fills["ser_sub"], font=dark)
                    for i, h in enumerate(["DO", "NO", "SC", "SO", "TOTAL"]):
                        th(3, 16 + i, h, fills["kms_sub"], font=dark)
                    th(3, 21, "COND", fills["crew_sub"], font=dark)
                    th(3, 22, "DRI", fills["crew_sub"], font=dark)
                    th(3, 23, "COND", fills["ot_sub"], font=dark)
                    th(3, 24, "DRI", fills["ot_sub"], font=dark)

                    def put_num(cell, v, is_ot=False):
                        try:
                            fv = float(v or 0)
                        except Exception:
                            fv = 0.0
                        if abs(fv) < 1e-9:
                            cell.value = None
                            return
                        if is_ot:
                            cell.value = _min_to_hhmm(fv)
                        else:
                            cell.value = int(round(fv)) if abs(fv - round(fv)) < 1e-6 else round(fv, 1)
                        cell.alignment = center

                    data_r = 4
                    for _, r in mdf.iterrows():
                        is_tot = str(r.get("DEPOT", "")) == "TOTAL"
                        row_fill = fills["tot"] if is_tot else None
                        bold = Font(bold=True, size=10) if is_tot else Font(size=10)

                        vals_left = [
                            r.get("SNo", ""),
                            r.get("DEPOT", ""),
                            r.get("ROUTE", ""),
                            r.get("PRODUCT", ""),
                            r.get("SERVICE_NUMBERS", ""),
                        ]
                        for c, v in enumerate(vals_left, 1):
                            cell = ws.cell(data_r, c, "" if (v is None or str(v) == "nan") else v)
                            cell.border = thin
                            cell.font = bold
                            cell.alignment = left_al if c == 5 else center
                            if row_fill:
                                cell.fill = row_fill
                            elif c <= 5:
                                cell.fill = PatternFill("solid", fgColor="E0F2FE")

                        # SCH DO NO SC SO TOTAL
                        for i, dt in enumerate(DTYPES3):
                            cell = ws.cell(data_r, 6 + i)
                            cell.border = thin
                            cell.fill = row_fill or fills["sch_bg"]
                            cell.font = bold
                            put_num(cell, r.get(f"SCH_{dt}", 0))
                        cell = ws.cell(data_r, 10)
                        cell.border = thin
                        cell.fill = row_fill or PatternFill("solid", fgColor="FEE2E2")
                        cell.font = Font(bold=True, size=10)
                        put_num(cell, r.get("SCH_TOTAL", 0))

                        # SER
                        for i, dt in enumerate(DTYPES3):
                            cell = ws.cell(data_r, 11 + i)
                            cell.border = thin
                            cell.fill = row_fill or fills["ser_bg"]
                            cell.font = bold
                            put_num(cell, r.get(f"SER_{dt}", 0))
                        cell = ws.cell(data_r, 15)
                        cell.border = thin
                        cell.fill = row_fill or PatternFill("solid", fgColor="EDE9FE")
                        cell.font = Font(bold=True, size=10)
                        put_num(cell, r.get("SER_TOTAL", 0))

                        # KMS
                        for i, dt in enumerate(DTYPES3):
                            cell = ws.cell(data_r, 16 + i)
                            cell.border = thin
                            cell.fill = row_fill or fills["kms_bg"]
                            cell.font = bold
                            put_num(cell, r.get(f"KMS_{dt}", 0))
                        cell = ws.cell(data_r, 20)
                        cell.border = thin
                        cell.fill = row_fill or PatternFill("solid", fgColor="DCFCE7")
                        cell.font = Font(bold=True, size=10)
                        put_num(cell, r.get("KMS_TOTAL", 0))

                        # CREW
                        for c, key in ((21, "CREW_COND"), (22, "CREW_DRI")):
                            cell = ws.cell(data_r, c)
                            cell.border = thin
                            cell.fill = row_fill or fills["crew_bg"]
                            cell.font = bold
                            put_num(cell, r.get(key, 0))
                        # OT
                        for c, key in ((23, "OT_COND"), (24, "OT_DRI")):
                            cell = ws.cell(data_r, c)
                            cell.border = thin
                            cell.fill = row_fill or fills["ot_bg"]
                            cell.font = bold
                            put_num(cell, r.get(key, 0), is_ot=True)

                        data_r += 1

                    widths = {
                        "A": 6, "B": 10, "C": 12, "D": 10, "E": 28,
                    }
                    for col in range(6, 25):
                        widths[get_column_letter(col)] = 9
                    for letter, w in widths.items():
                        ws.column_dimensions[letter].width = w
                    ws.row_dimensions[2].height = 20
                    ws.row_dimensions[3].height = 18
                    ws.freeze_panes = "F4"
                    bio = BytesIO()
                    wb.save(bio)
                    bio.seek(0)
                    return bio.getvalue()

                st.download_button(
                    "Download Summary of Operation Excel",
                    summary_operation_excel(df3, f_month),
                    f"Summary_Operation_{f_month}.xlsx",
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="dl9c",
                )


# ============================================================================
# MISSION RR — modification schedules (upload master list + metrics from ser_wise)
# Fixed periods: 01.04.2026–31.07.2026  |  01.08.2026–till date
# Filters: Depot / Product / Route only
# ============================================================================
elif section == "MISSION RR":
    st.markdown(
        '<div class="title-bar">MISSION RR — Modification Schedules</div>',
        unsafe_allow_html=True,
    )
    _src = df_ser_wise if (df_ser_wise is not None and len(df_ser_wise)) else df
    if _src is None or len(_src) == 0:
        st.error("ser_wise.parquet not available.")
        st.stop()
    mrr = _src.copy()
    mrr["Date"] = pd.to_datetime(mrr["Date"], errors="coerce")
    mrr = mrr.dropna(subset=["Date"])

    for std, alts in {
        "GE_TOT": ["Gross Total", "GE_TOT"],
        "NE_TOT": ["Net Total", "NE_TOT"],
        "Optd_KMs": ["OPD_KMS", "Optd_KMs"],
        "DEPOT": ["DEPOT"],
        "SER_NO": ["SER_NO", "ServiceNo"],
        "PRODUCT": ["PRODUCT"],
        "ROUTEE": ["ROUTEE", "ROUTE"],
    }.items():
        if std not in mrr.columns:
            src = next((a for a in alts if a in mrr.columns), None)
            if src is not None:
                mrr[std] = pd.to_numeric(mrr[src], errors="coerce").fillna(0) if std in ("GE_TOT", "NE_TOT", "Optd_KMs") else mrr[src]
            else:
                mrr[std] = 0 if std in ("GE_TOT", "NE_TOT", "Optd_KMs") else ""
    if "ROUTEE" not in mrr.columns and "ROUTE" in mrr.columns:
        mrr["ROUTEE"] = mrr["ROUTE"]

    # Fixed periods
    _p1_start, _p1_end = pd.Timestamp("2026-04-01"), pd.Timestamp("2026-07-31")
    _p2_start = pd.Timestamp("2026-08-01")
    _till = mrr["Date"].max()
    _p2_end = _till if pd.notna(_till) and _till >= _p2_start else pd.Timestamp("2026-08-31")
    _p1_label = "01.04.2026 TO 31.07.2026"
    _p2_label = f"01.08.2026 to TILL DATE ({_p2_end.strftime('%d.%m.%Y')})"
    title = "DETAILS OF THE RR REGION MODIFICATION SCHEDULES FOR AUG - 2026"

    def _norm_svc(v):
        if v is None or (isinstance(v, float) and np.isnan(v)):
            return ""
        s = str(v).strip()
        if not s or s.lower() in ("nan", "none"):
            return ""
        try:
            return str(int(float(s)))
        except Exception:
            return s

    def _clean_str(v):
        if v is None or (isinstance(v, float) and np.isnan(v)):
            return ""
        s = str(v).strip()
        return "" if s.lower() in ("nan", "none", "") else s

    # ---- Robust Excel reader ----
    def _read_mod_excel(uploaded):
        """Return DataFrame of services with standardized columns.
        Handles multi-row headers, merged titles, and varied column names.
        """
        raw = uploaded.read()
        uploaded.seek(0)
        bio = BytesIO(raw)
        # Try every sheet
        try:
            xl = pd.ExcelFile(bio)
            sheet_names = xl.sheet_names
        except Exception:
            sheet_names = [0]
            bio = BytesIO(raw)

        best = None
        best_score = -1
        debug_cols = []

        for sheet in sheet_names:
            for hdr in range(0, 8):
                try:
                    bio.seek(0)
                    df_try = pd.read_excel(bio, sheet_name=sheet, header=hdr)
                except Exception:
                    continue
                if df_try is None or len(df_try) == 0:
                    continue
                # Drop fully empty columns
                df_try = df_try.dropna(axis=1, how="all")
                cols_raw = [str(c).strip() for c in df_try.columns]
                debug_cols.append((sheet, hdr, cols_raw[:20]))

                # Normalize column names for matching
                def _nk(c):
                    return (
                        str(c).strip().lower()
                        .replace(" ", "").replace("_", "").replace(".", "")
                        .replace("/", "").replace("\n", "").replace("\r", "")
                    )

                colmap = {_nk(c): c for c in df_try.columns}

                def find(*cands):
                    for cand in cands:
                        k = _nk(cand)
                        if k in colmap:
                            return colmap[k]
                        # partial contains
                        for nk, orig in colmap.items():
                            if k and (k in nk or nk in k):
                                return orig
                    return None

                c_dep = find("DEPOT", "Depot", "depot")
                c_ser = find("SER NO", "SER_NO", "SERNO", "ServiceNo", "SERVICE NO", "SERVICENO", "SER.NO", "SNO", "Service")
                c_route = find("ROUTE", "ROUTEE", "Route")
                c_prod = find("PRODUCT", "PRODUCT T", "PRODUCTT", "Product", "PROD")
                c_bef = find("BEFORE", "BEF", "BEFORE E", "Before")
                c_aft = find("AFTER", "AFT", "AFTER R", "After")
                c_sers = find("NO. OF SERS", "NO_OF_SERS", "NO.OFSERS", "NOOFSERS", "SERS", "No of Sers")

                score = sum(1 for x in [c_dep, c_ser, c_route, c_prod] if x is not None)
                if score > best_score:
                    best_score = score
                    best = (df_try, c_dep, c_ser, c_route, c_prod, c_bef, c_aft, c_sers, cols_raw, sheet, hdr)

                if score >= 3:
                    break
            if best_score >= 3:
                break

        if best is None or best_score < 1:
            # Positional fallback: assume image layout
            # SL NO | DEPOT | SER NO | ROUTE | PRODUCT | BEFORE | AFTER | NO. OF SERS
            bio.seek(0)
            try:
                df_pos = pd.read_excel(bio, header=None)
            except Exception:
                return None, debug_cols, "Could not read Excel"
            # Find first row that looks like a header or data
            start = 0
            for i in range(min(15, len(df_pos))):
                row_vals = [str(x).strip().upper() for x in df_pos.iloc[i].tolist()]
                joined = " ".join(row_vals)
                if "DEPOT" in joined and ("SER" in joined or "SERVICE" in joined):
                    start = i + 1
                    # remap using this header row
                    hdr_vals = [str(x).strip() for x in df_pos.iloc[i].tolist()]
                    df_pos2 = df_pos.iloc[start:].copy()
                    df_pos2.columns = [f"C{j}" if not h or h.lower() == "nan" else h for j, h in enumerate(hdr_vals)]
                    return _read_mod_excel_from_df(df_pos2), debug_cols, f"positional-header-row-{i}"
                # data row: depot-like short code + service number
                if len(row_vals) >= 5 and row_vals[1] and row_vals[2]:
                    start = i
                    break
            df_pos = df_pos.iloc[start:].copy()
            ncols = df_pos.shape[1]
            rows_out = []
            for _, r in df_pos.iterrows():
                vals = [r.iloc[j] if j < ncols else "" for j in range(ncols)]
                # skip blank
                if all(_clean_str(v) == "" for v in vals[:5]):
                    continue
                # if first col is SL number, shift
                offset = 0
                try:
                    int(float(vals[0]))
                    offset = 1
                except Exception:
                    offset = 0
                rows_out.append({
                    "DEPOT": _clean_str(vals[offset]).upper() if offset < ncols else "",
                    "SER NO": _norm_svc(vals[offset + 1]) if offset + 1 < ncols else "",
                    "ROUTE": _clean_str(vals[offset + 2]) if offset + 2 < ncols else "",
                    "PRODUCT": _clean_str(vals[offset + 3]) if offset + 3 < ncols else "",
                    "BEFORE": _clean_str(vals[offset + 4]).upper() if offset + 4 < ncols else "",
                    "AFTER": _clean_str(vals[offset + 5]).upper() if offset + 5 < ncols else "",
                    "NO. OF SERS": 1,
                })
            if rows_out:
                return pd.DataFrame(rows_out), debug_cols, "positional"
            return None, debug_cols, "no data parsed"

        df_try, c_dep, c_ser, c_route, c_prod, c_bef, c_aft, c_sers, cols_raw, sheet, hdr = best
        rows_out = []
        for _, r in df_try.iterrows():
            dep = _clean_str(r[c_dep]).upper() if c_dep else ""
            ser = _norm_svc(r[c_ser]) if c_ser else ""
            if not dep and not ser:
                continue
            rows_out.append({
                "DEPOT": dep,
                "SER NO": ser,
                "ROUTE": _clean_str(r[c_route]) if c_route else "",
                "PRODUCT": _clean_str(r[c_prod]) if c_prod else "",
                "BEFORE": _clean_str(r[c_bef]).upper() if c_bef else "",
                "AFTER": _clean_str(r[c_aft]).upper() if c_aft else "",
                "NO. OF SERS": int(float(r[c_sers])) if c_sers and pd.notna(r.get(c_sers)) else 1,
            })
        return pd.DataFrame(rows_out), debug_cols, f"sheet={sheet} header={hdr} score={best_score} cols={cols_raw[:12]}"

    def _read_mod_excel_from_df(df_try):
        def _nk(c):
            return str(c).strip().lower().replace(" ", "").replace("_", "").replace(".", "").replace("/", "").replace("\n", "")
        colmap = {_nk(c): c for c in df_try.columns}
        def find(*cands):
            for cand in cands:
                k = _nk(cand)
                if k in colmap:
                    return colmap[k]
                for nk, orig in colmap.items():
                    if k and (k in nk or nk in k):
                        return orig
            return None
        c_dep = find("DEPOT")
        c_ser = find("SER NO", "SER_NO", "SERNO", "ServiceNo", "SERVICE")
        c_route = find("ROUTE", "ROUTEE")
        c_prod = find("PRODUCT", "PROD")
        c_bef = find("BEFORE", "BEF")
        c_aft = find("AFTER", "AFT")
        rows_out = []
        for _, r in df_try.iterrows():
            dep = _clean_str(r[c_dep]).upper() if c_dep else ""
            ser = _norm_svc(r[c_ser]) if c_ser else ""
            if not dep and not ser:
                continue
            rows_out.append({
                "DEPOT": dep, "SER NO": ser,
                "ROUTE": _clean_str(r[c_route]) if c_route else "",
                "PRODUCT": _clean_str(r[c_prod]) if c_prod else "",
                "BEFORE": _clean_str(r[c_bef]).upper() if c_bef else "",
                "AFTER": _clean_str(r[c_aft]).upper() if c_aft else "",
                "NO. OF SERS": 1,
            })
        return pd.DataFrame(rows_out)

    # ---- Upload ----
    st.caption("Upload the modification services Excel (DEPOT, SER NO, ROUTE, PRODUCT, BEFORE, AFTER).")
    _up = st.file_uploader("Modification services list", type=["xlsx", "xls", "csv"], key="mrr_upload")

    master = pd.DataFrame()
    if _up is not None:
        if _up.name.lower().endswith(".csv"):
            try:
                _ul = pd.read_csv(_up)
                master = _read_mod_excel_from_df(_ul)
                st.success(f"Loaded {len(master)} services from CSV.")
            except Exception as e:
                st.error(f"CSV read failed: {e}")
        else:
            master, debug_cols, info = _read_mod_excel(_up)
            if master is not None and len(master):
                st.success(f"Loaded {len(master)} services from upload. ({info})")
                # show sample so user can verify
                st.caption("Sample of parsed rows:")
                st.dataframe(master.head(5), width="stretch")
            else:
                st.error("Could not parse service columns from the Excel.")
                with st.expander("Detected column headers (debug)"):
                    for sheet, hdr, cols in (debug_cols or [])[:12]:
                        st.write(f"Sheet={sheet} header_row={hdr}: {cols}")
                st.stop()
    else:
        st.info("Upload the MODIFICATION services Excel to load the 253/257 service list.")
        st.stop()

    if len(master) == 0:
        st.warning("No services parsed from file.")
        st.stop()

    # Filters: Depot / Product / Route only
    c1, c2, c3 = st.columns(3)
    with c1:
        _deps = ["ALL"] + sorted({x for x in master["DEPOT"].dropna().unique() if str(x).strip()})
        mrr_depot = st.selectbox("Depot", _deps, index=0, key="mrr_depot")
    with c2:
        _prods = ["ALL"] + sorted({x for x in master["PRODUCT"].dropna().unique() if str(x).strip()})
        mrr_product = st.selectbox("Product", _prods, index=0, key="mrr_product")
    with c3:
        _routes = ["ALL"] + sorted({x for x in master["ROUTE"].dropna().unique() if str(x).strip()})
        mrr_route = st.selectbox("Route", _routes, index=0, key="mrr_route")

    if mrr_depot != "ALL":
        master = master[master["DEPOT"].astype(str).str.strip().str.upper() == mrr_depot.strip().upper()]
    if mrr_product != "ALL":
        master = master[master["PRODUCT"].astype(str).str.strip().str.upper() == mrr_product.strip().upper()]
    if mrr_route != "ALL":
        master = master[master["ROUTE"].astype(str).str.strip().str.upper() == mrr_route.strip().upper()]
    if len(master) == 0:
        st.info("No services match filters.")
        st.stop()

    # Metrics from ser_wise
    base = mrr.copy()
    base["SER_NO"] = base["SER_NO"].map(_norm_svc)
    base["DEPOT"] = base["DEPOT"].astype(str).str.strip().str.upper()
    p1 = base[(base["Date"] >= _p1_start) & (base["Date"] <= _p1_end)]
    p2 = base[(base["Date"] >= _p2_start) & (base["Date"] <= _p2_end)]

    def _agg(data):
        if len(data) == 0:
            return pd.DataFrame(columns=["DEPOT", "SER_NO", "kms", "ge", "ne"])
        return data.groupby(["DEPOT", "SER_NO"], dropna=False).agg(
            kms=("Optd_KMs", "sum"), ge=("GE_TOT", "sum"), ne=("NE_TOT", "sum"),
        ).reset_index()

    a1, a2 = _agg(p1), _agg(p2)
    orf_map, _, orf_err = load_orf_map(r"D:\\dashboard\\ORF.xlsx")

    def _orf(dep):
        d = str(dep).strip().upper()
        if orf_map and d in orf_map:
            try:
                return float(orf_map[d].get("cy", np.nan))
            except Exception:
                return np.nan
        if orf_map and "REGION" in orf_map:
            try:
                return float(orf_map["REGION"].get("cy", np.nan))
            except Exception:
                return np.nan
        return np.nan

    def _epk(e, k):
        try:
            e, k = float(e), float(k)
            return e / k if k > 0 else np.nan
        except Exception:
            return np.nan

    def _or(epk, orf):
        if pd.isna(epk) or pd.isna(orf) or not orf:
            return np.nan
        return epk * 10000 / orf

    rows = []
    for i, r in master.reset_index(drop=True).iterrows():
        dep, ser = str(r["DEPOT"]).strip().upper(), _norm_svc(r["SER NO"])
        r1 = a1[(a1["DEPOT"] == dep) & (a1["SER_NO"] == ser)]
        r2 = a2[(a2["DEPOT"] == dep) & (a2["SER_NO"] == ser)]
        kms1 = float(r1["kms"].sum()) if len(r1) else 0.0
        kms2 = float(r2["kms"].sum()) if len(r2) else 0.0
        ge1 = float(r1["ge"].sum()) if len(r1) else 0.0
        ge2 = float(r2["ge"].sum()) if len(r2) else 0.0
        ne1 = float(r1["ne"].sum()) if len(r1) else 0.0
        ne2 = float(r2["ne"].sum()) if len(r2) else 0.0
        orf_v = _orf(dep)
        ne1e, ge1e = _epk(ne1, kms1), _epk(ge1, kms1)
        ne2e, ge2e = _epk(ne2, kms2), _epk(ge2, kms2)
        no1, go1 = _or(ne1e, orf_v), _or(ge1e, orf_v)
        no2, go2 = _or(ne2e, orf_v), _or(ge2e, orf_v)

        def _v(a, b):
            if pd.isna(a) or pd.isna(b):
                return np.nan
            return float(b) - float(a)

        rows.append({
            "SL NO": i + 1,
            "DEPOT": dep,
            "SER NO": ser,
            "ROUTE": r["ROUTE"],
            "PRODUCT": r["PRODUCT"],
            "BEFORE": r["BEFORE"],
            "AFTER": r["AFTER"],
            "NO. OF SERS": r["NO. OF SERS"],
            "P1 NET EPK": round(ne1e, 2) if pd.notna(ne1e) else None,
            "P1 NET OR": round(no1, 0) if pd.notna(no1) else None,
            "P1 GROSS EPK": round(ge1e, 2) if pd.notna(ge1e) else None,
            "P1 GROSS OR": round(go1, 0) if pd.notna(go1) else None,
            "P2 NET EPK": round(ne2e, 2) if pd.notna(ne2e) else None,
            "P2 NET OR": round(no2, 0) if pd.notna(no2) else None,
            "P2 GROSS EPK": round(ge2e, 2) if pd.notna(ge2e) else None,
            "P2 GROSS OR": round(go2, 0) if pd.notna(go2) else None,
            "VAR NET EPK": round(_v(ne1e, ne2e), 2) if pd.notna(_v(ne1e, ne2e)) else None,
            "VAR NET OR": round(_v(no1, no2), 0) if pd.notna(_v(no1, no2)) else None,
            "VAR GROSS EPK": round(_v(ge1e, ge2e), 2) if pd.notna(_v(ge1e, ge2e)) else None,
            "VAR GROSS OR": round(_v(go1, go2), 0) if pd.notna(_v(go1, go2)) else None,
        })

    out = pd.DataFrame(rows)

    def _th_m(txt, rowspan=1, colspan=1, bg="#1e3a8a", color="#fff", top=0):
        return (
            f'<th rowspan="{rowspan}" colspan="{colspan}" style="background:{bg};color:{color};'
            f'position:sticky;top:{top}px;z-index:3;padding:6px 8px;border:1px solid #cbd5e1;'
            f'font-size:11px;text-align:center;white-space:nowrap;">{txt}</th>'
        )

    def _td_m(v, is_var=False):
        if v is None or (isinstance(v, float) and np.isnan(v)) or v == "":
            return '<td style="padding:4px 8px;border:1px solid #e2e8f0;text-align:center;"></td>'
        style = "padding:4px 8px;border:1px solid #e2e8f0;text-align:center;font-size:12px;"
        if is_var:
            try:
                fv = float(v)
                if fv > 0:
                    style += "color:#15803d;font-weight:700;background:#dcfce7;"
                elif fv < 0:
                    style += "color:#b91c1c;font-weight:700;background:#fee2e2;"
            except Exception:
                pass
        return f'<td style="{style}">{v}</td>'

    thead = (
        "<tr>"
        + _th_m("SL<br>NO", rowspan=3, bg="#0f172a")
        + _th_m("DEPOT", rowspan=3, bg="#0f172a")
        + _th_m("SER NO", rowspan=3, bg="#0f172a")
        + _th_m("ROUTE", rowspan=3, bg="#0f172a")
        + _th_m("PRODUCT", rowspan=3, bg="#0f172a")
        + _th_m("BEFORE", rowspan=3, bg="#7c3aed")
        + _th_m("AFTER", rowspan=3, bg="#7c3aed")
        + _th_m("NO.<br>OF<br>SERS", rowspan=3, bg="#0f172a")
        + _th_m(_p1_label, colspan=4, bg="#0e7490")
        + _th_m(_p2_label, colspan=4, bg="#0369a1")
        + _th_m("VARIANCE", colspan=4, bg="#9f1239")
        + "</tr><tr>"
        + _th_m("NET", colspan=2, bg="#14b8a6", top=28)
        + _th_m("GROSS", colspan=2, bg="#22c55e", top=28)
        + _th_m("NET", colspan=2, bg="#38bdf8", top=28)
        + _th_m("GROSS", colspan=2, bg="#4ade80", top=28)
        + _th_m("NET", colspan=2, bg="#fb7185", top=28)
        + _th_m("GROSS", colspan=2, bg="#f87171", top=28)
        + "</tr><tr>"
    )
    for _ in range(3):
        thead += _th_m("EPK", bg="#f1f5f9", color="#334155", top=56) + _th_m("OR", bg="#f1f5f9", color="#334155", top=56)
        thead += _th_m("EPK", bg="#f1f5f9", color="#334155", top=56) + _th_m("OR", bg="#f1f5f9", color="#334155", top=56)
    thead += "</tr>"

    body = []
    cols = [
        "SL NO", "DEPOT", "SER NO", "ROUTE", "PRODUCT", "BEFORE", "AFTER", "NO. OF SERS",
        "P1 NET EPK", "P1 NET OR", "P1 GROSS EPK", "P1 GROSS OR",
        "P2 NET EPK", "P2 NET OR", "P2 GROSS EPK", "P2 GROSS OR",
        "VAR NET EPK", "VAR NET OR", "VAR GROSS EPK", "VAR GROSS OR",
    ]
    var_cols = {"VAR NET EPK", "VAR NET OR", "VAR GROSS EPK", "VAR GROSS OR"}
    for _, r in out.iterrows():
        body.append("<tr>")
        for c in cols:
            body.append(_td_m(r.get(c), is_var=(c in var_cols)))
        body.append("</tr>")

    html = (
        f'<div style="font-weight:700;font-size:15px;text-align:center;padding:8px;'
        f'background:linear-gradient(90deg,#0f172a,#1e3a8a);color:#fff;border-radius:6px 6px 0 0;">{title}</div>'
        f'<div style="overflow:auto;max-height:70vh;border:1px solid #cbd5e1;border-radius:0 0 6px 6px;">'
        f'<table style="border-collapse:collapse;width:100%;min-width:1400px;">'
        f'<thead>{thead}</thead><tbody>{"".join(body)}</tbody></table></div>'
    )
    st.markdown(html, unsafe_allow_html=True)
    st.caption(f"Services: {len(out)} · P1={_p1_label} · P2={_p2_label} · Variance = P2−P1 (green ↑ / red ↓)")

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        wb = Workbook()
        ws = wb.active
        ws.title = "MISSION RR"
        thin = Border(
            left=Side(style="thin", color="CBD5E1"), right=Side(style="thin", color="CBD5E1"),
            top=Side(style="thin", color="CBD5E1"), bottom=Side(style="thin", color="CBD5E1"),
        )
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=20)
        ws.cell(1, 1, title).font = Font(bold=True, color="FFFFFF", size=12)
        ws.cell(1, 1).fill = PatternFill("solid", fgColor="0F172A")
        ws.cell(1, 1).alignment = Alignment(horizontal="center")
        headers = cols
        for i, h in enumerate(headers, 1):
            cell = ws.cell(2, i, h)
            cell.font = Font(bold=True, color="FFFFFF", size=9)
            cell.fill = PatternFill("solid", fgColor="1E3A8A")
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            cell.border = thin
        gf, rf = PatternFill("solid", fgColor="DCFCE7"), PatternFill("solid", fgColor="FEE2E2")
        gfont, rfont = Font(color="15803D", bold=True), Font(color="B91C1C", bold=True)
        for ri, row in enumerate(out.itertuples(index=False), 3):
            for ci, val in enumerate(row, 1):
                cell = ws.cell(ri, ci, val if val is not None and not (isinstance(val, float) and np.isnan(val)) else "")
                cell.border = thin
                cell.alignment = Alignment(horizontal="center")
                if ci >= 17:
                    try:
                        fv = float(val)
                        if fv > 0:
                            cell.fill, cell.font = gf, gfont
                        elif fv < 0:
                            cell.fill, cell.font = rf, rfont
                    except Exception:
                        pass
        bio = BytesIO()
        wb.save(bio)
        bio.seek(0)
        st.download_button(
            "Download MISSION RR Excel",
            bio.getvalue(),
            "MISSION_RR_Aug2026.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="mrr_dl",
        )
    except Exception as _ex:
        st.caption(f"Excel export unavailable: {_ex}")


st.caption("Cascading filters • Weighted EPK • Self-hosted on your PC")
